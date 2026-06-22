from __future__ import annotations

import csv
import re
import shutil
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "processed_data" / "hard_data_exports" / "hunt_tables" / "2026" / "CLEAN_XLXS_STAGED" / "2025-27 Conservation Permits.xlsx"
DATABASE_PATH = ROOT / "pipeline" / "RAW" / "hunt_unit_database" / "2026" / "csv" / "DATABASE.csv"
AUDIT_DIR = ROOT / "audits" / "2025_canonical_finalization"
BACKUP_DIR = AUDIT_DIR / "backups"
AUDIT_PATH = AUDIT_DIR / "2025_27_conservation_blank_season_database_fill.csv"


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def norm(value: object) -> str:
    return re.sub(r"\s+", " ", clean(value).lower()).strip()


def normalize_species(value: object) -> str:
    text = norm(value)
    if text == "antlerless elk":
        return "elk"
    if text == "bear":
        return "black bear"
    return text


def normalize_sex(value: object) -> str:
    text = norm(value).replace("'", "")
    aliases = {
        "hunter s choice": "either sex",
        "hunters choice": "either sex",
        "either sex": "either sex",
        "ram": "male only",
        "male only": "male only",
        "female only": "female only",
    }
    return aliases.get(text, text)


def normalize_boundary(value: object) -> str:
    text = clean(value)
    if text.endswith(".0"):
        text = text[:-2]
    return text


def normalize_weapon(value: object) -> str:
    text = norm(value).replace("muzz", "muzzleloader")
    text = text.replace("alw", "any legal weapon")
    text = re.sub(r"\s+", " ", text)
    return text


def read_database() -> dict[str, list[dict[str, str]]]:
    by_code: dict[str, list[dict[str, str]]] = defaultdict(list)
    with DATABASE_PATH.open("r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            code = clean(row.get("hunt_code")).upper()
            if code:
                by_code[code].append({key: clean(value) for key, value in row.items()})
    return by_code


def matching_candidate(workbook_row: dict[str, str], candidates: list[dict[str, str]]) -> tuple[dict[str, str] | None, str]:
    if not candidates:
        return None, "no_database_code_match"

    species = normalize_species(workbook_row["Species"])
    boundary = normalize_boundary(workbook_row["BOUNDARY ID"])
    sex = normalize_sex(workbook_row["SEX TYPE"])
    weapon = normalize_weapon(workbook_row["WEAPON"])

    matches: list[dict[str, str]] = []
    for candidate in candidates:
        if normalize_species(candidate.get("species")) != species:
            continue
        if normalize_boundary(candidate.get("boundary_id")) != boundary:
            continue
        if normalize_sex(candidate.get("sex_type")) != sex:
            continue
        if normalize_weapon(candidate.get("weapon")) != weapon:
            continue
        if not clean(candidate.get("season")):
            continue
        matches.append(candidate)

    if not matches:
        return None, "no_strict_species_boundary_sex_weapon_season_match"

    season_values = {clean(match.get("season")) for match in matches}
    if len(season_values) == 1:
        return matches[0], "strict_match"

    return None, "review_multiple_distinct_database_seasons"


def main() -> None:
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)

    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    backup_path = BACKUP_DIR / f"{WORKBOOK_PATH.stem}.before_blank_season_database_fill_{stamp}{WORKBOOK_PATH.suffix}"
    shutil.copy2(WORKBOOK_PATH, backup_path)

    database_by_code = read_database()
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb["Table 1"] if "Table 1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [clean(ws.cell(1, column).value) for column in range(1, ws.max_column + 1)]
    cols = {header: index + 1 for index, header in enumerate(headers) if header}

    required = ["HUNT CODE", "BOUNDARY ID", "Species", "HUNT NAME", "SEX TYPE", "WEAPON", "SEASON"]
    missing = [header for header in required if header not in cols]
    if missing:
        raise RuntimeError(f"Missing required headers: {missing}")

    audit_rows: list[dict[str, str]] = []
    for row_idx in range(2, ws.max_row + 1):
        if not any(clean(ws.cell(row_idx, column).value) for column in range(1, ws.max_column + 1)):
            continue
        old_season = clean(ws.cell(row_idx, cols["SEASON"]).value)
        if old_season:
            continue

        workbook_row = {header: clean(ws.cell(row_idx, col_idx).value) for header, col_idx in cols.items()}
        candidate, status = matching_candidate(workbook_row, database_by_code.get(workbook_row["HUNT CODE"].upper(), []))
        new_season = ""
        if candidate is not None:
            new_season = clean(candidate.get("season"))
            ws.cell(row_idx, cols["SEASON"]).value = new_season
            status = f"{status}_filled"

        audit_rows.append(
            {
                "row": str(row_idx),
                "status": status,
                "hunt_code": workbook_row["HUNT CODE"],
                "boundary_id": workbook_row["BOUNDARY ID"],
                "species": workbook_row["Species"],
                "hunt_name": workbook_row["HUNT NAME"],
                "sex_type": workbook_row["SEX TYPE"],
                "weapon": workbook_row["WEAPON"],
                "old_season": old_season,
                "new_season": new_season,
                "db_hunt_name": clean(candidate.get("hunt_name")) if candidate else "",
                "db_species": clean(candidate.get("species")) if candidate else "",
                "db_boundary_id": clean(candidate.get("boundary_id")) if candidate else "",
                "db_sex_type": clean(candidate.get("sex_type")) if candidate else "",
                "db_weapon": clean(candidate.get("weapon")) if candidate else "",
            }
        )

    wb.save(WORKBOOK_PATH)

    fieldnames = [
        "row",
        "status",
        "hunt_code",
        "boundary_id",
        "species",
        "hunt_name",
        "sex_type",
        "weapon",
        "old_season",
        "new_season",
        "db_hunt_name",
        "db_species",
        "db_boundary_id",
        "db_sex_type",
        "db_weapon",
    ]
    with AUDIT_PATH.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(audit_rows)

    filled = sum(1 for row in audit_rows if row["status"].endswith("_filled"))
    print(f"blank_season_rows_seen={len(audit_rows)}")
    print(f"filled={filled}")
    print(f"backup={backup_path}")
    print(f"audit={AUDIT_PATH}")


if __name__ == "__main__":
    main()
