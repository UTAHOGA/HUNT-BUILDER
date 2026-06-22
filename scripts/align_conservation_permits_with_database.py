"""Align 2025-27 conservation permit workbook rows to DATABASE.csv truth."""

from __future__ import annotations

import csv
import json
import re
import shutil
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "processed_data" / "hard_data_exports" / "hunt_tables" / "2026" / "CLEAN_XLXS_STAGED" / "2025-27 Conservation Permits.xlsx"
DATABASE_PATH = ROOT / "pipeline" / "RAW" / "hunt_unit_database" / "2026" / "csv" / "DATABASE.csv"
BOUNDARY_DIR = ROOT / "processed_data" / "boundaries"
AUDIT_DIR = ROOT / "audits" / "2025_canonical_finalization"
BACKUP_DIR = AUDIT_DIR / "backups"
AUDIT_CSV = AUDIT_DIR / "2025_27_conservation_database_geo_alignment_audit.csv"
SUMMARY_JSON = AUDIT_DIR / "2025_27_conservation_database_geo_alignment_summary.json"

EXPECTED_PREFIXES = {
    "Black Bear": {"BR"},
    "Bear": {"BR"},
    "Bison": {"BI"},
    "Deer": {"DB"},
    "Desert Bighorn Sheep": {"DS"},
    "Elk": {"EA", "EB"},
    "Moose": {"MB"},
    "Mountain Goat": {"GO"},
    "Pronghorn": {"PB"},
    "Rocky Mountain Bighorn Sheep": {"RS"},
    "Turkey": {"TK"},
}

STOPWORDS = {
    "a",
    "and",
    "any",
    "choice",
    "conservation",
    "early",
    "hunter",
    "hunters",
    "late",
    "le",
    "legal",
    "limited",
    "only",
    "permit",
    "premium",
    "season",
    "statewide",
    "weapon",
}


@dataclass(frozen=True)
class Candidate:
    row: dict[str, str]
    code: str
    boundary_id: str
    hunt_name: str
    species: str
    sex_type: str
    weapon: str
    hunt_type: str
    hunt_class: str


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def norm_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", clean(value)).strip()


def normalize_species(value: str) -> str:
    text = norm_spaces(value)
    low = text.lower()
    if low == "antlerless elk":
        return "Elk"
    if low == "bear":
        return "Black Bear"
    if low in {"rmb sheep", "rocky mtn bighorn sheep", "rocky mountain sheep"}:
        return "Rocky Mountain Bighorn Sheep"
    if low in {"desert sheep", "desert bighorn"}:
        return "Desert Bighorn Sheep"
    return text


def compact_text(value: str) -> str:
    text = clean(value).lower()
    text = text.replace("&", " and ")
    text = text.replace("'", "")
    text = re.sub(r"\([^)]*\)", " ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return norm_spaces(text)


def area_text(value: str) -> str:
    text = compact_text(value)
    text = re.sub(r"\b(hunters?|choice|cow|only|bull|ram|early|late|le)\b", " ", text)
    return norm_spaces(text)


def tokens(value: str) -> set[str]:
    return {part for part in area_text(value).split() if part and part not in STOPWORDS}


def token_score(left: str, right: str) -> float:
    lt = tokens(left)
    rt = tokens(right)
    if not lt or not rt:
        return 0.0
    inter = len(lt & rt)
    union = len(lt | rt)
    contain = inter / min(len(lt), len(rt))
    jaccard = inter / union
    return max(jaccard, contain * 0.85)


def sequence_score(left: str, right: str) -> float:
    ltext = area_text(left)
    rtext = area_text(right)
    if not ltext or not rtext:
        return 0.0
    return SequenceMatcher(None, ltext, rtext).ratio()


def area_score(left: str, right: str) -> float:
    return max(token_score(left, right), sequence_score(left, right))


def condition_score(condition: str, candidate: Candidate) -> tuple[float, list[str]]:
    cond = compact_text(condition)
    weapon = compact_text(candidate.weapon)
    sex = compact_text(candidate.sex_type)
    score = 0.0
    notes: list[str] = []

    if "statewide" in cond and "statewide" in compact_text(candidate.hunt_name):
        score += 0.25
        notes.append("statewide")
    if "hunter" in cond and "choice" in cond and ("hunter" in sex or "choice" in sex):
        score += 0.24
        notes.append("hunters_choice")
    if "multiseason" in cond and "multiseason" in weapon:
        score += 0.34
        notes.append("multiseason")
    if "archery" in cond and "archery" in weapon:
        score += 0.34
        notes.append("archery")
    if ("muzzleloader" in cond or "muzz" in cond) and ("muzzleloader" in weapon or "muzz" in weapon):
        score += 0.34
        notes.append("muzzleloader")
    if "any legal weapon" in condition.lower() and "any legal weapon" in candidate.weapon.lower():
        score += 0.34
        notes.append("any_legal_weapon")
    if "late" in cond and "late" in weapon:
        score += 0.18
        notes.append("late")
    if "early" in cond and "early" in weapon:
        score += 0.18
        notes.append("early")

    return min(score, 0.44), notes


def normalized_hunt_name(name: str) -> str:
    text = norm_spaces(name)
    text = re.sub(r"\s*-\s*Conservation/Expo\s*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s*-\s*Statewide Permit\s*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+Conservation\s*$", "", text, flags=re.IGNORECASE)
    if text.lower() in {"black bear", "bison", "mountain goat", "pronghorn"}:
        return "Statewide"
    return norm_spaces(text)


def read_database() -> list[Candidate]:
    with DATABASE_PATH.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = [{k: clean(v) for k, v in row.items()} for row in csv.DictReader(handle)]

    candidates: list[Candidate] = []
    for row in rows:
        code = clean(row.get("hunt_code")).upper()
        if not code:
            continue
        candidates.append(
            Candidate(
                row=row,
                code=code,
                boundary_id=clean(row.get("boundary_id")),
                hunt_name=clean(row.get("hunt_name")),
                species=normalize_species(row.get("species", "")),
                sex_type=clean(row.get("sex_type")),
                weapon=clean(row.get("weapon")),
                hunt_type=clean(row.get("hunt_type")),
                hunt_class=clean(row.get("hunt_class")),
            )
        )
    return candidates


def boundary_path(candidate: Candidate) -> str:
    code_path = BOUNDARY_DIR / f"{candidate.code}.geojson"
    if code_path.exists():
        return code_path.relative_to(ROOT).as_posix()
    if candidate.boundary_id:
        id_path = BOUNDARY_DIR / f"{candidate.boundary_id}.geojson"
        if id_path.exists():
            return id_path.relative_to(ROOT).as_posix()
    return ""


def conservation_weight(candidate: Candidate) -> float:
    conservation_total = clean(candidate.row.get("conservation_permits_2026_total"))
    text = " ".join([candidate.hunt_type, candidate.hunt_class, candidate.hunt_name]).lower()
    if conservation_total:
        return 0.16
    if "conservation" in text:
        return 0.12
    return 0.0


def expected_prefixes(species: str) -> set[str]:
    return EXPECTED_PREFIXES.get(normalize_species(species), set())


def score_candidate(row: dict[str, str], candidate: Candidate) -> tuple[float, list[str]]:
    row_species = normalize_species(row["Species"])
    if candidate.species != row_species:
        return -1.0, ["species_mismatch"]

    score = 0.50
    notes = ["species"]
    area = row["Area"]
    candidate_area = candidate.hunt_name

    a_score = area_score(area, candidate_area)
    score += a_score * 0.42
    notes.append(f"area={a_score:.3f}")

    c_score, c_notes = condition_score(row["Condition"], candidate)
    score += c_score
    notes.extend(c_notes)

    c_weight = conservation_weight(candidate)
    if c_weight:
        score += c_weight
        notes.append("conservation_weight")

    current_code = row["HUNT CODE"].upper()
    if current_code and current_code == candidate.code:
        score += 0.30
        notes.append("current_code_exact")

    if candidate.boundary_id and boundary_path(candidate):
        score += 0.04
        notes.append("geo")

    return score, notes


def best_match(row: dict[str, str], candidates: list[Candidate]) -> tuple[Candidate | None, float, str, list[str]]:
    scored: list[tuple[float, Candidate, list[str]]] = []
    for candidate in candidates:
        score, notes = score_candidate(row, candidate)
        if score >= 0:
            scored.append((score, candidate, notes))
    if not scored:
        return None, 0.0, "NO_SPECIES_CANDIDATE", []
    scored.sort(key=lambda item: item[0], reverse=True)
    best_score, best_candidate, notes = scored[0]
    second_score = scored[1][0] if len(scored) > 1 else 0.0
    margin = best_score - second_score
    current_code = row["HUNT CODE"].upper()

    if current_code:
        for score, candidate, candidate_notes in scored:
            if candidate.code != current_code:
                continue
            current_area_score = area_score(row["Area"], candidate.hunt_name)
            row_area = compact_text(row["Area"])
            candidate_name = compact_text(candidate.hunt_name)
            area_aligned = current_area_score >= 0.70 or ("statewide" in row_area and "statewide" in candidate_name)
            if area_aligned and score >= 0.86:
                return candidate, score, "PATCHED_CURRENT_CODE", candidate_notes + [f"second={second_score:.3f}"]

    if best_score < 0.92:
        return best_candidate, best_score, "REVIEW_LOW_SCORE", notes
    if margin < 0.08:
        return best_candidate, best_score, "REVIEW_TIE", notes + [f"second={second_score:.3f}"]
    return best_candidate, best_score, "PATCHED", notes + [f"second={second_score:.3f}"]


def workbook_rows(ws: openpyxl.worksheet.worksheet.Worksheet) -> tuple[list[str], dict[str, int]]:
    headers = [clean(cell.value) for cell in ws[1]]
    return headers, {header: index + 1 for index, header in enumerate(headers) if header}


def main() -> None:
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    backup_path = BACKUP_DIR / f"2025-27 Conservation Permits.before_database_geo_alignment_{timestamp}.xlsx"
    shutil.copy2(WORKBOOK_PATH, backup_path)

    candidates = read_database()
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    ws = wb["Table 1"]
    _headers, columns = workbook_rows(ws)

    no_values = []
    for excel_row in range(2, ws.max_row + 1):
        no_text = clean(ws.cell(excel_row, columns["No."]).value)
        if no_text:
            no_values.append(int(float(no_text)))
    if len(no_values) != 336 or sorted(no_values) != list(range(1, 337)):
        raise RuntimeError("Conservation workbook must be reconciled to PDF rows 1-336 before database alignment.")

    audit_rows: list[dict[str, str]] = []
    status_counts: Counter[str] = Counter()
    changed_rows = 0

    for excel_row in range(2, ws.max_row + 1):
        current = {
            "No.": clean(ws.cell(excel_row, columns["No."]).value),
            "Species": normalize_species(ws.cell(excel_row, columns["Species"]).value),
            "Area": clean(ws.cell(excel_row, columns["Area"]).value),
            "Condition": clean(ws.cell(excel_row, columns["Condition"]).value),
            "Organization": clean(ws.cell(excel_row, columns["Organization"]).value),
            "HUNT CODE": clean(ws.cell(excel_row, columns["HUNT CODE"]).value).upper(),
            "HUNT NAME": clean(ws.cell(excel_row, columns["HUNT NAME"]).value),
            "SEX": clean(ws.cell(excel_row, columns["SEX"]).value),
            "WEAPON": clean(ws.cell(excel_row, columns["WEAPON"]).value),
            "BOUNDARY ID": clean(ws.cell(excel_row, columns["BOUNDARY ID"]).value),
            "MAP GEOJSON": clean(ws.cell(excel_row, columns["MAP GEOJSON"]).value),
        }
        candidate, score, status, notes = best_match(current, candidates)
        status_counts[status] += 1

        after = dict(current)
        if status in {"PATCHED", "PATCHED_CURRENT_CODE"} and candidate is not None:
            after.update(
                {
                    "HUNT CODE": candidate.code,
                    "HUNT NAME": normalized_hunt_name(candidate.hunt_name),
                    "SEX": candidate.sex_type,
                    "WEAPON": candidate.weapon,
                    "BOUNDARY ID": candidate.boundary_id,
                    "MAP GEOJSON": boundary_path(candidate),
                }
            )
            if any(after[key] != current[key] for key in ["HUNT CODE", "HUNT NAME", "SEX", "WEAPON", "BOUNDARY ID", "MAP GEOJSON"]):
                changed_rows += 1
            ws.cell(excel_row, columns["HUNT CODE"]).value = after["HUNT CODE"]
            ws.cell(excel_row, columns["HUNT NAME"]).value = after["HUNT NAME"]
            ws.cell(excel_row, columns["SEX"]).value = after["SEX"]
            ws.cell(excel_row, columns["WEAPON"]).value = after["WEAPON"]
            ws.cell(excel_row, columns["BOUNDARY ID"]).value = after["BOUNDARY ID"]
            ws.cell(excel_row, columns["MAP GEOJSON"]).value = after["MAP GEOJSON"]

        audit_rows.append(
            {
                "excel_row": str(excel_row),
                "no": current["No."],
                "status": status,
                "score": f"{score:.3f}",
                "notes": ";".join(notes),
                "before_species": current["Species"],
                "after_species": after["Species"],
                "area": current["Area"],
                "condition": current["Condition"],
                "organization": current["Organization"],
                "before_hunt_code": current["HUNT CODE"],
                "after_hunt_code": after["HUNT CODE"],
                "before_hunt_name": current["HUNT NAME"],
                "after_hunt_name": after["HUNT NAME"],
                "before_sex": current["SEX"],
                "after_sex": after["SEX"],
                "before_weapon": current["WEAPON"],
                "after_weapon": after["WEAPON"],
                "before_boundary_id": current["BOUNDARY ID"],
                "after_boundary_id": after["BOUNDARY ID"],
                "before_map_geojson": current["MAP GEOJSON"],
                "after_map_geojson": after["MAP GEOJSON"],
            }
        )

    wb.save(WORKBOOK_PATH)

    audit_fields = list(audit_rows[0].keys()) if audit_rows else []
    with AUDIT_CSV.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=audit_fields)
        writer.writeheader()
        writer.writerows(audit_rows)

    summary = {
        "generated_utc": datetime.now(timezone.utc).isoformat(),
        "workbook_path": str(WORKBOOK_PATH),
        "database_path": str(DATABASE_PATH),
        "backup_path": str(backup_path),
        "rows": ws.max_row - 1,
        "changed_rows": changed_rows,
        "status_counts": dict(sorted(status_counts.items())),
        "audit_csv": str(AUDIT_CSV),
    }
    with SUMMARY_JSON.open("w", encoding="utf-8") as handle:
        json.dump(summary, handle, indent=2)
        handle.write("\n")
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
