from __future__ import annotations

import csv
import json
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]

DATABASE_PATH = ROOT / "pipeline" / "RAW" / "hunt_unit_database" / "2026" / "csv" / "DATABASE.csv"
MASTER_PATH = ROOT / "processed_data" / "hard_data_exports" / "hunt_tables" / "2026" / "CLEAN_XLXS_STAGED" / "MASTER.xlsx"

AUDIT_DIR = ROOT / "processed_data" / "audits"
REPORT_JSON = AUDIT_DIR / "master_database_2026_reconciliation_report.json"
REPORT_CSV = AUDIT_DIR / "master_database_2026_reconciliation_changes.csv"

SYNC_FIELDS = [
    ("HUNT NAME", "hunt_name"),
    ("SPECIES", "species"),
    ("SEX TYPE", "sex_type"),
    ("SEASON", "season"),
]

PERMIT_FIELDS = [
    ("2026 PERMITS RES", "res"),
    ("2026 PERMITS NR", "nr"),
    ("2026 PERMITS TOTAL", "total"),
]

NORMALIZATION_AUDIT_FIELDS = [
    ("HUNT TYPE", "hunt_type"),
    ("WEAPON", "weapon"),
    ("HUNT CLASS", "hunt_class"),
]


def clean(value: object) -> str:
    return str(value or "").strip()


def normalize_hunt_code(value: object) -> str:
    return "".join(ch for ch in clean(value).upper() if ch.isalnum())


def read_database_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def permit_truth(row: dict[str, str], suffix: str) -> str:
    return clean(row.get(f"permit_allotment_2026_{suffix}") or row.get(f"permits_2026_{suffix}"))


def build_database_index(rows: list[dict[str, str]]) -> dict[str, dict[str, str]]:
    out: dict[str, dict[str, str]] = {}
    for row in rows:
        code = normalize_hunt_code(row.get("hunt_code"))
        if code:
            out[code] = row
    return out


def load_master_sheet(path: Path):
    workbook = load_workbook(path)
    sheet = workbook.active
    headers = [clean(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    index = {header: i + 1 for i, header in enumerate(headers) if header}
    return workbook, sheet, headers, index


def audit_normalization(
    sheet,
    column_index: dict[str, int],
    db_index: dict[str, dict[str, str]],
) -> dict[str, dict[str, object]]:
    audit: dict[str, dict[str, object]] = {}
    for master_field, db_field in NORMALIZATION_AUDIT_FIELDS:
        same = 0
        different = 0
        blank = 0
        map_counter: Counter[tuple[str, str]] = Counter()
        for row_number in range(2, sheet.max_row + 1):
            hunt_code = normalize_hunt_code(sheet.cell(row=row_number, column=column_index["HUNT CODE"]).value)
            db_row = db_index.get(hunt_code)
            if not db_row:
                continue
            master_value = clean(sheet.cell(row=row_number, column=column_index[master_field]).value)
            database_value = clean(db_row.get(db_field))
            if not master_value and not database_value:
                blank += 1
                continue
            if master_value == database_value:
                same += 1
            else:
                different += 1
                map_counter[(master_value, database_value)] += 1
        top_mappings = [
            {
                "master_value": mv,
                "database_value": dv,
                "count": count,
            }
            for (mv, dv), count in map_counter.most_common(15)
        ]
        audit[master_field] = {
            "same_count": same,
            "different_count": different,
            "blank_both_count": blank,
            "top_mappings": top_mappings,
        }
    return audit


def reconcile_master_with_database() -> dict[str, object]:
    database_rows = read_database_rows(DATABASE_PATH)
    db_index = build_database_index(database_rows)

    workbook, sheet, headers, column_index = load_master_sheet(MASTER_PATH)
    required = {"HUNT CODE", *[x[0] for x in SYNC_FIELDS], *[x[0] for x in PERMIT_FIELDS], *[x[0] for x in NORMALIZATION_AUDIT_FIELDS]}
    missing_headers = sorted(required - set(headers))
    if missing_headers:
        raise RuntimeError(f"MASTER.xlsx is missing required headers: {missing_headers}")

    normalization_audit = audit_normalization(sheet, column_index, db_index)

    master_codes_seen: set[str] = set()
    changes: list[dict[str, object]] = []
    changes_by_field: Counter[str] = Counter()
    unmatched_master_codes: list[str] = []
    corrected_total_codes: list[str] = []
    code_change_counter: defaultdict[str, int] = defaultdict(int)

    for row_number in range(2, sheet.max_row + 1):
        hunt_code = normalize_hunt_code(sheet.cell(row=row_number, column=column_index["HUNT CODE"]).value)
        if not hunt_code:
            continue
        master_codes_seen.add(hunt_code)
        db_row = db_index.get(hunt_code)
        if not db_row:
            unmatched_master_codes.append(hunt_code)
            continue

        for master_field, db_field in SYNC_FIELDS:
            target = clean(db_row.get(db_field))
            cell = sheet.cell(row=row_number, column=column_index[master_field])
            current = clean(cell.value)
            if current != target:
                cell.value = target
                record = {
                    "row_number": row_number,
                    "hunt_code": hunt_code,
                    "field": master_field,
                    "old_value": current,
                    "new_value": target,
                }
                changes.append(record)
                changes_by_field[master_field] += 1
                code_change_counter[hunt_code] += 1

        for master_field, suffix in PERMIT_FIELDS:
            target = permit_truth(db_row, suffix)
            cell = sheet.cell(row=row_number, column=column_index[master_field])
            current = clean(cell.value)
            if current != target:
                cell.value = target
                record = {
                    "row_number": row_number,
                    "hunt_code": hunt_code,
                    "field": master_field,
                    "old_value": current,
                    "new_value": target,
                }
                changes.append(record)
                changes_by_field[master_field] += 1
                code_change_counter[hunt_code] += 1
                if master_field == "2026 PERMITS TOTAL" and not current and target:
                    corrected_total_codes.append(hunt_code)

    workbook.save(MASTER_PATH)

    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    with REPORT_CSV.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["row_number", "hunt_code", "field", "old_value", "new_value"])
        writer.writeheader()
        writer.writerows(changes)

    database_codes_missing_in_master = sorted(set(db_index.keys()) - master_codes_seen)

    report = {
        "database_path": str(DATABASE_PATH.relative_to(ROOT)).replace("\\", "/"),
        "master_path": str(MASTER_PATH.relative_to(ROOT)).replace("\\", "/"),
        "truth_rule": "DATABASE.csv authoritative; permit truth uses permit_allotment_2026_* with fallback to permits_2026_*.",
        "master_rows_processed": sheet.max_row - 1,
        "database_rows_available": len(database_rows),
        "unmatched_master_hunt_codes_count": len(unmatched_master_codes),
        "unmatched_master_hunt_codes": sorted(set(unmatched_master_codes)),
        "database_hunt_codes_missing_in_master_count": len(database_codes_missing_in_master),
        "database_hunt_codes_missing_in_master": database_codes_missing_in_master,
        "changes_total": len(changes),
        "changes_by_field": dict(changes_by_field),
        "changed_hunt_codes_count": len(code_change_counter),
        "changed_hunt_codes": sorted(code_change_counter.keys()),
        "corrected_missing_permits_total_hunt_codes_count": len(sorted(set(corrected_total_codes))),
        "corrected_missing_permits_total_hunt_codes": sorted(set(corrected_total_codes)),
        "normalization_audit": normalization_audit,
        "report_csv": str(REPORT_CSV.relative_to(ROOT)).replace("\\", "/"),
    }

    with REPORT_JSON.open("w", encoding="utf-8") as handle:
        json.dump(report, handle, indent=2)

    return report


def main() -> None:
    report = reconcile_master_with_database()
    print("MASTER reconciliation complete.")
    print(f"Changes: {report['changes_total']}")
    print(f"Changed hunt codes: {report['changed_hunt_codes_count']}")
    print(
        "Corrected missing 2026 PERMITS TOTAL rows: "
        f"{report['corrected_missing_permits_total_hunt_codes_count']}"
    )
    print(REPORT_JSON)
    print(REPORT_CSV)


if __name__ == "__main__":
    main()
