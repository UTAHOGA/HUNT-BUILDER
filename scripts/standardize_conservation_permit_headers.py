from __future__ import annotations

import csv
import shutil
from copy import copy
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "processed_data" / "hard_data_exports" / "hunt_tables" / "2026" / "CLEAN_XLXS_STAGED" / "2025-27 Conservation Permits.xlsx"
AUDIT_DIR = ROOT / "audits" / "2025_canonical_finalization"
BACKUP_DIR = AUDIT_DIR / "backups"
AUDIT_PATH = AUDIT_DIR / "2025_27_conservation_header_standardization_audit.csv"


FINAL_HEADERS = [
    "No.",
    "Species",
    "HUNT NAME",
    "SEX TYPE",
    "WEAPON",
    "Value",
    "Organization",
    "HUNT CODE",
    "BOUNDARY ID",
    "MAP GEOJSON",
]


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def header_map(headers: list[str]) -> dict[str, int]:
    mapped: dict[str, int] = {}
    for index, header in enumerate(headers, start=1):
        if header and header not in mapped:
            mapped[header] = index
    return mapped


def first_nonblank(*values: object) -> object:
    for value in values:
        if clean(value):
            return value
    return None


def main() -> None:
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)

    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    backup_path = BACKUP_DIR / f"{WORKBOOK_PATH.stem}.before_header_standardization_{stamp}{WORKBOOK_PATH.suffix}"
    shutil.copy2(WORKBOOK_PATH, backup_path)

    wb = load_workbook(WORKBOOK_PATH)
    ws = wb["Table 1"] if "Table 1" in wb.sheetnames else wb[wb.sheetnames[0]]
    original_headers = [clean(ws.cell(1, column).value) for column in range(1, ws.max_column + 1)]
    cols = header_map(original_headers)

    required = ["No.", "Species", "Area", "Condition", "Value", "Organization", "HUNT CODE", "BOUNDARY ID"]
    missing = [header for header in required if header not in cols]
    if missing:
        raise RuntimeError(f"Missing required source headers: {missing}. Existing headers: {original_headers}")

    row_numbers = []
    source_rows = []
    for row_idx in range(2, ws.max_row + 1):
        if not any(clean(ws.cell(row_idx, column).value) for column in range(1, ws.max_column + 1)):
            continue
        no_value = ws.cell(row_idx, cols["No."]).value
        row_numbers.append(int(float(clean(no_value))))
        source_rows.append(row_idx)

    if len(row_numbers) != 336 or sorted(row_numbers) != list(range(1, 337)):
        raise RuntimeError("Workbook must contain exactly PDF permit rows 1-336 before header standardization.")

    header_style = {
        "font": copy(ws.cell(1, 1).font),
        "fill": copy(ws.cell(1, 1).fill),
        "alignment": copy(ws.cell(1, 1).alignment),
        "border": copy(ws.cell(1, 1).border),
    }
    if header_style["font"] == Font():
        header_style["font"] = Font(bold=True)
    if header_style["fill"] == PatternFill():
        header_style["fill"] = PatternFill("solid", fgColor="D9EAF7")

    rows = []
    for row_idx in source_rows:
        rows.append(
            {
                "No.": ws.cell(row_idx, cols["No."]).value,
                "Species": ws.cell(row_idx, cols["Species"]).value,
                "HUNT NAME": ws.cell(row_idx, cols["Area"]).value,
                "SEX TYPE": first_nonblank(ws.cell(row_idx, cols.get("SEX", 0)).value if "SEX" in cols else None),
                "WEAPON": ws.cell(row_idx, cols["Condition"]).value,
                "Value": ws.cell(row_idx, cols["Value"]).value,
                "Organization": ws.cell(row_idx, cols["Organization"]).value,
                "HUNT CODE": ws.cell(row_idx, cols["HUNT CODE"]).value,
                "BOUNDARY ID": ws.cell(row_idx, cols["BOUNDARY ID"]).value,
                "MAP GEOJSON": ws.cell(row_idx, cols.get("MAP GEOJSON", 0)).value if "MAP GEOJSON" in cols else None,
            }
        )

    max_existing_cols = ws.max_column
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=max_existing_cols):
        for cell in row:
            cell.value = None

    for col_idx, header in enumerate(FINAL_HEADERS, start=1):
        cell = ws.cell(1, col_idx)
        cell.value = header
        cell.font = copy(header_style["font"])
        cell.fill = copy(header_style["fill"])
        cell.alignment = copy(header_style["alignment"])
        cell.border = copy(header_style["border"])

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(FINAL_HEADERS, start=1):
            ws.cell(row_idx, col_idx).value = row[header]

    if max_existing_cols > len(FINAL_HEADERS):
        ws.delete_cols(len(FINAL_HEADERS) + 1, max_existing_cols - len(FINAL_HEADERS))

    widths = {
        "A": 8,
        "B": 20,
        "C": 42,
        "D": 16,
        "E": 28,
        "F": 12,
        "G": 18,
        "H": 14,
        "I": 14,
        "J": 42,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{len(rows) + 1}"

    wb.save(WORKBOOK_PATH)

    with AUDIT_PATH.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["item", "value"])
        writer.writeheader()
        writer.writerow({"item": "backup_path", "value": str(backup_path)})
        writer.writerow({"item": "workbook_path", "value": str(WORKBOOK_PATH)})
        writer.writerow({"item": "rows", "value": len(rows)})
        writer.writerow({"item": "original_headers", "value": " | ".join(original_headers)})
        writer.writerow({"item": "final_headers", "value": " | ".join(FINAL_HEADERS)})

    print(f"rows={len(rows)}")
    print(f"backup={backup_path}")
    print(f"audit={AUDIT_PATH}")
    print("headers=" + " | ".join(FINAL_HEADERS))


if __name__ == "__main__":
    main()
