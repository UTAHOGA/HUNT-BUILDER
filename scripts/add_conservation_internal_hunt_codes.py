#!/usr/bin/env python3
"""Add internal C_ conservation codes while preserving DWR hunt codes.

Conservation permits can reuse DWR hunt codes that also appear in sportsman or
draw-result sources. The C_ CONSERVATION_CODE is for internal identity
only; the original HUNT CODE remains available for boundary_id joins and map
rendering.
"""

from __future__ import annotations

import csv
import re
import shutil
from datetime import datetime, timezone
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
WORKBOOKS = [
    ROOT / "data_truth" / "conservation_permit_truth" / "2025-27 Conservation Permits.xlsx",
    ROOT
    / "processed_data"
    / "hard_data_exports"
    / "hunt_tables"
    / "2026"
    / "CLEAN_XLXS_STAGED"
    / "2025-27 Conservation Permits.xlsx",
]
BACKUP_DIR = ROOT / "audits" / "2025_canonical_finalization" / "backups"
AUDIT_CSV = ROOT / "audits" / "2025_canonical_finalization" / "conservation_internal_hunt_code_audit.csv"


def clean(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split())


def internal_code(dwr_code: str) -> str:
    code = clean(dwr_code).upper()
    if not code:
        return ""
    if re.fullmatch(r"C_[A-Z]{2}\d{4}", code):
        return code
    if re.fullmatch(r"C[A-Z]{2}\d{4}", code):
        code = code[1:]
    return f"C_{code}"


def header_values(sheet) -> list[str]:
    return [clean(cell.value).upper() for cell in sheet[1]]


def ensure_column(sheet, header: str, after_header: str | None = None) -> int:
    headers = [clean(cell.value).upper() for cell in sheet[1]]
    target = header.upper()
    if target in headers:
        return headers.index(target) + 1

    insert_at = sheet.max_column + 1
    if after_header:
        after = after_header.upper()
        if after in headers:
            insert_at = headers.index(after) + 2
    sheet.insert_cols(insert_at)
    sheet.cell(row=1, column=insert_at).value = header
    return insert_at


def delete_columns_by_header(sheet, headers_to_delete: set[str]) -> None:
    """Remove obsolete helper columns from right to left so indexes stay valid."""
    headers = header_values(sheet)
    for idx in range(len(headers), 0, -1):
        if headers[idx - 1] in headers_to_delete:
            sheet.delete_cols(idx)


def find_column(sheet, names: list[str]) -> int | None:
    headers = [clean(cell.value).upper() for cell in sheet[1]]
    for name in names:
        target = name.upper()
        if target in headers:
            return headers.index(target) + 1
    return None


def process_workbook(path: Path) -> list[dict[str, str]]:
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    backup_path = BACKUP_DIR / f"{path.stem}.before_conservation_internal_codes_{timestamp}{path.suffix}"
    shutil.copy2(path, backup_path)

    workbook = openpyxl.load_workbook(path)
    audit_rows: list[dict[str, str]] = []

    for sheet in workbook.worksheets:
        if sheet.max_row < 2:
            continue
        hunt_code_col = find_column(sheet, ["HUNT CODE", "hunt_code"])
        if not hunt_code_col:
            continue

        delete_columns_by_header(
            sheet,
            {
                "CONSERVATION HUNT CODE",
                "DWR HUNT CODE",
                "BOUNDARY HUNT CODE",
            },
        )
        hunt_code_col = find_column(sheet, ["HUNT CODE", "hunt_code"])
        if not hunt_code_col:
            continue
        conservation_col = ensure_column(sheet, "CONSERVATION_CODE", "HUNT CODE")

        for row_idx in range(2, sheet.max_row + 1):
            dwr_code = clean(sheet.cell(row=row_idx, column=hunt_code_col).value).upper()
            if not dwr_code:
                continue
            c_code = internal_code(dwr_code)
            sheet.cell(row=row_idx, column=conservation_col).value = c_code
            audit_rows.append(
                {
                    "workbook": str(path.relative_to(ROOT)).replace("\\", "/"),
                    "sheet": sheet.title,
                    "row": str(row_idx),
                    "hunt_code": dwr_code,
                    "conservation_code": c_code,
                    "backup_path": str(backup_path.relative_to(ROOT)).replace("\\", "/"),
                }
            )

    workbook.save(path)
    return audit_rows


def main() -> int:
    all_rows: list[dict[str, str]] = []
    for workbook_path in WORKBOOKS:
        if not workbook_path.exists():
            continue
        all_rows.extend(process_workbook(workbook_path))

    AUDIT_CSV.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "workbook",
        "sheet",
        "row",
        "hunt_code",
        "conservation_code",
        "backup_path",
    ]
    with AUDIT_CSV.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        writer.writerows(all_rows)

    print(f"updated_rows={len(all_rows)}")
    print(f"audit={AUDIT_CSV}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
