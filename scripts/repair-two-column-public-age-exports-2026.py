from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
RESEARCH_CONTRACT = ROOT / "processed_data" / "research_page" / "hunt_application_outlook.json"
DATABASE_CSV = ROOT / "pipeline" / "RAW" / "hunt_unit_database" / "2026" / "csv" / "DATABASE.csv"
TABLE_EXPORT_DIR = ROOT / "processed_data" / "hard_data_exports" / "hunt_tables" / "2026" / "CLEAN_XLXS_STAGED"
SAMPLE_OUTPUT_DIR = ROOT / "processed_data" / "hard_data_exports" / "hunt_tables" / "2026" / "SAMPLES_TWO_COLUMN_AGE"

AUDIT_CSV = ROOT / "processed_data" / "audits" / "average_harvest_age_two_column_audit.csv"
FULL_PROMOTION_AUDIT_JSON = ROOT / "processed_data" / "audits" / "average_harvest_age_two_column_full_promotion_audit.json"
FULL_PROMOTION_AUDIT_CSV = ROOT / "processed_data" / "audits" / "average_harvest_age_two_column_full_promotion_audit.csv"
POLICY_MD = ROOT / "docs" / "average_harvest_age_two_column_policy.md"

AVERAGE_HEADER = "Average Harvest Age"
CURRENT_AGE_HEADER = "Current Age (3-Yr Avg)"

AVG_AGE_HEADER_ALIASES = {
    "average harvest age",
    "average harvest age prior year",
    "average age harvested",
    "average age harvested previous hunting season",
    "avg age harvested",
}

CURRENT_AGE_HEADER_ALIASES = {
    "current age 3 yr avg",
    "current age 3yr avg",
    "current age 3year avg",
    "currentage3yraverage",
    "current_age_3yr_average",
}

REQUIRED_SPECIES = [
    "Elk",
    "Black Bear",
    "Mountain Goat",
    "Moose",
    "Pronghorn",
    "Deer",
    "Desert Bighorn Sheep",
    "Rocky Mountain Bighorn Sheep",
    "Bison",
    "Turkey",
    "Cougar",
]


def clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def norm_code(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]+", "", clean(value).upper())


def norm_header(value: Any) -> str:
    text = clean(value).lower().replace("_", " ")
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def norm_name(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", clean(value).lower())


def to_positive_float(value: Any) -> float | None:
    text = clean(value)
    if not text:
        return None
    try:
        number = float(text)
    except (TypeError, ValueError):
        return None
    if number <= 0:
        return None
    return number


def display_number(value: float | None) -> str:
    if value is None:
        return ""
    return f"{value:.1f}".rstrip("0").rstrip(".")


def normalize_species(species: str) -> str:
    text = clean(species).lower()
    if "deer" in text:
        return "Deer"
    if "elk" in text:
        return "Elk"
    if "black bear" in text or text == "bear" or "bear" in text:
        return "Black Bear"
    if "goat" in text:
        return "Mountain Goat"
    if "moose" in text:
        return "Moose"
    if "pronghorn" in text:
        return "Pronghorn"
    if "desert bighorn" in text:
        return "Desert Bighorn Sheep"
    if "rocky mountain bighorn" in text or "rocky mtn bighorn" in text:
        return "Rocky Mountain Bighorn Sheep"
    if "bison" in text:
        return "Bison"
    if "turkey" in text:
        return "Turkey"
    if "cougar" in text or "mountain lion" in text:
        return "Cougar"
    return clean(species)


def annual_family_from_row(row: dict[str, Any]) -> str:
    avg = to_positive_float(row.get("average_harvest_age"))
    if avg is None:
        return "SOURCE_MISSING"
    status = clean(row.get("age_review_status")).upper()
    source_file = clean(row.get("age_source_file")).lower()
    if "UNIT_LEVEL" in status or "UNIT" in status:
        return "UNIT_LEVEL_REPEATED_ANNUAL_AGE"
    if source_file:
        return "ANNUAL_HARVEST_REPORT_AGE"
    return "FALLBACK_MERGED_AGE"


def load_research_rows() -> list[dict[str, Any]]:
    return json.loads(RESEARCH_CONTRACT.read_text(encoding="utf-8"))


def load_current_age_lookup() -> dict[str, float]:
    lookup: dict[str, float] = {}
    with DATABASE_CSV.open("r", encoding="utf-8-sig", newline="") as fh:
        for row in csv.DictReader(fh):
            code = norm_code(row.get("hunt_code"))
            age = to_positive_float(row.get("current_age_3yr_average"))
            if code and age is not None:
                lookup[code] = age
    return lookup


def build_average_age_lookup(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    rank = {
        "ANNUAL_HARVEST_REPORT_AGE": 3,
        "UNIT_LEVEL_REPEATED_ANNUAL_AGE": 2,
        "FALLBACK_MERGED_AGE": 1,
    }
    lookup: dict[str, dict[str, Any]] = {}
    for row in rows:
        code = norm_code(row.get("hunt_code"))
        age = to_positive_float(row.get("average_harvest_age"))
        if not code or age is None:
            continue
        family = annual_family_from_row(row)
        if family not in rank:
            continue
        current = lookup.get(code)
        payload = {
            "value": age,
            "family": family,
            "source_file": clean(row.get("age_source_file")),
            "review_status": clean(row.get("age_review_status")),
        }
        if not current or rank[family] > rank[current["family"]]:
            lookup[code] = payload
    return lookup


def find_header_row(ws) -> int | None:
    for row_idx in range(1, min(ws.max_row, 12) + 1):
        vals = [norm_header(ws.cell(row=row_idx, column=col).value) for col in range(1, ws.max_column + 1)]
        if "hunt code" in vals or "huntcode" in vals or "hunt no" in vals:
            return row_idx
    return None


def header_map(ws, header_row: int) -> dict[str, int]:
    result: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        key = norm_header(ws.cell(row=header_row, column=col).value)
        if key and key not in result:
            result[key] = col
    return result


def find_column(index: dict[str, int], aliases: set[str]) -> int | None:
    for alias in aliases:
        key = norm_header(alias)
        if key in index:
            return index[key]
    return None


def name_species_key(species: str, hunt_name: str) -> str:
    return f"{normalize_species(species)}|{norm_name(hunt_name)}"


def choose_sample_files() -> list[Path]:
    preferred = [
        "2026_BLACK_BEAR.xlsx",
        "2026_ELK_BULL_ALL.xlsx",
        "2026_DEER_BUCK_LIMITED_ENTRY.xlsx",
    ]
    files = []
    for name in preferred:
        path = TABLE_EXPORT_DIR / name
        if path.exists():
            files.append(path)
    if len(files) >= 3:
        return files
    for path in sorted(TABLE_EXPORT_DIR.glob("*.xlsx")):
        if path not in files and not path.name.startswith("~$"):
            files.append(path)
        if len(files) >= 3:
            break
    return files


def list_full_export_files() -> list[Path]:
    return sorted([p for p in TABLE_EXPORT_DIR.glob("*.xlsx") if not p.name.startswith("~$")])


def rewrite_workbooks(
    sources: list[Path],
    avg_lookup: dict[str, dict[str, Any]],
    current_lookup: dict[str, float],
    write_mode: str,
) -> list[dict[str, Any]]:
    if write_mode == "sample_copy":
        SAMPLE_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    reports: list[dict[str, Any]] = []
    for src in sources:
        dst = SAMPLE_OUTPUT_DIR / src.name if write_mode == "sample_copy" else src
        wb = load_workbook(src)
        ws = wb[wb.sheetnames[0]]
        header_row = find_header_row(ws)
        if header_row is None:
            wb.save(dst)
            wb.close()
            reports.append({"file": src.name, "status": "NO_HUNT_CODE_HEADER"})
            continue
        index = header_map(ws, header_row)
        code_col = find_column(index, {"hunt code", "hunt_code", "hunt no", "hunt_no"})
        age_col = find_column(index, AVG_AGE_HEADER_ALIASES)
        current_col = find_column(index, CURRENT_AGE_HEADER_ALIASES)
        if code_col is None:
            wb.save(dst)
            wb.close()
            reports.append({"file": src.name, "status": "NO_HUNT_CODE_COLUMN"})
            continue
        columns_added = 0
        old_header_replaced = 0
        if age_col is not None:
            old_header = clean(ws.cell(row=header_row, column=age_col).value).lower()
            if old_header in {"average harvest age prior year", "average age harvested (previous hunting season)"}:
                old_header_replaced = 1
        if age_col is None:
            age_col = ws.max_column + 1
            ws.cell(row=header_row, column=age_col).value = AVERAGE_HEADER
            columns_added += 1
        else:
            ws.cell(row=header_row, column=age_col).value = AVERAGE_HEADER
        if current_col is None:
            ws.insert_cols(age_col + 1)
            current_col = age_col + 1
            ws.cell(row=header_row, column=current_col).value = CURRENT_AGE_HEADER
            columns_added += 1
        else:
            ws.cell(row=header_row, column=current_col).value = CURRENT_AGE_HEADER

        rows_touched = 0
        avg_filled = 0
        current_filled = 0
        for row_idx in range(header_row + 1, ws.max_row + 1):
            code = norm_code(ws.cell(row=row_idx, column=code_col).value)
            if not code:
                continue
            rows_touched += 1
            avg_item = avg_lookup.get(code)
            if avg_item:
                ws.cell(row=row_idx, column=age_col).value = display_number(avg_item["value"])
                avg_filled += 1
            else:
                existing = to_positive_float(ws.cell(row=row_idx, column=age_col).value)
                if existing is None:
                    ws.cell(row=row_idx, column=age_col).value = ""
            cur_val = current_lookup.get(code)
            if cur_val is not None:
                ws.cell(row=row_idx, column=current_col).value = display_number(cur_val)
                current_filled += 1
            else:
                existing_cur = to_positive_float(ws.cell(row=row_idx, column=current_col).value)
                if existing_cur is None:
                    ws.cell(row=row_idx, column=current_col).value = ""

        wb.save(dst)
        wb.close()
        reports.append(
            {
                "file": src.name,
                "status": "UPDATED",
                "rows_touched": rows_touched,
                "average_harvest_age_filled": avg_filled,
                "current_age_3yr_filled": current_filled,
                "columns_added": columns_added,
                "old_header_replaced": old_header_replaced,
                "write_mode": write_mode,
            }
        )
    return reports


def build_two_column_audit(
    rows: list[dict[str, Any]],
    avg_lookup: dict[str, dict[str, Any]],
    current_lookup: dict[str, float],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    species_rows: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
    name_species_has_avg: set[str] = set()
    name_species_has_current: set[str] = set()

    for row in rows:
        code = norm_code(row.get("hunt_code"))
        species = normalize_species(clean(row.get("species")))
        hunt_name = clean(row.get("hunt_name"))
        key = name_species_key(species, hunt_name)
        if code in avg_lookup:
            name_species_has_avg.add(key)
        if code in current_lookup:
            name_species_has_current.add(key)
        species_rows[species].append(row)

    out_rows: list[dict[str, Any]] = []
    cause_counts_avg: Counter[str] = Counter()
    cause_counts_current: Counter[str] = Counter()

    for row in rows:
        code = norm_code(row.get("hunt_code"))
        species = normalize_species(clean(row.get("species")))
        hunt_name = clean(row.get("hunt_name"))
        key = name_species_key(species, hunt_name)
        residency = clean(row.get("residency"))

        public_avg = to_positive_float(row.get("average_harvest_age"))
        public_current = to_positive_float(row.get("current_age_3yr_average"))
        canonical_avg = avg_lookup.get(code)
        canonical_current = current_lookup.get(code)

        avg_cause = ""
        cur_cause = ""

        if public_avg is None:
            if canonical_avg is not None:
                avg_cause = "JOIN_FAILURE"
            elif key in name_species_has_avg:
                avg_cause = "MAPPING_FAILURE"
            elif canonical_current is not None and species in {"Turkey", "Cougar", "Bison"}:
                avg_cause = "NOT_SUPPORTED_FOR_SPECIES"
            else:
                avg_cause = "SOURCE_MISSING"
            cause_counts_avg[avg_cause] += 1

        if public_current is None:
            if canonical_current is not None:
                cur_cause = "JOIN_FAILURE"
            elif key in name_species_has_current:
                cur_cause = "MAPPING_FAILURE"
            elif species in {"Turkey", "Cougar", "Bison", "Desert Bighorn Sheep", "Rocky Mountain Bighorn Sheep", "Pronghorn", "Deer"}:
                cur_cause = "NOT_SUPPORTED_FOR_SPECIES"
            else:
                cur_cause = "SOURCE_MISSING"
            cause_counts_current[cur_cause] += 1

        out_rows.append(
            {
                "row_key": f"{code}|{residency}",
                "hunt_code": code,
                "residency": residency,
                "species": species,
                "hunt_name": hunt_name,
                "public_average_harvest_age": display_number(public_avg),
                "public_current_age_3yr_average": display_number(public_current),
                "canonical_average_harvest_age": display_number(canonical_avg["value"] if canonical_avg else None),
                "canonical_current_age_3yr_average": display_number(canonical_current),
                "average_age_source_family": canonical_avg["family"] if canonical_avg else "SOURCE_MISSING",
                "current_age_source_family": "HUNT_PLANNER_CURRENT_3YR_AVG" if canonical_current is not None else "SOURCE_MISSING",
                "average_age_blank_cause": avg_cause,
                "current_age_blank_cause": cur_cause,
                "harvest_success_pct": clean(row.get("harvest_success_pct")),
                "average_days_hunted": clean(row.get("average_days_hunted")),
                "average_age_source_file": clean(row.get("age_source_file")),
                "average_age_review_status": clean(row.get("age_review_status")),
            }
        )

    summary = {
        "rows": len(rows),
        "unique_hunt_codes": len({norm_code(r.get("hunt_code")) for r in rows if norm_code(r.get("hunt_code"))}),
        "public_average_populated_rows": sum(1 for r in out_rows if r["public_average_harvest_age"]),
        "public_current_age_3yr_populated_rows": sum(1 for r in out_rows if r["public_current_age_3yr_average"]),
        "canonical_average_populated_codes": len(avg_lookup),
        "canonical_current_age_3yr_populated_codes": len(current_lookup),
        "average_blank_cause_counts": dict(cause_counts_avg),
        "current_blank_cause_counts": dict(cause_counts_current),
    }
    return out_rows, summary


def species_support_section(out_rows: list[dict[str, Any]]) -> list[str]:
    by_species: defaultdict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in out_rows:
        by_species[row["species"]].append(row)

    lines = [
        "| Species | Average Harvest Age Support | Current Age (3-Yr Avg) Support |",
        "|---|---|---|",
    ]

    for species in REQUIRED_SPECIES:
        rows = by_species.get(species, [])
        total = len(rows)
        avg_pop = sum(1 for r in rows if r["public_average_harvest_age"])
        cur_pop = sum(1 for r in rows if r["public_current_age_3yr_average"])
        direct = sum(1 for r in rows if r["average_age_source_family"] == "ANNUAL_HARVEST_REPORT_AGE")
        unit = sum(1 for r in rows if r["average_age_source_family"] == "UNIT_LEVEL_REPEATED_ANNUAL_AGE")

        if total == 0:
            avg_support = "NOT_SUPPORTED"
            cur_support = "NOT_SUPPORTED"
        else:
            avg_cov = avg_pop / total
            cur_cov = cur_pop / total
            if direct > 0 and avg_cov >= 0.60:
                avg_support = "SUPPORTED_DIRECT"
            elif direct == 0 and unit > 0:
                avg_support = "SUPPORTED_UNIT_CROSSWALK"
            elif avg_cov > 0:
                avg_support = "PARTIAL_SUPPORT"
            else:
                avg_support = "NOT_SUPPORTED"

            if cur_cov >= 0.60:
                cur_support = "SUPPORTED_DIRECT"
            elif cur_cov > 0:
                cur_support = "PARTIAL_SUPPORT"
            else:
                cur_support = "NOT_SUPPORTED"

            if avg_support == "NOT_SUPPORTED" and species in {"Deer", "Pronghorn", "Moose"}:
                avg_support = "REVIEW_REQUIRED"

        lines.append(f"| {species} | {avg_support} | {cur_support} |")
    return lines


def write_outputs(
    out_rows: list[dict[str, Any]],
    summary: dict[str, Any],
    sample_reports: list[dict[str, Any]],
) -> None:
    AUDIT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with AUDIT_CSV.open("w", newline="", encoding="utf-8") as fh:
        fieldnames = [
            "row_key",
            "hunt_code",
            "residency",
            "species",
            "hunt_name",
            "public_average_harvest_age",
            "public_current_age_3yr_average",
            "canonical_average_harvest_age",
            "canonical_current_age_3yr_average",
            "average_age_source_family",
            "current_age_source_family",
            "average_age_blank_cause",
            "current_age_blank_cause",
            "harvest_success_pct",
            "average_days_hunted",
            "average_age_source_file",
            "average_age_review_status",
        ]
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(out_rows)

    source_rules = [
        "- `Average Harvest Age` uses only: `ANNUAL_HARVEST_REPORT_AGE`, `UNIT_LEVEL_REPEATED_ANNUAL_AGE`, `FALLBACK_MERGED_AGE`.",
        "- `Current Age (3-Yr Avg)` uses only: `HUNT_PLANNER_CURRENT_3YR_AVG` (from canonical 2026 DATABASE field `current_age_3yr_average`).",
        "- `Age Objective` is not used for either column.",
        "- Values `<= 0` are blanked.",
    ]

    policy_lines = [
        "# Average Harvest Age Two-Column Policy",
        "",
        "## Final Definitions",
        "",
        "- `Average Harvest Age`: observed harvested-age value from annual-report age evidence and validated carry-forward families.",
        "- `Current Age (3-Yr Avg)`: Utah DWR Hunt Planner current three-year average age context value.",
        "",
        "## Source-Family Rules",
        "",
        *source_rules,
        "",
        "## Species-by-Species Support Policy",
        "",
        *species_support_section(out_rows),
        "",
        "## Population Summary",
        "",
        f"- Rows audited: {summary['rows']}",
        f"- Unique hunt codes audited: {summary['unique_hunt_codes']}",
        f"- Public `Average Harvest Age` populated rows: {summary['public_average_populated_rows']}",
        f"- Public `Current Age (3-Yr Avg)` populated rows: {summary['public_current_age_3yr_populated_rows']}",
        f"- Canonical average-age code coverage: {summary['canonical_average_populated_codes']}",
        f"- Canonical current-age code coverage: {summary['canonical_current_age_3yr_populated_codes']}",
        "",
        "### Blank Cause Breakdown (Average Harvest Age)",
        "",
        *[f"- {k}: {v}" for k, v in sorted(summary["average_blank_cause_counts"].items())],
        "",
        "### Blank Cause Breakdown (Current Age 3-Yr Avg)",
        "",
        *[f"- {k}: {v}" for k, v in sorted(summary["current_blank_cause_counts"].items())],
        "",
        "## Regenerated Sample Public Outputs",
        "",
        *[f"- `{item.get('file')}`: {item.get('status')} (rows={item.get('rows_touched', 0)}, avg={item.get('average_harvest_age_filled', 0)}, current3yr={item.get('current_age_3yr_filled', 0)})" for item in sample_reports],
        "",
        "## Remaining Gaps",
        "",
        "- Species and hunt families without defensible annual age evidence remain blank by design.",
        "- Rows marked `JOIN_FAILURE` or `MAPPING_FAILURE` indicate pipeline alignment work still needed for fully consistent public rendering.",
        "",
        "## Further Source Work Needed",
        "",
        "- Expand reviewed annual-report age extraction for species currently classified as `REVIEW_REQUIRED` or `NOT_SUPPORTED`.",
        "- Improve hunt-code join coverage where canonical values exist but public rows remain blank.",
    ]
    POLICY_MD.write_text("\n".join(policy_lines) + "\n", encoding="utf-8")


def write_full_promotion_audit(full_reports: list[dict[str, Any]]) -> dict[str, Any]:
    updated = [r for r in full_reports if r.get("status") == "UPDATED"]
    with_hunt_table_headers = [r for r in updated if int(r.get("rows_touched", 0)) > 0]
    summary = {
        "files_scanned": len(full_reports),
        "files_updated": len(updated),
        "files_with_hunt_rows": len(with_hunt_table_headers),
        "columns_added_total": sum(int(r.get("columns_added", 0)) for r in updated),
        "old_header_replaced_files": sum(int(r.get("old_header_replaced", 0)) for r in updated),
        "average_harvest_age_filled_rows_total": sum(int(r.get("average_harvest_age_filled", 0)) for r in updated),
        "current_age_3yr_filled_rows_total": sum(int(r.get("current_age_3yr_filled", 0)) for r in updated),
    }
    payload = {
        "summary": summary,
        "files": full_reports,
    }
    FULL_PROMOTION_AUDIT_JSON.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    FULL_PROMOTION_AUDIT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with FULL_PROMOTION_AUDIT_CSV.open("w", newline="", encoding="utf-8") as fh:
        fieldnames = [
            "file",
            "status",
            "rows_touched",
            "average_harvest_age_filled",
            "current_age_3yr_filled",
            "columns_added",
            "old_header_replaced",
            "write_mode",
        ]
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for row in full_reports:
            writer.writerow({k: row.get(k, "") for k in fieldnames})
    return summary


def audit_full_export_headers() -> dict[str, int]:
    files = list_full_export_files()
    counts: Counter[str] = Counter()
    for path in files:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        header_row = find_header_row(ws)
        if header_row is None:
            counts["no_hunt_header_row"] += 1
            wb.close()
            continue
        index = header_map(ws, header_row)
        has_avg = find_column(index, {AVERAGE_HEADER, *AVG_AGE_HEADER_ALIASES}) is not None
        has_current = find_column(index, {CURRENT_AGE_HEADER, *CURRENT_AGE_HEADER_ALIASES}) is not None
        if has_avg:
            counts["has_average_header"] += 1
        else:
            counts["missing_average_header"] += 1
        if has_current:
            counts["has_current_3yr_header"] += 1
        else:
            counts["missing_current_3yr_header"] += 1
        wb.close()
    counts["files_scanned"] = len(files)
    return dict(counts)


def main() -> None:
    parser = argparse.ArgumentParser(description="Repair two-column public age exports for 2026 hunt tables.")
    parser.add_argument(
        "--promote-full",
        action="store_true",
        help="Apply two-column age headers/values in-place to the full CLEAN_XLXS_STAGED workbook set.",
    )
    args = parser.parse_args()

    rows = load_research_rows()
    avg_lookup = build_average_age_lookup(rows)
    current_lookup = load_current_age_lookup()
    if args.promote_full:
        full_reports = rewrite_workbooks(list_full_export_files(), avg_lookup, current_lookup, "in_place_full")
        full_summary = write_full_promotion_audit(full_reports)
    else:
        full_reports = []
        full_summary = {}
    sample_reports = rewrite_workbooks(choose_sample_files(), avg_lookup, current_lookup, "sample_copy")
    audit_rows, summary = build_two_column_audit(rows, avg_lookup, current_lookup)
    write_outputs(audit_rows, summary, sample_reports)
    header_summary = audit_full_export_headers()
    print(
        json.dumps(
            {
                "ok": True,
                "summary": summary,
                "sample_outputs": sample_reports,
                "full_promotion_summary": full_summary,
                "full_header_summary": header_summary,
                "promote_full": bool(args.promote_full),
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
