#!/usr/bin/env python3
"""Export memory-efficient yearly canonical review workbooks.

Each workbook contains:
- README
- YEARLY_DATABASE: editable/review-friendly rows with core columns first
- HUNT_CODE_SUMMARY: one row per hunt code
- RAW_CANONICAL: the full yearly canonical CSV
- LONG_FILE_COPY: draw_results_long.csv filtered to the same actual_draw_year
- Audit: row counts and code alignment checks

This uses openpyxl write-only mode because the all-data workbooks are too large
for the in-memory artifact workbook path.
"""

from __future__ import annotations

import argparse
import csv
import json
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
CANONICAL_DIR = ROOT / "data_truth" / "draw_results_truth" / "normalized" / "canonical_yearly"
LONG_FILE = ROOT / "data_truth" / "draw_results_truth" / "normalized" / "draw_results_long.csv"
DATABASE = ROOT / "pipeline" / "RAW" / "hunt_unit_database" / "2026" / "csv" / "DATABASE.csv"
OUTPUT_DIR = ROOT / "outputs" / "yearly_canonical_workbooks"
REPORT_PATH = OUTPUT_DIR / "yearly_canonical_workbooks_report.json"

SCORABLE_RECORD_TYPES = {"point_level_draw_result", "sportsman_total"}
NON_SCORABLE_RECORD_TYPES = {
    "hunt_total_draw_result",
    "permit_reference",
    "allocation_reference",
    "reference_only",
    "permit_number_only_not_draw_result",
}
TEXT_COLUMNS = {
    "hunt_code",
    "conservation_code",
    "source_file",
    "source_scope",
    "source_namespace",
    "draw_source_namespace",
    "hunt_name",
    "species",
    "sex_type",
    "draw_design",
    "weapon",
    "hunt_type",
    "season",
    "residency",
    "record_type",
    "algorithm_status",
    "source_dataset",
    "extraction_status",
    "parse_method",
    "qa_status",
    "notes",
    "success_ratio",
}
HARVEST_COLUMNS = [
    "percent_harvest_success_previous_hunting_season",
    "current_age_3yr_average",
    "average_harvest_age",
    "average_harvest_age_reported_hunt_year",
    "average_harvest_age_source_file",
    "average_harvest_age_review_status",
    "dwr_huntplanner_age_objective",
    "dwr_huntplanner_population_objective",
    "dwr_huntplanner_current_population_estimate",
]


def is_p_value_column(column: object) -> bool:
    text = clean(column).lower()
    return "p_draw" in text


def visible_header(header: Iterable[str]) -> list[str]:
    return [column for column in header if not is_p_value_column(column)]


def clean(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def lower_clean(value: object) -> str:
    return clean(value).lower()


def is_conservation_or_reference(row: dict[str, str]) -> bool:
    text = " ".join(
        lower_clean(row.get(field))
        for field in (
            "record_type",
            "row_type",
            "hunt_type",
            "draw_design",
            "hunt_class",
            "hunt_draw_class",
            "source_scope",
            "source_namespace",
            "draw_source_namespace",
            "source_file",
            "notes",
        )
    )
    source_scope = lower_clean(row.get("source_scope"))
    return (
        "permit_number_only_not_draw_result" in text
        or "reference_only" in text
        or "allocation_only" in text
        or "allocation/reference" in text
        or "point_purchase" in text
        or "point-only" in text
        or "point only" in text
        or "conservation" in source_scope
        or lower_clean(row.get("hunt_type")) == "conservation"
    )


def maybe_number(column: str, value: object) -> object:
    text = clean(value).replace(",", "")
    if not text or column in TEXT_COLUMNS:
        return clean(value)
    try:
        number = float(text)
    except ValueError:
        return clean(value)
    return int(number) if number.is_integer() else number


def canonical_path(year: int) -> Path:
    return CANONICAL_DIR / f"draw_results_{year}_for_{year + 1}_canonical_yearly_draw_results.csv"


def available_years() -> list[int]:
    years: list[int] = []
    for path in CANONICAL_DIR.glob("draw_results_*_for_*_canonical_yearly_draw_results.csv"):
        parts = path.name.split("_")
        if len(parts) > 2 and parts[2].isdigit():
            years.append(int(parts[2]))
    return sorted(set(years))


def read_csv_rows(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), list(reader)


def load_database_by_hunt_code() -> dict[str, dict[str, str]]:
    if not DATABASE.exists():
        return {}
    with DATABASE.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        return {clean(row.get("hunt_code")).upper(): row for row in reader if clean(row.get("hunt_code"))}


def long_rows_for_year(year: int) -> tuple[list[str], list[dict[str, str]]]:
    with LONG_FILE.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        header = list(reader.fieldnames or [])
        rows = [row for row in reader if clean(row.get("actual_draw_year")) == str(year)]
    return header, rows


def is_scorable(row: dict[str, str]) -> bool:
    if is_conservation_or_reference(row):
        return False
    record_type = lower_clean(row.get("record_type"))
    if record_type in NON_SCORABLE_RECORD_TYPES:
        return False
    return record_type in SCORABLE_RECORD_TYPES


def first_nonblank(rows: Iterable[dict[str, str]], column: str) -> str:
    for row in rows:
        value = clean(row.get(column))
        if value:
            return value
    return ""


def first_number(rows: Iterable[dict[str, str]], column: str) -> object:
    for row in rows:
        value = maybe_number(column, row.get(column))
        if isinstance(value, (int, float)):
            return value
    return ""


def unique_nonblank(rows: Iterable[dict[str, str]], column: str) -> list[str]:
    return sorted({clean(row.get(column)) for row in rows if clean(row.get(column))})


def build_summary_rows(rows: list[dict[str, str]], year: int) -> list[dict[str, object]]:
    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        code = clean(row.get("hunt_code")).upper()
        if code:
            grouped[code].append(row)

    res_col = f"permits_{year}_res"
    nr_col = f"permits_{year}_nr"
    total_col = f"permits_{year}_total"
    output: list[dict[str, object]] = []
    for hunt_code, group in grouped.items():
        permits_res = first_number(group, res_col)
        permits_nr = first_number(group, nr_col)
        permits_total = first_number(group, total_col)
        if permits_total == "" and (permits_res != "" or permits_nr != ""):
            permits_total = int(permits_res or 0) + int(permits_nr or 0)

        output.append(
            {
                "actual_draw_year": year,
                "model_target_year": year + 1,
                "hunt_code": hunt_code,
                "boundary_id": first_number(group, "boundary_id"),
                "species": first_nonblank(group, "species"),
                "sex_type": first_nonblank(group, "sex_type"),
                "weapon": first_nonblank(group, "weapon"),
                "hunt_name": first_nonblank(group, "hunt_name"),
                "hunt_type": first_nonblank(group, "hunt_type"),
                "draw_design": first_nonblank(group, "draw_design"),
                "season": first_nonblank(group, "season"),
                res_col: permits_res,
                nr_col: permits_nr,
                total_col: permits_total,
                "canonical_rows": len(group),
                "scorable_rows": sum(1 for row in group if is_scorable(row)),
                "display_reference_rows": sum(1 for row in group if not is_scorable(row)),
                "record_types": "; ".join(unique_nonblank(group, "record_type")),
                "source_files": "; ".join(unique_nonblank(group, "source_file")[:5]),
            }
        )
    return sorted(output, key=lambda row: str(row["hunt_code"]))


def yearly_database_header(raw_header: list[str], year: int) -> list[str]:
    return [
        "actual_draw_year",
        "model_target_year",
        "boundary_id",
        "hunt_code",
        "hunt_name",
        "sex_type",
        "species",
        "hunt_type",
        "weapon",
        "season",
        "draw_design",
        "points",
        "record_type",
        f"{year}_PERMITS_RES",
        f"{year}_PERMITS_NR",
        f"{year}_PERMITS_TOTAL",
        "resident_eligible_applicants",
        "resident_bonus_permits",
        "resident_regular_permits",
        "resident_total_permits",
        "resident_success_ratio",
        "nonresident_eligible_applicants",
        "nonresident_bonus_permits",
        "nonresident_regular_permits",
        "nonresident_total_permits",
        "nonresident_success_ratio",
        "total_eligible_applicants",
        "total_bonus_permits",
        "total_regular_permits",
        "total_permits",
        "total_success_ratio",
        *HARVEST_COLUMNS,
        "source_file",
        "source_scope",
        "source_namespace",
        "draw_source_namespace",
        "page_kind",
        "pdf_page",
        "algorithm_status",
        "source_dataset",
        "extraction_status",
        "parse_method",
        "qa_status",
        "notes",
    ]


def residency_bucket(row: dict[str, str]) -> str:
    text = clean(row.get("residency")).lower().replace("-", "")
    if text in {"resident", "res"}:
        return "resident"
    if text in {"nonresident", "non resident", "nr", "nonres"}:
        return "nonresident"
    return "total"


def collapse_key(row: dict[str, str]) -> tuple[str, ...]:
    return (
        clean(row.get("actual_draw_year")),
        clean(row.get("model_target_year")),
        clean(row.get("hunt_code")).upper(),
        clean(row.get("hunt_name")),
        clean(row.get("species")),
        clean(row.get("sex_type")),
        clean(row.get("hunt_type")),
        clean(row.get("weapon")),
        clean(row.get("season")),
        clean(row.get("draw_design")),
        clean(row.get("points")),
        clean(row.get("record_type")),
        clean(row.get("boundary_id")),
        clean(row.get("source_file")),
        clean(row.get("source_scope")),
        clean(row.get("source_namespace")),
        clean(row.get("draw_source_namespace")),
        clean(row.get("page_kind")),
        clean(row.get("pdf_page")),
    )


def fill_metric_columns(out: dict[str, object], row: dict[str, str], prefix: str) -> None:
    for column in [
        "eligible_applicants",
        "bonus_permits",
        "regular_permits",
        "total_permits",
        "success_ratio",
    ]:
        source_column = "total_permits" if prefix == "total" and column == "total_permits" else f"{prefix}_{column}"
        value = maybe_number(source_column, row.get(source_column))
        if value == "" and source_column not in row:
            value = maybe_number(column, row.get(column))
        if value != "":
            out[source_column] = value


def has_split_metric_columns(row: dict[str, str]) -> bool:
    for prefix in ("resident", "nonresident", "total"):
        for column in (
            "eligible_applicants",
            "bonus_permits",
            "regular_permits",
            "total_permits",
            "success_ratio",
        ):
            source_column = "total_permits" if prefix == "total" and column == "total_permits" else f"{prefix}_{column}"
            if clean(row.get(source_column)):
                return True
    return False


def collapsed_yearly_database_rows(
    rows: list[dict[str, str]],
    header: list[str],
    year: int,
    database_by_code: dict[str, dict[str, str]],
) -> Iterable[dict[str, object]]:
    res_col = f"permits_{year}_res"
    nr_col = f"permits_{year}_nr"
    total_col = f"permits_{year}_total"
    collapsed: dict[tuple[str, ...], dict[str, object]] = {}
    for row in rows:
        key = collapse_key(row)
        code = clean(row.get("hunt_code")).upper()
        if key not in collapsed:
            db_row = database_by_code.get(code, {})
            out: dict[str, object] = {
                "actual_draw_year": maybe_number("actual_draw_year", row.get("actual_draw_year")),
                "model_target_year": maybe_number("model_target_year", row.get("model_target_year")),
                "boundary_id": maybe_number("boundary_id", row.get("boundary_id")),
                "hunt_code": code,
                "hunt_name": clean(row.get("hunt_name")),
                "sex_type": clean(row.get("sex_type")),
                "species": clean(row.get("species")),
                "hunt_type": clean(row.get("hunt_type")),
                "weapon": clean(row.get("weapon")),
                "season": clean(row.get("season")),
                "draw_design": clean(row.get("draw_design")),
                "points": maybe_number("points", row.get("points")),
                "record_type": clean(row.get("record_type")),
                f"{year}_PERMITS_RES": maybe_number(res_col, row.get(res_col)),
                f"{year}_PERMITS_NR": maybe_number(nr_col, row.get(nr_col)),
                f"{year}_PERMITS_TOTAL": maybe_number(total_col, row.get(total_col)),
                "source_file": clean(row.get("source_file")),
                "source_scope": clean(row.get("source_scope")),
                "source_namespace": clean(row.get("source_namespace")),
                "draw_source_namespace": clean(row.get("draw_source_namespace")),
                "page_kind": clean(row.get("page_kind")),
                "pdf_page": maybe_number("pdf_page", row.get("pdf_page")),
                "algorithm_status": clean(row.get("algorithm_status")),
                "source_dataset": clean(row.get("source_dataset")),
                "extraction_status": clean(row.get("extraction_status")),
                "parse_method": clean(row.get("parse_method")),
                "qa_status": clean(row.get("qa_status")),
                "notes": clean(row.get("notes")),
            }
            for column in HARVEST_COLUMNS:
                out[column] = db_row.get(column, "")
            for column in header:
                out.setdefault(column, "")
            collapsed[key] = out
        out = collapsed[key]
        for permit_column, source_column in [
            (f"{year}_PERMITS_RES", res_col),
            (f"{year}_PERMITS_NR", nr_col),
            (f"{year}_PERMITS_TOTAL", total_col),
        ]:
            value = maybe_number(source_column, row.get(source_column))
            if value != "":
                out[permit_column] = value
        if has_split_metric_columns(row):
            for prefix in ("resident", "nonresident", "total"):
                fill_metric_columns(out, row, prefix)
        else:
            fill_metric_columns(out, row, residency_bucket(row))

    def point_sort_value(value: object) -> tuple[int, float]:
        text = clean(value)
        if not text:
            return (1, 0.0)
        try:
            return (0, -float(text))
        except ValueError:
            return (2, 0.0)

    for out in sorted(
        collapsed.values(),
        key=lambda item: (
            clean(item.get("hunt_code")),
            point_sort_value(item.get("points")),
            clean(item.get("record_type")),
        ),
    ):
        for column in HARVEST_COLUMNS:
            out.setdefault(column, "")
        yield out


def style_cell(cell, *, header: bool = False):
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    if header:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="183A37")
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    return cell


def append_row(sheet, values: Iterable[object], header: bool = False) -> None:
    if not header:
        sheet.append(list(values))
        return
    cells = [style_cell(WriteOnlyCell(sheet, value=value), header=True) for value in values]
    sheet.append(cells)


def write_table(sheet, header: list[str], rows: Iterable[dict[str, object]]) -> int:
    append_row(sheet, header, header=True)
    count = 0
    for row in rows:
        append_row(sheet, [maybe_number(column, row.get(column)) for column in header])
        count += 1
    return count


def set_widths(sheet, header: list[str]) -> None:
    widths = {
        "hunt_code": 14,
        "conservation_code": 16,
        "hunt_name": 36,
        "source_file": 38,
        "source_scope": 24,
        "source_namespace": 24,
        "draw_source_namespace": 26,
        "species": 20,
        "sex_type": 16,
        "draw_design": 20,
        "weapon": 22,
        "hunt_type": 20,
        "season": 28,
        "record_type": 22,
        "notes": 38,
        "source_files": 44,
        "record_types": 32,
        "percent_harvest_success_previous_hunting_season": 22,
        "current_age_3yr_average": 20,
        "average_harvest_age": 18,
        "average_harvest_age_reported_hunt_year": 18,
        "average_harvest_age_source_file": 28,
        "average_harvest_age_review_status": 24,
        "dwr_huntplanner_age_objective": 24,
        "dwr_huntplanner_population_objective": 24,
        "dwr_huntplanner_current_population_estimate": 24,
    }
    for index, column in enumerate(header, start=1):
        width = widths.get(column, 14 if column.startswith("permits_") else 13)
        sheet.column_dimensions[get_column_letter(index)].width = width


def audit_rows(
    *,
    year: int,
    canonical_header: list[str],
    canonical_rows: list[dict[str, str]],
    long_header: list[str],
    long_rows: list[dict[str, str]],
    summary_rows: list[dict[str, object]],
) -> list[dict[str, object]]:
    canonical_codes = {clean(row.get("hunt_code")).upper() for row in canonical_rows if clean(row.get("hunt_code"))}
    long_codes = {clean(row.get("hunt_code")).upper() for row in long_rows if clean(row.get("hunt_code"))}
    canonical_only = sorted(canonical_codes - long_codes)
    long_only = sorted(long_codes - canonical_codes)
    record_counts = Counter(clean(row.get("record_type")) or "(blank)" for row in canonical_rows)
    hunt_type_counts = Counter(clean(row.get("hunt_type")) or "(blank)" for row in canonical_rows)
    draw_design_counts = Counter(clean(row.get("draw_design")) or "(blank)" for row in canonical_rows)
    return [
        {"item": "actual_draw_year", "value": year},
        {"item": "model_target_year", "value": year + 1},
        {"item": "canonical_source", "value": str(canonical_path(year).relative_to(ROOT)).replace("\\", "/")},
        {"item": "long_source", "value": str(LONG_FILE.relative_to(ROOT)).replace("\\", "/")},
        {"item": "canonical_rows", "value": len(canonical_rows)},
        {"item": "canonical_columns", "value": len(canonical_header)},
        {"item": "long_file_slice_rows", "value": len(long_rows)},
        {"item": "long_file_columns", "value": len(long_header)},
        {"item": "summary_hunt_code_rows", "value": len(summary_rows)},
        {"item": "scorable_rows", "value": sum(1 for row in canonical_rows if is_scorable(row))},
        {"item": "display_reference_rows", "value": sum(1 for row in canonical_rows if not is_scorable(row))},
        {"item": "canonical_unique_hunt_codes", "value": len(canonical_codes)},
        {"item": "long_unique_hunt_codes", "value": len(long_codes)},
        {"item": "canonical_codes_missing_from_long_count", "value": len(canonical_only)},
        {"item": "long_codes_missing_from_canonical_count", "value": len(long_only)},
        {"item": "canonical_codes_missing_from_long_sample", "value": ", ".join(canonical_only[:40])},
        {"item": "long_codes_missing_from_canonical_sample", "value": ", ".join(long_only[:40])},
        {"item": "record_type_counts", "value": "; ".join(f"{key}: {value}" for key, value in record_counts.most_common())},
        {"item": "hunt_type_counts", "value": "; ".join(f"{key}: {value}" for key, value in hunt_type_counts.most_common())},
        {"item": "draw_design_counts", "value": "; ".join(f"{key}: {value}" for key, value in draw_design_counts.most_common())},
    ]


def export_year(year: int) -> dict[str, object]:
    canonical_header, canonical_rows = read_csv_rows(canonical_path(year))
    long_header, long_rows = long_rows_for_year(year)
    visible_canonical_header = visible_header(canonical_header)
    visible_long_header = visible_header(long_header)
    database_by_code = load_database_by_hunt_code()
    summary_rows = build_summary_rows(canonical_rows, year)
    database_header = yearly_database_header(canonical_header, year)

    workbook = Workbook(write_only=True)
    if workbook.worksheets:
        workbook.remove(workbook.worksheets[0])

    readme = workbook.create_sheet("README")
    write_table(
        readme,
        ["item", "value"],
        [
            {"item": "Workbook purpose", "value": "Generated yearly review workbook from canonical yearly truth plus matching draw_results_long.csv slice."},
            {"item": "YEARLY_DATABASE", "value": "Core review columns first, then raw/source columns. This is the main editable yearly sheet."},
            {"item": "HUNT_CODE_SUMMARY", "value": "One row per hunt code, useful for year-to-year comparison."},
            {"item": "RAW_CANONICAL", "value": "Full yearly canonical CSV, uncollapsed."},
            {"item": "LONG_FILE_COPY", "value": "Copy of draw_results_long.csv filtered to this actual_draw_year."},
        ],
    )
    set_widths(readme, ["item", "value"])

    summary_header = [
        "actual_draw_year",
        "model_target_year",
        "hunt_code",
        "boundary_id",
        "species",
        "sex_type",
        "weapon",
        "hunt_name",
        "hunt_type",
        "draw_design",
        "season",
        f"permits_{year}_res",
        f"permits_{year}_nr",
        f"permits_{year}_total",
        "canonical_rows",
        "scorable_rows",
        "display_reference_rows",
        "record_types",
        "source_files",
    ]
    yearly_sheet = workbook.create_sheet("YEARLY_DATABASE")
    collapsed_rows = list(collapsed_yearly_database_rows(canonical_rows, database_header, year, database_by_code))
    write_table(yearly_sheet, database_header, collapsed_rows)
    set_widths(yearly_sheet, database_header)

    summary = workbook.create_sheet("HUNT_CODE_SUMMARY")
    write_table(summary, summary_header, summary_rows)
    set_widths(summary, summary_header)

    canonical_sheet = workbook.create_sheet("RAW_CANONICAL")
    write_table(canonical_sheet, visible_canonical_header, canonical_rows)
    set_widths(canonical_sheet, visible_canonical_header)

    long_sheet = workbook.create_sheet("LONG_FILE_COPY")
    write_table(long_sheet, visible_long_header, long_rows)
    set_widths(long_sheet, visible_long_header)

    audit = workbook.create_sheet("Audit")
    audit_data = audit_rows(
        year=year,
        canonical_header=canonical_header,
        canonical_rows=canonical_rows,
        long_header=long_header,
        long_rows=long_rows,
        summary_rows=summary_rows,
    )
    write_table(audit, ["item", "value"], audit_data)
    set_widths(audit, ["item", "value"])

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = OUTPUT_DIR / f"{year}_PERMITS={year + 1}_MODEL__CANONICAL_WORKBOOK.xlsx"
    workbook.save(output_path)
    return {
        "year": year,
        "output_xlsx": str(output_path.relative_to(ROOT)).replace("\\", "/"),
        "canonical_rows": len(canonical_rows),
        "long_file_slice_rows": len(long_rows),
        "summary_hunt_code_rows": len(summary_rows),
        "yearly_database_collapsed_rows": len(collapsed_rows),
        "scorable_rows": sum(1 for row in canonical_rows if is_scorable(row)),
        "display_reference_rows": sum(1 for row in canonical_rows if not is_scorable(row)),
        "canonical_columns": len(canonical_header),
        "long_file_columns": len(long_header),
        "yearly_database_columns": len(database_header),
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--year", type=int, action="append", help="Actual draw year to export. Can be repeated.")
    parser.add_argument("--all", action="store_true", help="Export every available canonical year.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    years = available_years() if args.all else (args.year or [2025])
    report = {
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "output_dir": str(OUTPUT_DIR.relative_to(ROOT)).replace("\\", "/"),
        "years": [],
    }
    for year in years:
        report["years"].append(export_year(year))
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    REPORT_PATH.write_text(json.dumps(report, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
