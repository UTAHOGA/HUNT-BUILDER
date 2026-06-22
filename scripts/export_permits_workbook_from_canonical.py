#!/usr/bin/env python3
"""Export a public/review PERMITS workbook from canonical yearly draw results.

This replaces older workbook builders that used candidate/rollup files and
therefore missed canonicalized draw-results columns. The workbook keeps a
compact summary plus one point-level row per hunt code/point with resident,
nonresident, and total draw-result columns side by side.
"""

from __future__ import annotations

import argparse
import csv
import json
import shutil
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
CANONICAL_DIR = ROOT / "data_truth" / "draw_results_truth" / "normalized" / "canonical_yearly"
OUTPUT_DIR = ROOT / "outputs"
AUDIT_DIR = ROOT / "audits" / "database_alignment" / "identity_registry"

TEXT_COLUMNS = {
    "hunt_code",
    "hunt_name",
    "sex_type",
    "species",
    "hunt_type",
    "weapon",
    "season",
    "draw_design",
    "record_type",
    "source_file",
    "page_kind",
    "algorithm_status",
    "qa_status",
    "notes",
    "resident_success_ratio",
    "nonresident_success_ratio",
    "total_success_ratio",
}


def clean(value: object) -> str:
    return " ".join(str(value or "").strip().split())


def maybe_number(column: str, value: object) -> object:
    text = clean(value).replace(",", "")
    if not text or column in TEXT_COLUMNS:
        return clean(value)
    try:
        number = float(text)
    except ValueError:
        return clean(value)
    return int(number) if number.is_integer() else number


def probability_from_row(row: dict[str, str], prefix: str, percent_value: object) -> str:
    """Return p_draw as a 0-1 decimal string when source data can support it."""
    applicants = maybe_number(f"{prefix}_eligible_applicants", row.get(f"{prefix}_eligible_applicants"))
    permits = maybe_number(f"{prefix}_total_permits", row.get(f"{prefix}_total_permits"))
    if permits == "":
        bonus = maybe_number(f"{prefix}_bonus_permits", row.get(f"{prefix}_bonus_permits"))
        regular = maybe_number(f"{prefix}_regular_permits", row.get(f"{prefix}_regular_permits"))
        if isinstance(bonus, (int, float)) or isinstance(regular, (int, float)):
            permits = int(bonus or 0) + int(regular or 0)
    if isinstance(applicants, (int, float)) and isinstance(permits, (int, float)):
        if applicants > 0:
            return f"{min(1.0, max(0.0, permits / applicants)):.10g}"

    percent = clean(percent_value).replace("%", "")
    if percent:
        try:
            return f"{float(percent) / 100:.10g}"
        except ValueError:
            pass

    ratio = clean(row.get(f"{prefix}_success_ratio"))
    if ratio and ratio.upper() != "N/A" and "in" in ratio.lower():
        try:
            denominator = float(ratio.lower().split("in", 1)[1].strip())
            if denominator > 0:
                return f"{1 / denominator:.10g}"
        except ValueError:
            return ""
    return ""


def read_csv_rows(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), list(reader)


def canonical_path(year: int) -> Path:
    return CANONICAL_DIR / f"draw_results_{year}_for_{year + 1}_canonical_yearly_draw_results.csv"


def first_nonblank(rows: Iterable[dict[str, str]], column: str) -> str:
    for row in rows:
        value = clean(row.get(column))
        if value:
            return value
    return ""


def first_number(rows: Iterable[dict[str, str]], column: str) -> object:
    for row in rows:
        value = maybe_number(column, row.get(column))
        if isinstance(value, (int, float)):
            return value
    return ""


def sum_number(rows: Iterable[dict[str, str]], column: str) -> int:
    total = 0
    for row in rows:
        value = maybe_number(column, row.get(column))
        if isinstance(value, (int, float)):
            total += int(value)
    return total


def unique_join(rows: Iterable[dict[str, str]], column: str, limit: int = 8) -> str:
    values = sorted({clean(row.get(column)) for row in rows if clean(row.get(column))})
    suffix = "" if len(values) <= limit else f"; ... +{len(values) - limit} more"
    return "; ".join(values[:limit]) + suffix


def build_summary(rows: list[dict[str, str]], year: int) -> list[dict[str, object]]:
    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        code = clean(row.get("hunt_code")).upper()
        if code:
            grouped[code].append(row)

    res_col = f"permits_{year}_res"
    nr_col = f"permits_{year}_nr"
    total_col = f"permits_{year}_total"
    output: list[dict[str, object]] = []
    for code, group in grouped.items():
        output.append(
            {
                "ACTUAL DRAW YEAR": year,
                "MODEL TARGET YEAR": year + 1,
                "HUNT CODE": code,
                "BOUNDARY ID": first_number(group, "boundary_id"),
                "SPECIES": first_nonblank(group, "species"),
                "HUNT NAME": first_nonblank(group, "hunt_name"),
                "WEAPON": first_nonblank(group, "weapon"),
                "SEX TYPE": first_nonblank(group, "sex_type"),
                "HUNT TYPE": first_nonblank(group, "hunt_type"),
                "DRAW DESIGN": first_nonblank(group, "draw_design"),
                f"PERMITS {year} RES": first_number(group, res_col),
                f"PERMITS {year} NR": first_number(group, nr_col),
                f"PERMITS {year} TOTAL": first_number(group, total_col),
                "DRAW RESULT ROWS": len(group),
                "RECORD TYPES": unique_join(group, "record_type"),
                "SOURCE FILES": unique_join(group, "source_file", limit=5),
            }
        )
    return sorted(output, key=lambda row: str(row["HUNT CODE"]))


def long_header(year: int) -> list[str]:
    return [
        "ACTUAL DRAW YEAR",
        "MODEL TARGET YEAR",
        "BOUNDARY ID",
        "HUNT CODE",
        "SPECIES",
        "HUNT NAME",
        "WEAPON",
        "SEX TYPE",
        "HUNT TYPE",
        "SEASON",
        "DRAW DESIGN",
        "POINTS",
        "RECORD TYPE",
        f"PERMITS {year} RES",
        f"PERMITS {year} NR",
        f"PERMITS {year} TOTAL",
        "RES ELIGIBLE APPLICANTS",
        "RES BONUS PERMITS",
        "RES REGULAR PERMITS",
        "RES TOTAL PERMITS",
        "RES SUCCESS RATIO",
        "RES P DRAW",
        "NR ELIGIBLE APPLICANTS",
        "NR BONUS PERMITS",
        "NR REGULAR PERMITS",
        "NR TOTAL PERMITS",
        "NR SUCCESS RATIO",
        "NR P DRAW",
        "TOTAL ELIGIBLE APPLICANTS",
        "TOTAL BONUS PERMITS",
        "TOTAL REGULAR PERMITS",
        "TOTAL PERMITS",
        "TOTAL SUCCESS RATIO",
        "TOTAL P DRAW",
        "TOTAL P DRAW PERCENT",
        "SOURCE FILE",
        "PDF PAGE",
    ]


def long_row(row: dict[str, str], year: int) -> dict[str, object]:
    mapping = {
        "ACTUAL DRAW YEAR": ("actual_draw_year",),
        "MODEL TARGET YEAR": ("model_target_year",),
        "BOUNDARY ID": ("boundary_id",),
        "HUNT CODE": ("hunt_code",),
        "SPECIES": ("species",),
        "HUNT NAME": ("hunt_name",),
        "WEAPON": ("weapon",),
        "SEX TYPE": ("sex_type",),
        "HUNT TYPE": ("hunt_type",),
        "SEASON": ("season",),
        "DRAW DESIGN": ("draw_design",),
        "POINTS": ("points",),
        "RECORD TYPE": ("record_type",),
        f"PERMITS {year} RES": (f"permits_{year}_res",),
        f"PERMITS {year} NR": (f"permits_{year}_nr",),
        f"PERMITS {year} TOTAL": (f"permits_{year}_total",),
        "RES ELIGIBLE APPLICANTS": ("resident_eligible_applicants",),
        "RES BONUS PERMITS": ("resident_bonus_permits",),
        "RES REGULAR PERMITS": ("resident_regular_permits",),
        "RES TOTAL PERMITS": ("resident_total_permits",),
        "RES SUCCESS RATIO": ("resident_success_ratio",),
        "RES P DRAW": ("resident_p_draw",),
        "NR ELIGIBLE APPLICANTS": ("nonresident_eligible_applicants",),
        "NR BONUS PERMITS": ("nonresident_bonus_permits",),
        "NR REGULAR PERMITS": ("nonresident_regular_permits",),
        "NR TOTAL PERMITS": ("nonresident_total_permits",),
        "NR SUCCESS RATIO": ("nonresident_success_ratio",),
        "NR P DRAW": ("nonresident_p_draw",),
        "TOTAL ELIGIBLE APPLICANTS": ("total_eligible_applicants",),
        "TOTAL BONUS PERMITS": ("total_bonus_permits",),
        "TOTAL REGULAR PERMITS": ("total_regular_permits",),
        "TOTAL PERMITS": ("total_permits",),
        "TOTAL SUCCESS RATIO": ("total_success_ratio",),
        "TOTAL P DRAW": ("total_p_draw",),
        "TOTAL P DRAW PERCENT": ("total_p_draw_percent",),
        "SOURCE FILE": ("source_file",),
        "PDF PAGE": ("pdf_page",),
    }
    out: dict[str, object] = {}
    for label, (source,) in mapping.items():
        out[label] = maybe_number(source, row.get(source))
    out["RES P DRAW"] = maybe_number("resident_p_draw", probability_from_row(row, "resident", row.get("resident_p_draw_percent")))
    out["NR P DRAW"] = maybe_number("nonresident_p_draw", probability_from_row(row, "nonresident", row.get("nonresident_p_draw_percent")))
    out["TOTAL P DRAW"] = maybe_number("total_p_draw", probability_from_row(row, "total", row.get("total_p_draw_percent")))
    total_p = out.get("TOTAL P DRAW")
    if isinstance(total_p, (int, float)):
        out["TOTAL P DRAW PERCENT"] = total_p * 100
    out["HUNT CODE"] = clean(out.get("HUNT CODE")).upper()
    return out


def style_header(sheet, values: list[str]):
    cells = []
    for value in values:
        cell = WriteOnlyCell(sheet, value=value)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="183A37")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cells.append(cell)
    sheet.append(cells)


def append_table(sheet, header: list[str], rows: Iterable[dict[str, object]]) -> int:
    style_header(sheet, header)
    count = 0
    for row in rows:
        sheet.append([row.get(column, "") for column in header])
        count += 1
    return count


def set_widths(sheet, header: list[str]) -> None:
    widths = {
        "HUNT NAME": 34,
        "SOURCE FILE": 34,
        "SOURCE FILES": 44,
        "NOTES": 36,
        "DRAW DESIGN": 20,
        "HUNT TYPE": 20,
        "WEAPON": 22,
        "SEASON": 26,
        "RECORD TYPES": 32,
    }
    for index, column in enumerate(header, start=1):
        width = widths.get(column, 18 if "APPLICANTS" in column else 14)
        sheet.column_dimensions[get_column_letter(index)].width = width


def audit_rows(rows: list[dict[str, str]], year: int, summary_rows: list[dict[str, object]]) -> list[dict[str, object]]:
    permit_cols = [f"permits_{year}_res", f"permits_{year}_nr", f"permits_{year}_total"]
    record_counts = Counter(clean(row.get("record_type")) or "(blank)" for row in rows)
    return [
        {"item": "actual_draw_year", "value": year},
        {"item": "model_target_year", "value": year + 1},
        {"item": "canonical_source", "value": str(canonical_path(year).relative_to(ROOT)).replace("\\", "/")},
        {"item": "canonical_rows", "value": len(rows)},
        {"item": "summary_rows", "value": len(summary_rows)},
        {"item": "unique_hunt_codes", "value": len({clean(row.get("hunt_code")).upper() for row in rows if clean(row.get("hunt_code"))})},
        {"item": "permit_columns", "value": ", ".join(permit_cols)},
        {"item": "permit_blank_counts", "value": "; ".join(f"{col}: {sum(1 for row in rows if not clean(row.get(col)))}" for col in permit_cols)},
        {"item": "record_type_counts", "value": "; ".join(f"{key}: {value}" for key, value in record_counts.most_common())},
    ]


def export_year(year: int, output_name: str | None = None, report_name: str | None = None, write_backup: bool = True) -> dict[str, object]:
    _, rows = read_csv_rows(canonical_path(year))
    summary_rows = build_summary(rows, year)
    long_rows = [long_row(row, year) for row in rows]

    workbook = Workbook(write_only=True)
    if workbook.worksheets:
        workbook.remove(workbook.worksheets[0])

    summary_header = list(summary_rows[0].keys()) if summary_rows else []
    summary_sheet = workbook.create_sheet(f"{year} Summary")
    append_table(summary_sheet, summary_header, summary_rows)
    set_widths(summary_sheet, summary_header)

    long_sheet = workbook.create_sheet(f"{year} Long")
    lh = long_header(year)
    append_table(long_sheet, lh, long_rows)
    set_widths(long_sheet, lh)

    audit_sheet = workbook.create_sheet("Audit")
    audit_header = ["item", "value"]
    append_table(audit_sheet, audit_header, audit_rows(rows, year, summary_rows))
    set_widths(audit_sheet, audit_header)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output = OUTPUT_DIR / (output_name or f"{year} PERMITS.xlsx")
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    if write_backup and output.exists():
        backup = AUDIT_DIR / "backups" / f"{output.stem}.before_canonical_draw_results_export_{stamp}{output.suffix}"
        backup.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(output, backup)
    workbook.save(output)

    report = {
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "year": year,
        "canonical_source": str(canonical_path(year).relative_to(ROOT)).replace("\\", "/"),
        "output_xlsx": str(output.relative_to(ROOT)).replace("\\", "/"),
        "summary_rows": len(summary_rows),
        "long_rows": len(long_rows),
        "long_columns": len(lh),
        "record_type_counts": dict(Counter(clean(row.get("record_type")) or "(blank)" for row in rows)),
        "permit_blank_counts": {
            f"permits_{year}_res": sum(1 for row in rows if not clean(row.get(f"permits_{year}_res"))),
            f"permits_{year}_nr": sum(1 for row in rows if not clean(row.get(f"permits_{year}_nr"))),
            f"permits_{year}_total": sum(1 for row in rows if not clean(row.get(f"permits_{year}_total"))),
        },
    }
    report_path = OUTPUT_DIR / (report_name or f"{year}_PERMITS_report.json")
    report_path.write_text(json.dumps(report, indent=2) + "\n", encoding="utf-8")
    return report


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--year", type=int, required=True)
    parser.add_argument("--output-name", help="Workbook filename under outputs/. Defaults to '<year> PERMITS.xlsx'.")
    parser.add_argument("--report-name", help="Report filename under outputs/. Defaults to '<year>_PERMITS_report.json'.")
    parser.add_argument("--no-backup", action="store_true")
    args = parser.parse_args()
    report = export_year(args.year, output_name=args.output_name, report_name=args.report_name, write_backup=not args.no_backup)
    print(json.dumps(report, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
