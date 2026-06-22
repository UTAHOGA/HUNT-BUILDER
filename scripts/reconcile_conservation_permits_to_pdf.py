from __future__ import annotations

import csv
import re
import shutil
from datetime import datetime, timezone
from pathlib import Path

import pdfplumber
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PDF_PATH = Path(r"C:\Users\tyler\Desktop\2025-27 Conservation Permits.pdf")
WORKBOOK_PATH = ROOT / "processed_data" / "hard_data_exports" / "hunt_tables" / "2026" / "CLEAN_XLXS_STAGED" / "2025-27 Conservation Permits.xlsx"
AUDIT_DIR = ROOT / "audits" / "2025_canonical_finalization"
BACKUP_DIR = AUDIT_DIR / "backups"


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def parse_money(value: object) -> float | int | str:
    text = clean_text(value)
    if not text:
        return ""
    numeric = text.replace("$", "").replace(",", "")
    try:
        number = float(numeric)
    except ValueError:
        return text
    if number.is_integer():
        return int(number)
    return number


def extract_pdf_rows() -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    with pdfplumber.open(PDF_PATH) as pdf:
        for page_number, page in enumerate(pdf.pages, start=1):
            for table in page.extract_tables():
                for raw in table:
                    if not raw or not clean_text(raw[0]).isdigit():
                        continue
                    row_no = int(clean_text(raw[0]))
                    rows.append(
                        {
                            "pdf_page": page_number,
                            "No.": row_no,
                            "Species": clean_text(raw[1]) if len(raw) > 1 else "",
                            "Area": clean_text(raw[2]) if len(raw) > 2 else "",
                            "Condition": clean_text(raw[3]) if len(raw) > 3 else "",
                            "Value": parse_money(raw[4]) if len(raw) > 4 else "",
                            "blank_col_6": int(clean_text(raw[5])) if len(raw) > 5 and clean_text(raw[5]).isdigit() else clean_text(raw[5]) if len(raw) > 5 else "",
                            "Organization": clean_text(raw[6]) if len(raw) > 6 else "",
                        }
                    )
    return rows


def value_for_compare(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return clean_text(value).replace("$", "").replace(",", "")


def main() -> None:
    if not PDF_PATH.exists():
        raise FileNotFoundError(PDF_PATH)
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(WORKBOOK_PATH)

    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)

    pdf_rows = extract_pdf_rows()
    pdf_by_no = {int(row["No."]): row for row in pdf_rows}
    expected = list(range(1, 337))
    actual = sorted(pdf_by_no)
    if len(pdf_rows) != 336 or actual != expected:
        raise RuntimeError(f"PDF did not extract as rows 1-336. count={len(pdf_rows)} first_last={actual[:3]}...{actual[-3:]}")

    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    backup_path = BACKUP_DIR / f"{WORKBOOK_PATH.stem}.before_pdf_reconcile_{stamp}{WORKBOOK_PATH.suffix}"
    shutil.copy2(WORKBOOK_PATH, backup_path)

    wb = load_workbook(WORKBOOK_PATH)
    ws = wb["Table 1"] if "Table 1" in wb.sheetnames else wb[wb.sheetnames[0]]

    audit_path = AUDIT_DIR / "2025_27_conservation_pdf_reconcile_audit.csv"
    audit_rows: list[dict[str, object]] = []

    row_by_no: dict[int, int] = {}
    rows_to_delete: list[int] = []
    for excel_row in range(2, ws.max_row + 1):
        no_value = ws.cell(excel_row, 1).value
        no_text = clean_text(no_value)
        if not no_text:
            continue
        try:
            no_int = int(float(no_text))
        except ValueError:
            rows_to_delete.append(excel_row)
            audit_rows.append({"status": "removed_non_numeric_no", "excel_row": excel_row, "workbook_no": no_text})
            continue
        if no_int in pdf_by_no and no_int not in row_by_no:
            row_by_no[no_int] = excel_row
        else:
            rows_to_delete.append(excel_row)
            status = "removed_not_in_pdf" if no_int not in pdf_by_no else "removed_duplicate_no"
            audit_rows.append({"status": status, "excel_row": excel_row, "workbook_no": no_int})

    missing = [n for n in expected if n not in row_by_no]
    if missing:
        raise RuntimeError(f"Workbook is missing PDF row numbers: {missing[:25]} count={len(missing)}")

    source_columns = [
        ("No.", 1),
        ("Species", 2),
        ("Area", 3),
        ("Condition", 4),
        ("Value", 5),
        ("blank_col_6", 6),
        ("Organization", 7),
    ]

    changed_cells = 0
    for no_int in expected:
        excel_row = row_by_no[no_int]
        pdf_row = pdf_by_no[no_int]
        for pdf_key, col_idx in source_columns:
            old_value = ws.cell(excel_row, col_idx).value
            new_value = pdf_row[pdf_key]
            if value_for_compare(old_value) != value_for_compare(new_value):
                changed_cells += 1
                audit_rows.append(
                    {
                        "status": "source_cell_reset_to_pdf",
                        "excel_row": excel_row,
                        "workbook_no": no_int,
                        "column": ws.cell(1, col_idx).value if ws.cell(1, col_idx).value is not None else "blank_col_6",
                        "old_value": old_value,
                        "pdf_value": new_value,
                    }
                )
            ws.cell(excel_row, col_idx).value = new_value

    for excel_row in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(excel_row, 1)

    wb.save(WORKBOOK_PATH)

    with audit_path.open("w", newline="", encoding="utf-8") as f:
        fieldnames = ["status", "excel_row", "workbook_no", "column", "old_value", "pdf_value"]
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(audit_rows)

    print(f"pdf_rows={len(pdf_rows)}")
    print(f"rows_removed={len(rows_to_delete)}")
    print(f"source_cells_reset={changed_cells}")
    print(f"backup={backup_path}")
    print(f"audit={audit_path}")
    print(f"workbook={WORKBOOK_PATH}")


if __name__ == "__main__":
    main()
