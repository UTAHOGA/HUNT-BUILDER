from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
CLEAN_XLXS_DIR = ROOT / "processed_data" / "hard_data_exports" / "hunt_tables" / "2026" / "CLEAN_XLXS_STAGED"
OUTPUT_DIR = ROOT / "processed_data" / "hard_data_exports" / "hunt_tables" / "2026" / "GENERATED_CSV"
BUILD_REPORT = ROOT / "processed_data" / "hard_data_exports" / "hunt_tables" / "2026" / "_PRIMARY_TABLES_BUILD_REPORT.json"
AUDIT_PATH = ROOT / "processed_data" / "audits" / "hunt_tables_2026_clean_xlsx_to_csv_audit.json"

EXPECTED_FIELDS = {
    "hunt name": "hunt_name",
    "hunt_code": "hunt_code",
    "sex": "sex_type",
    "species": "species",
    "weapon": "weapon",
    "hunt type": "hunt_type",
    "season": "season",
    "res": "permits_2026_res",
    "non-res": "permits_2026_nr",
    "total": "permits_2026_total",
    "hunt class": "hunt_class",
    "harvest success prior year pct": "harvest_success_prior_year_pct",
    "average harvest age": "average_harvest_age",
    "current age (3-yr avg)": "current_age_3yr_avg",
    "avg days hunted prior year": "avg_days_hunted_prior_year",
}

OUTPUT_COLUMNS = [
    "source_file",
    "source_sheet",
    "hunt_name",
    "hunt_code",
    "sex_type",
    "species",
    "weapon",
    "hunt_type",
    "season",
    "permits_2026_res",
    "permits_2026_nr",
    "permits_2026_total",
    "hunt_class",
    "harvest_success_prior_year_pct",
    "average_harvest_age",
    "current_age_3yr_avg",
    "avg_days_hunted_prior_year",
]


def clean(v):
    return "" if v is None else str(v).strip()


def normalize_header_value(value):
    return clean(value).strip().lower()


def parse_header_row(values: List[str]):
    normalized = [normalize_header_value(v) for v in values]
    if not any(v for v in normalized):
        return None

    if "hunt code" not in normalized and "hunt_code" not in normalized:
        return None

    header_map: Dict[str, int] = {}
    for idx, header in enumerate(normalized):
        if not header:
            continue
        if header in EXPECTED_FIELDS:
            header_map[EXPECTED_FIELDS[header]] = idx
            continue
        if header == "hunt" and idx + 1 < len(normalized) and normalize_header_value(values[idx + 1]) == "name":
            # some odd merged headers; ignore in this context.
            continue

    # allow permissive header if core fields exist
    if "hunt_name" not in header_map and "hunt name" in normalized:
        header_map["hunt_name"] = normalized.index("hunt name")
    if "hunt_code" not in header_map and "hunt code" in normalized:
        header_map["hunt_code"] = normalized.index("hunt code")

    if "hunt_name" not in header_map and "hunt_code" not in header_map:
        return None
    return header_map


def detect_header_row(ws) -> tuple[int, Dict[str, int]]:
    max_rows = min(ws.max_row, 25)
    best_idx = None
    best_map = None
    for row_idx in range(1, max_rows + 1):
        values = [ws.cell(row=row_idx, column=col).value for col in range(1, min(ws.max_column, 30) + 1)]
        header_map = parse_header_row(values)
        if header_map is None:
            continue

        # prefer row containing Hunt Code + at least 4 of 6 required permit fields
        required = len(set(["hunt_name", "hunt_code", "hunt_type", "species", "weapon", "season"]).intersection(header_map))
        permit_like = len([k for k in ("permits_2026_res", "permits_2026_nr", "permits_2026_total") if k in header_map])
        score = required + permit_like
        if best_map is None or score > (len([k for k in ("hunt_name", "hunt_code", "hunt_type", "species", "weapon", "season") if k in best_map]) + len([k for k in ("permits_2026_res", "permits_2026_nr", "permits_2026_total") if k in best_map])):
            best_idx = row_idx
            best_map = header_map
        if score >= 7:
            return row_idx, header_map

    if best_idx is None:
        raise RuntimeError("No header row found")
    return best_idx, best_map or {}


def read_source_list() -> List[str]:
    # Prefer explicit source manifest; fallback to discovered xlsx files.
    source_xlsx_files: List[str] = []
    if BUILD_REPORT.exists():
        data = json.loads(BUILD_REPORT.read_text(encoding="utf-8"))
        for row in data:
            xlsx_name = row.get("xlsx") or ""
            if xlsx_name and xlsx_name.lower().endswith(".xlsx"):
                source_xlsx_files.append(xlsx_name)

    if not source_xlsx_files:
        for p in CLEAN_XLXS_DIR.glob("*.xlsx"):
            if p.name in {"MASTER.xlsx", "DISPLAY_READY.xlsx"}:
                continue
            if "SAMPLES" in p.name.upper():
                continue
            source_xlsx_files.append(p.name)

    return sorted(set(source_xlsx_files))


def write_rows_to_csv(path: Path, rows: List[Dict[str, str]]):
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow({col: row.get(col, "") for col in OUTPUT_COLUMNS})


def parse_xlsx(path: Path) -> List[Dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    header_row, header_map = detect_header_row(ws)

    rows: List[Dict[str, str]] = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        hunt_code = clean(ws.cell(row=row_idx, column=(header_map.get("hunt_code", 1) + 1)).value)
        hunt_name = clean(ws.cell(row=row_idx, column=(header_map.get("hunt_name", 0) + 1)).value)

        if not hunt_code and not hunt_name:
            continue

        record = {col: "" for col in OUTPUT_COLUMNS}
        record["source_file"] = path.name
        record["source_sheet"] = ws.title

        for target, col_idx in header_map.items():
            if target not in OUTPUT_COLUMNS:
                continue
            col = col_idx + 1
            record[target] = clean(ws.cell(row=row_idx, column=col).value)

        # Keep numeric-style columns present in legacy title rows as fallbacks.
        if not record["permits_2026_res"] and not record["permits_2026_nr"] and not record["permits_2026_total"]:
            # some legacy tables use alternate header labels
            for alt, target in (("res", "permits_2026_res"), ("non-res", "permits_2026_nr"), ("total", "permits_2026_total")):
                idx = header_map.get(alt)
                if idx is not None:
                    value = clean(ws.cell(row=row_idx, column=idx + 1).value)
                    if value:
                        record[target] = value

        rows.append(record)

    wb.close()
    return rows


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    source_names = read_source_list()

    summary = {
        "created_at": "",
        "rows_written": 0,
        "files": {},
    }

    total_rows = 0
    for name in source_names:
        path = CLEAN_XLXS_DIR / name
        if not path.exists():
            summary["files"][name] = {"status": "missing", "rows": 0}
            continue

        rows = parse_xlsx(path)
        out_path = OUTPUT_DIR / (path.stem + ".csv")
        write_rows_to_csv(out_path, rows)
        summary["files"][name] = {"status": "ok", "rows": len(rows), "csv": str(out_path)}
        total_rows += len(rows)

    summary["rows_written"] = total_rows
    summary["created_at"] = __import__("datetime").datetime.utcnow().isoformat()
    with AUDIT_PATH.open("w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2)

    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
