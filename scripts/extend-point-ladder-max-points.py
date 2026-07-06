#!/usr/bin/env python3
"""Extend point_ladder_view.csv to a minimum displayed point ceiling.

The Research page renders point rungs from the runtime ladder feed. This script
adds missing high-point display rows without inventing probabilities. If a
matching UtahDraws XLSX row exists, official applicant/success fields are copied
into the new rung; otherwise the row remains structural/display-only.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


REPO = Path(__file__).resolve().parents[1]
DEFAULT_LADDER = REPO / "processed_data" / "point_ladder_view.csv"
DEFAULT_DRAW_ODDS_DIR = REPO / "outputs" / "20260626_fresh_2026_source_species_docs" / "draw_odds_xlsx"
DEFAULT_MIN_MAX_POINTS = 35
DEFAULT_MIN_EXISTING_MAX_POINT = 32

GROUP_FIELDS = [
    "hunt_code",
    "residency",
    "draw_pool",
    "draw_system_type",
    "hunt_category",
    "hunt_type",
    "species",
    "hunt_class",
    "probability_model",
    "algorithm_status",
]

CLEAR_FOR_STRUCTURAL_ROWS = {
    "applicants",
    "applicants_above",
    "applicants_at_level",
    "bonus_permits",
    "regular_permits",
    "total_permits",
    "success_ratio",
    "dwr_result_display",
    "display_2025_draw_results",
    "display_2026_max_point_pool",
    "display_2026_random_draw",
    "random_draw_model_source",
    "p_max_pool_mean",
    "p_random_mean",
    "p_draw_mean",
    "p_prior_year_baseline",
    "p_quota_adjusted",
    "p_rollover_adjusted",
    "p_harvest_adjusted",
    "p_preference_mean",
    "p_sportsman_draw",
    "p_draw_p10",
    "p_draw_p50",
    "p_draw_p90",
    "display_odds_pct",
    "display_odds_text",
    "forecast_applicants_at_level",
    "forecast_applicants_above",
    "rolled_applicants",
    "projected_applicants",
    "projected_nonwinners_from_prior_year",
    "projected_new_or_returning_applicants",
    "prior_year_applicants",
    "prior_year_success_count",
    "prior_year_success_rate",
    "prior_year_draw_odds_pct",
}


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def to_int(value: object) -> int | None:
    text = clean(value).replace(",", "")
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def number_text(value: float | int | None, digits: int = 6) -> str:
    if value is None:
        return ""
    value = float(value)
    if abs(value - round(value)) < 1e-9:
        return str(int(round(value)))
    return f"{value:.{digits}f}".rstrip("0").rstrip(".")


def display_odds(successful: int, eligible: int) -> str:
    if successful <= 0 or eligible <= 0:
        return ""
    pct = successful / eligible * 100
    one_in = eligible / successful
    return f"~1 in {round(one_in, 1)} or {number_text(pct, 1)}%"


def safe_rel(path: Path) -> str:
    try:
        return path.relative_to(REPO).as_posix()
    except ValueError:
        return str(path)


def normalize_residency(value: object) -> str:
    text = clean(value).lower()
    if text.startswith("res"):
        return "Resident"
    if text.startswith("non"):
        return "Nonresident"
    return clean(value)


def row_key(row: dict[str, str]) -> tuple[str, str, str, str]:
    return (
        clean(row.get("hunt_code")).upper(),
        normalize_residency(row.get("residency")),
        clean(row.get("points") or row.get("point")),
        clean(row.get("draw_pool") or "standard"),
    )


def group_key(row: dict[str, str]) -> tuple[str, ...]:
    return tuple(clean(row.get(field)).upper() if field == "hunt_code" else clean(row.get(field)) for field in GROUP_FIELDS)


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), [dict(row) for row in reader]


def write_csv(path: Path, fields: list[str], rows: list[dict[str, str]]) -> None:
    tmp = path.with_suffix(path.suffix + ".tmp")
    with tmp.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore", lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)
    os.replace(tmp, path)


def read_draw_odds(draw_odds_dir: Path) -> dict[tuple[str, str, str], dict[str, str]]:
    by_key: dict[tuple[str, str, str], dict[str, str]] = {}
    duplicates: list[tuple[str, str, str]] = []
    for path in sorted(draw_odds_dir.glob("*.xlsx")):
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            if "Table 1" not in wb.sheetnames:
                continue
            ws = wb["Table 1"]
            rows = ws.iter_rows(values_only=True)
            try:
                headers = [clean(cell) for cell in next(rows)]
            except StopIteration:
                continue
            for raw in rows:
                row = {headers[index]: raw[index] if index < len(raw) else "" for index in range(len(headers))}
                code = clean(row.get("hunt_code")).upper()
                residency = normalize_residency(row.get("residency"))
                points = number_text(to_int(row.get("points")))
                category = clean(row.get("hunt_category"))
                if not code or not residency or not points:
                    continue
                if category.lower() in {"bonus point", "preference point"}:
                    continue
                key = (code, residency, points)
                item = {
                    "hunt_code": code,
                    "residency": residency,
                    "points": points,
                    "eligible_applicants": number_text(to_int(row.get("eligible_applicants"))),
                    "successful_bonus": number_text(to_int(row.get("successful_bonus"))),
                    "successful_regular": number_text(to_int(row.get("successful_regular"))),
                    "successful_total": number_text(to_int(row.get("successful_total"))),
                    "total_quota": number_text(to_int(row.get("total_quota"))),
                    "source_file": safe_rel(path),
                    "source_json_file": clean(row.get("source_json_file")),
                    "hunt_category": category,
                    "master_hunt_type_name": clean(row.get("master_hunt_type_name")),
                }
                if key in by_key:
                    duplicates.append(key)
                by_key[key] = item
        finally:
            wb.close()
    return by_key


def apply_draw_source(row: dict[str, str], source: dict[str, str]) -> None:
    eligible = to_int(source.get("eligible_applicants")) or 0
    successful_bonus = to_int(source.get("successful_bonus")) or 0
    successful_regular = to_int(source.get("successful_regular")) or 0
    successful_total = to_int(source.get("successful_total")) or successful_bonus + successful_regular
    p_draw = (successful_total / eligible) if eligible > 0 else None

    row["applicants"] = number_text(eligible)
    row["eligible_applicants"] = number_text(eligible)
    row["bonus_permits"] = number_text(successful_bonus)
    row["regular_permits"] = number_text(successful_regular)
    row["total_permits"] = number_text(successful_total)
    row["success_ratio"] = f"1 in {number_text(eligible / successful_total, 1)}" if successful_total > 0 and eligible > 0 else ""
    row["dwr_result_display"] = display_odds(successful_total, eligible)
    row["display_2025_draw_results"] = row["dwr_result_display"]
    row["p_draw_mean"] = number_text(p_draw)
    row["display_odds_pct"] = number_text(p_draw * 100 if p_draw is not None else None, 4)
    row["display_odds_text"] = row["dwr_result_display"]
    row["truth_source_file"] = source["source_file"]
    row["truth_source_status"] = "OFFICIAL_UTAHDRAWS_2026_POINT_ROW"
    row["data_quality_grade"] = "A"
    row["reason_codes"] = "LADDER_MAX_POINT_EXTENSION|OFFICIAL_UTAHDRAWS_POINT_ROW"
    row["data_status"] = "COMPLETE"


def structuralize(row: dict[str, str], point: int) -> None:
    row["points"] = str(point)
    row["point"] = str(point)
    for field in CLEAR_FOR_STRUCTURAL_ROWS:
        if field in row:
            row[field] = ""
    if "data_status" in row:
        row["data_status"] = "STRUCTURAL_POINT_DISPLAY_ROW"
    if "truth_source_status" in row:
        row["truth_source_status"] = "STRUCTURAL_POINT_DISPLAY_ROW"
    if "reason_codes" in row:
        row["reason_codes"] = "LADDER_MAX_POINT_EXTENSION|NO_OFFICIAL_POINT_ROW"
    if "status" in row:
        row["status"] = "DISPLAY ONLY"
    if "draw_outlook" in row:
        row["draw_outlook"] = "DISPLAY ONLY"
    if "trend" in row:
        row["trend"] = ""


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--ladder", type=Path, default=DEFAULT_LADDER)
    parser.add_argument("--draw-odds-dir", type=Path, default=DEFAULT_DRAW_ODDS_DIR)
    parser.add_argument("--min-max-points", type=int, default=DEFAULT_MIN_MAX_POINTS)
    parser.add_argument("--min-existing-max-point", type=int, default=DEFAULT_MIN_EXISTING_MAX_POINT)
    parser.add_argument("--audit-dir", type=Path)
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    audit_dir = args.audit_dir or (REPO / "audits" / "point_ladder_max_point_extension" / stamp)
    audit_dir.mkdir(parents=True, exist_ok=True)

    fields, rows = read_csv(args.ladder)
    draw_sources = read_draw_odds(args.draw_odds_dir)
    existing_keys = {row_key(row) for row in rows if row_key(row)[0] and row_key(row)[2]}

    grouped: dict[tuple[str, ...], list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        point = to_int(row.get("points"))
        if point is None:
            continue
        grouped[group_key(row)].append(row)

    additions: list[dict[str, str]] = []
    audit_rows: list[dict[str, str]] = []
    for key, group_rows in grouped.items():
        numeric_points = [to_int(row.get("points")) for row in group_rows]
        numeric_points = [point for point in numeric_points if point is not None]
        if not numeric_points:
            continue
        max_existing = max(numeric_points)
        if max_existing >= args.min_max_points:
            continue
        if max_existing < args.min_existing_max_point:
            continue
        template = max(group_rows, key=lambda row: to_int(row.get("points")) or -1)
        for point in range(max_existing + 1, args.min_max_points + 1):
            new_row = dict(template)
            structuralize(new_row, point)
            key4 = row_key(new_row)
            if key4 in existing_keys:
                continue
            source_key = (key4[0], key4[1], key4[2])
            source = draw_sources.get(source_key)
            source_status = "STRUCTURAL_DISPLAY_ONLY"
            if source:
                apply_draw_source(new_row, source)
                source_status = "OFFICIAL_UTAHDRAWS_POINT_ROW"
            additions.append(new_row)
            existing_keys.add(key4)
            audit_rows.append(
                {
                    "hunt_code": key4[0],
                    "residency": key4[1],
                    "draw_pool": key4[3],
                    "point": str(point),
                    "source_status": source_status,
                    "source_file": source.get("source_file", "") if source else "",
                    "template_max_existing_point": str(max_existing),
                    "hunt_name": clean(new_row.get("hunt_name")),
                    "species": clean(new_row.get("species")),
                    "draw_system_type": clean(new_row.get("draw_system_type")),
                }
            )

    output_rows = rows + additions
    output_rows.sort(
        key=lambda row: (
            clean(row.get("hunt_code")).upper(),
            normalize_residency(row.get("residency")),
            clean(row.get("draw_pool") or "standard"),
            to_int(row.get("points")) if to_int(row.get("points")) is not None else -1,
        )
    )

    if args.apply and additions:
        write_csv(args.ladder, fields, output_rows)

    with (audit_dir / "POINT_LADDER_MAX_POINT_EXTENSION_ROWS.csv").open("w", newline="", encoding="utf-8") as handle:
        fieldnames = [
            "hunt_code",
            "residency",
            "draw_pool",
            "point",
            "source_status",
            "source_file",
            "template_max_existing_point",
            "hunt_name",
            "species",
            "draw_system_type",
        ]
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(audit_rows)

    counts = Counter(row["source_status"] for row in audit_rows)
    summary = {
        "classification": "POINT_LADDER_MAX_POINT_EXTENSION_APPLIED" if args.apply else "POINT_LADDER_MAX_POINT_EXTENSION_DRY_RUN",
        "apply": args.apply,
        "ladder": safe_rel(args.ladder),
        "draw_odds_dir": safe_rel(args.draw_odds_dir),
        "min_max_points": args.min_max_points,
        "min_existing_max_point": args.min_existing_max_point,
        "rows_before": len(rows),
        "rows_added": len(additions),
        "rows_after": len(output_rows),
        "groups_seen": len(grouped),
        "source_status_counts": dict(counts),
        "audit_dir": safe_rel(audit_dir),
    }
    (audit_dir / "POINT_LADDER_MAX_POINT_EXTENSION_SUMMARY.json").write_text(
        json.dumps(summary, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    (audit_dir / "POINT_LADDER_MAX_POINT_EXTENSION_SUMMARY.md").write_text(
        "\n".join(
            [
                "# Point Ladder Max Point Extension",
                "",
                f"- Classification: `{summary['classification']}`",
                f"- Target max point: `{args.min_max_points}`",
                f"- Rows before: `{summary['rows_before']}`",
                f"- Rows added: `{summary['rows_added']}`",
                f"- Rows after: `{summary['rows_after']}`",
                f"- Official UtahDraws point rows added: `{counts.get('OFFICIAL_UTAHDRAWS_POINT_ROW', 0)}`",
                f"- Structural display-only rows added: `{counts.get('STRUCTURAL_DISPLAY_ONLY', 0)}`",
                "",
                "Structural rows do not invent probabilities. Official values are populated only when a matching UtahDraws point row exists.",
                "",
            ]
        ),
        encoding="utf-8",
    )
    print(json.dumps(summary, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
