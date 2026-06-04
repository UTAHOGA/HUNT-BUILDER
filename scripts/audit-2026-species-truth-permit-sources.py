"""Normalize user-supplied 2026 species truth permit files and compare them.

This is an audit-only pass. It does not modify DATABASE.csv.
"""

from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = Path("C:/Users/tyler/Desktop/species truth data")
CODE_RE = re.compile(r"^[A-Z]{2,3}\d{3,4}$")

EXPO_DRAW = SOURCE_DIR / "2026 EXPO DRAW RESULTS.csv"
EXPO_PERMIT = SOURCE_DIR / "2026 EXPO PERMIT DRAW.csv"
CONSERVATION = SOURCE_DIR / "2026 CONSERVATION  PERMITS.xlsx"
DEER_BUCK = SOURCE_DIR / "2026 deer buck db.csv"
DEER_DOE = SOURCE_DIR / "2026 deer doe.csv"

DATABASE = ROOT / "pipeline/RAW/hunt_unit_database/2026/csv/DATABASE.csv"
RECON = ROOT / "processed_data/audits/current_2026_hunt_code_permit_reconciliation.csv"

OUT_NORMALIZED = ROOT / "processed_data/audits/permit_2026_species_truth_sources_normalized.csv"
OUT_COMPARE = ROOT / "processed_data/audits/permit_2026_species_truth_sources_vs_current_reconciliation.csv"
OUT_SUMMARY = ROOT / "processed_data/audits/permit_2026_species_truth_sources_summary.json"
OUT_DOC = ROOT / "docs/permit_2026_species_truth_sources_audit.md"


def clean(value: object) -> str:
    text = str(value or "").replace("\ufeff", "")
    return re.sub(r"\s+", " ", text).strip()


def code(value: object) -> str:
    return re.sub(r"[^A-Z0-9]", "", clean(value).upper())


def valid_hunt_code(value: object) -> bool:
    return bool(CODE_RE.match(code(value)))


def int_text(value: object) -> str:
    text = clean(value).replace(",", "")
    if not text:
        return ""
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return ""
    number = float(match.group(0))
    return str(int(number)) if number.is_integer() else str(number)


def has_value(res: str, nr: str, total: str) -> bool:
    return any(v not in {"", "0"} for v in (res, nr, total))


def normalize_total_shape(res: object = "", nr: object = "", total: object = "") -> tuple[str, str, str, str]:
    r = int_text(res)
    n = int_text(nr)
    t = int_text(total)
    res_label = clean(res).lower()
    nr_label = clean(nr).lower()
    total_label = clean(total).lower()

    if "total" in res_label and not t:
        t = r
        r = ""
    if "nonres" in res_label and not n:
        n = r
        r = ""
    if "res" in res_label and "nonres" not in res_label:
        pass
    if "total" in nr_label and not t:
        t = n
        n = ""
    if "total" in total_label:
        t = int_text(total)

    if not t and (r or n):
        t = str(int(r or 0) + int(n or 0))
    if r == "0":
        r = ""
    if n == "0":
        n = ""
    if t == "0":
        t = ""
    if r or n:
        return r, n, t, "RES_NR_SPLIT"
    if t:
        return "", "", t, "TOTAL_ONLY"
    return "", "", "", "NO_PERMIT_VALUE"


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, rows: list[dict[str, object]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def source_row(
    *,
    source_family: str,
    source_file: Path,
    source_row_count: int,
    hunt_code: str = "",
    hunt_name: str = "",
    species: str = "",
    sex_type: str = "",
    weapon: str = "",
    hunt_type: str = "",
    hunt_class: str = "",
    res: str = "",
    nr: str = "",
    total: str = "",
    permit_shape: str = "",
    mapping_status: str = "DIRECT_HUNT_CODE",
    notes: str = "",
) -> dict[str, str]:
    return {
        "source_family": source_family,
        "source_file": str(source_file).replace("\\", "/"),
        "source_row_count": str(source_row_count),
        "hunt_code": code(hunt_code),
        "source_hunt_name": clean(hunt_name),
        "species": clean(species),
        "sex_type": clean(sex_type),
        "weapon": clean(weapon),
        "hunt_type": clean(hunt_type),
        "hunt_class": clean(hunt_class),
        "source_res": res,
        "source_nr": nr,
        "source_total": total,
        "source_permit_shape": permit_shape,
        "mapping_status": mapping_status,
        "notes": notes,
    }


def normalize_deer_buck() -> list[dict[str, str]]:
    out = []
    for row in read_csv(DEER_BUCK):
        r, n, t, shape = normalize_total_shape(
            row.get("2026 permits res"),
            row.get("2026 permits non-res"),
            row.get("2026 permits total"),
        )
        out.append(
            source_row(
                source_family="DEER_BUCK_DB_DIRECT",
                source_file=DEER_BUCK,
                source_row_count=1,
                hunt_code=row.get("hunt_code", ""),
                hunt_name=row.get("hunt_name", ""),
                species=row.get("species", ""),
                sex_type=row.get("sex_type", ""),
                weapon=row.get("weapon", ""),
                hunt_type=row.get("hunt_type", ""),
                hunt_class=row.get("hunt_class", ""),
                res=r,
                nr=n,
                total=t,
                permit_shape=shape,
            )
        )
    return out


def normalize_deer_doe() -> list[dict[str, str]]:
    grouped: dict[str, dict[str, str]] = {}
    last_code = ""
    for row in read_csv(DEER_DOE):
        current_code = code(row.get("hunt_code"))
        value = clean(row.get("permits_2026_res"))
        if current_code:
            last_code = current_code
            r, n, t, shape = normalize_total_shape(value, "", row.get("permits_2026_total", ""))
            grouped[current_code] = {
                "hunt_code": current_code,
                "hunt_name": row.get("hunt_name", ""),
                "species": row.get("species", ""),
                "sex_type": row.get("sex_type", ""),
                "weapon": row.get("weapon", ""),
                "hunt_type": row.get("hunt_type", ""),
                "hunt_class": row.get("hunt_class", ""),
                "res": r,
                "nr": n,
                "total": t,
                "shape": shape,
                "continuation_rows": "0",
            }
        elif last_code and "nonres" in value.lower():
            n = int_text(value)
            rec = grouped[last_code]
            rec["nr"] = "" if n == "0" else n
            if rec["res"] or rec["nr"]:
                rec["total"] = str(int(rec["res"] or 0) + int(rec["nr"] or 0))
                rec["shape"] = "RES_NR_SPLIT"
            rec["continuation_rows"] = str(int(rec["continuation_rows"]) + 1)

    out = []
    for rec in grouped.values():
        out.append(
            source_row(
                source_family="DEER_DOE_DIRECT",
                source_file=DEER_DOE,
                source_row_count=1 + int(rec["continuation_rows"]),
                hunt_code=rec["hunt_code"],
                hunt_name=rec["hunt_name"],
                species=rec["species"],
                sex_type=rec["sex_type"],
                weapon=rec["weapon"],
                hunt_type=rec["hunt_type"],
                hunt_class=rec["hunt_class"],
                res=rec["res"],
                nr=rec["nr"],
                total=rec["total"],
                permit_shape=rec["shape"],
                notes=f"continuation_rows={rec['continuation_rows']}",
            )
        )
    return out


def normalize_conservation() -> list[dict[str, str]]:
    wb = load_workbook(CONSERVATION, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = [clean(v).replace("\n", " ") for v in rows[1]]
    idx = {name: i for i, name in enumerate(header)}
    grouped: dict[str, dict[str, object]] = {}
    for raw in rows[2:]:
        c = code(raw[idx["HUNT CODE"]])
        if not valid_hunt_code(c):
            continue
        rec = grouped.setdefault(
            c,
            {
                "hunt_code": c,
                "hunt_name": clean(raw[idx["HUNT NAME"]]),
                "sex_type": clean(raw[idx["SEX"]]).title(),
                "species": clean(raw[idx["Species"]]),
                "weapon": clean(raw[idx["WEAPON"]]),
                "hunt_type": clean(raw[idx["HUNT TYPE"]]),
                "hunt_class": set(),
                "count": 0,
            },
        )
        rec["count"] = int(rec["count"]) + 1
        rec["hunt_class"].add(clean(raw[idx["HUNT CLASS"]]))
    out = []
    for rec in grouped.values():
        total = str(rec["count"])
        out.append(
            source_row(
                source_family="CONSERVATION_PERMITS_DIRECT",
                source_file=CONSERVATION,
                source_row_count=int(rec["count"]),
                hunt_code=str(rec["hunt_code"]),
                hunt_name=str(rec["hunt_name"]),
                species=str(rec["species"]),
                sex_type=str(rec["sex_type"]),
                weapon=str(rec["weapon"]),
                hunt_type=str(rec["hunt_type"]),
                hunt_class="|".join(sorted(x for x in rec["hunt_class"] if x)),
                total=total,
                permit_shape="TOTAL_ONLY",
                notes="one workbook row counted as one conservation permit; Value column is not used as permit count",
            )
        )
    return out


def normalize_expo_draw() -> list[dict[str, str]]:
    grouped: dict[str, dict[str, object]] = defaultdict(lambda: {"permits": set(), "count": 0})
    for row in read_csv(EXPO_DRAW):
        name = clean(row.get("HUNT NAME"))
        if not name:
            continue
        grouped[name]["permits"].add(int_text(row.get("PERMITS")))
        grouped[name]["count"] = int(grouped[name]["count"]) + 1
    out = []
    for name, rec in grouped.items():
        values = sorted(v for v in rec["permits"] if v)
        permit_total = values[0] if len(values) == 1 else str(rec["count"])
        notes = f"winner_rows={rec['count']}; permits_values={'|'.join(values)}"
        if len(values) == 1 and int(values[0]) != int(rec["count"]):
            notes += "; winner_row_count_differs_from_permit_value"
        out.append(
            source_row(
                source_family="EXPO_DRAW_RESULTS_NAME_ONLY",
                source_file=EXPO_DRAW,
                source_row_count=int(rec["count"]),
                hunt_name=name,
                total=permit_total,
                permit_shape="TOTAL_ONLY",
                mapping_status="NAME_ONLY_NO_HUNT_CODE",
                notes=notes,
            )
        )
    return out


def normalize_expo_permit_draw() -> list[dict[str, str]]:
    out = []
    current_name = ""
    current_permits = ""
    current_count = 0
    pattern = re.compile(r"^(.*) - Permits:\s*(\d+)\s*$")
    for row in read_csv(EXPO_PERMIT):
        text = clean(row.get("HUNT UNIT / PERMITS / NAME"))
        match = pattern.match(text)
        if match:
            if current_name:
                out.append(
                    source_row(
                        source_family="EXPO_PERMIT_DRAW_NAME_ONLY",
                        source_file=EXPO_PERMIT,
                        source_row_count=current_count,
                        hunt_name=current_name,
                        total=current_permits,
                        permit_shape="TOTAL_ONLY",
                        mapping_status="NAME_ONLY_NO_HUNT_CODE",
                        notes=f"winner_rows={current_count}",
                    )
                )
            current_name = match.group(1)
            current_permits = match.group(2)
            current_count = 0
            continue
        if text and text.lower() != "name" and clean(row.get("CITY, STATE")):
            current_count += 1
    if current_name:
        out.append(
            source_row(
                source_family="EXPO_PERMIT_DRAW_NAME_ONLY",
                source_file=EXPO_PERMIT,
                source_row_count=current_count,
                hunt_name=current_name,
                total=current_permits,
                permit_shape="TOTAL_ONLY",
                mapping_status="NAME_ONLY_NO_HUNT_CODE",
                notes=f"winner_rows={current_count}",
            )
        )
    return out


def compare_value(source: tuple[str, str, str], target: tuple[str, str, str]) -> str:
    if not has_value(*source):
        return "SOURCE_BLANK"
    if not has_value(*target):
        return "TARGET_BLANK"
    if source == target:
        return "MATCH"
    if source[2] and target[2] and source[2] == target[2]:
        return "TOTAL_MATCH_SPLIT_DIFFERS"
    return "MISMATCH"


def main() -> int:
    normalized = []
    normalized.extend(normalize_deer_buck())
    normalized.extend(normalize_deer_doe())
    normalized.extend(normalize_conservation())
    normalized.extend(normalize_expo_draw())
    normalized.extend(normalize_expo_permit_draw())

    fields = [
        "source_family",
        "source_file",
        "source_row_count",
        "hunt_code",
        "source_hunt_name",
        "species",
        "sex_type",
        "weapon",
        "hunt_type",
        "hunt_class",
        "source_res",
        "source_nr",
        "source_total",
        "source_permit_shape",
        "mapping_status",
        "notes",
    ]
    write_csv(OUT_NORMALIZED, normalized, fields)

    recon = {code(r.get("hunt_code")): r for r in read_csv(RECON)}
    database = {code(r.get("hunt_code")): r for r in read_csv(DATABASE)}
    compare_rows = []
    for row in normalized:
        c = row["hunt_code"]
        source_values = (row["source_res"], row["source_nr"], row["source_total"])
        rec = recon.get(c, {})
        db = database.get(c, {})
        recommended = (
            clean(rec.get("recommended_res")),
            clean(rec.get("recommended_nr")),
            clean(rec.get("recommended_total")),
        )
        database_values = (
            int_text(db.get("permit_allotment_2026_res")),
            int_text(db.get("permit_allotment_2026_nr")),
            int_text(db.get("permit_allotment_2026_total")),
        )
        if not c:
            status = "UNMAPPED_NAME_ONLY_SOURCE"
        else:
            rec_status = compare_value(source_values, recommended)
            db_status = compare_value(source_values, database_values)
            if rec_status == "MATCH":
                status = "SOURCE_MATCHES_RECOMMENDED"
            elif rec_status == "TOTAL_MATCH_SPLIT_DIFFERS":
                status = "SOURCE_TOTAL_MATCHES_RECOMMENDED"
            elif db_status == "MATCH":
                status = "SOURCE_MATCHES_DATABASE"
            elif db_status == "TOTAL_MATCH_SPLIT_DIFFERS":
                status = "SOURCE_TOTAL_MATCHES_DATABASE"
            elif rec_status == "TARGET_BLANK" and db_status == "TARGET_BLANK":
                status = "SOURCE_HAS_VALUE_NO_CURRENT_COMPARISON_VALUE"
            elif rec_status == "SOURCE_BLANK":
                status = "SOURCE_NO_PERMIT_VALUE"
            else:
                status = "SOURCE_DIFFERS_FROM_CURRENT_RECONCILIATION"
        compare_rows.append(
            {
                **row,
                "comparison_status": status,
                "current_reconciliation_confidence": rec.get("confidence", ""),
                "current_reconciliation_winner": rec.get("winner_source", ""),
                "recommended_res": recommended[0],
                "recommended_nr": recommended[1],
                "recommended_total": recommended[2],
                "database_res": database_values[0],
                "database_nr": database_values[1],
                "database_total": database_values[2],
                "database_hunt_name": db.get("hunt_name", ""),
                "database_hunt_type": db.get("hunt_type", ""),
                "database_hunt_class": db.get("hunt_class", ""),
                "database_source": db.get("permit_allotment_2026_source", ""),
            }
        )

    compare_fields = [
        *fields,
        "comparison_status",
        "current_reconciliation_confidence",
        "current_reconciliation_winner",
        "recommended_res",
        "recommended_nr",
        "recommended_total",
        "database_res",
        "database_nr",
        "database_total",
        "database_hunt_name",
        "database_hunt_type",
        "database_hunt_class",
        "database_source",
    ]
    write_csv(OUT_COMPARE, compare_rows, compare_fields)

    source_counts = Counter(r["source_family"] for r in normalized)
    source_value_counts = Counter(r["source_family"] for r in normalized if has_value(r["source_res"], r["source_nr"], r["source_total"]))
    status_counts = Counter(r["comparison_status"] for r in compare_rows)
    status_by_source = Counter((r["source_family"], r["comparison_status"]) for r in compare_rows)
    summary = {
        "created_at_utc": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        "source_files": {
            "expo_draw_results": str(EXPO_DRAW),
            "expo_permit_draw": str(EXPO_PERMIT),
            "conservation_permits": str(CONSERVATION),
            "deer_buck": str(DEER_BUCK),
            "deer_doe": str(DEER_DOE),
        },
        "normalized_rows": len(normalized),
        "mapped_hunt_code_rows": sum(1 for r in normalized if r["hunt_code"]),
        "name_only_rows": sum(1 for r in normalized if not r["hunt_code"]),
        "rows_with_permit_value": sum(1 for r in normalized if has_value(r["source_res"], r["source_nr"], r["source_total"])),
        "source_family_counts": dict(sorted(source_counts.items())),
        "source_family_value_counts": dict(sorted(source_value_counts.items())),
        "comparison_status_counts": dict(sorted(status_counts.items())),
        "status_by_source": {f"{k[0]}|{k[1]}": v for k, v in sorted(status_by_source.items())},
        "outputs": {
            "normalized_csv": OUT_NORMALIZED.relative_to(ROOT).as_posix(),
            "comparison_csv": OUT_COMPARE.relative_to(ROOT).as_posix(),
            "summary_json": OUT_SUMMARY.relative_to(ROOT).as_posix(),
            "report_md": OUT_DOC.relative_to(ROOT).as_posix(),
        },
        "guardrail": "Audit only. DATABASE.csv was not modified.",
    }
    OUT_SUMMARY.write_text(json.dumps(summary, indent=2, sort_keys=True) + "\n", encoding="utf-8")

    lines = [
        "# 2026 Species Truth Permit Source Audit",
        "",
        "## Scope",
        "",
        "Normalized the user-supplied 2026 species truth permit files and compared direct hunt-code rows against the current reconciliation and DATABASE allotment reference fields.",
        "",
        "Expo files were kept as name-only evidence because they do not carry hunt codes. They are not promotion-ready until mapped by an approved hunt-code crosswalk.",
        "",
        "## Key Counts",
        "",
        f"- Normalized source rows: `{len(normalized)}`",
        f"- Direct hunt-code rows: `{summary['mapped_hunt_code_rows']}`",
        f"- Name-only rows: `{summary['name_only_rows']}`",
        f"- Rows with permit values: `{summary['rows_with_permit_value']}`",
        "",
        "## Source Family Counts",
        "",
    ]
    for source, count in sorted(source_counts.items()):
        lines.append(f"- `{source}`: `{count}` rows, `{source_value_counts[source]}` with permit values")
    lines.extend(["", "## Comparison Status Counts", ""])
    for status, count in sorted(status_counts.items()):
        lines.append(f"- `{status}`: `{count}`")
    lines.extend(
        [
            "",
            "## Outputs",
            "",
            f"- Normalized CSV: `{OUT_NORMALIZED.relative_to(ROOT).as_posix()}`",
            f"- Comparison CSV: `{OUT_COMPARE.relative_to(ROOT).as_posix()}`",
            f"- Summary JSON: `{OUT_SUMMARY.relative_to(ROOT).as_posix()}`",
            "",
            "## Guardrail",
            "",
            "`DATABASE.csv` was not modified.",
        ]
    )
    OUT_DOC.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(json.dumps(summary, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
