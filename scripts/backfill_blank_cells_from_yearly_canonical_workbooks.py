#!/usr/bin/env python3
"""Backfill safe blank cells from yearly canonical workbooks.

This pass intentionally uses only HUNT_CODE_SUMMARY sheets. Those sheets are
one row per hunt code, which avoids point-row drift and keeps the join exact on
actual_draw_year + hunt_code.
"""

from __future__ import annotations

import csv
import json
import os
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_DIR = ROOT / "outputs" / "yearly_canonical_workbooks"
CANONICAL_DIR = ROOT / "data_truth" / "draw_results_truth" / "normalized" / "canonical_yearly"
LONG_PATH = ROOT / "data_truth" / "draw_results_truth" / "normalized" / "draw_results_long.csv"
AUDIT_DIR = ROOT / "pipeline" / "R2_OFFLOAD" / "incoming"
AUDIT_CSV = AUDIT_DIR / "yearly_workbook_blank_cell_backfill_audit.csv"
SUMMARY_JSON = AUDIT_DIR / "yearly_workbook_blank_cell_backfill_summary.json"

# Keep this narrow. A broader field list can reintroduce old hunt-name bleed.
FIELDS_TO_BACKFILL = ["season"]


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), [dict(row) for row in reader]


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    with tmp.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore", lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fieldnames})
    os.replace(tmp, path)


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    os.replace(tmp, path)


def workbook_year(path: Path) -> str:
    match = re.match(r"(20\d{2})_PERMITS=", path.name)
    return match.group(1) if match else ""


def build_workbook_values() -> tuple[dict[tuple[str, str, str], str], dict[tuple[str, str, str], list[str]], dict[str, Any]]:
    values: dict[tuple[str, str, str], set[str]] = defaultdict(set)
    source_refs: dict[tuple[str, str, str], set[str]] = defaultdict(set)
    workbook_rows: dict[str, int] = {}

    for path in sorted(WORKBOOK_DIR.glob("*_CANONICAL_WORKBOOK.xlsx")):
        year = workbook_year(path)
        if not year:
            continue
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            if "HUNT_CODE_SUMMARY" not in workbook.sheetnames:
                continue
            sheet = workbook["HUNT_CODE_SUMMARY"]
            rows = sheet.iter_rows(values_only=True)
            headers = [clean(value).lower() for value in next(rows)]
            index = {header: offset for offset, header in enumerate(headers) if header}
            if "hunt_code" not in index:
                continue
            available_fields = [field for field in FIELDS_TO_BACKFILL if field in index]
            row_count = 0
            for row in rows:
                code = clean(row[index["hunt_code"]]).upper() if index["hunt_code"] < len(row) else ""
                if not code:
                    continue
                actual_year = (
                    clean(row[index["actual_draw_year"]])
                    if "actual_draw_year" in index and index["actual_draw_year"] < len(row)
                    else year
                )
                if actual_year != year:
                    continue
                row_count += 1
                for field in available_fields:
                    value = clean(row[index[field]]) if index[field] < len(row) else ""
                    if not value:
                        continue
                    key = (year, code, field)
                    values[key].add(value)
                    source_refs[key].add(f"{path.name}:HUNT_CODE_SUMMARY")
            workbook_rows[path.name] = row_count
        finally:
            workbook.close()

    single: dict[tuple[str, str, str], str] = {}
    conflicts: dict[str, list[str]] = {}
    for key, field_values in values.items():
        if len(field_values) == 1:
            single[key] = next(iter(field_values))
        else:
            conflicts["|".join(key)] = sorted(field_values)
    return single, {key: sorted(refs) for key, refs in source_refs.items()}, {
        "workbook_summary_rows": workbook_rows,
        "single_value_count": len(single),
        "conflict_count": len(conflicts),
        "conflict_sample": dict(list(sorted(conflicts.items()))[:25]),
    }


def patch_file(
    path: Path,
    workbook_values: dict[tuple[str, str, str], str],
    source_refs: dict[tuple[str, str, str], list[str]],
    fixed_year: str | None,
) -> tuple[int, list[dict[str, Any]]]:
    headers, rows = read_csv(path)
    changed = 0
    audit_rows: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows, start=2):
        year = fixed_year or clean(row.get("actual_draw_year"))
        if not year:
            continue
        code = clean(row.get("hunt_code")).upper()
        if not code:
            continue
        for field in FIELDS_TO_BACKFILL:
            if field not in headers or clean(row.get(field)):
                continue
            key = (year, code, field)
            value = workbook_values.get(key, "")
            if not value:
                continue
            row[field] = value
            changed += 1
            audit_rows.append(
                {
                    "target_file": str(path),
                    "row_number": row_number,
                    "actual_draw_year": year,
                    "hunt_code": code,
                    "hunt_name": clean(row.get("hunt_name")),
                    "field": field,
                    "old_value": "",
                    "new_value": value,
                    "source_refs": "|".join(source_refs.get(key, [])),
                    "action": "FILLED",
                    "reason": "blank cell filled from exact actual_draw_year + hunt_code HUNT_CODE_SUMMARY workbook value",
                }
            )

    if changed:
        write_csv(path, headers, rows)
    return changed, audit_rows


def main() -> None:
    workbook_values, source_refs, source_summary = build_workbook_values()
    audit_rows: list[dict[str, Any]] = []
    changed_by_file: dict[str, int] = {}

    for year in range(2019, 2027):
        path = CANONICAL_DIR / f"draw_results_{year}_for_{year + 1}_canonical_yearly_draw_results.csv"
        if not path.exists():
            continue
        changed, file_audit = patch_file(path, workbook_values, source_refs, fixed_year=str(year))
        changed_by_file[str(path)] = changed
        audit_rows.extend(file_audit)

    changed, file_audit = patch_file(LONG_PATH, workbook_values, source_refs, fixed_year=None)
    changed_by_file[str(LONG_PATH)] = changed
    audit_rows.extend(file_audit)

    write_csv(
        AUDIT_CSV,
        [
            "target_file",
            "row_number",
            "actual_draw_year",
            "hunt_code",
            "hunt_name",
            "field",
            "old_value",
            "new_value",
            "source_refs",
            "action",
            "reason",
        ],
        audit_rows,
    )
    summary = {
        "fields_backfilled": FIELDS_TO_BACKFILL,
        "source_workbook_dir": str(WORKBOOK_DIR),
        "rows_filled_total": len(audit_rows),
        "rows_filled_by_file": changed_by_file,
        "source_summary": source_summary,
        "audit_csv": str(AUDIT_CSV),
        "note": "Only blank season cells were filled. Existing values were not overwritten.",
    }
    write_json(SUMMARY_JSON, summary)
    print(json.dumps(summary, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
