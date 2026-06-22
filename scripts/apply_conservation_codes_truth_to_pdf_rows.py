from __future__ import annotations

import csv
import re
import shutil
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "processed_data" / "hard_data_exports" / "hunt_tables" / "2026" / "CLEAN_XLXS_STAGED" / "2025-27 Conservation Permits.xlsx"
TRUTH_PATH = Path(r"C:\Users\tyler\Desktop\conservation codes.xlsx")
DATABASE_PATH = ROOT / "pipeline" / "RAW" / "hunt_unit_database" / "2026" / "csv" / "DATABASE.csv"
BOUNDARY_DIR = ROOT / "processed_data" / "boundaries"
AUDIT_DIR = ROOT / "audits" / "2025_canonical_finalization"
BACKUP_DIR = AUDIT_DIR / "backups"
AUDIT_PATH = AUDIT_DIR / "2025_27_conservation_codes_truth_pdf_row_apply_audit.csv"


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def compact(value: object) -> str:
    text = clean(value).lower().replace("&", " and ")
    text = re.sub(r"\([^)]*\)", " ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def tokens(value: object) -> set[str]:
    stop = {"and", "the", "area"}
    return {token for token in compact(value).split() if token and token not in stop}


def normalize_species(value: object) -> str:
    text = compact(value)
    if text in {"bear", "black bear"}:
        return "black bear"
    if text == "antlerless elk":
        return "elk"
    return text


def weapon_matches(condition: str, weapon: str) -> bool:
    cond = compact(condition)
    weap = compact(weapon)
    if not weap:
        return True
    if weap == "any legal weapon":
        return "any legal weapon" in cond
    if weap == "multiseason":
        return "multiseason" in cond or "hunter s choice" in cond
    if weap == "muzzleloader":
        return "muzzleloader" in cond or "muzz" in cond
    return all(part in cond for part in weap.split())


def area_score(left: object, right: object) -> float:
    lt = tokens(left)
    rt = tokens(right)
    if not lt or not rt:
        return 0.0
    inter = len(lt & rt)
    contain = inter / min(len(lt), len(rt))
    jaccard = inter / len(lt | rt)
    return max(contain, jaccard)


def read_database() -> dict[str, dict[str, str]]:
    with DATABASE_PATH.open("r", encoding="utf-8-sig", newline="") as f:
        by_code = {}
        for row in csv.DictReader(f):
            code = clean(row.get("hunt_code")).upper()
            if code and code not in by_code:
                by_code[code] = {k: clean(v) for k, v in row.items()}
        return by_code


def geojson_for(code: str, boundary_id: str) -> str:
    code_path = BOUNDARY_DIR / f"{code}.geojson"
    if code_path.exists():
        return code_path.relative_to(ROOT).as_posix()
    if boundary_id:
        id_path = BOUNDARY_DIR / f"{boundary_id}.geojson"
        if id_path.exists():
            return id_path.relative_to(ROOT).as_posix()
    return ""


def read_truth_rows() -> list[dict[str, str]]:
    wb = load_workbook(TRUTH_PATH, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not clean(row[1] if len(row) > 1 else ""):
            continue
        rows.append(
            {
                "hunt_name": clean(row[0]),
                "hunt_code": clean(row[1]).upper(),
                "sex": clean(row[2]),
                "species": clean(row[3]),
                "weapon": clean(row[4]),
                "hunt_type": clean(row[5]),
                "season": clean(row[6]),
            }
        )
    return rows


def main() -> None:
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)

    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    backup_path = BACKUP_DIR / f"{WORKBOOK_PATH.stem}.before_conservation_codes_truth_apply_{stamp}{WORKBOOK_PATH.suffix}"
    shutil.copy2(WORKBOOK_PATH, backup_path)

    db = read_database()
    truth_rows = read_truth_rows()

    wb = load_workbook(WORKBOOK_PATH)
    ws = wb["Table 1"] if "Table 1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [clean(cell.value) for cell in ws[1]]
    cols = {header: index + 1 for index, header in enumerate(headers) if header}

    applied = []
    consumed_excel_rows: set[int] = set()
    for truth in truth_rows:
        best = None
        best_score = 0.0
        for excel_row in range(2, ws.max_row + 1):
            if excel_row in consumed_excel_rows:
                continue
            row_species = normalize_species(ws.cell(excel_row, cols["Species"]).value)
            if row_species != normalize_species(truth["species"]):
                continue
            if not weapon_matches(clean(ws.cell(excel_row, cols["Condition"]).value), truth["weapon"]):
                continue
            score = area_score(ws.cell(excel_row, cols["Area"]).value, truth["hunt_name"])
            if score > best_score:
                best_score = score
                best = excel_row

        if best is None or best_score < 0.74:
            applied.append({**truth, "status": "review_no_confident_pdf_row", "excel_row": "", "score": f"{best_score:.3f}"})
            continue
        consumed_excel_rows.add(best)

        code = truth["hunt_code"]
        db_row = db.get(code, {})
        boundary_id = clean(db_row.get("boundary_id")) or clean(ws.cell(best, cols["BOUNDARY ID"]).value)
        map_geojson = geojson_for(code, boundary_id) or clean(ws.cell(best, cols["MAP GEOJSON"]).value)

        before = {
            "before_hunt_code": clean(ws.cell(best, cols["HUNT CODE"]).value),
            "before_hunt_name": clean(ws.cell(best, cols["HUNT NAME"]).value),
            "before_sex": clean(ws.cell(best, cols["SEX"]).value),
            "before_weapon": clean(ws.cell(best, cols["WEAPON"]).value),
            "before_boundary_id": clean(ws.cell(best, cols["BOUNDARY ID"]).value),
            "before_map_geojson": clean(ws.cell(best, cols["MAP GEOJSON"]).value),
        }

        ws.cell(best, cols["HUNT CODE"]).value = code
        ws.cell(best, cols["HUNT NAME"]).value = truth["hunt_name"]
        ws.cell(best, cols["SEX"]).value = truth["sex"]
        ws.cell(best, cols["WEAPON"]).value = truth["weapon"]
        ws.cell(best, cols["BOUNDARY ID"]).value = boundary_id
        ws.cell(best, cols["MAP GEOJSON"]).value = map_geojson

        applied.append(
            {
                **truth,
                **before,
                "status": "applied",
                "excel_row": best,
                "pdf_no": ws.cell(best, cols["No."]).value,
                "pdf_species": ws.cell(best, cols["Species"]).value,
                "pdf_area": ws.cell(best, cols["Area"]).value,
                "pdf_condition": ws.cell(best, cols["Condition"]).value,
                "score": f"{best_score:.3f}",
                "after_boundary_id": boundary_id,
                "after_map_geojson": map_geojson,
            }
        )

    wb.save(WORKBOOK_PATH)

    fieldnames = sorted({key for row in applied for key in row})
    with AUDIT_PATH.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(applied)

    print(f"truth_rows={len(truth_rows)}")
    print(f"applied={sum(1 for row in applied if row['status'] == 'applied')}")
    print(f"review={sum(1 for row in applied if row['status'] != 'applied')}")
    print(f"backup={backup_path}")
    print(f"audit={AUDIT_PATH}")


if __name__ == "__main__":
    main()
