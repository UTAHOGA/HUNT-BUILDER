#!/usr/bin/env python3
"""Append/fill boundary reconciliation columns in yearly BIBLE CSV/XLSX files."""

from __future__ import annotations

import csv
import json
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
YEAR_DIR = ROOT / "processed_data/audits/bible_hunt_code_year_documents"
RECONCILIATION = YEAR_DIR / "bible_hunt_code_year_boundary_reconciliation_2020_2026.csv"
SUMMARY = YEAR_DIR / "bible_hunt_code_year_boundary_columns_applied_summary.json"

BOUNDARY_COLUMNS = [
    "resolved_boundary_id",
    "boundary_resolution_status",
    "current_database_presence",
    "current_database_boundary_id",
    "display_boundary_id",
    "hunt_boundary_crosswalk_id",
    "split_index_boundary_id",
    "direct_hunt_code_geojson",
    "direct_boundary_id_geojson",
]


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        return [{key: (value or "").strip() for key, value in row.items()} for row in csv.DictReader(handle)]


def write_csv(path: Path, rows: list[dict[str, str]], columns: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column, "") for column in columns})


def build_reconciliation_lookup() -> dict[tuple[str, str], dict[str, str]]:
    lookup: dict[tuple[str, str], dict[str, str]] = {}
    for row in read_csv(RECONCILIATION):
        key = (row.get("report_year", ""), row.get("comparison_hunt_code", ""))
        if key[0] and key[1]:
            lookup[key] = row
    return lookup


def apply_to_csv(path: Path, lookup: dict[tuple[str, str], dict[str, str]]) -> dict[str, int]:
    rows = read_csv(path)
    if not rows:
        return {"rows": 0, "matched": 0, "resolved_boundary_populated": 0}

    columns = list(rows[0].keys())
    for column in BOUNDARY_COLUMNS:
        if column not in columns:
            columns.append(column)

    matched = 0
    populated = 0
    for row in rows:
        key = (row.get("report_year", ""), row.get("comparison_hunt_code", ""))
        source = lookup.get(key)
        if not source:
            continue
        matched += 1
        for column in BOUNDARY_COLUMNS:
            row[column] = source.get(column, "")
        if row.get("resolved_boundary_id"):
            populated += 1

    write_csv(path, rows, columns)
    return {"rows": len(rows), "matched": matched, "resolved_boundary_populated": populated}


def apply_to_xlsx(path: Path, lookup: dict[tuple[str, str], dict[str, str]]) -> dict[str, int]:
    if not path.exists():
        return {"rows": 0, "matched": 0, "resolved_boundary_populated": 0}
    workbook = load_workbook(path)
    sheet = workbook.active
    headers = [str(sheet.cell(row=1, column=col).value or "").strip() for col in range(1, sheet.max_column + 1)]
    header_to_col = {header: idx + 1 for idx, header in enumerate(headers) if header}

    for column in BOUNDARY_COLUMNS:
        if column not in header_to_col:
            sheet.cell(row=1, column=sheet.max_column + 1, value=column)
            header_to_col[column] = sheet.max_column

    report_year_col = header_to_col.get("report_year")
    hunt_code_col = header_to_col.get("comparison_hunt_code")
    if not report_year_col or not hunt_code_col:
        raise RuntimeError(f"{path} is missing report_year or comparison_hunt_code")

    matched = 0
    populated = 0
    rows = max(0, sheet.max_row - 1)
    for row_idx in range(2, sheet.max_row + 1):
        key = (
            str(sheet.cell(row=row_idx, column=report_year_col).value or "").strip(),
            str(sheet.cell(row=row_idx, column=hunt_code_col).value or "").strip(),
        )
        source = lookup.get(key)
        if not source:
            continue
        matched += 1
        for column in BOUNDARY_COLUMNS:
            sheet.cell(row=row_idx, column=header_to_col[column], value=source.get(column, ""))
        if source.get("resolved_boundary_id"):
            populated += 1

    for column in BOUNDARY_COLUMNS:
        sheet.column_dimensions[sheet.cell(row=1, column=header_to_col[column]).column_letter].width = 24

    workbook.save(path)
    return {"rows": rows, "matched": matched, "resolved_boundary_populated": populated}


def main() -> int:
    lookup = build_reconciliation_lookup()
    results: dict[str, dict[str, dict[str, int]]] = {}
    for csv_path in sorted(YEAR_DIR.glob("bible_hunt_code_year_document_*.csv")):
        if "summary" in csv_path.name:
            continue
        year = csv_path.stem.rsplit("_", 1)[-1]
        xlsx_path = csv_path.with_suffix(".xlsx")
        results[year] = {
            "csv": apply_to_csv(csv_path, lookup),
            "xlsx": apply_to_xlsx(xlsx_path, lookup),
        }

    csv_populated = sum(result["csv"]["resolved_boundary_populated"] for result in results.values())
    xlsx_populated = sum(result["xlsx"]["resolved_boundary_populated"] for result in results.values())
    summary = {
        "created_at_utc": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        "source_reconciliation": RECONCILIATION.relative_to(ROOT).as_posix(),
        "boundary_columns_applied": BOUNDARY_COLUMNS,
        "year_results": results,
        "totals": {
            "years_updated": len(results),
            "csv_rows": sum(result["csv"]["rows"] for result in results.values()),
            "csv_rows_matched_to_reconciliation": sum(result["csv"]["matched"] for result in results.values()),
            "csv_rows_with_resolved_boundary_id": csv_populated,
            "xlsx_rows": sum(result["xlsx"]["rows"] for result in results.values()),
            "xlsx_rows_matched_to_reconciliation": sum(result["xlsx"]["matched"] for result in results.values()),
            "xlsx_rows_with_resolved_boundary_id": xlsx_populated,
        },
        "boundary_resolution_status_counts": dict(
            Counter(row.get("boundary_resolution_status", "") for row in read_csv(RECONCILIATION))
        ),
        "notes": [
            "Original yearly extraction fields were preserved.",
            "Boundary columns were filled from the audited reconciliation output by report_year + comparison_hunt_code.",
        ],
    }
    SUMMARY.write_text(json.dumps(summary, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(json.dumps(summary, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
