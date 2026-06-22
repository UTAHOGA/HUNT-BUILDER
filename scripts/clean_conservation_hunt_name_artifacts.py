from __future__ import annotations

import csv
import re
import shutil
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "processed_data" / "hard_data_exports" / "hunt_tables" / "2026" / "CLEAN_XLXS_STAGED" / "2025-27 Conservation Permits.xlsx"
AUDIT_DIR = ROOT / "audits" / "2025_canonical_finalization"
BACKUP_DIR = AUDIT_DIR / "backups"
AUDIT_PATH = AUDIT_DIR / "2025_27_conservation_hunt_name_artifact_cleanup.csv"


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def append_timing_to_weapon(weapon: str, timing: str) -> str:
    weapon_text = clean(weapon)
    timing_text = clean(timing).lower()
    if not timing_text:
        return weapon_text
    if timing_text in weapon_text.lower():
        return weapon_text
    if weapon_text:
        return f"{weapon_text}, {timing_text}"
    return timing_text.title()


def cleanup_name(species: str, name: str, sex_type: str, weapon: str) -> tuple[str, str, str, list[str]]:
    species_text = clean(species)
    name_text = clean(name)
    sex_text = clean(sex_type)
    weapon_text = clean(weapon)
    notes: list[str] = []

    timing_match = re.search(r"\((early|late|mid)\)", name_text, flags=re.IGNORECASE)
    if timing_match:
        timing = timing_match.group(1).lower()
        weapon_text = append_timing_to_weapon(weapon_text, timing)
        name_text = re.sub(r"\s*\((early|late|mid)\)", "", name_text, flags=re.IGNORECASE).strip()
        notes.append(f"moved_{timing}_to_weapon")

    bison_sex_patterns = [
        (r",?\s*Hunter'?s Choice\b", "Hunter's Choice"),
        (r",?\s*Hunters Choice\b", "Hunter's Choice"),
        (r",?\s*Cow Only\b", "Cow Only"),
        (r",?\s*Female Only\b", "Female Only"),
    ]
    if species_text.lower() == "bison":
        for pattern, normalized in bison_sex_patterns:
            if re.search(pattern, name_text, flags=re.IGNORECASE):
                name_text = re.sub(pattern, "", name_text, flags=re.IGNORECASE).strip(" ,-/")
                if not sex_text or sex_text.lower() in {"hunter's choice", "hunters choice", "cow only", "female only"}:
                    sex_text = normalized
                note_label = normalized.lower().replace(" ", "_").replace("'", "")
                notes.append(f"moved_{note_label}_to_sex_type")

    if species_text.lower() == "elk":
        new_name = re.sub(r"\s+Bull\s+Elk\s*$", "", name_text, flags=re.IGNORECASE).strip(" ,-/")
        if new_name != name_text:
            name_text = new_name
            sex_text = "Bull"
            notes.append("moved_bull_elk_to_species_sex")

    species_artifact_patterns = [
        r"\s+-?\s*Statewide Permit\s*$",
        r"\s+Permit\s*$",
    ]
    for pattern in species_artifact_patterns:
        new_name = re.sub(pattern, "", name_text, flags=re.IGNORECASE).strip(" ,-/")
        if new_name != name_text:
            name_text = new_name
            notes.append("removed_permit_artifact")

    name_text = re.sub(r"\s+", " ", name_text).strip(" ,-/")
    return name_text, sex_text, weapon_text, notes


def main() -> None:
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)

    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    backup_path = BACKUP_DIR / f"{WORKBOOK_PATH.stem}.before_hunt_name_artifact_cleanup_{stamp}{WORKBOOK_PATH.suffix}"
    shutil.copy2(WORKBOOK_PATH, backup_path)

    wb = load_workbook(WORKBOOK_PATH)
    ws = wb["Table 1"] if "Table 1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [clean(ws.cell(1, column).value) for column in range(1, ws.max_column + 1)]
    cols = {header: index + 1 for index, header in enumerate(headers) if header}

    required = ["No.", "Species", "HUNT NAME", "SEX TYPE", "WEAPON"]
    missing = [header for header in required if header not in cols]
    if missing:
        raise RuntimeError(f"Missing required headers: {missing}")

    audit_rows = []
    for row_idx in range(2, ws.max_row + 1):
        if not clean(ws.cell(row_idx, cols["No."]).value):
            continue
        old_name = clean(ws.cell(row_idx, cols["HUNT NAME"]).value)
        old_sex = clean(ws.cell(row_idx, cols["SEX TYPE"]).value)
        old_weapon = clean(ws.cell(row_idx, cols["WEAPON"]).value)
        new_name, new_sex, new_weapon, notes = cleanup_name(
            ws.cell(row_idx, cols["Species"]).value,
            old_name,
            old_sex,
            old_weapon,
        )
        if not notes:
            continue
        ws.cell(row_idx, cols["HUNT NAME"]).value = new_name
        ws.cell(row_idx, cols["SEX TYPE"]).value = new_sex
        ws.cell(row_idx, cols["WEAPON"]).value = new_weapon
        audit_rows.append(
            {
                "row": row_idx,
                "no": ws.cell(row_idx, cols["No."]).value,
                "species": ws.cell(row_idx, cols["Species"]).value,
                "old_hunt_name": old_name,
                "new_hunt_name": new_name,
                "old_sex_type": old_sex,
                "new_sex_type": new_sex,
                "old_weapon": old_weapon,
                "new_weapon": new_weapon,
                "notes": ";".join(notes),
            }
        )

    wb.save(WORKBOOK_PATH)

    with AUDIT_PATH.open("w", encoding="utf-8", newline="") as f:
        fieldnames = [
            "row",
            "no",
            "species",
            "old_hunt_name",
            "new_hunt_name",
            "old_sex_type",
            "new_sex_type",
            "old_weapon",
            "new_weapon",
            "notes",
        ]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(audit_rows)

    print(f"changed_rows={len(audit_rows)}")
    print(f"backup={backup_path}")
    print(f"audit={AUDIT_PATH}")


if __name__ == "__main__":
    main()
