from __future__ import annotations

import csv
import re
import shutil
from copy import copy
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "processed_data" / "hard_data_exports" / "hunt_tables" / "2026" / "CLEAN_XLXS_STAGED" / "2025-27 Conservation Permits.xlsx"
AUDIT_DIR = ROOT / "audits" / "2025_canonical_finalization"
BACKUP_DIR = AUDIT_DIR / "backups"
AUDIT_PATH = AUDIT_DIR / "2025_27_conservation_season_label_moves.csv"


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def add_season(existing: str, value: str) -> str:
    existing_text = clean(existing)
    value_text = clean(value)
    if not value_text:
        return existing_text
    parts = [part.strip() for part in existing_text.split(";") if part.strip()]
    if value_text.lower() not in {part.lower() for part in parts}:
        parts.append(value_text)
    return "; ".join(parts)


def strip_timing_from_weapon(weapon: str) -> tuple[str, list[str]]:
    weapon_text = clean(weapon)
    seasons: list[str] = []
    if not weapon_text:
        return weapon_text, seasons

    for timing in ["early", "mid", "late"]:
        pattern = rf"(?i)(?:,\s*|\s+){timing}\b"
        if re.search(pattern, weapon_text):
            seasons.append(timing.title())
            weapon_text = re.sub(pattern, "", weapon_text).strip(" ,")

    return weapon_text, seasons


def main() -> None:
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)

    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    backup_path = BACKUP_DIR / f"{WORKBOOK_PATH.stem}.before_season_label_moves_{stamp}{WORKBOOK_PATH.suffix}"
    shutil.copy2(WORKBOOK_PATH, backup_path)

    wb = load_workbook(WORKBOOK_PATH)
    ws = wb["Table 1"] if "Table 1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [clean(ws.cell(1, column).value) for column in range(1, ws.max_column + 1)]
    cols = {header: index + 1 for index, header in enumerate(headers) if header}

    required = ["No.", "Species", "HUNT NAME", "SEX TYPE", "WEAPON"]
    missing = [header for header in required if header not in cols]
    if missing:
        raise RuntimeError(f"Missing required headers: {missing}")

    if "SEASON" not in cols:
        insert_at = cols["WEAPON"] + 1
        ws.insert_cols(insert_at)
        ws.cell(1, insert_at).value = "SEASON"
        for style_attr in ["font", "fill", "border", "alignment", "number_format", "protection"]:
            setattr(ws.cell(1, insert_at), style_attr, copy(getattr(ws.cell(1, cols["WEAPON"]), style_attr)))
        headers = [clean(ws.cell(1, column).value) for column in range(1, ws.max_column + 1)]
        cols = {header: index + 1 for index, header in enumerate(headers) if header}
        ws.column_dimensions[ws.cell(1, insert_at).column_letter].width = 18

    audit_rows = []
    for row_idx in range(2, ws.max_row + 1):
        if not clean(ws.cell(row_idx, cols["No."]).value):
            continue

        old_sex = clean(ws.cell(row_idx, cols["SEX TYPE"]).value)
        old_weapon = clean(ws.cell(row_idx, cols["WEAPON"]).value)
        old_season = clean(ws.cell(row_idx, cols["SEASON"]).value)

        new_sex = old_sex
        new_weapon = old_weapon
        new_season = old_season
        notes: list[str] = []

        if old_sex.lower() in {"hunter's choice", "hunters choice"}:
            new_season = add_season(new_season, "Hunter's Choice")
            new_sex = "Either Sex"
            notes.append("moved_hunters_choice_from_sex_type_to_season")

        if old_weapon.lower() in {"hunter's choice", "hunters choice"}:
            new_season = add_season(new_season, "Hunter's Choice")
            new_weapon = ""
            notes.append("moved_hunters_choice_from_weapon_to_season")
        else:
            stripped_weapon, timing_seasons = strip_timing_from_weapon(old_weapon)
            if timing_seasons:
                new_weapon = stripped_weapon
                for timing in timing_seasons:
                    new_season = add_season(new_season, timing)
                notes.append("moved_timing_from_weapon_to_season")

        if not notes:
            continue

        ws.cell(row_idx, cols["SEX TYPE"]).value = new_sex
        ws.cell(row_idx, cols["WEAPON"]).value = new_weapon
        ws.cell(row_idx, cols["SEASON"]).value = new_season
        audit_rows.append(
            {
                "row": row_idx,
                "no": ws.cell(row_idx, cols["No."]).value,
                "species": ws.cell(row_idx, cols["Species"]).value,
                "hunt_name": ws.cell(row_idx, cols["HUNT NAME"]).value,
                "old_sex_type": old_sex,
                "new_sex_type": new_sex,
                "old_weapon": old_weapon,
                "new_weapon": new_weapon,
                "old_season": old_season,
                "new_season": new_season,
                "notes": ";".join(notes),
            }
        )

    if ws.auto_filter and ws.auto_filter.ref:
        last_col = ws.cell(1, ws.max_column).column_letter
        ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

    wb.save(WORKBOOK_PATH)

    with AUDIT_PATH.open("w", encoding="utf-8", newline="") as f:
        fieldnames = [
            "row",
            "no",
            "species",
            "hunt_name",
            "old_sex_type",
            "new_sex_type",
            "old_weapon",
            "new_weapon",
            "old_season",
            "new_season",
            "notes",
        ]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(audit_rows)

    print(f"changed_rows={len(audit_rows)}")
    print(f"backup={backup_path}")
    print(f"audit={AUDIT_PATH}")


if __name__ == "__main__":
    main()
