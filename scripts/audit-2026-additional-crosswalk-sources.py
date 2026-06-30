#!/usr/bin/env python3
"""Audit additional 2026 crosswalk sources against remaining unresolved permit rows."""

from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path


def _repo_root() -> Path:
    repo_root = Path(__file__).resolve()
    while repo_root.name != "HUNT-BUILDER" and repo_root.parent != repo_root:
        repo_root = repo_root.parent
    if repo_root.name != "HUNT-BUILDER":
        raise RuntimeError("Could not locate HUNT-BUILDER repo root")
    return repo_root

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
UNRESOLVED = ROOT / "processed_data/audits/current_2026_permit_unresolved_split/remaining_unresolved_after_crosswalk_hunts_upload_rule.csv"
REGULATION_AUDIT = ROOT / "processed_data/audits/current_2026_permit_unresolved_split/regulation_2026_unresolved_code_presence_audit.csv"
DATABASE = ROOT / "pipeline/RAW/hunt_unit_database/2026/csv/DATABASE.csv"
OUT_DIR = ROOT / "processed_data/audits/current_2026_permit_unresolved_split"

SOURCES = {
    "retired_current": Path(str(_repo_root() / "data_truth/crosswalk_truth/normalized/retired_current_hunt_codes_2026.csv")),
    "black_bear_primary": Path(str(_repo_root() / "data_truth/crosswalk_truth/normalized/black_bear_BR_2024_2025_2026_crosswalk.csv")),
    "current_historical_primary": Path(str(_repo_root() / "data_truth/crosswalk_truth/normalized/current_to_historical_hunt_code_crosswalk_2026.csv")),
    "current_historical_processed": Path(str(_repo_root() / "processed_data/current_to_historical_hunt_code_crosswalk_2026.csv")),
    "historical_le_to_eb": Path(str(_repo_root() / "processed_data/historical_le_to_eb_crosswalk_2024.csv")),
    "boundary_unit_fill": Path(str(_repo_root() / "processed_data/hunt_boundary_crosswalk_2026_unit_fill_sources.csv")),
    "historical_retired": Path(str(_repo_root() / "processed_data/hunt_boundary_crosswalk_historical_retired_codes.csv")),
    "boundary_xlsx": Path(str(_repo_root() / "processed_data/hunt_boundary_crosswalk_2026.xlsx")),
    "boundary_id_to_hunt_code": Path(str(_repo_root() / "processed_data/audits/boundary_id_to_hunt_code_crosswalk_2026.csv")),
    "same_code_promotions": Path(str(_repo_root() / "data_truth/crosswalk_truth/validation/same_code_2025_pdf_crosswalk_active_promotions.csv")),
    "model_year_crosswalk": Path(str(_repo_root() / "data_truth/crosswalk_truth/validation/hunt_code_crosswalk_2024_pdf_to_2025_pdf_model_years.csv")),
    "model_year_dropped_review": Path(str(_repo_root() / "data_truth/crosswalk_truth/validation/hunt_code_crosswalk_2024_pdf_to_2025_pdf_dropped_review.csv")),
    "remaining_boundary_closeout": Path(str(_repo_root() / "data_truth/comparison_outputs/validation/remaining_2025_history_crosswalk_boundary_closeout.csv")),
    "black_bear_secondary": Path(str(_repo_root() / "data_truth/crosswalk_truth/normalized/black_bear_BR_2024_2025_2026_crosswalk.csv")),
    "conservation_area": Path(str(_repo_root() / "processed_data/conservation_area_crosswalk_2026.csv")),
    "runtime_boundary_draft": Path(str(_repo_root() / "data_model/runtime_drafts/hunt_boundary_crosswalk_v2.csv")),
    "bighorn_location_crosswalk": Path(str(_repo_root() / "data_truth/harvest_results_truth/raw_packages/2023_for_2024_harvest_results_2023_all_species_database/harvest_location_hunt_code_crosswalk_2023_bighorn_sheep.csv")),
    "bighorn_measurements_crosswalked": Path(str(_repo_root() / "data_truth/harvest_results_truth/raw_packages/2023_for_2024_harvest_results_2023_all_species_database/harvest_results_2023_bighorn_sheep_measurements_crosswalked.csv")),
}

MATCH_COLUMNS = [
    "current_hunt_code",
    "historical_hunt_code",
    "draw_year",
    "model_year",
    "historical_hunt_name",
    "current_hunt_name",
    "species",
    "sex_type",
    "weapon",
    "hunt_type",
    "historical_unit_name",
    "current_unit_name",
    "historical_boundary_id",
    "current_boundary_id",
    "boundary_match",
    "permits_res_historical",
    "permits_nr_historical",
    "permits_total_historical",
    "permits_res_current",
    "permits_nr_current",
    "permits_total_current",
    "crosswalk_status",
    "confidence",
    "evidence_source",
    "evidence_notes",
    "source_file_historical",
    "source_file_current",
    "source_priority_historical",
    "source_priority_current",
    "merge_split_flag",
    "statewide_flag",
    "unlimited_flag",
    "source_permit_conflict_flag",
    "unresolved_split_bucket",
    "regulation_presence_status",
    "recommended_action",
]

ROLLUP_COLUMNS = [
    "hunt_code",
    "hunt_name",
    "species",
    "sex_type",
    "weapon",
    "hunt_type",
    "best_crosswalk_status",
    "best_confidence",
    "recommended_resolution_bucket",
    "boundary_confirmed",
    "regulation_present",
    "permit_conflict",
    "evidence_source_count",
    "evidence_sources",
    "status_counts",
    "notes",
]


def read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        return [{k: (v or "").strip() for k, v in row.items()} for row in csv.DictReader(handle)]


def write_csv(path: Path, rows: list[dict[str, str]], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column, "") for column in columns})


def split_codes(value: str) -> list[str]:
    return re.findall(r"\b[A-Z]{1,3}\d{4}\b", value or "")


def norm(value: object) -> str:
    return str(value or "").strip()


def read_boundary_xlsx(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [norm(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    rows: list[dict[str, str]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {headers[idx]: norm(value) for idx, value in enumerate(values) if idx < len(headers)}
        if any(row.values()):
            rows.append(row)
    return rows


def load_indexes() -> tuple[dict[str, dict[str, str]], dict[str, dict[str, str]], dict[str, list[dict[str, str]]]]:
    db = {row["hunt_code"]: row for row in read_csv(DATABASE) if row.get("hunt_code")}
    reg = {row["hunt_code"]: row for row in read_csv(REGULATION_AUDIT) if row.get("hunt_code")}
    evidence: dict[str, list[dict[str, str]]] = defaultdict(list)

    for row in read_csv(SOURCES["retired_current"]):
        if row.get("hunt_code"):
            evidence[row["hunt_code"]].append({"source_key": "retired_current", **row})

    for key in ("black_bear_primary", "black_bear_secondary"):
        for row in read_csv(SOURCES[key]):
            for code_key in ("current_2026_code", "historical_2025_code", "historical_2024_code"):
                if row.get(code_key):
                    evidence[row[code_key]].append({"source_key": key, "matched_code_field": code_key, **row})

    for key in ("current_historical_primary", "current_historical_processed"):
        for row in read_csv(SOURCES[key]):
            if row.get("current_hunt_code"):
                evidence[row["current_hunt_code"]].append({"source_key": key, "matched_code_field": "current_hunt_code", **row})
            if row.get("historical_hunt_code") and row.get("historical_hunt_code") != row.get("current_hunt_code"):
                evidence[row["historical_hunt_code"]].append({"source_key": key, "matched_code_field": "historical_hunt_code", **row})

    for row in read_csv(SOURCES["historical_le_to_eb"]):
        for code_key in ("source_hunt_number", "recommended_eb_code"):
            if row.get(code_key):
                evidence[row[code_key]].append({"source_key": "historical_le_to_eb", "matched_code_field": code_key, **row})
        for code in split_codes(row.get("recommended_eb_candidates", "")):
            evidence[code].append({"source_key": "historical_le_to_eb", "matched_code_field": "recommended_eb_candidates", **row})

    for row in read_csv(SOURCES["boundary_unit_fill"]):
        if row.get("hunt_code"):
            evidence[row["hunt_code"]].append({"source_key": "boundary_unit_fill", "matched_code_field": "hunt_code", **row})

    for row in read_csv(SOURCES["historical_retired"]):
        for code_key in ("historical_hunt_code", "current_hunt_code_candidate"):
            if row.get(code_key):
                evidence[row[code_key]].append({"source_key": "historical_retired", "matched_code_field": code_key, **row})

    for row in read_boundary_xlsx(SOURCES["boundary_xlsx"]):
        if row.get("hunt_code"):
            evidence[row["hunt_code"]].append({"source_key": "boundary_xlsx", "matched_code_field": "hunt_code", **row})

    for row in read_csv(SOURCES["boundary_id_to_hunt_code"]):
        for code in split_codes(row.get("hunt_codes", "")):
            evidence[code].append({"source_key": "boundary_id_to_hunt_code", "matched_code_field": "hunt_codes", **row})

    for row in read_csv(SOURCES["same_code_promotions"]):
        if row.get("hunt_code"):
            evidence[row["hunt_code"]].append({"source_key": "same_code_promotions", "matched_code_field": "hunt_code", **row})

    for row in read_csv(SOURCES["model_year_crosswalk"]):
        for code_key in ("source_hunt_code", "mapped_hunt_code"):
            if row.get(code_key):
                evidence[row[code_key]].append({"source_key": "model_year_crosswalk", "matched_code_field": code_key, **row})
        for code in split_codes(row.get("candidate_hunt_codes", "")):
            evidence[code].append({"source_key": "model_year_crosswalk", "matched_code_field": "candidate_hunt_codes", **row})

    for row in read_csv(SOURCES["model_year_dropped_review"]):
        for code_key in ("source_hunt_code", "mapped_hunt_code"):
            if row.get(code_key):
                evidence[row[code_key]].append({"source_key": "model_year_dropped_review", "matched_code_field": code_key, **row})
        for code in split_codes(row.get("candidate_hunt_codes", "")):
            evidence[code].append({"source_key": "model_year_dropped_review", "matched_code_field": "candidate_hunt_codes", **row})

    for row in read_csv(SOURCES["remaining_boundary_closeout"]):
        for code_key in ("hunt_code", "candidate_hunt_code"):
            if row.get(code_key):
                evidence[row[code_key]].append({"source_key": "remaining_boundary_closeout", "matched_code_field": code_key, **row})

    for row in read_csv(SOURCES["conservation_area"]):
        code_fields = [
            "primary_hunt_code",
            "bundle_conservation_hunt_codes",
            "bundle_expo_hunt_codes",
            "included_hunt_codes",
            "boundary_reference_hunt_codes",
        ]
        for field in code_fields:
            for code in split_codes(row.get(field, "")):
                evidence[code].append({"source_key": "conservation_area", "matched_code_field": field, **row})

    for row in read_csv(SOURCES["runtime_boundary_draft"]):
        if row.get("hunt_code"):
            evidence[row["hunt_code"]].append({"source_key": "runtime_boundary_draft", "matched_code_field": "hunt_code", **row})

    for key in ("bighorn_location_crosswalk", "bighorn_measurements_crosswalked"):
        for row in read_csv(SOURCES[key]):
            for field in ("selected_hunt_code", "hunt_code", "possible_hunt_codes"):
                for code in split_codes(row.get(field, "")):
                    evidence[code].append({"source_key": key, "matched_code_field": field, **row})

    return db, reg, evidence


def same_family(unresolved: dict[str, str], evidence_row: dict[str, str]) -> bool:
    ev_species = norm(evidence_row.get("species") or evidence_row.get("conservation_species") or evidence_row.get("source_species"))
    if not ev_species:
        return True
    species = norm(unresolved.get("species"))
    return ev_species.lower() in species.lower() or species.lower() in ev_species.lower()


def classify_match(unresolved: dict[str, str], evidence_row: dict[str, str], db_row: dict[str, str]) -> tuple[str, str, str, str, str]:
    source_key = evidence_row.get("source_key", "")
    code = unresolved.get("hunt_code", "")
    db_boundary = norm(db_row.get("boundary_id"))

    if source_key in {"boundary_xlsx", "boundary_id_to_hunt_code", "runtime_boundary_draft"}:
        boundary = norm(evidence_row.get("boundary_id"))
        boundary_match = "YES" if boundary and boundary == db_boundary else ("NO" if boundary and db_boundary else "UNKNOWN")
        status = "CODE_MATCH_BOUNDARY_CONFIRMED" if boundary_match == "YES" else "EXACT_MATCH"
        confidence = "HIGH" if boundary_match == "YES" and same_family(unresolved, evidence_row) else "MEDIUM"
        source_note = {
            "boundary_xlsx": "Current code exists in 2026 boundary crosswalk workbook.",
            "boundary_id_to_hunt_code": "Current code appears in boundary-id-to-hunt-code crosswalk.",
            "runtime_boundary_draft": "Current code appears in runtime draft boundary crosswalk.",
        }[source_key]
        return status, confidence, boundary_match, "", source_note

    if source_key == "retired_current":
        return "HISTORICAL_ONLY", "MANUAL_REVIEW", "UNKNOWN", "", "Code appears in retired-current hunt-code file; treat as retirement evidence, not a permit-value promotion."

    if source_key.startswith("black_bear"):
        status = "CURRENT_ONLY" if evidence_row.get("mapping_status") == "CURRENT_ADMIN_OR_NON_DRAW_ROW" else "EXACT_MATCH"
        confidence = evidence_row.get("mapping_confidence") or "MEDIUM"
        return status, confidence, "UNKNOWN", "", f"Black bear 2024-2026 crosswalk status: {evidence_row.get('mapping_status', '')}."

    if source_key in {"current_historical_primary", "current_historical_processed"}:
        cw_status = evidence_row.get("crosswalk_status", "")
        if cw_status == "PROMOTED_EXACT_HISTORY":
            return "EXACT_MATCH", evidence_row.get("mapping_confidence") or "HIGH", "UNKNOWN", "", "Current-to-historical crosswalk promotes exact history."
        if cw_status == "PROMOTED_PREFIX_SWAP_CANDIDATE":
            return "NEEDS_MANUAL_REMAP", "MANUAL_REVIEW", "UNKNOWN", "", "Prefix-swap candidate needs boundary/family validation before promotion."
        if cw_status == "PROMOTED_PARALLEL_PUBLIC_UNIT_REFERENCE":
            return "BOUNDARY_SUCCESSOR_MATCH", evidence_row.get("mapping_confidence") or "MEDIUM", "UNKNOWN", "MERGED_OR_PARALLEL_REFERENCE", "Parallel public unit reference candidate."
        return "NEEDS_MANUAL_REMAP", "MANUAL_REVIEW", "UNKNOWN", "", f"Crosswalk status {cw_status or 'blank'} requires review."

    if source_key == "historical_le_to_eb":
        status = "RENAMED_MATCH" if evidence_row.get("recommended_eb_code") == code else "NEEDS_MANUAL_REMAP"
        source_confidence = evidence_row.get("mapping_confidence", "")
        confidence = "HIGH" if source_confidence in {"OWNER_RESOLVED", "OWNER_RULE_RESOLVED", "HIGH"} else "MEDIUM"
        return status, confidence, "UNKNOWN", "", f"Historical LE-to-current family mapping: {evidence_row.get('mapping_status', '')}."

    if source_key == "historical_retired":
        if evidence_row.get("current_hunt_code_candidate") == code:
            return "BOUNDARY_SUCCESSOR_MATCH", "LOW", "UNKNOWN", "SUCCESSOR_CANDIDATE", "Historical retired code names this row as current candidate."
        return "HISTORICAL_ONLY", "MANUAL_REVIEW", "UNKNOWN", "", "Historical retired-code source touches this code."

    if source_key == "same_code_promotions":
        return "EXACT_MATCH", "MEDIUM", "UNKNOWN", "", "Same-code active promotion evidence from 2025 PDF crosswalk validation."

    if source_key in {"model_year_crosswalk", "model_year_dropped_review"}:
        if source_key == "model_year_dropped_review" and not evidence_row.get("mapped_hunt_code"):
            return "HISTORICAL_ONLY", "MEDIUM", "UNKNOWN", "", f"Dropped-review status: {evidence_row.get('crosswalk_status', '')}."
        if evidence_row.get("mapped_hunt_code") == code and evidence_row.get("mapped_confidence") == "HIGH":
            boundary_match = "YES" if db_boundary and db_boundary == norm(evidence_row.get("mapped_boundary_id")) else "UNKNOWN"
            status = "CODE_MATCH_BOUNDARY_CONFIRMED" if boundary_match == "YES" else "EXACT_MATCH"
            return status, "HIGH", boundary_match, "", f"Model-year crosswalk status: {evidence_row.get('crosswalk_status', '')}."
        return "NEEDS_MANUAL_REMAP", evidence_row.get("mapped_confidence") or "MANUAL_REVIEW", "UNKNOWN", "", "Model-year crosswalk candidate needs review."

    if source_key == "remaining_boundary_closeout":
        if evidence_row.get("candidate_hunt_code") == code:
            return "BOUNDARY_SUCCESSOR_MATCH", evidence_row.get("confidence") or "MEDIUM", "UNKNOWN", "BOUNDARY_CLOSEOUT_CANDIDATE", f"Boundary closeout review status: {evidence_row.get('review_status', '')}."
        if evidence_row.get("hunt_code") == code and "HISTORICAL_ONLY" in evidence_row.get("review_status", ""):
            return "HISTORICAL_ONLY", evidence_row.get("confidence") or "MEDIUM", "UNKNOWN", "", f"Boundary closeout review status: {evidence_row.get('review_status', '')}."
        return "NEEDS_MANUAL_REMAP", evidence_row.get("confidence") or "MANUAL_REVIEW", "UNKNOWN", "", f"Boundary closeout review status: {evidence_row.get('review_status', '')}."

    if source_key == "conservation_area":
        field = evidence_row.get("matched_code_field", "")
        if field == "primary_hunt_code":
            return "BOUNDARY_SUCCESSOR_MATCH", "MEDIUM", "UNKNOWN", "CONSERVATION_AREA_PRIMARY", "Conservation area primary hunt-code evidence."
        if field in {"included_hunt_codes", "boundary_reference_hunt_codes"}:
            return "MERGED_UNITS", "LOW", "UNKNOWN", "CONSERVATION_AREA_BUNDLE", "Conservation area bundle includes or references this hunt code."
        return "STRUCTURE_CHANGED", "LOW", "UNKNOWN", "CONSERVATION_AREA_BUNDLE", "Conservation-area bundle touches this hunt code."

    if source_key == "boundary_unit_fill":
        return "EXACT_MATCH", "LOW", "UNKNOWN", "", "Unit-fill source provides unit text only."

    if source_key in {"bighorn_location_crosswalk", "bighorn_measurements_crosswalked"}:
        return "NEEDS_MANUAL_REMAP", "LOW", "UNKNOWN", "", "Bighorn harvest-location crosswalk evidence is historical measurement support only; do not use as quota truth."

    return "NEEDS_MANUAL_REMAP", "MANUAL_REVIEW", "UNKNOWN", "", "Unclassified source match."


def permits_from_evidence(evidence_row: dict[str, str], historical: bool) -> tuple[str, str, str]:
    if evidence_row.get("source_key", "").startswith("black_bear"):
        prefix = "permits_2025" if historical else "permits_2026"
        return evidence_row.get(f"{prefix}_res", ""), evidence_row.get(f"{prefix}_nr", ""), evidence_row.get(f"{prefix}_total", "")
    if evidence_row.get("source_key") == "model_year_crosswalk" and historical:
        return "", "", evidence_row.get("source_total_public_draw_permits", "")
    if evidence_row.get("source_key") == "runtime_boundary_draft" and not historical:
        return (
            evidence_row.get("permits_2026_res", ""),
            evidence_row.get("permits_2026_nr", ""),
            evidence_row.get("permits_2026_total", ""),
        )
    return "", "", ""


def make_match_row(unresolved: dict[str, str], evidence_row: dict[str, str], db_row: dict[str, str], reg_row: dict[str, str]) -> dict[str, str]:
    status, confidence, boundary_match, merge_split, notes = classify_match(unresolved, evidence_row, db_row)
    h_res, h_nr, h_total = permits_from_evidence(evidence_row, historical=True)
    c_res, c_nr, c_total = permits_from_evidence(evidence_row, historical=False)
    if not c_total:
        c_res = unresolved.get("recommended_res") or unresolved.get("database_res_reference", "")
        c_nr = unresolved.get("recommended_nr") or unresolved.get("database_nr_reference", "")
        c_total = unresolved.get("recommended_total") or unresolved.get("database_total_reference", "")

    current_unit = evidence_row.get("unit") or evidence_row.get("current_unit_name") or evidence_row.get("hunt_name") or unresolved.get("hunt_name", "")
    historical_unit = evidence_row.get("historical_unit") or evidence_row.get("source_hunt_name") or evidence_row.get("hunt_name_2024", "")

    statewide_flag = "true" if "statewide" in " ".join([unresolved.get("hunt_name", ""), unresolved.get("hunt_type", ""), notes]).lower() else "false"
    unlimited_flag = "true" if "unlimited" in " ".join([unresolved.get("database_total_reference", ""), unresolved.get("recommended_total", ""), notes]).lower() else "false"

    return {
        "current_hunt_code": unresolved.get("hunt_code", ""),
        "historical_hunt_code": evidence_row.get("historical_hunt_code") or evidence_row.get("historical_2025_code") or evidence_row.get("historical_2024_code") or evidence_row.get("source_hunt_code") or evidence_row.get("source_hunt_number", ""),
        "draw_year": evidence_row.get("source_model_target_year", ""),
        "model_year": "2026",
        "historical_hunt_name": evidence_row.get("historical_hunt_name") or evidence_row.get("hunt_name_2025") or evidence_row.get("hunt_name_2024") or evidence_row.get("source_hunt_name", ""),
        "current_hunt_name": unresolved.get("hunt_name", ""),
        "species": unresolved.get("species", ""),
        "sex_type": unresolved.get("sex_type", ""),
        "weapon": unresolved.get("weapon", ""),
        "hunt_type": unresolved.get("hunt_type", ""),
        "historical_unit_name": historical_unit,
        "current_unit_name": current_unit,
        "historical_boundary_id": evidence_row.get("historical_boundary_id_2024") or evidence_row.get("source_boundary_id", ""),
        "current_boundary_id": db_row.get("boundary_id") or evidence_row.get("boundary_id") or evidence_row.get("mapped_boundary_id", ""),
        "boundary_match": boundary_match,
        "permits_res_historical": h_res,
        "permits_nr_historical": h_nr,
        "permits_total_historical": h_total,
        "permits_res_current": c_res,
        "permits_nr_current": c_nr,
        "permits_total_current": c_total,
        "crosswalk_status": status,
        "confidence": confidence,
        "evidence_source": evidence_row.get("source_key", ""),
        "evidence_notes": notes,
        "source_file_historical": evidence_row.get("source_file") or evidence_row.get("source_pdf") or evidence_row.get("mapped_source", ""),
        "source_file_current": str(SOURCES.get(evidence_row.get("source_key", ""), "")),
        "source_priority_historical": "prior_verified_crosswalk_file",
        "source_priority_current": "current_database_or_verified_crosswalk",
        "merge_split_flag": merge_split,
        "statewide_flag": statewide_flag,
        "unlimited_flag": unlimited_flag,
        "source_permit_conflict_flag": "true" if unresolved.get("split_bucket") == "true_source_conflicts" else "false",
        "unresolved_split_bucket": unresolved.get("split_bucket", ""),
        "regulation_presence_status": reg_row.get("regulation_presence_status", ""),
        "recommended_action": unresolved.get("recommended_action", ""),
    }


def make_discrepancy_rows(unresolved_rows: list[dict[str, str]], match_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    by_code = defaultdict(list)
    for row in match_rows:
        by_code[row["current_hunt_code"]].append(row)

    discrepancies: list[dict[str, str]] = []
    for row in unresolved_rows:
        values = {
            "hanumber": row.get("hanumber_total", ""),
            "hunttable": row.get("hunttable_total", ""),
            "utahdraws": row.get("utahdraws_total", ""),
            "database": row.get("database_total_reference", ""),
            "recommended": row.get("recommended_total", ""),
        }
        nonblank = {k: v for k, v in values.items() if v}
        unique = sorted(set(nonblank.values()))
        if row.get("split_bucket") == "true_source_conflicts" or len(unique) > 1:
            discrepancies.append(
                {
                    "hunt_code": row.get("hunt_code", ""),
                    "hunt_name": row.get("hunt_name", ""),
                    "species": row.get("species", ""),
                    "hanumber_total": values["hanumber"],
                    "hunttable_total": values["hunttable"],
                    "utahdraws_total": values["utahdraws"],
                    "database_total": values["database"],
                    "recommended_total": values["recommended"],
                    "evidence_sources": "|".join(sorted({m["evidence_source"] for m in by_code.get(row.get("hunt_code", ""), [])})),
                    "discrepancy_status": "PERMIT_VALUE_CONFLICT",
                    "notes": "Permit totals conflict across current evidence sources; crosswalk evidence must not override the value conflict by itself.",
                }
            )
    return discrepancies


def make_rollup_rows(unresolved_rows: list[dict[str, str]], match_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    by_code: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in match_rows:
        by_code[row["current_hunt_code"]].append(row)

    status_rank = {
        "CODE_MATCH_BOUNDARY_CONFIRMED": 7,
        "RENAMED_MATCH": 6,
        "EXACT_MATCH": 5,
        "BOUNDARY_SUCCESSOR_MATCH": 4,
        "MERGED_UNITS": 3,
        "STRUCTURE_CHANGED": 3,
        "CURRENT_ONLY": 2,
        "NEEDS_MANUAL_REMAP": 1,
        "HISTORICAL_ONLY": 1,
    }
    confidence_rank = {"HIGH": 4, "MEDIUM": 3, "LOW": 2, "MANUAL_REVIEW": 1}

    rollups: list[dict[str, str]] = []
    for unresolved in unresolved_rows:
        code = unresolved.get("hunt_code", "")
        rows = by_code.get(code, [])
        if rows:
            best = max(
                rows,
                key=lambda row: (
                    status_rank.get(row.get("crosswalk_status", ""), 0),
                    confidence_rank.get(row.get("confidence", ""), 0),
                ),
            )
            status_counts = Counter(row.get("crosswalk_status", "") for row in rows)
            sources = sorted({row.get("evidence_source", "") for row in rows if row.get("evidence_source")})
            boundary_confirmed = any(row.get("boundary_match") == "YES" for row in rows)
            has_merge = any(row.get("crosswalk_status") in {"MERGED_UNITS", "SPLIT_UNIT", "STRUCTURE_CHANGED"} or row.get("merge_split_flag") for row in rows)
            has_manual = any(row.get("crosswalk_status") == "NEEDS_MANUAL_REMAP" for row in rows)
        else:
            best = {}
            status_counts = Counter({"NEEDS_MANUAL_REMAP": 1})
            sources = []
            boundary_confirmed = False
            has_merge = False
            has_manual = True

        permit_conflict = unresolved.get("split_bucket") == "true_source_conflicts"
        regulation_present = any(row.get("regulation_presence_status") == "REGULATION_CODE_PRESENT" for row in rows)
        if permit_conflict and boundary_confirmed:
            bucket = "ACTIVE_CODE_CONFIRMED_BUT_PERMIT_CONFLICT_REMAINS"
        elif boundary_confirmed and not permit_conflict:
            bucket = "BOUNDARY_CONFIRMED_IDENTITY_READY"
        elif has_merge:
            bucket = "MERGE_SPLIT_OR_STRUCTURE_REVIEW"
        elif has_manual:
            bucket = "MANUAL_REMAP_REQUIRED"
        elif rows:
            bucket = "WEAK_OR_PARTIAL_EVIDENCE_ONLY"
        else:
            bucket = "NO_SUPPLIED_CROSSWALK_EVIDENCE"

        notes = []
        if boundary_confirmed:
            notes.append("Boundary workbook/current database confirms current-code boundary.")
        if permit_conflict:
            notes.append("Permit values still conflict; identity evidence does not settle numeric winner.")
        if has_merge:
            notes.append("Merge/split/structure evidence is present.")
        if has_manual:
            notes.append("At least one source still calls for manual remap/review.")

        rollups.append(
            {
                "hunt_code": code,
                "hunt_name": unresolved.get("hunt_name", ""),
                "species": unresolved.get("species", ""),
                "sex_type": unresolved.get("sex_type", ""),
                "weapon": unresolved.get("weapon", ""),
                "hunt_type": unresolved.get("hunt_type", ""),
                "best_crosswalk_status": best.get("crosswalk_status", "NEEDS_MANUAL_REMAP"),
                "best_confidence": best.get("confidence", "MANUAL_REVIEW"),
                "recommended_resolution_bucket": bucket,
                "boundary_confirmed": "true" if boundary_confirmed else "false",
                "regulation_present": "true" if regulation_present else "false",
                "permit_conflict": "true" if permit_conflict else "false",
                "evidence_source_count": str(len(sources)),
                "evidence_sources": "|".join(sources),
                "status_counts": "|".join(f"{status}:{count}" for status, count in sorted(status_counts.items())),
                "notes": " ".join(notes),
            }
        )
    return rollups


def main() -> int:
    unresolved_rows = read_csv(UNRESOLVED)
    db, reg, evidence = load_indexes()

    match_rows: list[dict[str, str]] = []
    manual_rows: list[dict[str, str]] = []
    for unresolved in unresolved_rows:
        code = unresolved.get("hunt_code", "")
        db_row = db.get(code, {})
        reg_row = reg.get(code, {})
        matches = evidence.get(code, [])
        if matches:
            for evidence_row in matches:
                match_rows.append(make_match_row(unresolved, evidence_row, db_row, reg_row))
        else:
            manual_rows.append(
                {
                    "current_hunt_code": code,
                    "historical_hunt_code": "",
                    "draw_year": "",
                    "model_year": "2026",
                    "historical_hunt_name": "",
                    "current_hunt_name": unresolved.get("hunt_name", ""),
                    "species": unresolved.get("species", ""),
                    "sex_type": unresolved.get("sex_type", ""),
                    "weapon": unresolved.get("weapon", ""),
                    "hunt_type": unresolved.get("hunt_type", ""),
                    "historical_unit_name": "",
                    "current_unit_name": unresolved.get("hunt_name", ""),
                    "historical_boundary_id": "",
                    "current_boundary_id": db_row.get("boundary_id", ""),
                    "boundary_match": "UNKNOWN",
                    "permits_res_historical": "",
                    "permits_nr_historical": "",
                    "permits_total_historical": "",
                    "permits_res_current": unresolved.get("recommended_res") or unresolved.get("database_res_reference", ""),
                    "permits_nr_current": unresolved.get("recommended_nr") or unresolved.get("database_nr_reference", ""),
                    "permits_total_current": unresolved.get("recommended_total") or unresolved.get("database_total_reference", ""),
                    "crosswalk_status": "NEEDS_MANUAL_REMAP",
                    "confidence": "MANUAL_REVIEW",
                    "evidence_source": "",
                    "evidence_notes": "No supplied crosswalk source matched this unresolved code.",
                    "source_file_historical": "",
                    "source_file_current": "",
                    "source_priority_historical": "",
                    "source_priority_current": "current_unresolved_queue",
                    "merge_split_flag": "",
                    "statewide_flag": "false",
                    "unlimited_flag": "false",
                    "unresolved_split_bucket": unresolved.get("split_bucket", ""),
                    "regulation_presence_status": reg_row.get("regulation_presence_status", ""),
                    "recommended_action": unresolved.get("recommended_action", ""),
                }
            )

    merge_split_rows = [
        row for row in match_rows if row.get("crosswalk_status") in {"MERGED_UNITS", "SPLIT_UNIT", "STRUCTURE_CHANGED"} or row.get("merge_split_flag")
    ]
    statewide_rows = [
        row for row in match_rows if row.get("crosswalk_status") in {"STATEWIDE_SUCCESSOR", "UNLIMITED_SUCCESSOR"} or row.get("statewide_flag") == "true" or row.get("unlimited_flag") == "true"
    ]
    discrepancy_rows = make_discrepancy_rows(unresolved_rows, match_rows)
    rollup_rows = make_rollup_rows(unresolved_rows, match_rows)
    manual_remap_rows = [
        row for row in rollup_rows if row.get("recommended_resolution_bucket") == "MANUAL_REMAP_REQUIRED"
    ]
    no_source_manual_rows = manual_rows

    write_csv(OUT_DIR / "crosswalk_2026_additional_source_candidates.csv", match_rows, MATCH_COLUMNS)
    write_csv(OUT_DIR / "crosswalk_2026_code_resolution_rollup.csv", rollup_rows, ROLLUP_COLUMNS)
    write_csv(OUT_DIR / "crosswalk_2026_manual_remap_required.csv", manual_remap_rows, ROLLUP_COLUMNS)
    write_csv(OUT_DIR / "crosswalk_2026_merge_split_structure_changes.csv", merge_split_rows, MATCH_COLUMNS)
    write_csv(OUT_DIR / "crosswalk_2026_statewide_unlimited_successors.csv", statewide_rows, MATCH_COLUMNS)
    write_csv(
        OUT_DIR / "crosswalk_2026_permit_or_boundary_discrepancies.csv",
        discrepancy_rows,
        [
            "hunt_code",
            "hunt_name",
            "species",
            "hanumber_total",
            "hunttable_total",
            "utahdraws_total",
            "database_total",
            "recommended_total",
            "evidence_sources",
            "discrepancy_status",
            "notes",
        ],
    )

    matched_codes = {row["current_hunt_code"] for row in match_rows}
    summary = {
        "created_at_utc": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        "directive": "Evidence-first crosswalk: code + boundary + family meaning + source evidence; no name-only promotion.",
        "input_rows": len(unresolved_rows),
        "unique_input_codes": len({row.get("hunt_code", "") for row in unresolved_rows}),
        "matched_unique_codes": len(matched_codes),
        "manual_unique_codes": len({row["hunt_code"] for row in manual_remap_rows}),
        "match_evidence_rows": len(match_rows),
        "manual_rows": len(manual_remap_rows),
        "no_source_manual_rows": len(no_source_manual_rows),
        "merge_split_rows": len(merge_split_rows),
        "statewide_unlimited_rows": len(statewide_rows),
        "discrepancy_rows": len(discrepancy_rows),
        "rollup_bucket_counts": dict(Counter(row["recommended_resolution_bucket"] for row in rollup_rows)),
        "match_status_counts": dict(Counter(row["crosswalk_status"] for row in match_rows)),
        "match_confidence_counts": dict(Counter(row["confidence"] for row in match_rows)),
        "match_source_counts": dict(Counter(row["evidence_source"] for row in match_rows)),
        "manual_by_species": dict(Counter(row["species"] for row in manual_remap_rows)),
        "source_files": {key: str(path) for key, path in SOURCES.items()},
        "outputs": {
            "crosswalk_table": "processed_data/audits/current_2026_permit_unresolved_split/crosswalk_2026_additional_source_candidates.csv",
            "code_resolution_rollup": "processed_data/audits/current_2026_permit_unresolved_split/crosswalk_2026_code_resolution_rollup.csv",
            "unresolved_manual_remap_table": "processed_data/audits/current_2026_permit_unresolved_split/crosswalk_2026_manual_remap_required.csv",
            "merge_split_table": "processed_data/audits/current_2026_permit_unresolved_split/crosswalk_2026_merge_split_structure_changes.csv",
            "statewide_unlimited_table": "processed_data/audits/current_2026_permit_unresolved_split/crosswalk_2026_statewide_unlimited_successors.csv",
            "discrepancy_table": "processed_data/audits/current_2026_permit_unresolved_split/crosswalk_2026_permit_or_boundary_discrepancies.csv",
            "validation_summary": "processed_data/audits/current_2026_permit_unresolved_split/crosswalk_2026_additional_sources_summary.json",
        },
        "validation": {
            "no_database_values_modified": True,
            "row_preserving_note": "Candidate table may contain multiple evidence rows per hunt code; manual table contains codes whose rollup still requires manual remap.",
            "boundary_rule_note": "HIGH confidence is only assigned where source confidence was high or current boundary workbook confirmed the current code boundary.",
        },
    }
    (OUT_DIR / "crosswalk_2026_additional_sources_summary.json").write_text(
        json.dumps(summary, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    print(json.dumps(summary, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
