from __future__ import annotations

import argparse
import json
import re
import shutil
from copy import copy
from datetime import datetime, timezone
from pathlib import Path

import openpyxl


IDENTITY_COLS = range(1, 8)
PERMIT_COLS = (8, 9, 10)
DEFAULT_HEADERS = [
    "hunt_name",
    "hunt_code",
    "sex_type",
    "species",
    "weapon",
    "hunt_type",
    "season",
    "2026_permits_res",
    "2026_permits_nr",
    "2026_permits_total",
]


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def parse_labeled_number(value: object) -> tuple[str | None, int | None]:
    text = clean(value).replace(",", "")
    if not text:
        return None, None
    patterns = [
        ("res", r"^(Res|Resident)\s*:\s*(-?\d+(?:\.\d+)?)$"),
        ("nr", r"^(Non\s*Res|NonRes|Nonresident|Nonresident permits)\s*:\s*(-?\d+(?:\.\d+)?)$"),
        ("total", r"^(Total|Total permits)\s*:\s*(-?\d+(?:\.\d+)?)$"),
        ("number", r"^(-?\d+(?:\.\d+)?)$"),
    ]
    for label, pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if not match:
            continue
        raw = match.group(2) if label != "number" else match.group(1)
        return label, int(float(raw))
    return None, None


def copy_cell_style(source: object, target: object, number_format: str | None = None) -> None:
    if source.has_style:
        target._style = copy(source._style)
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.number_format = number_format or source.number_format


def identity_values(sheet: object, row_num: int) -> list[object]:
    return [sheet.cell(row_num, column).value for column in IDENTITY_COLS]


def has_identity(sheet: object, row_num: int) -> bool:
    return any(clean(sheet.cell(row_num, column).value) for column in IDENTITY_COLS)


def has_permit_value(sheet: object, row_num: int) -> bool:
    return any(clean(sheet.cell(row_num, column).value) for column in PERMIT_COLS)


def is_nonresident_continuation(sheet: object, row_num: int) -> bool:
    if row_num > sheet.max_row or has_identity(sheet, row_num):
        return False
    label, number = parse_labeled_number(sheet.cell(row_num, 8).value)
    return label == "nr" and number is not None and not clean(sheet.cell(row_num, 9).value) and not clean(sheet.cell(row_num, 10).value)


def extract_permits(sheet: object, row_num: int) -> tuple[int | None, int | None, int | None]:
    res: int | None = None
    nr: int | None = None
    total: int | None = None
    for column in PERMIT_COLS:
        label, number = parse_labeled_number(sheet.cell(row_num, column).value)
        if number is None:
            continue
        if label == "res" or (label == "number" and column == 8):
            res = number
        elif label == "nr" or (label == "number" and column == 9):
            nr = number
        elif label == "total" or (label == "number" and column == 10):
            total = number
    return res, nr, total


def normalize_workbook(workbook_path: Path, summary_out: Path | None = None, sheet_name: str | None = None) -> dict[str, object]:
    if not workbook_path.exists():
        raise FileNotFoundError(workbook_path)

    backup_path = workbook_path.with_name(f"{workbook_path.stem}.before_normalize_permits{workbook_path.suffix}")
    if not backup_path.exists():
        shutil.copy2(workbook_path, backup_path)

    source_workbook = openpyxl.load_workbook(backup_path)
    source_sheet = source_workbook[sheet_name] if sheet_name else source_workbook.active

    output_workbook = openpyxl.Workbook()
    output_sheet = output_workbook.active
    output_sheet.title = source_sheet.title
    output_sheet.freeze_panes = "A2"

    for column in range(1, source_sheet.max_column + 1):
        letter = openpyxl.utils.get_column_letter(column)
        output_sheet.column_dimensions[letter].width = source_sheet.column_dimensions[letter].width

    output_sheet.append(DEFAULT_HEADERS)
    for column in range(1, len(DEFAULT_HEADERS) + 1):
        copy_cell_style(source_sheet.cell(1, column), output_sheet.cell(1, column))

    collapsed = 0
    skipped_continuations = 0
    total_only = 0
    output_rows = 0
    row_num = 2
    while row_num <= source_sheet.max_row:
        if not has_identity(source_sheet, row_num) and not has_permit_value(source_sheet, row_num):
            row_num += 1
            continue
        if is_nonresident_continuation(source_sheet, row_num):
            skipped_continuations += 1
            row_num += 1
            continue

        res, nr, total = extract_permits(source_sheet, row_num)
        skip_next = False
        if is_nonresident_continuation(source_sheet, row_num + 1):
            _, nr_value = parse_labeled_number(source_sheet.cell(row_num + 1, 8).value)
            nr = nr_value
            collapsed += 1
            skip_next = True

        if res is not None and nr is not None:
            total = res + nr
        elif total is None and res is not None:
            total = res
        elif total is None and nr is not None:
            total = nr
        if res is None and nr is None and total is not None:
            total_only += 1

        output_sheet.append([*identity_values(source_sheet, row_num), res, nr, total])
        output_rows += 1
        destination_row = output_sheet.max_row
        for column in range(1, len(DEFAULT_HEADERS) + 1):
            number_format = "0" if column in PERMIT_COLS else None
            copy_cell_style(source_sheet.cell(row_num, column), output_sheet.cell(destination_row, column), number_format)

        row_num += 2 if skip_next else 1

    output_sheet.auto_filter.ref = f"A1:J{output_sheet.max_row}"
    output_workbook.save(workbook_path)

    validation = validate_workbook(workbook_path)
    summary: dict[str, object] = {
        "created_at_utc": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        "workbook_path": str(workbook_path),
        "backup_path": str(backup_path),
        "sheet_name": source_sheet.title,
        "source_rows": source_sheet.max_row,
        "rows_after": validation["rows_after"],
        "data_rows_after": validation["rows_after"] - 1,
        "output_rows_from_source": output_rows,
        "nonresident_continuation_rows_collapsed": collapsed,
        "stray_continuation_rows_skipped": skipped_continuations,
        "rows_with_permit_values_after": validation["rows_with_permit_values"],
        "total_only_rows_after": total_only,
        "blank_identity_rows_with_permit_values_after": validation["blank_identity_rows_with_permit_values"],
        "sum_validation_failure_count": len(validation["sum_validation_failures"]),
        "sum_validation_failures": validation["sum_validation_failures"],
    }
    if summary_out:
        summary_out.parent.mkdir(parents=True, exist_ok=True)
        summary_out.write_text(json.dumps(summary, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return summary


def validate_workbook(workbook_path: Path) -> dict[str, object]:
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    sheet = workbook.active
    blank_identity_rows: list[int] = []
    sum_failures: list[dict[str, object]] = []
    rows_with_values = 0
    for row_num in range(2, sheet.max_row + 1):
        identity = [clean(sheet.cell(row_num, column).value) for column in IDENTITY_COLS]
        values = [sheet.cell(row_num, column).value for column in PERMIT_COLS]
        if any(value is not None for value in values):
            rows_with_values += 1
        if not any(identity) and any(value is not None for value in values):
            blank_identity_rows.append(row_num)
        res, nr, total = values
        if res is None or nr is None:
            continue
        try:
            if int(res) + int(nr) != int(total):
                sum_failures.append({"row": row_num, "hunt_code": sheet.cell(row_num, 2).value, "res": res, "nr": nr, "total": total})
        except Exception:
            sum_failures.append({"row": row_num, "hunt_code": sheet.cell(row_num, 2).value, "res": res, "nr": nr, "total": total})
    if blank_identity_rows or sum_failures:
        raise RuntimeError(
            json.dumps(
                {
                    "blank_identity_rows_with_permit_values": blank_identity_rows[:20],
                    "sum_validation_failures": sum_failures[:20],
                    "sum_validation_failure_count": len(sum_failures),
                },
                indent=2,
            )
        )
    return {
        "rows_after": sheet.max_row,
        "rows_with_permit_values": rows_with_values,
        "blank_identity_rows_with_permit_values": blank_identity_rows,
        "sum_validation_failures": sum_failures,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Normalize permit workbooks that store NonRes values on blank continuation rows.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--sheet-name", default=None)
    parser.add_argument("--summary-out", type=Path, default=None)
    args = parser.parse_args()
    summary = normalize_workbook(args.workbook, args.summary_out, args.sheet_name)
    print(json.dumps(summary, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
