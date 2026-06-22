from __future__ import annotations

import csv
import shutil
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "processed_data" / "hard_data_exports" / "hunt_tables" / "2026" / "CLEAN_XLXS_STAGED" / "2025-27 Conservation Permits.xlsx"
DATABASE_PATH = ROOT / "pipeline" / "RAW" / "hunt_unit_database" / "2026" / "csv" / "DATABASE.csv"
BOUNDARY_DIR = ROOT / "processed_data" / "boundaries"
AUDIT_DIR = ROOT / "audits" / "2025_canonical_finalization"
BACKUP_DIR = AUDIT_DIR / "backups"
AUDIT_PATH = AUDIT_DIR / "2025_27_conservation_visible_prefix_mismatch_fixes.csv"


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def read_db() -> dict[str, dict[str, str]]:
    with DATABASE_PATH.open("r", encoding="utf-8-sig", newline="") as f:
        return {clean(row.get("hunt_code")).upper(): {k: clean(v) for k, v in row.items()} for row in csv.DictReader(f) if clean(row.get("hunt_code"))}


def geojson_for(code: str, boundary_id: str) -> str:
    code_path = BOUNDARY_DIR / f"{code}.geojson"
    if code_path.exists():
        return code_path.relative_to(ROOT).as_posix()
    if boundary_id:
        id_path = BOUNDARY_DIR / f"{boundary_id}.geojson"
        if id_path.exists():
            return id_path.relative_to(ROOT).as_posix()
    return ""


def explicit_code(species: str, hunt_name: str, sex_type: str) -> str:
    key = (clean(species), clean(hunt_name), clean(sex_type))
    explicit = {
        ("Bison", "Book Cliffs, Bitter Creek, Hunter's Choice", "Hunter's Choice"): "BI6534",
        ("Bison", "Book Cliffs, Little Creek/South, Hunter's Choice", "Hunter's Choice"): "BI6531",
        ("Bison", "Henry Mtns, Hunter's Choice (early)", "Hunter's Choice"): "BI6503",
        ("Bison", "Henry Mtns, Cow Only (late)", "Cow Only"): "BI6505",
        ("Desert Bighorn Sheep", "La Sal, Potash/South Cisco", "Male Only"): "DS6604",
        ("Desert Bighorn Sheep", "Statewide", "Male Only"): "DS1000",
        ("Deer", "West Desert, Vernon", "Buck"): "DB1048",
        ("Deer", "Statewide", "Buck"): "DB0007",
        ("Pronghorn", "Book Cliffs, South", "Buck"): "PB5027",
        ("Pronghorn", "Box Elder, Snowville", "Buck"): "PB5030",
        ("Pronghorn", "Box Elder, West", "Buck"): "PB5031",
        ("Turkey", "Northeastern Area", "Bearded"): "TK1013",
        ("Turkey", "Northern Area", "Bearded"): "TK1014",
        ("Turkey", "Southeastern Area", "Bearded"): "TK1015",
    }
    return explicit.get(key, "")


def main() -> None:
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)

    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    backup_path = BACKUP_DIR / f"{WORKBOOK_PATH.stem}.before_visible_prefix_mismatch_fixes_{stamp}{WORKBOOK_PATH.suffix}"
    shutil.copy2(WORKBOOK_PATH, backup_path)

    db = read_db()
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb["Table 1"] if "Table 1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [clean(ws.cell(1, column).value) for column in range(1, ws.max_column + 1)]
    cols = {header: index + 1 for index, header in enumerate(headers) if header}

    audit_rows = []
    for row_idx in range(2, ws.max_row + 1):
        species = clean(ws.cell(row_idx, cols["Species"]).value)
        hunt_name = clean(ws.cell(row_idx, cols["HUNT NAME"]).value)
        sex_type = clean(ws.cell(row_idx, cols["SEX TYPE"]).value)
        code = explicit_code(species, hunt_name, sex_type)
        if not code:
            continue
        db_row = db.get(code, {})
        if not db_row:
            audit_rows.append({"row": row_idx, "status": "code_not_in_database", "new_hunt_code": code})
            continue

        old_code = clean(ws.cell(row_idx, cols["HUNT CODE"]).value)
        old_boundary = clean(ws.cell(row_idx, cols["BOUNDARY ID"]).value)
        old_map = clean(ws.cell(row_idx, cols["MAP GEOJSON"]).value)
        boundary_id = clean(db_row.get("boundary_id")) or old_boundary
        map_geojson = geojson_for(code, boundary_id) or old_map

        if old_code == code and old_boundary == boundary_id and old_map == map_geojson:
            continue

        ws.cell(row_idx, cols["HUNT CODE"]).value = code
        ws.cell(row_idx, cols["BOUNDARY ID"]).value = boundary_id
        ws.cell(row_idx, cols["MAP GEOJSON"]).value = map_geojson
        audit_rows.append(
            {
                "row": row_idx,
                "no": ws.cell(row_idx, cols["No."]).value,
                "species": species,
                "hunt_name": hunt_name,
                "sex_type": sex_type,
                "old_hunt_code": old_code,
                "new_hunt_code": code,
                "old_boundary_id": old_boundary,
                "new_boundary_id": boundary_id,
                "old_map_geojson": old_map,
                "new_map_geojson": map_geojson,
                "status": "patched",
            }
        )

    wb.save(WORKBOOK_PATH)

    fieldnames = [
        "row",
        "no",
        "species",
        "hunt_name",
        "sex_type",
        "old_hunt_code",
        "new_hunt_code",
        "old_boundary_id",
        "new_boundary_id",
        "old_map_geojson",
        "new_map_geojson",
        "status",
    ]
    with AUDIT_PATH.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(audit_rows)

    print(f"patched={sum(1 for row in audit_rows if row.get('status') == 'patched')}")
    print(f"backup={backup_path}")
    print(f"audit={AUDIT_PATH}")


if __name__ == "__main__":
    main()
