"""Reconcile permit values from the 2026 PERMITS workbook into DATABASE.csv.

Rules from Tyler:
- Align by the year-specific hunt code, not workbook row position.
- Workbook values take priority for replacement/fill.
- If any numeric field for a code/year differs by more than 10 permits, hold
  that code/year and report it instead of applying it.
"""

from __future__ import annotations

import csv
import json
import shutil
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATABASE = ROOT / "pipeline/RAW/hunt_unit_database/2026/csv/DATABASE.csv"
WORKBOOK = Path(r"C:\Users\tyler\Desktop\2026 PERMITS.xlsx")

OUT_AUDIT = ROOT / "processed_data/audits/2026_permits_workbook_database_reconciliation.csv"
OUT_SUMMARY = ROOT / "processed_data/audits/2026_permits_workbook_database_reconciliation_summary.json"
OUT_DOC = ROOT / "processed_data/audits/2026_permits_workbook_database_reconciliation.md"

WORKBOOK_SOURCE_LABEL = "2026_PERMITS_WORKBOOK_BY_HUNT_CODE"

YEAR_MAP = {
    2026: {"code": 2, "res": 10, "nr": 11, "total": 12},
    2025: {"code": 14, "res": 17, "nr": 18, "total": 19},
    2024: {"code": 23, "res": 24, "nr": 25, "total": 26},
    2023: {"code": 30, "res": 31, "nr": 32, "total": 33},
    2022: {"code": 37, "res": 38, "nr": 39, "total": 40},
    2021: {"code": 44, "res": 45, "nr": 46, "total": 47},
    2020: {"code": 51, "res": 52, "nr": 53, "total": 54},
}


def clean(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\ufeff", "").strip().split())


def txt(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    text = clean(value)
    return "" if text.lower() in {"nan", "none"} else text


def to_int(value: str) -> int | None:
    value = clean(value).replace(",", "")
    if not value:
        return None
    try:
        return int(float(value))
    except ValueError:
        return None


def read_csv(path: Path) -> tuple[list[dict[str, str]], list[str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        rows = [{key: clean(val) for key, val in row.items()} for row in reader]
        return rows, list(reader.fieldnames or [])


def write_csv(path: Path, rows: list[dict[str, object]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def backup_database() -> Path:
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    backup = ROOT / "processed_data/backups" / f"DATABASE_before_2026_permits_workbook_patch_{stamp}.csv"
    backup.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(DATABASE, backup)
    return backup


def workbook_rows() -> list[tuple]:
    wb = load_workbook(WORKBOOK, data_only=True, read_only=True)
    ws = wb["2026 Summary"]
    return list(ws.iter_rows(values_only=True))


def classify_code_year(
    year: int,
    code: str,
    wb_vals: dict[str, str],
    db_row: dict[str, str] | None,
) -> dict[str, object]:
    result = {
        "year": year,
        "hunt_code": code,
        "wb_res": wb_vals["res"],
        "wb_nr": wb_vals["nr"],
        "wb_total": wb_vals["total"],
        "db_res": "",
        "db_nr": "",
        "db_total": "",
        "hunt_name": "",
        "species": "",
        "action": "MATCH",
        "changed_fields": [],
        "notes": [],
        "gt10_fields": [],
    }

    if db_row is None:
        result["action"] = "MISSING_DB_CODE"
        result["notes"].append("Year-specific workbook hunt code is not present in DATABASE.csv.")
        return result

    result["hunt_name"] = clean(db_row.get("hunt_name"))
    result["species"] = clean(db_row.get("species"))
    result["db_res"] = clean(db_row.get(f"permits_{year}_res"))
    result["db_nr"] = clean(db_row.get(f"permits_{year}_nr"))
    result["db_total"] = clean(db_row.get(f"permits_{year}_total"))

    for field in ("res", "nr", "total"):
        wb_value = wb_vals[field]
        db_value = result[f"db_{field}"]
        if not wb_value or wb_value == db_value:
            continue
        wb_num = to_int(wb_value)
        db_num = to_int(db_value)
        if wb_num is not None and db_num is not None:
            delta = abs(wb_num - db_num)
            if delta > 10:
                result["gt10_fields"].append({"field": field, "delta": delta})

    if result["gt10_fields"]:
        result["action"] = "HOLD_GT10_DELTA"
        result["notes"].append("At least one year-field delta is greater than 10 permits.")
        return result

    changed_fields: list[str] = []
    for field in ("res", "nr", "total"):
        wb_value = wb_vals[field]
        db_value = result[f"db_{field}"]
        if wb_value and wb_value != db_value:
            changed_fields.append(field)

    if not changed_fields:
        result["action"] = "MATCH"
        return result

    result["changed_fields"] = changed_fields
    workbook_split_blank = not wb_vals["res"] and not wb_vals["nr"] and bool(wb_vals["total"])
    if workbook_split_blank and changed_fields == ["total"]:
        result["action"] = "APPLY_TOTAL_ONLY"
        result["notes"].append("Workbook provides total only; applying total without overwriting res/nr blanks.")
    elif all(not result[f"db_{field}"] and wb_vals[field] for field in changed_fields):
        result["action"] = "APPLY_FILL"
        result["notes"].append("Workbook fills previously blank database permit fields.")
    else:
        result["action"] = "APPLY_REPLACE"
        result["notes"].append("Workbook replaces database values with no field delta above 10 permits.")
    return result


def apply_changes(db_row: dict[str, str], year: int, wb_vals: dict[str, str], action: str) -> None:
    changed_any = False
    for field in ("res", "nr", "total"):
        wb_value = wb_vals[field]
        if not wb_value:
            continue
        target_field = f"permits_{year}_{field}"
        if clean(db_row.get(target_field)) != wb_value:
            db_row[target_field] = wb_value
            changed_any = True
    source_field = f"permits_{year}_source"
    if changed_any and source_field in db_row:
        db_row[source_field] = WORKBOOK_SOURCE_LABEL
    if changed_any and year == 2026 and "permits_2026_draw_source" in db_row:
        db_row["permits_2026_draw_source"] = WORKBOOK_SOURCE_LABEL


def main() -> int:
    db_rows, db_fields = read_csv(DATABASE)
    db_by_code = {clean(row.get("hunt_code")).upper(): row for row in db_rows if clean(row.get("hunt_code"))}
    wb_rows = workbook_rows()

    audit_rows: list[dict[str, object]] = []
    apply_counter: Counter[str] = Counter()

    backup = backup_database()

    for row in wb_rows[1:]:
        for year, idx in YEAR_MAP.items():
            code = txt(row[idx["code"]]).upper()
            if not code:
                continue
            wb_vals = {field: txt(row[idx[field]]) for field in ("res", "nr", "total")}
            if not any(wb_vals.values()):
                continue
            db_row = db_by_code.get(code)
            classified = classify_code_year(year, code, wb_vals, db_row)
            action = classified["action"]
            apply_counter[action] += 1
            if db_row is not None and action in {"APPLY_FILL", "APPLY_REPLACE", "APPLY_TOTAL_ONLY"}:
                apply_changes(db_row, year, wb_vals, action)
            audit_rows.append(
                {
                    "year": year,
                    "hunt_code": code,
                    "hunt_name": classified["hunt_name"],
                    "species": classified["species"],
                    "db_res_before": classified["db_res"],
                    "db_nr_before": classified["db_nr"],
                    "db_total_before": classified["db_total"],
                    "wb_res": classified["wb_res"],
                    "wb_nr": classified["wb_nr"],
                    "wb_total": classified["wb_total"],
                    "action": action,
                    "changed_fields": "|".join(classified["changed_fields"]),
                    "gt10_fields": json.dumps(classified["gt10_fields"], separators=(",", ":")),
                    "notes": " ".join(classified["notes"]),
                }
            )

    with DATABASE.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=db_fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(db_rows)

    audit_fields = [
        "year",
        "hunt_code",
        "hunt_name",
        "species",
        "db_res_before",
        "db_nr_before",
        "db_total_before",
        "wb_res",
        "wb_nr",
        "wb_total",
        "action",
        "changed_fields",
        "gt10_fields",
        "notes",
    ]
    write_csv(OUT_AUDIT, audit_rows, audit_fields)

    gt10_rows = [row for row in audit_rows if row["action"] == "HOLD_GT10_DELTA"]
    missing_rows = [row for row in audit_rows if row["action"] == "MISSING_DB_CODE"]
    applied_rows = [row for row in audit_rows if row["action"] in {"APPLY_FILL", "APPLY_REPLACE", "APPLY_TOTAL_ONLY"}]

    gt10_by_year = defaultdict(int)
    for row in gt10_rows:
        gt10_by_year[int(row["year"])] += 1

    summary = {
        "created_at_utc": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        "workbook_path": str(WORKBOOK),
        "database_path": DATABASE.relative_to(ROOT).as_posix(),
        "backup_path": backup.relative_to(ROOT).as_posix(),
        "source_label": WORKBOOK_SOURCE_LABEL,
        "action_counts": dict(sorted(apply_counter.items())),
        "applied_code_year_rows": len(applied_rows),
        "hold_gt10_code_year_rows": len(gt10_rows),
        "missing_db_code_year_rows": len(missing_rows),
        "hold_gt10_by_year": dict(sorted(gt10_by_year.items())),
        "outputs": {
            "audit_csv": OUT_AUDIT.relative_to(ROOT).as_posix(),
            "summary_json": OUT_SUMMARY.relative_to(ROOT).as_posix(),
            "report_md": OUT_DOC.relative_to(ROOT).as_posix(),
        },
        "guardrail": "Any code/year with a numeric permit delta greater than 10 was held out of the database patch and only reported.",
    }
    OUT_SUMMARY.write_text(json.dumps(summary, indent=2, sort_keys=True) + "\n", encoding="utf-8")

    gt10_preview = [
        f"- `{row['year']}` `{row['hunt_code']}` {row['hunt_name']} | DB `{row['db_res_before']}/{row['db_nr_before']}/{row['db_total_before']}` | WB `{row['wb_res']}/{row['wb_nr']}/{row['wb_total']}` | {row['gt10_fields']}"
        for row in gt10_rows[:25]
    ]
    missing_preview = [
        f"- `{row['year']}` `{row['hunt_code']}` | WB `{row['wb_res']}/{row['wb_nr']}/{row['wb_total']}`"
        for row in missing_rows[:25]
    ]
    report_lines = [
        "# 2026 PERMITS Workbook Reconciliation To DATABASE",
        "",
        "Applied workbook permit values to DATABASE.csv by year-specific hunt code, not workbook row order.",
        "",
        "## Applied",
        "",
        f"- Applied code/year rows: `{len(applied_rows)}`",
        f"- Action counts: `{dict(sorted(apply_counter.items()))}`",
        "",
        "## Held For Review (>10 Delta)",
        "",
    ]
    report_lines.extend(gt10_preview if gt10_preview else ["- None"])
    report_lines.extend(
        [
            "",
            "## Missing DB Codes",
            "",
        ]
    )
    report_lines.extend(missing_preview if missing_preview else ["- None"])
    report_lines.extend(
        [
            "",
            "Guardrail: workbook values were not auto-applied when any numeric field delta exceeded 10 permits for that code/year.",
        ]
    )
    OUT_DOC.write_text("\n".join(report_lines) + "\n", encoding="utf-8")

    print(json.dumps(summary, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
