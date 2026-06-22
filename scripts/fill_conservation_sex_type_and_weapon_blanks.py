from __future__ import annotations

import csv
import shutil
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "processed_data" / "hard_data_exports" / "hunt_tables" / "2026" / "CLEAN_XLXS_STAGED" / "2025-27 Conservation Permits.xlsx"
PRE_HEADER_BACKUP = ROOT / "audits" / "2025_canonical_finalization" / "backups" / "2025-27 Conservation Permits.before_header_standardization_20260621T211542Z.xlsx"
AUDIT_DIR = ROOT / "audits" / "2025_canonical_finalization"
BACKUP_DIR = AUDIT_DIR / "backups"
AUDIT_PATH = AUDIT_DIR / "2025_27_conservation_sex_type_weapon_fill_audit.csv"


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def sex_type_for(species: str, hunt_name: str) -> str:
    species_text = clean(species).lower()
    hunt_text = clean(hunt_name).lower()
    if species_text == "antlerless elk":
        return "Antlerless"
    if species_text == "elk":
        return "Bull"
    if species_text == "deer":
        return "Buck"
    if species_text == "bear":
        return "Either Sex"
    if species_text == "bison":
        if "cow only" in hunt_text:
            return "Cow Only"
        return "Hunter's Choice"
    if species_text == "moose":
        return "Bull"
    if species_text == "mountain goat":
        return "Either Sex"
    if species_text == "pronghorn":
        return "Buck"
    if species_text in {"desert bighorn sheep", "rocky mountain bighorn sheep"}:
        return "Male Only"
    if species_text == "turkey":
        return "Bearded"
    return ""


def read_backup_weapon_by_no() -> dict[int, str]:
    if not PRE_HEADER_BACKUP.exists():
        return {}
    wb = load_workbook(PRE_HEADER_BACKUP, read_only=True, data_only=True)
    ws = wb["Table 1"] if "Table 1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [clean(ws.cell(1, column).value) for column in range(1, ws.max_column + 1)]
    cols = {header: index + 1 for index, header in enumerate(headers) if header}
    if "No." not in cols or "WEAPON" not in cols:
        return {}
    result: dict[int, str] = {}
    for row_idx in range(2, ws.max_row + 1):
        no_text = clean(ws.cell(row_idx, cols["No."]).value)
        if not no_text:
            continue
        weapon = clean(ws.cell(row_idx, cols["WEAPON"]).value)
        if weapon:
            result[int(float(no_text))] = weapon
    return result


def main() -> None:
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)

    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    backup_path = BACKUP_DIR / f"{WORKBOOK_PATH.stem}.before_sex_type_weapon_fill_{stamp}{WORKBOOK_PATH.suffix}"
    shutil.copy2(WORKBOOK_PATH, backup_path)

    backup_weapon_by_no = read_backup_weapon_by_no()
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb["Table 1"] if "Table 1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [clean(ws.cell(1, column).value) for column in range(1, ws.max_column + 1)]
    cols = {header: index + 1 for index, header in enumerate(headers) if header}

    required = ["No.", "Species", "HUNT NAME", "SEX TYPE", "WEAPON"]
    missing = [header for header in required if header not in cols]
    if missing:
        raise RuntimeError(f"Missing required headers: {missing}")

    audit_rows = []
    for row_idx in range(2, ws.max_row + 1):
        no_text = clean(ws.cell(row_idx, cols["No."]).value)
        if not no_text:
            continue
        no_value = int(float(no_text))
        species = clean(ws.cell(row_idx, cols["Species"]).value)
        hunt_name = clean(ws.cell(row_idx, cols["HUNT NAME"]).value)

        old_sex = clean(ws.cell(row_idx, cols["SEX TYPE"]).value)
        new_sex = sex_type_for(species, hunt_name)
        if new_sex and old_sex != new_sex:
            ws.cell(row_idx, cols["SEX TYPE"]).value = new_sex
            audit_rows.append(
                {
                    "row": row_idx,
                    "no": no_value,
                    "field": "SEX TYPE",
                    "old_value": old_sex,
                    "new_value": new_sex,
                    "reason": "species_rule",
                }
            )

        old_weapon = clean(ws.cell(row_idx, cols["WEAPON"]).value)
        fallback_weapon = backup_weapon_by_no.get(no_value, "")
        if not old_weapon and fallback_weapon:
            ws.cell(row_idx, cols["WEAPON"]).value = fallback_weapon
            audit_rows.append(
                {
                    "row": row_idx,
                    "no": no_value,
                    "field": "WEAPON",
                    "old_value": old_weapon,
                    "new_value": fallback_weapon,
                    "reason": "pre_header_enrichment_fallback",
                }
            )

    wb.save(WORKBOOK_PATH)

    with AUDIT_PATH.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["row", "no", "field", "old_value", "new_value", "reason"])
        writer.writeheader()
        writer.writerows(audit_rows)

    print(f"changes={len(audit_rows)}")
    print(f"backup={backup_path}")
    print(f"audit={AUDIT_PATH}")


if __name__ == "__main__":
    main()
