from __future__ import annotations

import csv
import hashlib
import json
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
XLSX_SOURCE = Path(r"C:\Users\tyler\Desktop\BIBLE HUNT CODES\buck deer.xlsx")
PASTED_AUDIT = ROOT / "processed_data/audits/buck_deer_pasted_permit_source_2026.csv"
HANUMBER = ROOT / "processed_data/dwr_huntplanner_hanumber_2026.csv"
HUNTTABLE = ROOT / "data_truth/crosswalk_truth/validation/live_dwr_permit_numbers_comprehensive_vs_DATABASE_2026.csv"
UTAHDRAWS = ROOT / "processed_data/audits/dwr_2026_draw_results_vs_database_allotments.csv"

OUT_CORRECTED = ROOT / "processed_data/audits/buck_deer_current_permit_source_2026_corrected.csv"
OUT_RECONCILIATION = ROOT / "processed_data/audits/buck_deer_xlsx_pasted_permit_reconciliation_2026.csv"
OUT_SUMMARY = ROOT / "processed_data/audits/buck_deer_xlsx_pasted_permit_reconciliation_2026_summary.json"

CODE_RE = re.compile(r"^[A-Z]{2}\d{4}$")


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def int_text(value: object) -> str:
    text = clean(value).replace(",", "")
    if text in {"", "-", "nan", "None"}:
        return ""
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return ""
    number = float(match.group(0))
    return str(int(number)) if number.is_integer() else str(number)


def extract_labeled_number(value: object, label: str) -> str:
    match = re.search(rf"\b{re.escape(label)}:\s*([0-9,]+)", clean(value), re.I)
    return int_text(match.group(1)) if match else ""


def permit_total(res: object, nr: object, total: object) -> str:
    res_text = int_text(res)
    nr_text = int_text(nr)
    total_text = int_text(total)
    if total_text not in {"", "0"}:
        return total_text
    if res_text not in {"", "0"} or nr_text not in {"", "0"}:
        return str(int(res_text or 0) + int(nr_text or 0))
    return ""


def triple(res: object, nr: object, total: object) -> tuple[str, str, str]:
    return int_text(res), int_text(nr), permit_total(res, nr, total)


def has_value(values: tuple[str, str, str]) -> bool:
    return any(value not in {"", "0"} for value in values)


def read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return [{key: clean(value) for key, value in row.items()} for row in csv.DictReader(handle)]


def parse_workbook() -> tuple[list[dict[str, str]], dict[str, int]]:
    if not XLSX_SOURCE.exists():
        raise FileNotFoundError(XLSX_SOURCE)
    workbook = openpyxl.load_workbook(XLSX_SOURCE, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows: list[dict[str, str]] = []
    current: dict[str, str] | None = None
    stats = {
        "worksheet_rows_including_header": sheet.max_row,
        "worksheet_columns": sheet.max_column,
        "code_rows": 0,
        "continuation_rows_collapsed": 0,
        "ignored_blank_rows": 0,
    }
    for raw_index, cells in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        values = [clean(value) for value in cells[:10]]
        if not any(values):
            stats["ignored_blank_rows"] += 1
            continue
        hunt_code = values[1].upper()
        if CODE_RE.fullmatch(hunt_code):
            permit_cell = values[7]
            res_value = extract_labeled_number(permit_cell, "Res")
            nr_value = extract_labeled_number(permit_cell, "NonRes")
            total_value = extract_labeled_number(permit_cell, "Total")
            if not any([res_value, nr_value, total_value]):
                res_value = int_text(values[7])
                nr_value = int_text(values[8])
                total_value = int_text(values[9])
            current = {
                "hunt_code": hunt_code,
                "hunt_name": values[0],
                "sex_type": values[2],
                "species": values[3],
                "weapon": values[4],
                "hunt_type": values[5],
                "season": values[6],
                "xlsx_res": res_value,
                "xlsx_nr": nr_value,
                "xlsx_total_printed": total_value,
                "xlsx_row_start": str(raw_index),
                "xlsx_continuation_rows": "",
            }
            rows.append(current)
            stats["code_rows"] += 1
            continue
        if current is not None:
            row_text = " ".join(value for value in values if value)
            nr_value = extract_labeled_number(row_text, "NonRes")
            total_value = extract_labeled_number(row_text, "Total")
            if nr_value or total_value:
                if nr_value:
                    current["xlsx_nr"] = nr_value
                if total_value:
                    current["xlsx_total_printed"] = total_value
                current["xlsx_continuation_rows"] = ",".join(
                    value for value in [current["xlsx_continuation_rows"], str(raw_index)] if value
                )
                stats["continuation_rows_collapsed"] += 1
    for row in rows:
        row["xlsx_total"] = permit_total(row["xlsx_res"], row["xlsx_nr"], row["xlsx_total_printed"])
        if row["xlsx_res"] or row["xlsx_nr"]:
            row["xlsx_permit_shape"] = "RES_NR_SPLIT_TOTAL_COMPUTED"
        elif row["xlsx_total"]:
            row["xlsx_permit_shape"] = "TOTAL_ONLY_PRINTED"
        else:
            row["xlsx_permit_shape"] = "NO_PERMIT_NUMBER_IN_XLSX"
    return rows, stats


def pasted_map() -> dict[str, dict[str, str]]:
    return {row["hunt_code"]: row for row in read_csv(PASTED_AUDIT) if row.get("hunt_code")}


def source_maps() -> tuple[dict[str, tuple[str, str, str]], dict[str, tuple[str, str, str]], dict[str, tuple[str, str, str]]]:
    hanumber = {
        row["hunt_code"]: triple(row.get("permits_2026_res"), row.get("permits_2026_nr"), row.get("permits_2026_total"))
        for row in read_csv(HANUMBER)
        if row.get("hunt_code")
    }
    hunttable = {
        row["hunt_code"]: triple(row.get("live_res"), row.get("live_nr"), row.get("live_total"))
        for row in read_csv(HUNTTABLE)
        if row.get("hunt_code") and row.get("presence_status") != "DATABASE_ONLY"
    }
    utahdraws = {
        row["hunt_code"]: triple(row.get("source_res"), row.get("source_nr"), row.get("source_total"))
        for row in read_csv(UTAHDRAWS)
        if row.get("hunt_code") and row.get("source_presence") == "SOURCE_AND_DATABASE"
    }
    return hanumber, hunttable, utahdraws


def compare_triples(left: tuple[str, str, str], right: tuple[str, str, str]) -> str:
    if left == right:
        return "MATCH_EXACT"
    if left[2] and right[2] and left[2] == right[2] and not left[0] and not left[1]:
        return "MATCH_TOTAL_ONLY"
    if left[2] and right[2] and left[2] == right[2] and not right[0] and not right[1]:
        return "MATCH_TOTAL_ONLY"
    if not has_value(left) and not has_value(right):
        return "BOTH_BLANK"
    if has_value(left) and not has_value(right):
        return "MISSING_IN_PASTED_AUDIT"
    if not has_value(left) and has_value(right):
        return "MISSING_IN_XLSX"
    return "REVIEW_MISMATCH"


def external_agreement(values: tuple[str, str, str], sources: dict[str, tuple[str, str, str]]) -> str:
    exact = [name for name, source_values in sources.items() if has_value(source_values) and source_values == values]
    total = [
        name
        for name, source_values in sources.items()
        if source_values[2] and values[2] and source_values[2] == values[2] and name not in exact
    ]
    if len(exact) == 3:
        return "MATCHES_ALL_EXTERNAL_SOURCES_EXACT"
    if exact:
        return "MATCHES_" + "_AND_".join(exact) + "_EXACT"
    if total:
        return "TOTAL_MATCHES_" + "_AND_".join(total)
    if has_value(values):
        return "VALUE_UNCONFIRMED_OR_CONFLICTS"
    return "NO_PERMIT_VALUE"


def main() -> int:
    workbook_rows, workbook_stats = parse_workbook()
    pasted = pasted_map()
    hanumber, hunttable, utahdraws = source_maps()
    xlsx_hash = hashlib.sha256(XLSX_SOURCE.read_bytes()).hexdigest()
    reconciliation_rows: list[dict[str, str]] = []
    corrected_rows: list[dict[str, str]] = []
    for row in workbook_rows:
        code = row["hunt_code"]
        pasted_row = pasted.get(code, {})
        xlsx_values = triple(row["xlsx_res"], row["xlsx_nr"], row["xlsx_total"])
        pasted_values = triple(
            pasted_row.get("pasted_res", ""),
            pasted_row.get("pasted_nr", ""),
            pasted_row.get("pasted_total", ""),
        )
        xlsx_pasted_status = compare_triples(xlsx_values, pasted_values)
        external_sources = {
            "HANUMBER": hanumber.get(code, ("", "", "")),
            "HUNTTABLE": hunttable.get(code, ("", "", "")),
            "UTAHDRAWS": utahdraws.get(code, ("", "", "")),
        }
        external_status = external_agreement(xlsx_values, external_sources)
        if xlsx_pasted_status in {"MATCH_EXACT", "MATCH_TOTAL_ONLY", "BOTH_BLANK"} and external_status.startswith(
            ("MATCHES_", "TOTAL_MATCHES_")
        ):
            action = "XLSX_PERMIT_ROWS_REPAIRED_AND_EXTERNALLY_CONFIRMED"
        elif xlsx_pasted_status in {"MATCH_EXACT", "MATCH_TOTAL_ONLY"}:
            action = "XLSX_PERMIT_ROWS_REPAIRED_MATCH_PASTED_SOURCE"
        elif xlsx_pasted_status == "BOTH_BLANK":
            action = "NO_PERMIT_VALUE_TO_PROMOTE"
        else:
            action = "REVIEW_XLSX_VS_PASTED_MISMATCH"
        common = {
            "hunt_code": code,
            "hunt_name": row["hunt_name"],
            "sex_type": row["sex_type"],
            "species": row["species"],
            "weapon": row["weapon"],
            "hunt_type": row["hunt_type"],
            "season": row["season"],
        }
        corrected_rows.append(
            {
                **common,
                "permits_2026_res": xlsx_values[0],
                "permits_2026_nr": xlsx_values[1],
                "permits_2026_total": xlsx_values[2],
                "permit_shape": row["xlsx_permit_shape"],
                "source_file": str(XLSX_SOURCE),
                "source_sha256": xlsx_hash,
                "source_row_start": row["xlsx_row_start"],
                "source_continuation_rows_collapsed": row["xlsx_continuation_rows"],
                "validation_status": action,
            }
        )
        reconciliation_rows.append(
            {
                **common,
                "xlsx_res": xlsx_values[0],
                "xlsx_nr": xlsx_values[1],
                "xlsx_total": xlsx_values[2],
                "xlsx_total_printed": row["xlsx_total_printed"],
                "xlsx_permit_shape": row["xlsx_permit_shape"],
                "xlsx_row_start": row["xlsx_row_start"],
                "xlsx_continuation_rows_collapsed": row["xlsx_continuation_rows"],
                "pasted_res": pasted_values[0],
                "pasted_nr": pasted_values[1],
                "pasted_total": pasted_values[2],
                "pasted_permit_shape": pasted_row.get("pasted_permit_shape", "MISSING_PASTED_ROW"),
                "xlsx_vs_pasted_status": xlsx_pasted_status,
                "hanumber_res": external_sources["HANUMBER"][0],
                "hanumber_nr": external_sources["HANUMBER"][1],
                "hanumber_total": external_sources["HANUMBER"][2],
                "hunttable_res": external_sources["HUNTTABLE"][0],
                "hunttable_nr": external_sources["HUNTTABLE"][1],
                "hunttable_total": external_sources["HUNTTABLE"][2],
                "utahdraws_res": external_sources["UTAHDRAWS"][0],
                "utahdraws_nr": external_sources["UTAHDRAWS"][1],
                "utahdraws_total": external_sources["UTAHDRAWS"][2],
                "external_agreement_status": external_status,
                "recommended_action": action,
                "source_file": str(XLSX_SOURCE),
                "source_sha256": xlsx_hash,
            }
        )
    OUT_CORRECTED.parent.mkdir(parents=True, exist_ok=True)
    corrected_fields = [
        "hunt_code",
        "hunt_name",
        "sex_type",
        "species",
        "weapon",
        "hunt_type",
        "season",
        "permits_2026_res",
        "permits_2026_nr",
        "permits_2026_total",
        "permit_shape",
        "source_file",
        "source_sha256",
        "source_row_start",
        "source_continuation_rows_collapsed",
        "validation_status",
    ]
    with OUT_CORRECTED.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=corrected_fields)
        writer.writeheader()
        writer.writerows(corrected_rows)
    reconciliation_fields = list(reconciliation_rows[0].keys()) if reconciliation_rows else []
    with OUT_RECONCILIATION.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=reconciliation_fields)
        writer.writeheader()
        writer.writerows(reconciliation_rows)
    summary = {
        "created_at_utc": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        "xlsx_source_file": str(XLSX_SOURCE),
        "xlsx_source_sha256": xlsx_hash,
        "pasted_audit_file": PASTED_AUDIT.relative_to(ROOT).as_posix(),
        "workbook_stats": workbook_stats,
        "corrected_rows": len(corrected_rows),
        "unique_hunt_codes": len({row["hunt_code"] for row in corrected_rows}),
        "permit_shape_counts": dict(Counter(row["permit_shape"] for row in corrected_rows)),
        "xlsx_vs_pasted_status_counts": dict(Counter(row["xlsx_vs_pasted_status"] for row in reconciliation_rows)),
        "external_agreement_status_counts": dict(Counter(row["external_agreement_status"] for row in reconciliation_rows)),
        "recommended_action_counts": dict(Counter(row["recommended_action"] for row in reconciliation_rows)),
        "outputs": {
            "corrected_csv": OUT_CORRECTED.relative_to(ROOT).as_posix(),
            "reconciliation_csv": OUT_RECONCILIATION.relative_to(ROOT).as_posix(),
            "summary_json": OUT_SUMMARY.relative_to(ROOT).as_posix(),
        },
        "notes": [
            "The workbook source has NonRes continuation rows; this script collapses those rows into the preceding hunt-code row.",
            "Totals are computed as resident plus nonresident only when the source has a split row and no printed total.",
            "DATABASE.csv is not used as a winner source in this reconciliation.",
        ],
    }
    OUT_SUMMARY.write_text(json.dumps(summary, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(json.dumps(summary, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
