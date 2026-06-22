import csv
import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
STAGED_XLSX = ROOT / "processed_data" / "hard_data_exports" / "hunt_tables" / "2026" / "CLEAN_XLXS_STAGED"
STAGED_PDFS = ROOT / "processed_data" / "hard_data_exports" / "hunt_tables" / "2026" / "PDF'S"
RAW_TRUTH_PDFS = ROOT / "data_truth" / "draw_results_truth" / "raw_pdfs" / "2026_PERMITS=2027_MODEL"
FRESH = ROOT / "audits" / "2025_canonical_finalization" / "fresh_live_pulls_20260621_192945"
CANONICAL = ROOT / "data_truth" / "draw_results_truth" / "normalized" / "canonical_yearly" / "draw_results_2026_for_2027_canonical_yearly_draw_results.csv"
AUDIT_DIR = ROOT / "audits" / "2026_live_source_comparison"

HUNT_CODE_RE = re.compile(r"\b[A-Z]{2}\d{4}\b")


def clean(value):
    return " ".join(str(value or "").replace("\r", "\n").split())


def is_hunt_code(value):
    return bool(HUNT_CODE_RE.fullmatch(clean(value).upper()))


def add_code(target, code, source_file):
    code = clean(code).upper()
    if is_hunt_code(code):
        target[code].add(source_file)


def extract_codes_from_xlsx(path):
    codes = set()
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            header = []
            hunt_code_indexes = set()
            for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
                values = [clean(value) for value in row]
                if row_index <= 5:
                    for index, value in enumerate(values):
                        if value.lower().replace("_", " ") in {"hunt code", "hunt codes", "hunt number", "hunt nbr"}:
                            hunt_code_indexes.add(index)
                    if hunt_code_indexes:
                        header = values
                if hunt_code_indexes:
                    for index in hunt_code_indexes:
                        if index < len(values) and is_hunt_code(values[index].upper()):
                            codes.add(values[index].upper())
                else:
                    for value in values:
                        if is_hunt_code(value.upper()):
                            codes.add(value.upper())
    except Exception as exc:
        return codes, str(exc)
    return codes, ""


def extract_codes_from_pdf(path):
    codes = set()
    error = ""
    try:
        reader = PdfReader(str(path))
        for page in reader.pages:
            text = page.extract_text() or ""
            codes.update(match.group(0).upper() for match in HUNT_CODE_RE.finditer(text.upper()))
    except Exception as exc:
        error = str(exc)
    return codes, error


def extract_fresh_dwr_codes():
    by_code = defaultdict(set)
    for path in FRESH.glob("dwr_huntboundary_*.json"):
        if path.name in {"dwr_huntboundary_hasetup.json"}:
            continue
        try:
            data = json.loads(path.read_text(encoding="utf-8-sig"))
        except Exception:
            continue
        if not isinstance(data, list):
            continue
        for row in data:
            add_code(by_code, row.get("HUNT_NBR"), path.name)
    return by_code


def extract_fresh_utahdraws_codes():
    by_code = defaultdict(set)
    for path in FRESH.glob("utahdraws_*_2026_*.json"):
        try:
            data = json.loads(path.read_text(encoding="utf-8-sig"))
        except Exception:
            continue
        rows = data.get("Data") if isinstance(data, dict) else data
        if not isinstance(rows, list):
            continue
        for row in rows:
            if not isinstance(row, dict):
                continue
            add_code(by_code, row.get("HuntCode"), path.name)
            add_code(by_code, row.get("huntCode"), path.name)
    return by_code


def extract_canonical_codes():
    by_code = defaultdict(set)
    with CANONICAL.open(newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            add_code(by_code, row.get("hunt_code"), CANONICAL.name)
    return by_code


def code_family(code):
    return code[:2]


def write_code_sources(path, code_sources):
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["hunt_code", "prefix", "source_count", "sources"], lineterminator="\n")
        writer.writeheader()
        for code in sorted(code_sources):
            writer.writerow(
                {
                    "hunt_code": code,
                    "prefix": code_family(code),
                    "source_count": len(code_sources[code]),
                    "sources": "; ".join(sorted(code_sources[code])),
                }
            )


def main():
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    staged_xlsx = defaultdict(set)
    staged_pdf = defaultdict(set)
    raw_truth_pdf = defaultdict(set)
    file_rows = []

    for path in sorted(STAGED_XLSX.glob("*.xlsx")):
        codes, error = extract_codes_from_xlsx(path)
        for code in codes:
            staged_xlsx[code].add(path.name)
        file_rows.append({"source_group": "staged_xlsx", "file": path.name, "hunt_code_count": len(codes), "error": error})

    for path in sorted(STAGED_PDFS.glob("*.pdf")):
        codes, error = extract_codes_from_pdf(path)
        for code in codes:
            staged_pdf[code].add(path.name)
        file_rows.append({"source_group": "staged_pdf", "file": path.name, "hunt_code_count": len(codes), "error": error})

    for path in sorted(RAW_TRUTH_PDFS.glob("*.pdf")):
        codes, error = extract_codes_from_pdf(path)
        for code in codes:
            raw_truth_pdf[code].add(path.name)
        file_rows.append({"source_group": "raw_truth_pdf", "file": path.name, "hunt_code_count": len(codes), "error": error})

    fresh_dwr = extract_fresh_dwr_codes()
    fresh_utahdraws = extract_fresh_utahdraws_codes()
    canonical = extract_canonical_codes()
    fresh_all = defaultdict(set)
    for source in (fresh_dwr, fresh_utahdraws):
        for code, files in source.items():
            fresh_all[code].update(files)

    write_code_sources(AUDIT_DIR / "staged_xlsx_2026_hunt_code_sources.csv", staged_xlsx)
    write_code_sources(AUDIT_DIR / "staged_pdf_2026_hunt_code_sources.csv", staged_pdf)
    write_code_sources(AUDIT_DIR / "raw_truth_pdf_2026_hunt_code_sources.csv", raw_truth_pdf)
    write_code_sources(AUDIT_DIR / "fresh_dwr_2026_hunt_code_sources.csv", fresh_dwr)
    write_code_sources(AUDIT_DIR / "fresh_utahdraws_2026_hunt_code_sources.csv", fresh_utahdraws)

    with (AUDIT_DIR / "staged_raw_2026_file_inventory.csv").open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["source_group", "file", "hunt_code_count", "error"], lineterminator="\n")
        writer.writeheader()
        writer.writerows(file_rows)

    comparisons = [
        ("staged_xlsx_missing_from_fresh_downloads", set(staged_xlsx) - set(fresh_all), staged_xlsx),
        ("fresh_downloads_missing_from_staged_xlsx", set(fresh_all) - set(staged_xlsx), fresh_all),
        ("staged_pdf_missing_from_fresh_downloads", set(staged_pdf) - set(fresh_all), staged_pdf),
        ("fresh_downloads_missing_from_staged_pdf", set(fresh_all) - set(staged_pdf), fresh_all),
        ("raw_truth_pdf_missing_from_fresh_downloads", set(raw_truth_pdf) - set(fresh_all), raw_truth_pdf),
        ("fresh_downloads_missing_from_raw_truth_pdf", set(fresh_all) - set(raw_truth_pdf), fresh_all),
        ("canonical_missing_from_raw_truth_pdf", set(canonical) - set(raw_truth_pdf), canonical),
        ("raw_truth_pdf_missing_from_canonical", set(raw_truth_pdf) - set(canonical), raw_truth_pdf),
        ("canonical_missing_from_fresh_downloads", set(canonical) - set(fresh_all), canonical),
        ("fresh_downloads_missing_from_canonical", set(fresh_all) - set(canonical), fresh_all),
        ("staged_xlsx_missing_from_canonical", set(staged_xlsx) - set(canonical), staged_xlsx),
        ("canonical_missing_from_staged_xlsx", set(canonical) - set(staged_xlsx), canonical),
    ]

    with (AUDIT_DIR / "staged_raw_2026_vs_fresh_downloads_code_gap_audit.csv").open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["comparison", "hunt_code", "prefix", "sources"], lineterminator="\n")
        writer.writeheader()
        for comparison, codes, source_map in comparisons:
            for code in sorted(codes):
                writer.writerow(
                    {
                        "comparison": comparison,
                        "hunt_code": code,
                        "prefix": code_family(code),
                        "sources": "; ".join(sorted(source_map.get(code, []))),
                    }
                )

    summary = {
        "staged_xlsx_file_count": len(list(STAGED_XLSX.glob("*.xlsx"))),
        "staged_pdf_file_count": len(list(STAGED_PDFS.glob("*.pdf"))),
        "raw_truth_pdf_file_count": len(list(RAW_TRUTH_PDFS.glob("*.pdf"))),
        "fresh_dwr_file_count": len(list(FRESH.glob("dwr_huntboundary_*.json"))) - 1,
        "fresh_utahdraws_file_count": len(list(FRESH.glob("utahdraws_*_2026_*.json"))),
        "staged_xlsx_unique_hunt_codes": len(staged_xlsx),
        "staged_pdf_unique_hunt_codes": len(staged_pdf),
        "raw_truth_pdf_unique_hunt_codes": len(raw_truth_pdf),
        "fresh_dwr_unique_hunt_codes": len(fresh_dwr),
        "fresh_utahdraws_unique_hunt_codes": len(fresh_utahdraws),
        "fresh_all_unique_hunt_codes": len(fresh_all),
        "canonical_unique_hunt_codes": len(canonical),
        "gap_counts": {name: len(codes) for name, codes, _ in comparisons},
        "staged_xlsx_prefix_counts": dict(sorted(Counter(map(code_family, staged_xlsx)).items())),
        "raw_truth_pdf_prefix_counts": dict(sorted(Counter(map(code_family, raw_truth_pdf)).items())),
        "fresh_all_prefix_counts": dict(sorted(Counter(map(code_family, fresh_all)).items())),
        "canonical_prefix_counts": dict(sorted(Counter(map(code_family, canonical)).items())),
        "audit_files": {
            "file_inventory": str(AUDIT_DIR / "staged_raw_2026_file_inventory.csv"),
            "code_gap_audit": str(AUDIT_DIR / "staged_raw_2026_vs_fresh_downloads_code_gap_audit.csv"),
            "staged_xlsx_sources": str(AUDIT_DIR / "staged_xlsx_2026_hunt_code_sources.csv"),
            "staged_pdf_sources": str(AUDIT_DIR / "staged_pdf_2026_hunt_code_sources.csv"),
            "raw_truth_pdf_sources": str(AUDIT_DIR / "raw_truth_pdf_2026_hunt_code_sources.csv"),
            "fresh_dwr_sources": str(AUDIT_DIR / "fresh_dwr_2026_hunt_code_sources.csv"),
            "fresh_utahdraws_sources": str(AUDIT_DIR / "fresh_utahdraws_2026_hunt_code_sources.csv"),
        },
    }
    (AUDIT_DIR / "staged_raw_2026_vs_fresh_downloads_summary.json").write_text(
        json.dumps(summary, indent=2, sort_keys=True),
        encoding="utf-8",
    )
    print(json.dumps(summary, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
