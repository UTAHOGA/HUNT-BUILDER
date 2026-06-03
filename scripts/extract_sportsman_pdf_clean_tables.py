import argparse
import csv
import json
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_BIBLE_ROOT = Path(r"C:\Users\tyler\Desktop\BIBLE HUNT CODES")
OUT_DIR = ROOT / "processed_data" / "audits"
OUT_CSV = OUT_DIR / "sportsman_pdf_clean_extract.csv"
OUT_SCRIPT_FEED = OUT_DIR / "sportsman_pdf_clean_script_feed.csv"
OUT_XLSX = OUT_DIR / "sportsman_pdf_clean_extract.xlsx"
OUT_AUDIT = OUT_DIR / "sportsman_pdf_clean_extract_audit.csv"
OUT_SUMMARY = OUT_DIR / "sportsman_pdf_clean_extract_summary.json"

OFFICIAL_NAMES = {
    "BI1000": "Sportsman Bison",
    "BR1000": "Sportsman Black Bear",
    "CG1000": "Sportsman Cougar",
    "CG9999": "Sportsman Cougar 2023*",
    "DB0007": "Sportsman Deer",
    "DS1000": "Sportsman Desert Bighorn Sheep",
    "EB1000": "Sportsman Elk",
    "GO1000": "Sportsman Mountain Goat",
    "MB1000": "Sportsman Moose",
    "PB1000": "Sportsman Pronghorn",
    "RS0001": "Sportsman Rocky Mtn Bighorn Sheep",
    "TK0001": "Sportsman Bearded Turkey",
}

SPECIES_BY_CODE = {
    "BI1000": "Bison",
    "BR1000": "Black Bear",
    "CG1000": "Cougar",
    "CG9999": "Cougar",
    "DB0007": "Deer",
    "DS1000": "Desert Bighorn Sheep",
    "EB1000": "Elk",
    "GO1000": "Mountain Goat",
    "MB1000": "Moose",
    "PB1000": "Pronghorn",
    "RS0001": "Rocky Mountain Bighorn Sheep",
    "TK0001": "Turkey",
}

FIELDNAMES = [
    "source_scope",
    "source_file",
    "source_page",
    "storage_year",
    "draw_results_year",
    "permit_draw_year",
    "model_year",
    "year_basis",
    "report_title",
    "raw_extracted_code",
    "normalized_hunt_code",
    "artifact_status",
    "hunt_name",
    "species",
    "resident_successful",
    "nonresident_successful",
    "resident_unsuccessful",
    "nonresident_unsuccessful",
    "total_applications",
    "resident_quota",
    "nonresident_quota",
    "total_quota",
    "resident_success_ratio",
    "nonresident_success_ratio",
    "resident_bonus_permits",
    "resident_regular_permits",
    "nonresident_bonus_permits",
    "nonresident_regular_permits",
    "odds_denominator",
    "parse_method",
    "parse_status",
    "notes",
]


def norm(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def clean_int(value: object) -> str:
    text = norm(value).replace(",", "")
    if not text or text.upper() == "N/A":
        return ""
    try:
        return str(int(float(text)))
    except ValueError:
        return text


def extract_denominator(success_ratio: str) -> str:
    match = re.search(r"1\s+in\s+([\d,]+(?:\.\d+)?)", success_ratio or "", flags=re.I)
    return clean_int(match.group(1)) if match else ""


def normalize_code(raw_code: str) -> tuple[str, str]:
    raw = norm(raw_code).upper()
    if raw in OFFICIAL_NAMES:
        return raw, "RAW_VALID"
    if raw.startswith("A") and raw[1:] in OFFICIAL_NAMES:
        return raw[1:], "NORMALIZED_NA_GLUE_ARTIFACT"
    return raw, "REVIEW_UNKNOWN_CODE"


def source_scope(path: Path, bible_root: Path) -> str:
    try:
        path.relative_to(bible_root)
        return "BIBLE_HUNT_CODES"
    except ValueError:
        return "REPO_PIPELINE"


def storage_year(path: Path, bible_root: Path) -> str:
    parts = list(path.parts)
    try:
        rel = path.relative_to(bible_root)
        if rel.parts and re.fullmatch(r"\d{4}", rel.parts[0]):
            return rel.parts[0]
    except ValueError:
        pass
    if "hunt_unit_database" in parts:
        idx = parts.index("hunt_unit_database")
        if idx + 1 < len(parts) and re.fullmatch(r"\d{4}", parts[idx + 1]):
            return parts[idx + 1]
    for part in parts:
        if re.fullmatch(r"\d{4}", part):
            return part
    return ""


def report_title(text: str) -> str:
    for line in text.splitlines():
        line = norm(line)
        if "Sportsman" in line and ("Draw Odds Report" in line or "Permit Draw Results" in line):
            return line
    return ""


def infer_draw_year(path: Path, text: str, bible_root: Path) -> tuple[str, str]:
    scope = source_scope(path, bible_root)
    stored = storage_year(path, bible_root)
    if scope == "BIBLE_HUNT_CODES" and stored:
        return stored, "BIBLE_FOLDER_YEAR"
    title = report_title(text)
    match = re.search(r"\b(20\d{2})(?:-\d{2,4})?\s+Sportsman", title)
    if match:
        return match.group(1), "REPORT_TITLE"
    match = re.search(r"\b(20\d{2})[_ -]*(?:sportsman|SPORTSMAN)", path.name)
    if match:
        return match.group(1), "FILENAME"
    return stored, "STORAGE_YEAR_FALLBACK"


def base_row(path: Path, page: int, text: str, bible_root: Path) -> dict[str, str]:
    draw_year, basis = infer_draw_year(path, text, bible_root)
    model_year = str(int(draw_year) + 1) if draw_year.isdigit() else ""
    try:
        source_file = str(path.relative_to(ROOT)).replace("\\", "/")
    except ValueError:
        source_file = str(path)
    return {
        "source_scope": source_scope(path, bible_root),
        "source_file": source_file,
        "source_page": str(page),
        "storage_year": storage_year(path, bible_root),
        "draw_results_year": draw_year,
        "permit_draw_year": draw_year,
        "model_year": model_year,
        "year_basis": basis,
        "report_title": report_title(text),
    }


def complete_row(row: dict[str, str], raw_code: str, hunt_name: str) -> dict[str, str]:
    code, artifact_status = normalize_code(raw_code)
    row["raw_extracted_code"] = raw_code
    row["normalized_hunt_code"] = code
    row["artifact_status"] = artifact_status
    row["hunt_name"] = OFFICIAL_NAMES.get(code, norm(hunt_name))
    row["species"] = SPECIES_BY_CODE.get(code, "")
    return row


def parse_inline_line(path: Path, page: int, full_text: str, line: str, bible_root: Path) -> dict[str, str] | None:
    line = norm(line)
    normal = re.match(
        r"^(?P<code>[A-Z]{2}\d{4}|CG9999)\s+(?P<name>Sportsman.+?)\s+"
        r"(?P<rs>\d+)\s+(?P<ns>\d+)\s+(?P<ru>[\d,]+)\s+(?P<nu>[\d,]+)\s+"
        r"(?P<apps>[\d,]+)\s+(?P<rq>\d+)\s+(?P<nq>N/A|\d+)\s+(?P<tq>\d+)\s+"
        r"(?P<rr>1\s+in\s+[\d,]+(?:\.\d+)?)\s+(?P<nr>N/A|1\s+in\s+[\d,]+(?:\.\d+)?)$",
        line,
        flags=re.I,
    )
    glued = re.match(
        r"^(?P<rs>\d+)\s+(?P<ns>\d+)\s+(?P<ru>[\d,]+)\s+(?P<nu>[\d,]+)\s+"
        r"(?P<apps>[\d,]+)\s+(?P<rq>\d+)\s+(?P<tq>\d+)\s+"
        r"(?P<rr>1\s+in\s+[\d,]+(?:\.\d+)?)\s+N/AN/A(?P<code>A?[A-Z]{2}\d{4}|A?CG9999)\s+"
        r"(?P<name>Sportsman.+)$",
        line,
        flags=re.I,
    )
    match = normal or glued
    if not match:
        return None

    row = base_row(path, page, full_text, bible_root)
    row.update(
        {
            "resident_successful": clean_int(match.group("rs")),
            "nonresident_successful": clean_int(match.group("ns")),
            "resident_unsuccessful": clean_int(match.group("ru")),
            "nonresident_unsuccessful": clean_int(match.group("nu")),
            "total_applications": clean_int(match.group("apps")),
            "resident_quota": clean_int(match.group("rq")),
            "nonresident_quota": clean_int(match.group("nq")) if "nq" in match.groupdict() else "",
            "total_quota": clean_int(match.group("tq")),
            "resident_success_ratio": norm(match.group("rr")),
            "nonresident_success_ratio": norm(match.group("nr")) if "nr" in match.groupdict() else "N/A",
            "resident_bonus_permits": "",
            "resident_regular_permits": "",
            "nonresident_bonus_permits": "",
            "nonresident_regular_permits": "",
            "parse_method": "inline_table" if normal else "na_glued_code_table",
            "parse_status": "OK",
            "notes": "Extracted from single-line Sportsman table.",
        }
    )
    row["odds_denominator"] = extract_denominator(row["resident_success_ratio"])
    raw_code = match.group("code")
    if glued and not raw_code.startswith("A"):
        raw_code = f"A{raw_code}"
    return complete_row(row, raw_code, match.group("name"))


def split_hunt_entries(lines: list[str], start: int, stop: int) -> list[tuple[str, str]]:
    entries: list[tuple[str, str]] = []
    i = start
    while i < stop:
        line = norm(lines[i])
        code_match = re.match(r"^([A-Z]{2}\d{4}|CG9999)(?:\s+(.*))?$", line)
        if not code_match:
            i += 1
            continue
        code = code_match.group(1)
        name = norm(code_match.group(2))
        if not name and i + 1 < stop:
            i += 1
            name = norm(lines[i])
        if i + 1 < stop and norm(lines[i + 1]) in {"Sheep", "Bighorn Sheep", "Turkey"}:
            name = norm(f"{name} {lines[i + 1]}")
            i += 1
        entries.append((code, name))
        i += 1
    return entries


def block_values(lines: list[str], label: str, start: int, count: int) -> tuple[list[str], int]:
    idx = start
    while idx < len(lines) and norm(lines[idx]) != label:
        idx += 1
    if idx >= len(lines):
        return [], idx
    values = [norm(value) for value in lines[idx + 1 : idx + 1 + count]]
    return values, idx + 1 + count


def parse_stacked_table(path: Path, page: int, text: str, bible_root: Path) -> list[dict[str, str]]:
    lines = [norm(line) for line in text.splitlines() if norm(line)]
    try:
        start = lines.index("Hunter Choice") + 1
    except ValueError:
        return []
    try:
        stop = next(i for i in range(start, len(lines)) if lines[i] == "Resident")
    except StopIteration:
        return []

    entries = split_hunt_entries(lines, start, stop)
    if len(entries) < 5:
        return []

    count = len(entries)
    pointer = stop
    resident_successful, pointer = block_values(lines, "Resident", pointer, count)
    nonresident_successful, pointer = block_values(lines, "Resident", pointer, count)
    resident_unsuccessful, pointer = block_values(lines, "Resident", pointer, count)
    nonresident_unsuccessful, pointer = block_values(lines, "Resident", pointer, count)
    total_applications, pointer = block_values(lines, "Applications", pointer, count)
    resident_quota, pointer = block_values(lines, "Quota", pointer, count)
    nonresident_quota, pointer = block_values(lines, "Quota", pointer, count)
    total_quota, pointer = block_values(lines, "Quota", pointer, count)
    resident_success_ratio, pointer = block_values(lines, "Success", pointer, count)
    nonresident_success_ratio, pointer = block_values(lines, "Success", pointer, count)

    blocks = [
        resident_successful,
        nonresident_successful,
        resident_unsuccessful,
        nonresident_unsuccessful,
        total_applications,
        resident_quota,
        nonresident_quota,
        total_quota,
        resident_success_ratio,
        nonresident_success_ratio,
    ]
    if any(len(block) != count for block in blocks):
        return []

    rows = []
    for i, (raw_code, hunt_name) in enumerate(entries):
        row = base_row(path, page, text, bible_root)
        row.update(
            {
                "resident_successful": clean_int(resident_successful[i]),
                "nonresident_successful": clean_int(nonresident_successful[i]),
                "resident_unsuccessful": clean_int(resident_unsuccessful[i]),
                "nonresident_unsuccessful": clean_int(nonresident_unsuccessful[i]),
                "total_applications": clean_int(total_applications[i]),
                "resident_quota": clean_int(resident_quota[i]),
                "nonresident_quota": clean_int(nonresident_quota[i]),
                "total_quota": clean_int(total_quota[i]),
                "resident_success_ratio": resident_success_ratio[i],
                "nonresident_success_ratio": nonresident_success_ratio[i],
                "resident_bonus_permits": "",
                "resident_regular_permits": "",
                "nonresident_bonus_permits": "",
                "nonresident_regular_permits": "",
                "parse_method": "stacked_column_table",
                "parse_status": "OK",
                "notes": "Extracted from stacked Sportsman table layout.",
            }
        )
        row["odds_denominator"] = extract_denominator(row["resident_success_ratio"])
        rows.append(complete_row(row, raw_code, hunt_name))
    return rows


def parse_official_style_pages(path: Path, text_by_page: list[str], bible_root: Path) -> list[dict[str, str]]:
    rows = []
    for page, text in enumerate(text_by_page, start=1):
        match = re.search(r"Hunt:\s+(?P<code>[A-Z]{2}\d{4}|CG9999)\s+(?P<name>Sportsman[^\n]+)", text)
        if not match:
            continue
        totals = re.findall(
            r"Totals\s+([\d,]+)\s+(\d+)\s+(\d+)\s+(\d+)\s+(1\s+in\s+[\d,]+(?:\.\d+)?|N/A)",
            text,
            flags=re.I,
        )
        if len(totals) < 2:
            continue
        resident = totals[0]
        nonresident = totals[1]
        row = base_row(path, page, text, bible_root)
        row.update(
            {
                "resident_successful": "",
                "nonresident_successful": "",
                "resident_unsuccessful": "",
                "nonresident_unsuccessful": "",
                "total_applications": clean_int(resident[0]),
                "resident_quota": clean_int(resident[3]),
                "nonresident_quota": clean_int(nonresident[3]),
                "total_quota": str(int(clean_int(resident[3]) or 0) + int(clean_int(nonresident[3]) or 0)),
                "resident_success_ratio": norm(resident[4]),
                "nonresident_success_ratio": norm(nonresident[4]),
                "resident_bonus_permits": clean_int(resident[1]),
                "resident_regular_permits": clean_int(resident[2]),
                "nonresident_bonus_permits": clean_int(nonresident[1]),
                "nonresident_regular_permits": clean_int(nonresident[2]),
                "parse_method": "one_hunt_per_page_totals",
                "parse_status": "OK",
                "notes": "Extracted from generated one-hunt-per-page Sportsman workbook/PDF totals.",
            }
        )
        row["odds_denominator"] = extract_denominator(row["resident_success_ratio"])
        rows.append(complete_row(row, match.group("code"), match.group("name")))
    return rows


def parse_utahdraws_summary_pdf(path: Path, text_by_page: list[str], bible_root: Path) -> list[dict[str, str]]:
    rows = []
    for page, text in enumerate(text_by_page, start=1):
        lines = [norm(line) for line in text.splitlines() if norm(line)]
        if "Odds Ladder Rows" not in lines or "AllChoicesSuccessfulCount" not in lines:
            continue
        start = lines.index("AllChoicesSuccessfulCount") + 1
        i = start
        while i + 9 < len(lines):
            if not re.fullmatch(r"[A-Z]{2}\d{4}|CG9999", lines[i]):
                i += 1
                continue
            raw_code = lines[i]
            hunt_name = lines[i + 1]
            category = lines[i + 2]
            residency = lines[i + 3]
            participants = lines[i + 6]
            successful = lines[i + 7]
            bonus_permits = lines[i + 8]
            regular_permits = lines[i + 9]
            total_permits = str(int(clean_int(bonus_permits) or 0) + int(clean_int(regular_permits) or 0))
            if category != "Sportsman" or residency != "Resident":
                i += 10
                continue
            row = base_row(path, page, text, bible_root)
            row.update(
                {
                    "resident_successful": clean_int(successful),
                    "nonresident_successful": "",
                    "resident_unsuccessful": "",
                    "nonresident_unsuccessful": "",
                    "total_applications": clean_int(participants),
                    "resident_quota": total_permits,
                    "nonresident_quota": "0",
                    "total_quota": total_permits,
                    "resident_success_ratio": f"1 in {int(clean_int(participants) or 0):,}.0" if clean_int(participants) else "",
                    "nonresident_success_ratio": "N/A",
                    "resident_bonus_permits": clean_int(bonus_permits),
                    "resident_regular_permits": clean_int(regular_permits),
                    "nonresident_bonus_permits": "0",
                    "nonresident_regular_permits": "0",
                    "parse_method": "utahdraws_summary_ladder_pdf",
                    "parse_status": "OK",
                    "notes": "Extracted from UtahDraws summary PDF odds-ladder rows.",
                }
            )
            row["odds_denominator"] = clean_int(participants)
            rows.append(complete_row(row, raw_code, hunt_name))
            i += 10
    return rows


def extract_pdf(path: Path, bible_root: Path) -> tuple[list[dict[str, str]], dict[str, str]]:
    reader = PdfReader(str(path))
    text_by_page = [page.extract_text() or "" for page in reader.pages]
    full_text = "\n".join(text_by_page)
    rows: list[dict[str, str]] = []
    for page, text in enumerate(text_by_page, start=1):
        for line in text.splitlines():
            parsed = parse_inline_line(path, page, full_text, line, bible_root)
            if parsed:
                rows.append(parsed)
    if not rows:
        rows = parse_stacked_table(path, 1, full_text, bible_root)
    if not rows:
        rows = parse_official_style_pages(path, text_by_page, bible_root)
    if not rows:
        rows = parse_utahdraws_summary_pdf(path, text_by_page, bible_root)

    audit = {
        "source_file": str(path),
        "source_scope": source_scope(path, bible_root),
        "storage_year": storage_year(path, bible_root),
        "report_title": report_title(full_text),
        "rows_extracted": str(len(rows)),
        "normalized_artifact_rows": str(sum(1 for row in rows if row.get("artifact_status") == "NORMALIZED_NA_GLUE_ARTIFACT")),
        "unknown_code_rows": str(sum(1 for row in rows if row.get("artifact_status") == "REVIEW_UNKNOWN_CODE")),
        "parse_methods": ";".join(sorted({row.get("parse_method", "") for row in rows if row.get("parse_method")})),
        "status": "OK" if rows else "NO_ROWS_EXTRACTED",
    }
    return rows, audit


def discover_pdfs(bible_root: Path, include_pipeline: bool) -> list[Path]:
    paths: list[Path] = []
    if bible_root.exists():
        paths.extend(path for path in bible_root.rglob("*.pdf") if "sportsman" in path.name.lower())
    if include_pipeline:
        pipeline_root = ROOT / "pipeline" / "RAW" / "hunt_unit_database"
        paths.extend(path for path in pipeline_root.rglob("*.pdf") if "sportsman" in path.name.lower())
    seen = set()
    unique = []
    for path in sorted(paths, key=lambda p: str(p).lower()):
        key = str(path.resolve()).lower()
        if key not in seen:
            unique.append(path)
            seen.add(key)
    return unique


def write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_workbook(path: Path, rows: list[dict[str, str]], audit_rows: list[dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Clean Sportsman Rows"
    feed = wb.create_sheet("Preferred Script Feed")
    audit = wb.create_sheet("Source Audit")

    header_fill = PatternFill("solid", fgColor="8FB7D5")
    audit_fill = PatternFill("solid", fgColor="D9EAD3")
    artifact_fill = PatternFill("solid", fgColor="FFF2CC")
    issue_fill = PatternFill("solid", fgColor="F4CCCC")
    thin = Side(style="thin", color="9E9E9E")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(FIELDNAMES)
    for row in rows:
        ws.append([row.get(field, "") for field in FIELDNAMES])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        artifact_status = row[FIELDNAMES.index("artifact_status")].value
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if artifact_status == "NORMALIZED_NA_GLUE_ARTIFACT":
                cell.fill = artifact_fill
            elif artifact_status == "REVIEW_UNKNOWN_CODE":
                cell.fill = issue_fill

    audit_fields = list(audit_rows[0].keys()) if audit_rows else []
    audit.append(audit_fields)
    for row in audit_rows:
        audit.append([row.get(field, "") for field in audit_fields])
    for cell in audit[1]:
        cell.font = Font(bold=True)
        cell.fill = audit_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    audit.freeze_panes = "A2"
    audit.auto_filter.ref = audit.dimensions
    for row in audit.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        "A": 18,
        "B": 58,
        "C": 10,
        "D": 12,
        "E": 14,
        "F": 14,
        "G": 12,
        "H": 22,
        "I": 32,
        "J": 18,
        "K": 18,
        "L": 28,
        "M": 34,
        "N": 22,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for idx in range(15, len(FIELDNAMES) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 16

    preferred_rows = preferred_script_rows(rows)
    feed.append(FIELDNAMES)
    for row in preferred_rows:
        feed.append([row.get(field, "") for field in FIELDNAMES])
    for cell in feed[1]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    feed.freeze_panes = "A2"
    feed.auto_filter.ref = feed.dimensions
    for row in feed.iter_rows(min_row=2):
        artifact_status = row[FIELDNAMES.index("artifact_status")].value
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if artifact_status == "NORMALIZED_NA_GLUE_ARTIFACT":
                cell.fill = artifact_fill
            elif artifact_status == "REVIEW_UNKNOWN_CODE":
                cell.fill = issue_fill
    for col, width in widths.items():
        feed.column_dimensions[col].width = width
    for idx in range(15, len(FIELDNAMES) + 1):
        feed.column_dimensions[get_column_letter(idx)].width = 16

    for idx in range(1, len(audit_fields) + 1):
        audit.column_dimensions[get_column_letter(idx)].width = 24 if idx != 1 else 58
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def preferred_rank(row: dict[str, str]) -> tuple[int, int, int, str]:
    scope_rank = 0 if row.get("source_scope") == "BIBLE_HUNT_CODES" else 1
    method_rank = {
        "one_hunt_per_page_totals": 0,
        "utahdraws_summary_ladder_pdf": 1,
        "inline_table": 2,
        "stacked_column_table": 3,
        "na_glued_code_table": 4,
    }.get(row.get("parse_method", ""), 9)
    artifact_rank = 1 if row.get("artifact_status") == "NORMALIZED_NA_GLUE_ARTIFACT" else 0
    return (scope_rank, method_rank, artifact_rank, row.get("source_file", ""))


def preferred_script_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    selected: dict[tuple[str, str], dict[str, str]] = {}
    for row in rows:
        key = (row.get("draw_results_year", ""), row.get("normalized_hunt_code", ""))
        if not key[0] or not key[1]:
            continue
        current = selected.get(key)
        if current is None or preferred_rank(row) < preferred_rank(current):
            selected[key] = row
    return [selected[key] for key in sorted(selected)]


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract clean Sportsman draw-result rows from Sportsman PDFs.")
    parser.add_argument("--bible-root", default=str(DEFAULT_BIBLE_ROOT))
    parser.add_argument("--include-pipeline", action="store_true", help="Also scan repo pipeline Sportsman PDFs.")
    parser.add_argument("--output-csv", default=str(OUT_CSV))
    parser.add_argument("--script-feed-csv", default=str(OUT_SCRIPT_FEED))
    parser.add_argument("--output-xlsx", default=str(OUT_XLSX))
    parser.add_argument("--audit-csv", default=str(OUT_AUDIT))
    parser.add_argument("--summary-json", default=str(OUT_SUMMARY))
    args = parser.parse_args()

    bible_root = Path(args.bible_root)
    pdfs = discover_pdfs(bible_root, include_pipeline=args.include_pipeline)
    all_rows: list[dict[str, str]] = []
    audit_rows: list[dict[str, str]] = []
    for path in pdfs:
        rows, audit = extract_pdf(path, bible_root)
        all_rows.extend(rows)
        audit_rows.append(audit)

    all_rows.sort(key=lambda row: (row.get("draw_results_year", ""), row.get("normalized_hunt_code", ""), row.get("source_file", "")))
    script_rows = preferred_script_rows(all_rows)
    write_csv(Path(args.output_csv), all_rows, FIELDNAMES)
    write_csv(Path(args.script_feed_csv), script_rows, FIELDNAMES)
    write_csv(Path(args.audit_csv), audit_rows, list(audit_rows[0].keys()) if audit_rows else [])
    write_workbook(Path(args.output_xlsx), all_rows, audit_rows)

    summary = {
        "generated_at_utc": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "source_pdf_count": len(pdfs),
        "rows_extracted": len(all_rows),
        "preferred_script_feed_rows": len(script_rows),
        "unique_normalized_hunt_codes": len({row.get("normalized_hunt_code") for row in all_rows if row.get("normalized_hunt_code")}),
        "preferred_script_feed_year_counts": dict(sorted(Counter(row.get("draw_results_year", "") for row in script_rows).items())),
        "draw_results_year_counts": dict(sorted(Counter(row.get("draw_results_year", "") for row in all_rows).items())),
        "artifact_status_counts": dict(sorted(Counter(row.get("artifact_status", "") for row in all_rows).items())),
        "parse_method_counts": dict(sorted(Counter(row.get("parse_method", "") for row in all_rows).items())),
        "audit_status_counts": dict(sorted(Counter(row.get("status", "") for row in audit_rows).items())),
        "outputs": {
            "csv": str(Path(args.output_csv).relative_to(ROOT)).replace("\\", "/") if Path(args.output_csv).is_relative_to(ROOT) else args.output_csv,
            "script_feed_csv": str(Path(args.script_feed_csv).relative_to(ROOT)).replace("\\", "/") if Path(args.script_feed_csv).is_relative_to(ROOT) else args.script_feed_csv,
            "xlsx": str(Path(args.output_xlsx).relative_to(ROOT)).replace("\\", "/") if Path(args.output_xlsx).is_relative_to(ROOT) else args.output_xlsx,
            "audit_csv": str(Path(args.audit_csv).relative_to(ROOT)).replace("\\", "/") if Path(args.audit_csv).is_relative_to(ROOT) else args.audit_csv,
        },
    }
    Path(args.summary_json).write_text(json.dumps(summary, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
