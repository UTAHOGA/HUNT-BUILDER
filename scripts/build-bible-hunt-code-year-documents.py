from __future__ import annotations

import csv
import difflib
import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
BIBLE = Path(r"C:\Users\tyler\Desktop\BIBLE HUNT CODES")
OUT_DIR = ROOT / "processed_data" / "audits" / "bible_hunt_code_year_documents"
DOC_INDEX = ROOT / "docs" / "bible_hunt_code_year_documents_2018_2026.md"

YEARS = list(range(2018, 2027))
CODE_RE = re.compile(r"\b[A-Z]{2,3}\d{3,4}\b")
MODEL_RE = re.compile(r"PERMITS=(\d{4})_MODEL", re.I)

PREFIX_SPECIES = {
    "BI": "Bison",
    "BR": "Black Bear",
    "CG": "Cougar",
    "DA": "Antlerless Deer",
    "DB": "Buck Deer",
    "DS": "Desert Bighorn Sheep",
    "EA": "Antlerless Elk",
    "EB": "Bull Elk",
    "GO": "Mountain Goat",
    "MA": "Doe Pronghorn",
    "MB": "Bull Moose",
    "PB": "Buck Pronghorn",
    "PD": "Pronghorn",
    "RS": "Rocky Mountain Bighorn Sheep",
    "TK": "Turkey",
}

RAW_FIELDS = [
    "report_year",
    "draw_year",
    "model_year",
    "source_file",
    "source_extension",
    "source_location",
    "source_label",
    "source_family",
    "hunt_code_raw",
    "comparison_hunt_code",
    "prefix",
    "species_from_prefix",
    "hunt_title_or_row_text",
    "unit_or_area_inferred",
    "weapon_or_last_segment_inferred",
    "hunt_code_review_status",
    "candidate_normalized_hunt_code",
    "source_priority",
]

YEAR_FIELDS = [
    "report_year",
    "draw_year",
    "model_year",
    "comparison_hunt_code",
    "prefix",
    "species_from_prefix",
    "best_hunt_title_or_row_text",
    "unit_or_area_inferred",
    "weapon_or_last_segment_inferred",
    "source_family_values",
    "raw_hunt_codes_observed",
    "source_file_count",
    "source_files",
    "source_locations",
    "hunt_code_review_statuses",
    "candidate_normalized_hunt_codes",
    "year_document_status",
    "notes",
]

COMPARE_FIELDS = [
    "from_report_year",
    "to_report_year",
    "from_model_year",
    "to_model_year",
    "from_hunt_code",
    "to_hunt_code",
    "from_prefix",
    "to_prefix",
    "compare_status",
    "candidate_confidence",
    "identity_score",
    "from_species",
    "to_species",
    "from_unit_or_area",
    "to_unit_or_area",
    "from_weapon",
    "to_weapon",
    "from_title",
    "to_title",
    "from_source_files",
    "to_source_files",
    "recommended_next_action",
    "notes",
]


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def norm_text(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", " ", clean(value).lower()).strip()


def rel(path: Path) -> str:
    try:
        return path.relative_to(BIBLE).as_posix()
    except ValueError:
        return path.as_posix()


def prefix_of(code: str) -> str:
    match = re.match(r"^([A-Z]+)", code or "")
    return match.group(1) if match else ""


def report_label(path: Path) -> str:
    label = path.stem
    return label.split("__", 1)[1] if "__" in label else label


def report_family(label: str) -> str:
    upper = label.upper()
    if "SPORTSMAN" in upper:
        return "SPORTSMAN"
    if "COUGAR" in upper:
        return "COUGAR"
    if "BEAR" in upper:
        return "BEAR"
    if "TURKEY" in upper:
        return "TURKEY"
    if "YOUTH" in upper:
        return "YOUTH"
    if "D.H." in upper or "DEDICATED" in upper:
        return "DEDICATED_HUNTER_DEER"
    if "G.S." in upper or "GENERAL" in upper:
        return "GENERAL_SEASON"
    if "ANTLERLESS" in upper:
        return "ANTLERLESS"
    if "O.I.L." in upper or "OIL" in upper or "BISON" in upper or "MOOSE" in upper or "GOAT" in upper or "SHEEP" in upper:
        return "ONCE_IN_A_LIFETIME"
    if "L.E." in upper or "LE " in upper or "LIMITED" in upper:
        return "LIMITED_ENTRY"
    return "REVIEW"


def model_year_for(path: Path, year: int) -> str:
    match = MODEL_RE.search(path.name)
    return match.group(1) if match else str(year + 1)


def source_priority(path: Path) -> str:
    ext = path.suffix.lower()
    if ext == ".pdf":
        return "TIER_1_OR_2_PDF_SOURCE"
    if ext == ".xlsx":
        return "TIER_4_STRUCTURED_XLSX_HELPER"
    if ext == ".csv":
        return "TIER_4_STRUCTURED_CSV_HELPER"
    return "REVIEW_SOURCE_TYPE"


def normalize_code_for_compare(code: str) -> tuple[str, str, str]:
    prefix = prefix_of(code)
    if prefix.startswith("A") and prefix[1:] in PREFIX_SPECIES:
        candidate = code[1:]
        return candidate, "POSSIBLE_A_PREFIX_OCR_ARTIFACT", candidate
    return code, "AS_EXTRACTED", ""


def infer_identity(context: str, code: str) -> tuple[str, str]:
    text = clean(context)
    if code in text:
        text = clean(text.split(code, 1)[1])
    text = re.sub(r"^[-: ]+", "", text)
    text = re.sub(r"\bPage\s+\d+\b.*$", "", text, flags=re.I)
    parts = [clean(part) for part in re.split(r"\s+-\s+", text) if clean(part)]
    if len(parts) >= 3:
        return " - ".join(parts[1:-1]), parts[-1]
    if len(parts) == 2:
        return parts[0], parts[1]
    return "", ""


def raw_row(year: int, path: Path, location: str, code: str, context: str) -> dict[str, str]:
    comparison_code, review_status, candidate = normalize_code_for_compare(code)
    prefix = prefix_of(comparison_code)
    label = report_label(path)
    unit, weapon = infer_identity(context, code)
    return {
        "report_year": str(year),
        "draw_year": str(year),
        "model_year": model_year_for(path, year),
        "source_file": rel(path),
        "source_extension": path.suffix.lower(),
        "source_location": location,
        "source_label": label,
        "source_family": report_family(label),
        "hunt_code_raw": code,
        "comparison_hunt_code": comparison_code,
        "prefix": prefix,
        "species_from_prefix": PREFIX_SPECIES.get(prefix, ""),
        "hunt_title_or_row_text": clean(context),
        "unit_or_area_inferred": unit,
        "weapon_or_last_segment_inferred": weapon,
        "hunt_code_review_status": review_status,
        "candidate_normalized_hunt_code": candidate,
        "source_priority": source_priority(path),
    }


def extract_pdf(year: int, path: Path) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    try:
        reader = PdfReader(str(path))
    except Exception as exc:
        return [raw_error(year, path, f"PDF_READ_ERROR: {exc}")]
    for page_num, page in enumerate(reader.pages, start=1):
        try:
            text = page.extract_text() or ""
        except Exception as exc:
            rows.append(raw_error(year, path, f"PDF_PAGE_{page_num}_EXTRACT_ERROR: {exc}"))
            continue
        lines = [clean(line) for line in text.splitlines() if clean(line)]
        if not lines:
            continue
        for idx, line in enumerate(lines):
            codes = CODE_RE.findall(line.upper())
            if not codes:
                continue
            context = line
            if "Hunt:" in line or "Hunt #" in line or len(line) > 20:
                context = line
            else:
                context = clean(" ".join(lines[max(0, idx - 1) : min(len(lines), idx + 2)]))
            for code in sorted(set(codes)):
                rows.append(raw_row(year, path, f"pdf_page_{page_num}", code, context))
    return rows


def extract_xlsx(year: int, path: Path) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return [raw_error(year, path, f"XLSX_READ_ERROR: {exc}")]
    for ws in wb.worksheets:
        for row_num, cells in enumerate(ws.iter_rows(values_only=True), start=1):
            values = [clean(value) for value in cells if clean(value)]
            if not values:
                continue
            text = " | ".join(values)
            codes = CODE_RE.findall(text.upper())
            for code in sorted(set(codes)):
                rows.append(raw_row(year, path, f"{ws.title}!row_{row_num}", code, text))
    return rows


def extract_csv_file(year: int, path: Path) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    try:
        with path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle)
            for row_num, cells in enumerate(reader, start=1):
                text = " | ".join(clean(cell) for cell in cells if clean(cell))
                codes = CODE_RE.findall(text.upper())
                for code in sorted(set(codes)):
                    rows.append(raw_row(year, path, f"csv_row_{row_num}", code, text))
    except Exception as exc:
        return [raw_error(year, path, f"CSV_READ_ERROR: {exc}")]
    return rows


def raw_error(year: int, path: Path, message: str) -> dict[str, str]:
    return {
        "report_year": str(year),
        "draw_year": str(year),
        "model_year": model_year_for(path, year),
        "source_file": rel(path),
        "source_extension": path.suffix.lower(),
        "source_location": "",
        "source_label": report_label(path),
        "source_family": report_family(report_label(path)),
        "hunt_code_raw": "",
        "comparison_hunt_code": "",
        "prefix": "",
        "species_from_prefix": "",
        "hunt_title_or_row_text": "",
        "unit_or_area_inferred": "",
        "weapon_or_last_segment_inferred": "",
        "hunt_code_review_status": "SOURCE_READ_ERROR",
        "candidate_normalized_hunt_code": "",
        "source_priority": source_priority(path),
        "notes": message,
    }


def source_files_for_year(year: int) -> list[Path]:
    year_dir = BIBLE / str(year)
    if not year_dir.exists():
        return []
    files = [
        path
        for path in year_dir.rglob("*")
        if path.is_file() and path.suffix.lower() in {".pdf", ".xlsx", ".csv"}
    ]
    if year == 2026:
        xlsx_stems = {path.stem for path in files if path.suffix.lower() == ".xlsx"}
        files = [
            path
            for path in files
            if not (path.suffix.lower() == ".pdf" and path.stem in xlsx_stems)
        ]
    return sorted(files, key=lambda path: rel(path).lower())


def extract_year(year: int) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for path in source_files_for_year(year):
        ext = path.suffix.lower()
        if ext == ".pdf":
            rows.extend(extract_pdf(year, path))
        elif ext == ".xlsx":
            rows.extend(extract_xlsx(year, path))
        elif ext == ".csv":
            rows.extend(extract_csv_file(year, path))
    return [row for row in rows if row.get("comparison_hunt_code") or row.get("hunt_code_review_status") == "SOURCE_READ_ERROR"]


def choose_best(rows: list[dict[str, str]]) -> dict[str, str]:
    def score(row: dict[str, str]) -> tuple[int, int, int]:
        priority = 3 if row["source_extension"] == ".pdf" else 2 if row["source_extension"] == ".xlsx" else 1
        title_len = len(row["hunt_title_or_row_text"])
        clean_status = 1 if row["hunt_code_review_status"] == "AS_EXTRACTED" else 0
        return priority, clean_status, title_len

    return max(rows, key=score)


def year_document(year: int, raw_rows: list[dict[str, str]]) -> list[dict[str, object]]:
    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in raw_rows:
        code = row.get("comparison_hunt_code", "")
        if code:
            grouped[code].append(row)
    out: list[dict[str, object]] = []
    for code, rows in sorted(grouped.items()):
        best = choose_best(rows)
        statuses = sorted({row["hunt_code_review_status"] for row in rows})
        source_files = sorted({row["source_file"] for row in rows})
        locations = sorted({row["source_location"] for row in rows if row["source_location"]})
        raw_codes = sorted({row["hunt_code_raw"] for row in rows})
        candidates = sorted({row["candidate_normalized_hunt_code"] for row in rows if row["candidate_normalized_hunt_code"]})
        families = sorted({row["source_family"] for row in rows if row["source_family"]})
        if "SOURCE_READ_ERROR" in statuses:
            doc_status = "SOURCE_READ_REVIEW"
        elif "POSSIBLE_A_PREFIX_OCR_ARTIFACT" in statuses:
            doc_status = "NORMALIZED_FOR_COMPARISON_REVIEW"
        else:
            doc_status = "OK"
        out.append(
            {
                "report_year": year,
                "draw_year": year,
                "model_year": best["model_year"],
                "comparison_hunt_code": code,
                "prefix": best["prefix"],
                "species_from_prefix": best["species_from_prefix"],
                "best_hunt_title_or_row_text": best["hunt_title_or_row_text"],
                "unit_or_area_inferred": best["unit_or_area_inferred"],
                "weapon_or_last_segment_inferred": best["weapon_or_last_segment_inferred"],
                "source_family_values": "|".join(families),
                "raw_hunt_codes_observed": "|".join(raw_codes),
                "source_file_count": len(source_files),
                "source_files": "|".join(source_files),
                "source_locations": "|".join(locations[:80]),
                "hunt_code_review_statuses": "|".join(statuses),
                "candidate_normalized_hunt_codes": "|".join(candidates),
                "year_document_status": doc_status,
                "notes": "Independent BIBLE year-folder extraction; compare/crosswalk is downstream.",
            }
        )
    return out


def identity_key(row: dict[str, object]) -> str:
    return " | ".join(
        [
            norm_text(row.get("species_from_prefix")),
            norm_text(row.get("unit_or_area_inferred")),
            norm_text(row.get("weapon_or_last_segment_inferred")),
            norm_text(row.get("best_hunt_title_or_row_text")),
            norm_text(row.get("source_family_values")),
        ]
    )


def identity_score(left: dict[str, object], right: dict[str, object]) -> float:
    return round(difflib.SequenceMatcher(None, identity_key(left), identity_key(right)).ratio(), 4)


def compare_years(year_docs: dict[int, list[dict[str, object]]]) -> tuple[list[dict[str, object]], dict[str, list[dict[str, object]]]]:
    all_rows: list[dict[str, object]] = []
    by_pair: dict[str, list[dict[str, object]]] = {}
    for from_year, to_year in zip(YEARS[:-1], YEARS[1:]):
        left_by_code = {str(row["comparison_hunt_code"]): row for row in year_docs.get(from_year, [])}
        right_by_code = {str(row["comparison_hunt_code"]): row for row in year_docs.get(to_year, [])}
        retained = sorted(set(left_by_code) & set(right_by_code))
        dropped = sorted(set(left_by_code) - set(right_by_code))
        added = sorted(set(right_by_code) - set(left_by_code))
        added_unmatched = set(added)
        pair_rows: list[dict[str, object]] = []
        for code in retained:
            pair_rows.append(compare_row(from_year, to_year, left_by_code[code], right_by_code[code], "EXACT_CODE_RETAINED", "EXACT_CODE", 1.0, "Carry forward exact code continuity."))
        for code in dropped:
            left = left_by_code[code]
            candidates = []
            for right_code in added:
                right = right_by_code[right_code]
                if left["prefix"] != right["prefix"]:
                    continue
                score = identity_score(left, right)
                if score >= 0.74:
                    candidates.append((score, right_code, right))
            candidates.sort(reverse=True, key=lambda item: item[0])
            if candidates:
                score, right_code, right = candidates[0]
                if right_code in added_unmatched:
                    added_unmatched.remove(right_code)
                confidence = "HIGH_REVIEW" if score >= 0.93 else "MEDIUM_REVIEW" if score >= 0.84 else "LOW_REVIEW"
                pair_rows.append(compare_row(from_year, to_year, left, right, "CANDIDATE_SUCCESSOR_BY_YEAR_DOCUMENT", confidence, score, "Review official year documents before promoting this candidate."))
            else:
                pair_rows.append(compare_row(from_year, to_year, left, {}, "DROPPED_NO_YEAR_DOCUMENT_SUCCESSOR", "NO_SAFE_MATCH", "", "Resolve as retired, renamed outside threshold, source gap, or extraction issue."))
        for code in sorted(added_unmatched):
            right = right_by_code[code]
            pair_rows.append(compare_row(from_year, to_year, {}, right, "ADDED_NO_YEAR_DOCUMENT_PREDECESSOR", "NO_SAFE_MATCH", "", "Resolve as new, reactivated, renamed outside threshold, or prior-year extraction gap."))
        pair_rows.sort(key=lambda row: (str(row["compare_status"]), str(row["from_prefix"] or row["to_prefix"]), str(row["from_hunt_code"]), str(row["to_hunt_code"])))
        by_pair[f"{from_year}_to_{to_year}"] = pair_rows
        all_rows.extend(pair_rows)
    return all_rows, by_pair


def compare_row(from_year: int, to_year: int, left: dict[str, object], right: dict[str, object], status: str, confidence: str, score: object, action: str) -> dict[str, object]:
    return {
        "from_report_year": from_year,
        "to_report_year": to_year,
        "from_model_year": from_year + 1,
        "to_model_year": to_year + 1,
        "from_hunt_code": left.get("comparison_hunt_code", ""),
        "to_hunt_code": right.get("comparison_hunt_code", ""),
        "from_prefix": left.get("prefix", ""),
        "to_prefix": right.get("prefix", ""),
        "compare_status": status,
        "candidate_confidence": confidence,
        "identity_score": score,
        "from_species": left.get("species_from_prefix", ""),
        "to_species": right.get("species_from_prefix", ""),
        "from_unit_or_area": left.get("unit_or_area_inferred", ""),
        "to_unit_or_area": right.get("unit_or_area_inferred", ""),
        "from_weapon": left.get("weapon_or_last_segment_inferred", ""),
        "to_weapon": right.get("weapon_or_last_segment_inferred", ""),
        "from_title": left.get("best_hunt_title_or_row_text", ""),
        "to_title": right.get("best_hunt_title_or_row_text", ""),
        "from_source_files": left.get("source_files", ""),
        "to_source_files": right.get("source_files", ""),
        "recommended_next_action": action,
        "notes": "Built from independent BIBLE year documents.",
    }


def write_csv(path: Path, rows: list[dict[str, object]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def load_existing_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return [{key: clean(value) for key, value in row.items()} for row in csv.DictReader(handle)]


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="244062")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        width = max(len(clean(cell.value)) for cell in column_cells)
        ws.column_dimensions[letter].width = min(max(width + 2, 10), 45)


def write_workbook(path: Path, sheets: dict[str, tuple[list[str], list[dict[str, object]]]]) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    for sheet_name, (fields, rows) in sheets.items():
        ws = wb.create_sheet(sheet_name[:31])
        ws.append(fields)
        for row in rows:
            ws.append([row.get(field, "") for field in fields])
        style_sheet(ws)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_index(summary: dict[str, object], year_summaries: list[dict[str, object]], pair_summaries: list[dict[str, object]]) -> None:
    lines = [
        "# BIBLE Hunt Code Year Documents 2018-2026",
        "",
        "## Purpose",
        "",
        "These documents are generated independently from each `BIBLE HUNT CODES` year folder. The compare/crosswalk outputs are downstream and use the year documents as inputs.",
        "",
        "## Year Documents",
        "",
    ]
    for row in year_summaries:
        year = row["report_year"]
        lines.append(f"- `{year}`: `{row['unique_hunt_codes']}` unique comparison hunt codes, `{row['raw_source_hits']}` raw source hits")
    lines.extend(["", "## Year-To-Year Compare", ""])
    for row in pair_summaries:
        lines.append(
            f"- `{row['from_report_year']}->{row['to_report_year']}`: exact `{row['exact_retained']}`, candidate `{row['candidate_successor']}`, dropped `{row['dropped_no_successor']}`, added `{row['added_no_predecessor']}`"
        )
    lines.extend(["", "## Outputs", ""])
    for key, value in summary["outputs"].items():
        if isinstance(value, list):
            lines.append(f"- `{key}`:")
            for item in value:
                lines.append(f"- `{item}`")
        else:
            lines.append(f"- `{value}`")
    lines.extend(
        [
            "",
            "## Guardrail",
            "",
            "The year documents preserve source-folder evidence. Candidate crosswalk rows are review evidence only and do not modify `DATABASE.csv` or promote permit truth.",
        ]
    )
    DOC_INDEX.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    raw_by_year: dict[int, list[dict[str, str]]] = {}
    docs_by_year: dict[int, list[dict[str, object]]] = {}
    year_summaries: list[dict[str, object]] = []
    workbook_sheets: dict[str, tuple[list[str], list[dict[str, object]]]] = {}

    for year in YEARS:
        raw_path = OUT_DIR / f"bible_hunt_code_source_hits_{year}.csv"
        doc_path = OUT_DIR / f"bible_hunt_code_year_document_{year}.csv"
        if raw_path.exists() and doc_path.exists():
            raw_rows = load_existing_csv(raw_path)
            doc_rows = load_existing_csv(doc_path)
        else:
            raw_rows = extract_year(year)
            doc_rows = year_document(year, raw_rows)
            write_csv(raw_path, raw_rows, RAW_FIELDS)
            write_csv(doc_path, doc_rows, YEAR_FIELDS)
        raw_by_year[year] = raw_rows
        docs_by_year[year] = doc_rows
        write_workbook(OUT_DIR / f"bible_hunt_code_year_document_{year}.xlsx", {str(year): (YEAR_FIELDS, doc_rows)})
        workbook_sheets[str(year)] = (YEAR_FIELDS, doc_rows)
        year_summaries.append(
            {
                "report_year": year,
                "draw_year": year,
                "model_year": year + 1,
                "source_files_scanned": len(source_files_for_year(year)),
                "raw_source_hits": len(raw_rows),
                "unique_hunt_codes": len(doc_rows),
                "review_status_counts": json.dumps(dict(Counter(row["year_document_status"] for row in doc_rows)), sort_keys=True),
                "output_csv": f"processed_data/audits/bible_hunt_code_year_documents/bible_hunt_code_year_document_{year}.csv",
                "output_xlsx": f"processed_data/audits/bible_hunt_code_year_documents/bible_hunt_code_year_document_{year}.xlsx",
            }
        )

    all_compare, compare_by_pair = compare_years(docs_by_year)
    pair_summaries: list[dict[str, object]] = []
    for key, rows in compare_by_pair.items():
        from_year, to_year = [int(part) for part in key.split("_to_")]
        write_csv(OUT_DIR / f"bible_hunt_code_year_compare_{key}.csv", rows, COMPARE_FIELDS)
        workbook_sheets[f"{from_year}-{to_year}"] = (COMPARE_FIELDS, rows)
        counts = Counter(row["compare_status"] for row in rows)
        pair_summaries.append(
            {
                "from_report_year": from_year,
                "to_report_year": to_year,
                "exact_retained": counts.get("EXACT_CODE_RETAINED", 0),
                "candidate_successor": counts.get("CANDIDATE_SUCCESSOR_BY_YEAR_DOCUMENT", 0),
                "dropped_no_successor": counts.get("DROPPED_NO_YEAR_DOCUMENT_SUCCESSOR", 0),
                "added_no_predecessor": counts.get("ADDED_NO_YEAR_DOCUMENT_PREDECESSOR", 0),
                "total_rows": len(rows),
                "output_csv": f"processed_data/audits/bible_hunt_code_year_documents/bible_hunt_code_year_compare_{key}.csv",
            }
        )

    write_csv(OUT_DIR / "bible_hunt_code_year_document_summary_2018_2026.csv", year_summaries, list(year_summaries[0].keys()))
    write_csv(OUT_DIR / "bible_hunt_code_year_compare_summary_2018_2026.csv", pair_summaries, list(pair_summaries[0].keys()))
    write_csv(OUT_DIR / "bible_hunt_code_year_compare_all_2018_2026.csv", all_compare, COMPARE_FIELDS)
    write_workbook(OUT_DIR / "bible_hunt_code_year_documents_and_compare_2018_2026.xlsx", workbook_sheets)

    summary = {
        "created_at_utc": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        "scope": "Independent BIBLE HUNT CODES year documents followed by downstream adjacent-year compare/crosswalk.",
        "bible_root": str(BIBLE),
        "years": YEARS,
        "total_unique_year_document_rows": sum(int(row["unique_hunt_codes"]) for row in year_summaries),
        "total_raw_source_hits": sum(int(row["raw_source_hits"]) for row in year_summaries),
        "total_compare_rows": len(all_compare),
        "year_summaries": year_summaries,
        "pair_summaries": pair_summaries,
        "outputs": {
            "year_summary_csv": "processed_data/audits/bible_hunt_code_year_documents/bible_hunt_code_year_document_summary_2018_2026.csv",
            "compare_summary_csv": "processed_data/audits/bible_hunt_code_year_documents/bible_hunt_code_year_compare_summary_2018_2026.csv",
            "all_compare_csv": "processed_data/audits/bible_hunt_code_year_documents/bible_hunt_code_year_compare_all_2018_2026.csv",
            "combined_workbook": "processed_data/audits/bible_hunt_code_year_documents/bible_hunt_code_year_documents_and_compare_2018_2026.xlsx",
            "index_doc": "docs/bible_hunt_code_year_documents_2018_2026.md",
            "per_year_csv_xlsx": [
                f"processed_data/audits/bible_hunt_code_year_documents/bible_hunt_code_year_document_{year}.csv/.xlsx"
                for year in YEARS
            ],
        },
        "guardrail": "No DATABASE.csv changes; year docs are source evidence and candidate compare rows remain review evidence.",
    }
    (OUT_DIR / "bible_hunt_code_year_documents_2018_2026_summary.json").write_text(json.dumps(summary, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    write_index(summary, year_summaries, pair_summaries)
    print(json.dumps(summary, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
