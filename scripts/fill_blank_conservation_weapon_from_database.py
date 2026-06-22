from __future__ import annotations

import csv
import re
import shutil
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "processed_data" / "hard_data_exports" / "hunt_tables" / "2026" / "CLEAN_XLXS_STAGED" / "2025-27 Conservation Permits.xlsx"
DATABASE_PATH = ROOT / "pipeline" / "RAW" / "hunt_unit_database" / "2026" / "csv" / "DATABASE.csv"
AUDIT_DIR = ROOT / "audits" / "2025_canonical_finalization"
BACKUP_DIR = AUDIT_DIR / "backups"
AUDIT_PATH = AUDIT_DIR / "2025_27_conservation_blank_weapon_database_fill.csv"


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def norm(value: object) -> str:
    return re.sub(r"\s+", " ", clean(value).lower()).strip()


def normalize_species(value: object) -> str:
    text = norm(value)
    if text == "antlerless elk":
        return "elk"
    if text == "bear":
        return "black bear"
    return text


def normalize_sex(value: object) -> str:
    text = norm(value).replace("'", "")
    if text in {"hunters choice", "hunter s choice", "either sex"}:
        return "either sex"
    if text == "male only":
        return "male only"
    if text == "female only":
        return "female only"
    return text


def normalize_boundary(value: object) -> str:
    text = clean(value)
    if text.endswith(".0"):
        text = text[:-2]
    return text


def add_season(existing: str, value: str) -> str:
    existing_text = clean(existing)
    value_text = clean(value)
    if not value_text:
        return existing_text
    parts = [part.strip() for part in existing_text.split(";") if part.strip()]
    if value_text.lower() not in {part.lower() for part in parts}:
        parts.append(value_text)
    return "; ".join(parts)


def normalize_db_weapon_for_workbook(db_weapon: str, existing_season: str) -> tuple[str, str, list[str]]:
    weapon = clean(db_weapon)
    season = clean(existing_season)
    notes: list[str] = []

    if norm(weapon) in {"hunters choice", "hunter's choice", "hunter s choice"}:
        season = add_season(season, "Hunter's Choice")
        notes.append("database_hunters_choice_moved_to_season")
        return "", season, notes

    match = re.match(r"(?i)^(early|mid|late)\s+(.+)$", weapon)
    if match:
        season = add_season(season, match.group(1).title())
        weapon = match.group(2).strip()
        notes.append("database_timing_moved_to_season")

    match = re.match(r"(?i)^(.+?),\s*(early|mid|late)$", weapon)
    if match:
        season = add_season(season, match.group(2).title())
        weapon = match.group(1).strip()
        notes.append("database_timing_moved_to_season")

    return weapon, season, notes


def read_database() -> dict[str, list[dict[str, str]]]:
    by_code: dict[str, list[dict[str, str]]] = defaultdict(list)
    with DATABASE_PATH.open("r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            code = clean(row.get("hunt_code")).upper()
            if code:
                by_code[code].append({k: clean(v) for k, v in row.items()})
    return by_code


def choose_database_match(workbook_row: dict[str, str], candidates: list[dict[str, str]]) -> tuple[dict[str, str] | None, str]:
    if not candidates:
        return None, "no_database_code_match"

    species = normalize_species(workbook_row["Species"])
    boundary = normalize_boundary(workbook_row["BOUNDARY ID"])
    sex = normalize_sex(workbook_row["SEX TYPE"])

    scored: list[tuple[int, dict[str, str], list[str]]] = []
    for candidate in candidates:
        score = 0
        notes: list[str] = []
        if normalize_species(candidate.get("species")) == species:
            score += 4
            notes.append("species")
        if normalize_boundary(candidate.get("boundary_id")) == boundary:
            score += 4
            notes.append("boundary_id")
        if normalize_sex(candidate.get("sex_type")) == sex:
            score += 3
            notes.append("sex_type")
        if clean(candidate.get("weapon")):
            score += 1
            notes.append("weapon")
        scored.append((score, candidate, notes))

    scored.sort(key=lambda item: item[0], reverse=True)
    best_score, best_candidate, best_notes = scored[0]
    second_score = scored[1][0] if len(scored) > 1 else -1
    if best_score < 8:
        return best_candidate, "low_confidence_" + "_".join(best_notes)
    if second_score == best_score:
        best_weapons = {clean(item[1].get("weapon")) for item in scored if item[0] == best_score}
        if len(best_weapons) > 1:
            return best_candidate, "review_tied_distinct_weapons"
    return best_candidate, "matched_" + "_".join(best_notes)


def main() -> None:
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)

    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    backup_path = BACKUP_DIR / f"{WORKBOOK_PATH.stem}.before_blank_weapon_database_fill_{stamp}{WORKBOOK_PATH.suffix}"
    shutil.copy2(WORKBOOK_PATH, backup_path)

    database_by_code = read_database()

    wb = load_workbook(WORKBOOK_PATH)
    ws = wb["Table 1"] if "Table 1" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [clean(ws.cell(1, column).value) for column in range(1, ws.max_column + 1)]
    cols = {header: index + 1 for index, header in enumerate(headers) if header}
    required = ["HUNT CODE", "BOUNDARY ID", "Species", "SEX TYPE", "WEAPON", "SEASON"]
    missing = [header for header in required if header not in cols]
    if missing:
        raise RuntimeError(f"Missing required headers: {missing}")

    audit_rows = []
    for row_idx in range(2, ws.max_row + 1):
        old_weapon = clean(ws.cell(row_idx, cols["WEAPON"]).value)
        if old_weapon:
            continue

        workbook_row = {header: clean(ws.cell(row_idx, col_idx).value) for header, col_idx in cols.items()}
        candidates = database_by_code.get(workbook_row["HUNT CODE"].upper(), [])
        candidate, match_status = choose_database_match(workbook_row, candidates)
        if candidate is None:
            audit_rows.append(
                {
                    "row": row_idx,
                    "status": match_status,
                    "hunt_code": workbook_row["HUNT CODE"],
                    "boundary_id": workbook_row["BOUNDARY ID"],
                    "species": workbook_row["Species"],
                    "sex_type": workbook_row["SEX TYPE"],
                }
            )
            continue

        db_weapon = clean(candidate.get("weapon"))
        new_weapon = db_weapon
        new_season = workbook_row["SEASON"]
        transform_notes: list[str] = ["database_weapon_written_directly"]
        status = match_status
        if not new_weapon:
            status += "_no_weapon_after_normalization"
        else:
            ws.cell(row_idx, cols["WEAPON"]).value = new_weapon
            ws.cell(row_idx, cols["SEASON"]).value = new_season
            status += "_filled"

        audit_rows.append(
            {
                "row": row_idx,
                "status": status,
                "hunt_code": workbook_row["HUNT CODE"],
                "boundary_id": workbook_row["BOUNDARY ID"],
                "species": workbook_row["Species"],
                "hunt_name": workbook_row.get("HUNT NAME", ""),
                "sex_type": workbook_row["SEX TYPE"],
                "old_weapon": old_weapon,
                "new_weapon": new_weapon,
                "old_season": workbook_row["SEASON"],
                "new_season": new_season,
                "db_weapon": db_weapon,
                "db_hunt_name": candidate.get("hunt_name", ""),
                "db_species": candidate.get("species", ""),
                "db_boundary_id": candidate.get("boundary_id", ""),
                "db_sex_type": candidate.get("sex_type", ""),
                "notes": ";".join(transform_notes),
            }
        )

    wb.save(WORKBOOK_PATH)

    fieldnames = [
        "row",
        "status",
        "hunt_code",
        "boundary_id",
        "species",
        "hunt_name",
        "sex_type",
        "old_weapon",
        "new_weapon",
        "old_season",
        "new_season",
        "db_weapon",
        "db_hunt_name",
        "db_species",
        "db_boundary_id",
        "db_sex_type",
        "notes",
    ]
    with AUDIT_PATH.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(audit_rows)

    filled = sum(1 for row in audit_rows if row["status"].endswith("_filled"))
    print(f"blank_weapon_rows_seen={len(audit_rows)}")
    print(f"filled={filled}")
    print(f"backup={backup_path}")
    print(f"audit={AUDIT_PATH}")


if __name__ == "__main__":
    main()
