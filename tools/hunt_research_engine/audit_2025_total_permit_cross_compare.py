#!/usr/bin/env python3
"""Small 2025 draw-vs-harvest total permit cross-compare.

This read-only audit focuses only on total permit numbers. It compares 2025
draw-result totals from the normalized draw truth table against the 2025
preliminary big-game harvest permit counts. It does not mutate source files,
DATABASE.csv, runtime files, or model outputs.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path


def _repo_root() -> Path:
    repo_root = Path(__file__).resolve()
    while repo_root.name != "HUNT-BUILDER" and repo_root.parent != repo_root:
        repo_root = repo_root.parent
    if repo_root.name != "HUNT-BUILDER":
        raise RuntimeError("Could not locate HUNT-BUILDER repo root")
    return repo_root


from typing import Iterable

from openpyxl import load_workbook


DEFAULT_DRAW = "data_truth/draw_results_truth/normalized/draw_results_long.csv"
DEFAULT_HARVEST = "data_truth/harvest_results_truth/normalized/harvest_results_all_years_long.csv"
DEFAULT_DATABASE = "pipeline/RAW/hunt_unit_database/2026/csv/DATABASE.csv"
DEFAULT_OUT_DIR = "audits/hunt_research_engine"
DEFAULT_EXPO_XLSX = "pipeline/RAW/hunt_unit_database/2026/xlsx/2025  Expo Permits Draw Results.xlsx"
DEFAULT_EXTERNAL_EXPO_PDF = (
    str(
        _repo_root()
        / "pipeline/RAW/hunt_unit_database/2025/pdf/HUNT EXPO/2025 Hunt Expo Draw Permit Successful Applicants - Western Hunting & Conservation Expo.pdf"
    )
)
DRAW_YEAR = 2025
HARVEST_YEAR = 2025
HARVEST_SOURCE = "2026-03-06-2025-preliminary-bg-harvest.xlsx"

LISTED_PDFS = [
    "pipeline/RAW/hunt_unit_database/2025/bible_truth/pdf/2025_PERMITS=2026_MODEL__L.E. PRONGHORN DRAW RESULTS.pdf",
    "pipeline/RAW/hunt_unit_database/2025/bible_truth/pdf/2025_PERMITS=2026_MODEL__O.I.L. BISON DRAW RESULTS.pdf",
    "pipeline/RAW/hunt_unit_database/2025/bible_truth/pdf/2025_PERMITS=2026_MODEL__O.I.L. BULL MOOSE DRAW RESULTS.pdf",
    "pipeline/RAW/hunt_unit_database/2025/bible_truth/pdf/2025_PERMITS=2026_MODEL__O.I.L. DESERT BIGHORN SHEEP DRAW RESULTS.pdf",
    "pipeline/RAW/hunt_unit_database/2025/bible_truth/pdf/2025_PERMITS=2026_MODEL__O.I.L. DRAW RESULTS.pdf",
    "pipeline/RAW/hunt_unit_database/2025/bible_truth/pdf/2025_PERMITS=2026_MODEL__O.I.L. MTN GOAT DRAW RESULTS.pdf",
    "pipeline/RAW/hunt_unit_database/2025/bible_truth/pdf/2025_PERMITS=2026_MODEL__O.I.L. ROCKY MTN SHEEP DRAW RESULTS.pdf",
    "pipeline/RAW/hunt_unit_database/2025/bible_truth/pdf/2025_PERMITS=2026_MODEL__L.E. BIG GAME DRAW RESULTS.pdf",
    "pipeline/RAW/hunt_unit_database/2025/bible_truth/pdf/2025_PERMITS=2026_MODEL__L.E. DEER DRAW RESULTS.pdf",
    "pipeline/RAW/hunt_unit_database/2025/bible_truth/pdf/2025_PERMITS=2026_MODEL__L.E. ELK DRAW RESULTS.pdf",
    "pipeline/RAW/hunt_unit_database/2026/pdf/draw_odds/2025 Big Game Draw Results.pdf",
    "pipeline/RAW/hunt_unit_database/2026/pdf/harvest_report/2026-03-06-2025-preliminary-bg-harvest.pdf",
]

SOURCE_HINTS = [
    "2025 LE Pronghorn Draw Results",
    "2025 O.I.L. Draw Results",
    "2025 LE Deer Draw Results",
    "2025 LE Elk Draw Results",
    "2025 Big Game Draw Results",
]

REPORT_COLUMNS = [
    "hunt_code",
    "species",
    "sex_type",
    "hunt_name",
    "hunt_type",
    "weapon",
    "source_scope",
    "comparison_status",
    "draw_total_permits",
    "expo_permits_matched",
    "draw_plus_expo_total",
    "harvest_permits",
    "permit_delta_harvest_minus_draw",
    "permit_delta_harvest_minus_draw_plus_expo",
    "draw_plus_expo_status",
    "expo_match_status",
    "expo_match_sources",
    "draw_source_files",
    "harvest_source_files",
    "draw_row_count",
    "harvest_row_count",
    "draw_residencies",
    "draw_points_count",
    "harvest_unique_permit_values",
    "in_2026_database",
    "same_total_match_ready",
    "notes",
]


@dataclass(frozen=True)
class Paths:
    root: Path
    draw: Path
    harvest: Path
    database: Path
    expo_xlsx: Path
    external_expo_pdf: Path
    out_dir: Path


def norm(value: object) -> str:
    return "" if value is None else str(value).strip()


def key_text(value: object) -> str:
    text = norm(value).lower().replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def number(value: object) -> float | None:
    text = norm(value).replace(",", "")
    if not text or text.upper() in {"NA", "N/A", "UNLIMITED"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def format_number(value: float | None) -> str:
    if value is None:
        return ""
    if float(value).is_integer():
        return str(int(value))
    return f"{value:.4f}".rstrip("0").rstrip(".")


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), [dict(row) for row in reader]


def write_csv(path: Path, rows: list[dict[str, object]], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column, "") for column in columns})


def write_json(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def most_common(values: Iterable[object]) -> str:
    counts = Counter(norm(value) for value in values if norm(value))
    if not counts:
        return ""
    return counts.most_common(1)[0][0]


def source_in_scope(source_file: object) -> bool:
    source = key_text(source_file)
    return any(key_text(hint) in source for hint in SOURCE_HINTS)


def find_external_expo_pdf_matches(path: Path) -> list[dict[str, object]]:
    if path.exists():
        return [{"path": str(path), "size_bytes": path.stat().st_size}]
    parent = path.parent
    if not parent.exists():
        return []
    matches = sorted(parent.glob("*Expo*Successful*Applicants*.pdf"))
    return [{"path": str(match), "size_bytes": match.stat().st_size} for match in matches]


def source_scope(source_files: Iterable[str]) -> str:
    files = list(source_files)
    if not files:
        return "NO_DRAW_SOURCE"
    if any("o i l" in key_text(file) for file in files):
        return "OIL_OR_BIG_GAME_SCOPE"
    if any("le" in key_text(file) or "limited entry" in key_text(file) for file in files):
        return "LE_OR_BIG_GAME_SCOPE"
    if any("big game" in key_text(file) for file in files):
        return "BIG_GAME_SCOPE"
    return "OTHER_2025_DRAW_SCOPE"


def species_family(value: object) -> str:
    text = key_text(value)
    if "buck deer" in text or "deer" in text or "mule deer" in text:
        return "deer"
    if "elk" in text:
        return "elk"
    if "pronghorn" in text:
        return "pronghorn"
    if "bison" in text:
        return "bison"
    if "moose" in text:
        return "moose"
    if "desert bighorn" in text:
        return "desert bighorn sheep"
    if "rocky" in text and "bighorn" in text:
        return "rocky mountain bighorn sheep"
    if "bighorn" in text:
        return "bighorn sheep"
    if "goat" in text:
        return "mountain goat"
    if "bear" in text:
        return "black bear"
    if "turkey" in text:
        return "turkey"
    return text


def comparable_tokens(value: object) -> set[str]:
    stop = {
        "and",
        "any",
        "legal",
        "weapon",
        "limited",
        "entry",
        "premium",
        "once",
        "lifetime",
        "general",
        "season",
        "permit",
        "permits",
        "deer",
        "buck",
        "elk",
        "bull",
        "cow",
        "bison",
        "moose",
        "pronghorn",
        "sheep",
        "goat",
        "bear",
        "turkey",
        "mtn",
        "mtns",
        "mountain",
        "mountains",
        "north",
        "south",
        "east",
        "west",
    }
    return {part for part in key_text(value).split() if len(part) > 2 and part not in stop}


def weapon_family(value: object) -> str:
    text = key_text(value)
    if "archery" in text:
        return "archery"
    if "muzzle" in text:
        return "muzzleloader"
    if "multi" in text:
        return "multi season"
    if "any weapon" in text or "any legal" in text or "alw" in text:
        return "any legal weapon"
    return text


def hunt_type_family(value: object) -> str:
    text = key_text(value)
    if "premium" in text:
        return "premium limited entry"
    if "limited entry" in text:
        return "limited entry"
    if "once" in text or "lifetime" in text:
        return "once in a lifetime"
    if "general" in text:
        return "general season"
    if "conservation" in text:
        return "conservation"
    if "expo" in text:
        return "expo"
    if "sports" in text:
        return "sportsman"
    return text


def parse_expo_heading(value: object) -> dict[str, object] | None:
    raw = norm(value)
    if "Permits:" not in raw:
        return None
    match = re.search(r"Permits:\s*(\d+)", raw, flags=re.IGNORECASE)
    if not match:
        return None
    count = int(match.group(1))
    title = raw[: match.start()].strip(" -")
    parts = [part.strip() for part in title.split(" - ") if part.strip()]
    if len(parts) < 2:
        return None
    return {
        "raw_heading": raw,
        "species": parts[0],
        "hunt_type": parts[1] if len(parts) > 2 else "",
        "weapon": parts[2] if len(parts) > 3 else "",
        "area": parts[-1],
        "permits": count,
    }


def load_expo_rows(path: Path) -> list[dict[str, object]]:
    if not path.exists():
        return []
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict[str, object]] = []
    for sheet in workbook.worksheets:
        for cells in sheet.iter_rows(values_only=True):
            parsed = parse_expo_heading(cells[0] if cells else "")
            if not parsed:
                continue
            parsed["source_file"] = str(path)
            parsed["source_sheet"] = sheet.title
            rows.append(parsed)
    return rows


def aggregate_draw(rows: Iterable[dict[str, str]]) -> dict[str, dict[str, object]]:
    groups: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        if norm(row.get("year")) != str(DRAW_YEAR):
            continue
        if not source_in_scope(row.get("source_file")):
            continue
        code = norm(row.get("hunt_code"))
        if code:
            groups[code].append(row)

    out: dict[str, dict[str, object]] = {}
    for code, group in groups.items():
        by_residency: dict[str, float] = defaultdict(float)
        for row in group:
            value = number(row.get("total_drawn"))
            if value is None:
                value = number(row.get("total_permits"))
            if value is None:
                continue
            by_residency[norm(row.get("residency")) or "UNSPECIFIED"] += value

        if by_residency:
            residency_keys = {key.lower() for key in by_residency}
            if {"resident", "nonresident"} & residency_keys:
                total = sum(value for key, value in by_residency.items() if key.lower() in {"resident", "nonresident"})
            else:
                total = sum(by_residency.values())
        else:
            total = None

        source_files = sorted({norm(row.get("source_file")) for row in group if norm(row.get("source_file"))})
        out[code] = {
            "hunt_code": code,
            "species": most_common(row.get("species") for row in group),
            "sex_type": most_common(row.get("sex_type") for row in group),
            "hunt_name": most_common(row.get("hunt_name") for row in group),
            "hunt_type": most_common(row.get("hunt_type") for row in group),
            "weapon": most_common(row.get("weapon") for row in group),
            "draw_total_permits": total,
            "draw_source_files": source_files,
            "draw_row_count": len(group),
            "draw_residencies": sorted({norm(row.get("residency")) for row in group if norm(row.get("residency"))}),
            "draw_points_count": len({norm(row.get("points")) for row in group if norm(row.get("points"))}),
            "match_species_family": species_family(most_common(row.get("species") for row in group) or most_common(row.get("hunt_name") for row in group)),
            "match_hunt_type_family": hunt_type_family(most_common(row.get("hunt_type") for row in group) or most_common(row.get("hunt_name") for row in group)),
            "match_weapon_family": weapon_family(most_common(row.get("weapon") for row in group) or most_common(row.get("hunt_name") for row in group)),
            "match_tokens": comparable_tokens(most_common(row.get("hunt_name") for row in group)),
        }
    return out


def match_expo_to_draw(
    expo_rows: list[dict[str, object]],
    draw: dict[str, dict[str, object]],
    harvest: dict[str, dict[str, object]],
) -> tuple[dict[str, dict[str, object]], list[dict[str, object]]]:
    matched: dict[str, dict[str, object]] = defaultdict(lambda: {"permits": 0, "rows": [], "status": "NO_EXPO_MATCH"})
    decisions: list[dict[str, object]] = []
    for expo in expo_rows:
        expo_species = species_family(expo.get("species"))
        expo_tokens = comparable_tokens(expo.get("area"))
        expo_weapon = weapon_family(expo.get("weapon"))
        expo_type = hunt_type_family(expo.get("hunt_type"))
        candidates: list[tuple[int, str, dict[str, object], set[str], str]] = []
        for code, draw_item in draw.items():
            if draw_item.get("match_species_family") != expo_species:
                continue
            draw_tokens = draw_item.get("match_tokens")
            if not isinstance(draw_tokens, set):
                draw_tokens = set()
            overlap = expo_tokens & draw_tokens
            if not overlap:
                continue
            score = len(overlap) * 4
            if expo_weapon and expo_weapon == draw_item.get("match_weapon_family"):
                score += 3
            if expo_type and expo_type == draw_item.get("match_hunt_type_family"):
                score += 2
            gap_status = ""
            harvest_item = harvest.get(code, {})
            draw_total = draw_item.get("draw_total_permits")
            harvest_total = harvest_item.get("harvest_permits")
            if isinstance(draw_total, float) and isinstance(harvest_total, float):
                gap = harvest_total - draw_total
                if abs(gap - float(expo.get("permits") or 0)) < 0.0001:
                    gap_status = "EXPO_PERMITS_EQUAL_HARVEST_MINUS_DRAW_GAP"
                    score += 10
            candidates.append((score, code, draw_item, overlap, gap_status))
        candidates.sort(key=lambda item: (item[0], item[1]), reverse=True)
        if not candidates:
            decisions.append(
                {
                    "raw_heading": expo.get("raw_heading", ""),
                    "matched_hunt_code": "",
                    "matched_hunt_name": "",
                    "permits": expo.get("permits", 0),
                    "match_status": "NO_TOKEN_MATCH",
                    "score": 0,
                    "matched_tokens": "",
                }
            )
            continue
        gap_candidates = [candidate for candidate in candidates if candidate[4] == "EXPO_PERMITS_EQUAL_HARVEST_MINUS_DRAW_GAP"]
        if len(gap_candidates) == 1:
            top = gap_candidates[0]
            match_status = "EXPO_MATCHED_BY_EXACT_GAP_AND_SPECIES_UNIT"
        else:
            top = candidates[0]
            second_score = candidates[1][0] if len(candidates) > 1 else -1
            match_status = "EXPO_MATCHED_BY_SPECIES_UNIT_TOKENS"
        if top[0] < 4 or (match_status != "EXPO_MATCHED_BY_EXACT_GAP_AND_SPECIES_UNIT" and top[0] == second_score):
            decisions.append(
                {
                    "raw_heading": expo.get("raw_heading", ""),
                    "matched_hunt_code": top[1],
                    "matched_hunt_name": top[2].get("hunt_name", ""),
                    "permits": expo.get("permits", 0),
                    "match_status": "AMBIGUOUS_EXPO_MATCH",
                    "score": top[0],
                    "matched_tokens": "|".join(sorted(top[3])),
                }
            )
            continue
        code = top[1]
        matched[code]["permits"] += int(expo.get("permits") or 0)
        matched[code]["status"] = match_status
        matched[code]["rows"].append(expo)
        decisions.append(
            {
                "raw_heading": expo.get("raw_heading", ""),
                "matched_hunt_code": code,
                "matched_hunt_name": top[2].get("hunt_name", ""),
                "permits": expo.get("permits", 0),
                "match_status": matched[code]["status"],
                "score": top[0],
                "matched_tokens": "|".join(sorted(top[3])),
            }
        )
    return matched, decisions


def aggregate_harvest(rows: Iterable[dict[str, str]]) -> dict[str, dict[str, object]]:
    groups: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        if norm(row.get("reported_hunt_year")) != str(HARVEST_YEAR):
            continue
        if norm(row.get("source_file")) != HARVEST_SOURCE:
            continue
        code = norm(row.get("hunt_code"))
        if code:
            groups[code].append(row)

    out: dict[str, dict[str, object]] = {}
    for code, group in groups.items():
        values = [number(row.get("permits")) for row in group]
        values = [value for value in values if value is not None]
        unique_values = sorted({format_number(value) for value in values}, key=lambda item: float(item))
        out[code] = {
            "hunt_code": code,
            "species": most_common(row.get("species") for row in group),
            "sex_type": most_common(row.get("sex_type") for row in group),
            "hunt_name": most_common(row.get("hunt_name") for row in group),
            "hunt_type": most_common(row.get("hunt_type") for row in group),
            "weapon": most_common(row.get("weapon") for row in group),
            "harvest_permits": max(values) if values else None,
            "harvest_unique_permit_values": unique_values,
            "harvest_source_files": sorted({norm(row.get("source_file")) for row in group if norm(row.get("source_file"))}),
            "harvest_row_count": len(group),
        }
    return out


def comparison_status(draw_value: float | None, harvest_value: float | None) -> str:
    if draw_value is None and harvest_value is None:
        return "NO_TOTAL_PERMIT_VALUES"
    if draw_value is None:
        return "HARVEST_ONLY_TOTAL"
    if harvest_value is None:
        return "DRAW_ONLY_TOTAL"
    if abs(draw_value - harvest_value) < 0.0001:
        return "TOTAL_PERMIT_MATCH"
    if harvest_value > draw_value:
        return "HARVEST_GREATER_THAN_DRAW"
    return "DRAW_GREATER_THAN_HARVEST"


def status_from_values(source_value: float | None, harvest_value: float | None) -> str:
    if source_value is None and harvest_value is None:
        return "NO_TOTAL_PERMIT_VALUES"
    if source_value is None:
        return "HARVEST_ONLY_TOTAL"
    if harvest_value is None:
        return "SOURCE_ONLY_TOTAL"
    if abs(source_value - harvest_value) < 0.0001:
        return "TOTAL_PERMIT_MATCH"
    if harvest_value > source_value:
        return "HARVEST_GREATER_THAN_SOURCE"
    return "SOURCE_GREATER_THAN_HARVEST"


def build_audit(paths: Paths) -> tuple[dict[str, object], list[dict[str, object]]]:
    _, draw_rows = read_csv(paths.draw)
    _, harvest_rows = read_csv(paths.harvest)
    _, database_rows = read_csv(paths.database)

    draw = aggregate_draw(draw_rows)
    harvest = aggregate_harvest(harvest_rows)
    expo_rows = load_expo_rows(paths.expo_xlsx)
    expo_by_code, expo_decisions = match_expo_to_draw(expo_rows, draw, harvest)
    database_codes = {norm(row.get("hunt_code")) for row in database_rows if norm(row.get("hunt_code"))}

    all_codes = sorted(set(draw) | set(harvest))
    rows: list[dict[str, object]] = []
    counts = Counter()

    for code in all_codes:
        d = draw.get(code, {})
        h = harvest.get(code, {})
        draw_total = d.get("draw_total_permits")
        harvest_total = h.get("harvest_permits")
        status = comparison_status(draw_total if isinstance(draw_total, float) else None, harvest_total if isinstance(harvest_total, float) else None)
        expo_item = expo_by_code.get(code, {})
        expo_total = float(expo_item.get("permits", 0) or 0)
        draw_plus_expo = draw_total + expo_total if isinstance(draw_total, float) else None
        draw_plus_expo_status = status_from_values(draw_plus_expo, harvest_total if isinstance(harvest_total, float) else None)
        delta = None
        if isinstance(draw_total, float) and isinstance(harvest_total, float):
            delta = harvest_total - draw_total
        delta_plus_expo = None
        if isinstance(draw_plus_expo, float) and isinstance(harvest_total, float):
            delta_plus_expo = harvest_total - draw_plus_expo
        notes = ""
        if status == "TOTAL_PERMIT_MATCH":
            notes = "Exact hunt-code total match between normalized draw result and preliminary harvest permit count."
        elif status == "HARVEST_GREATER_THAN_DRAW":
            notes = "Harvest permit total is higher; likely needs overlay/source-channel review such as expo, conservation, CWMU, or other field-issued permits."
        elif status == "DRAW_GREATER_THAN_HARVEST":
            notes = "Draw-result total is higher; review extraction grain, hunt status, or whether harvest report filtered the permit universe."
        elif status == "HARVEST_ONLY_TOTAL":
            notes = "Harvest source has permit total but selected 2025 draw-result source scope has no matching hunt-code total."
        elif status == "DRAW_ONLY_TOTAL":
            notes = "Selected 2025 draw-result source scope has permit total but preliminary harvest source has no matching hunt-code total."

        row = {
            "hunt_code": code,
            "species": d.get("species") or h.get("species") or "",
            "sex_type": d.get("sex_type") or h.get("sex_type") or "",
            "hunt_name": d.get("hunt_name") or h.get("hunt_name") or "",
            "hunt_type": d.get("hunt_type") or h.get("hunt_type") or "",
            "weapon": d.get("weapon") or h.get("weapon") or "",
            "source_scope": source_scope(d.get("draw_source_files", [])),
            "comparison_status": status,
            "draw_total_permits": format_number(draw_total if isinstance(draw_total, float) else None),
            "expo_permits_matched": format_number(expo_total),
            "draw_plus_expo_total": format_number(draw_plus_expo),
            "harvest_permits": format_number(harvest_total if isinstance(harvest_total, float) else None),
            "permit_delta_harvest_minus_draw": format_number(delta),
            "permit_delta_harvest_minus_draw_plus_expo": format_number(delta_plus_expo),
            "draw_plus_expo_status": draw_plus_expo_status,
            "expo_match_status": expo_item.get("status", "NO_EXPO_MATCH"),
            "expo_match_sources": "|".join(norm(item.get("raw_heading")) for item in expo_item.get("rows", []) if norm(item.get("raw_heading"))),
            "draw_source_files": "|".join(d.get("draw_source_files", [])),
            "harvest_source_files": "|".join(h.get("harvest_source_files", [])),
            "draw_row_count": d.get("draw_row_count", 0),
            "harvest_row_count": h.get("harvest_row_count", 0),
            "draw_residencies": "|".join(d.get("draw_residencies", [])),
            "draw_points_count": d.get("draw_points_count", 0),
            "harvest_unique_permit_values": "|".join(h.get("harvest_unique_permit_values", [])),
            "in_2026_database": str(code in database_codes).upper(),
            "same_total_match_ready": str(status == "TOTAL_PERMIT_MATCH").upper(),
            "notes": notes,
        }
        rows.append(row)
        counts[status] += 1

    draw_plus_expo_counts = Counter(row["draw_plus_expo_status"] for row in rows)
    expo_decision_counts = Counter(row["match_status"] for row in expo_decisions)

    draw_source_counts = Counter()
    for row in draw_rows:
        if norm(row.get("year")) == str(DRAW_YEAR) and source_in_scope(row.get("source_file")):
            draw_source_counts[norm(row.get("source_file"))] += 1

    listed_pdf_status = [
        {
            "path": pdf,
            "exists": (paths.root / pdf).exists(),
            "size_bytes": (paths.root / pdf).stat().st_size if (paths.root / pdf).exists() else 0,
        }
        for pdf in LISTED_PDFS
    ]
    external_expo_pdf_matches = find_external_expo_pdf_matches(paths.external_expo_pdf)

    summary = {
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "scope": "2025 selected draw-source total permits compared to 2025 preliminary big-game harvest permit totals by exact hunt_code.",
        "draw_year": DRAW_YEAR,
        "harvest_reported_hunt_year": HARVEST_YEAR,
        "draw_file": str(paths.draw),
        "harvest_file": str(paths.harvest),
        "database_file": str(paths.database),
        "expo_xlsx": str(paths.expo_xlsx),
        "external_expo_pdf_reference": str(paths.external_expo_pdf),
        "external_expo_pdf_exists": bool(external_expo_pdf_matches),
        "external_expo_pdf_matches": external_expo_pdf_matches,
        "harvest_source_filter": HARVEST_SOURCE,
        "draw_source_hints": SOURCE_HINTS,
        "listed_pdf_status": listed_pdf_status,
        "draw_input_rows": len(draw_rows),
        "harvest_input_rows": len(harvest_rows),
        "selected_draw_hunt_codes": len(draw),
        "selected_harvest_hunt_codes": len(harvest),
        "comparison_hunt_codes": len(rows),
        "comparison_status_counts": dict(sorted(counts.items())),
        "draw_plus_expo_status_counts": dict(sorted(draw_plus_expo_counts.items())),
        "expo_rows_loaded": len(expo_rows),
        "expo_decision_counts": dict(sorted(expo_decision_counts.items())),
        "expo_matched_hunt_codes": sum(1 for value in expo_by_code.values() if value.get("permits")),
        "selected_draw_source_row_counts": dict(sorted(draw_source_counts.items())),
        "match_rate_among_codes_with_both_totals": "",
        "interpretation": [
            "TOTAL_PERMIT_MATCH rows are clean same-code total permit matches.",
            "HARVEST_GREATER_THAN_DRAW rows are the likely places where field permits include overlays beyond public draw-result permits.",
            "DRAW_GREATER_THAN_HARVEST rows need extraction-grain or source-scope review before promotion.",
            "This audit intentionally does not alter draw truth, harvest truth, DATABASE.csv, or runtime files.",
        ],
    }

    both = counts["TOTAL_PERMIT_MATCH"] + counts["HARVEST_GREATER_THAN_DRAW"] + counts["DRAW_GREATER_THAN_HARVEST"]
    if both:
        summary["match_rate_among_codes_with_both_totals"] = f"{counts['TOTAL_PERMIT_MATCH']}/{both} ({counts['TOTAL_PERMIT_MATCH'] / both:.1%})"
        summary["draw_plus_expo_match_rate_among_codes_with_both_totals"] = (
            f"{draw_plus_expo_counts['TOTAL_PERMIT_MATCH']}/{both} ({draw_plus_expo_counts['TOTAL_PERMIT_MATCH'] / both:.1%})"
        )
    summary["harvest_greater_rows_resolved_by_expo"] = sum(
        1
        for row in rows
        if row["comparison_status"] == "HARVEST_GREATER_THAN_DRAW" and row["draw_plus_expo_status"] == "TOTAL_PERMIT_MATCH"
    )

    return summary, rows, expo_decisions


def write_markdown(path: Path, summary: dict[str, object], rows: list[dict[str, object]]) -> None:
    counts = summary["comparison_status_counts"]
    source_counts = summary["selected_draw_source_row_counts"]
    top_gaps = [
        row
        for row in rows
        if row["comparison_status"] in {"HARVEST_GREATER_THAN_DRAW", "DRAW_GREATER_THAN_HARVEST"}
    ][:25]
    resolved_by_expo = [
        row
        for row in rows
        if row["comparison_status"] == "HARVEST_GREATER_THAN_DRAW" and row["draw_plus_expo_status"] == "TOTAL_PERMIT_MATCH"
    ][:40]
    lines = [
        "# 2025 Total Permit Cross-Compare",
        "",
        "Read-only audit comparing selected 2025 draw-result total permits against the 2025 preliminary big-game harvest permit counts by exact `hunt_code`.",
        "",
        "## Summary",
        "",
        f"- Draw year: `{summary['draw_year']}`.",
        f"- Harvest reported hunt year: `{summary['harvest_reported_hunt_year']}`.",
        f"- Selected draw hunt codes: `{summary['selected_draw_hunt_codes']}`.",
        f"- Selected harvest hunt codes: `{summary['selected_harvest_hunt_codes']}`.",
        f"- Compared hunt codes: `{summary['comparison_hunt_codes']}`.",
        f"- Expo rows loaded: `{summary['expo_rows_loaded']}`.",
        f"- Expo matched hunt codes: `{summary['expo_matched_hunt_codes']}`.",
        f"- Draw-only match rate where both sides have totals: `{summary['match_rate_among_codes_with_both_totals']}`.",
        f"- Draw plus Expo match rate where both sides have totals: `{summary['draw_plus_expo_match_rate_among_codes_with_both_totals']}`.",
        f"- Harvest-greater-than-draw rows resolved exactly by Expo: `{summary['harvest_greater_rows_resolved_by_expo']}`.",
        "",
        "## Status Counts",
        "",
        "| Status | Rows |",
        "| --- | ---: |",
    ]
    for status, count in counts.items():
        lines.append(f"| {status} | {count} |")
    lines.extend(["", "## Draw Plus Expo Status Counts", "", "| Status | Rows |", "| --- | ---: |"])
    for status, count in summary["draw_plus_expo_status_counts"].items():
        lines.append(f"| {status} | {count} |")
    lines.extend(["", "## Expo Match Decisions", "", "| Decision | Rows |", "| --- | ---: |"])
    for status, count in summary["expo_decision_counts"].items():
        lines.append(f"| {status} | {count} |")
    lines.extend(["", "## Selected Draw Source Rows", "", "| Source file | Rows |", "| --- | ---: |"])
    for source, count in source_counts.items():
        lines.append(f"| {source} | {count} |")
    lines.extend(["", "## What It Takes To Get Matches", ""])
    lines.append("- Exact same `hunt_code` plus summed 2025 draw totals already gives the clean matches.")
    lines.append("- Adding Expo permits can explain additional harvest-field permit totals when the Expo unit heading maps cleanly to the same draw hunt code.")
    lines.append("- Rows where harvest is higher than draw need an overlay channel test before they should be called conflicts; likely candidates are expo, conservation, CWMU, landowner, or other field-issued permits.")
    lines.append("- Rows where draw is higher than harvest should not be filled automatically; those need source-scope/extraction-grain review.")
    lines.append("- This pass only tests totals. It does not prove source authority for overwrites.")
    lines.extend(["", "## Rows Resolved By Expo", "", "| Hunt code | Hunt name | Species | Draw total | Expo | Draw + Expo | Harvest permits | Expo source |", "| --- | --- | --- | ---: | ---: | ---: | ---: | --- |"])
    for row in resolved_by_expo:
        lines.append(
            f"| {row['hunt_code']} | {row['hunt_name']} | {row['species']} | {row['draw_total_permits']} | {row['expo_permits_matched']} | {row['draw_plus_expo_total']} | {row['harvest_permits']} | {row['expo_match_sources']} |"
        )
    lines.extend(["", "## Top Total Gaps", "", "| Hunt code | Hunt name | Species | Draw total | Expo | Draw + Expo | Harvest permits | Delta after Expo | Status |", "| --- | --- | --- | ---: | ---: | ---: | ---: | ---: | --- |"])
    for row in top_gaps:
        lines.append(
            f"| {row['hunt_code']} | {row['hunt_name']} | {row['species']} | {row['draw_total_permits']} | {row['expo_permits_matched']} | {row['draw_plus_expo_total']} | {row['harvest_permits']} | {row['permit_delta_harvest_minus_draw_plus_expo']} | {row['draw_plus_expo_status']} |"
        )
    lines.extend(["", "## Guardrails", ""])
    lines.append("- `DATABASE.csv` was read only for current-code membership checks.")
    lines.append("- Raw PDFs were not edited.")
    lines.append("- Normalized draw and harvest truth tables were not edited.")
    lines.append("- Runtime manifests and website files were not edited.")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--root", default=".", help="Repository root.")
    parser.add_argument("--draw", default=DEFAULT_DRAW, help="Normalized draw results CSV.")
    parser.add_argument("--harvest", default=DEFAULT_HARVEST, help="Normalized harvest results CSV.")
    parser.add_argument("--database", default=DEFAULT_DATABASE, help="Current DATABASE.csv for membership checks only.")
    parser.add_argument("--expo-xlsx", default=DEFAULT_EXPO_XLSX, help="Structured 2025 Expo permit workbook.")
    parser.add_argument("--external-expo-pdf", default=DEFAULT_EXTERNAL_EXPO_PDF, help="Optional external Expo PDF reference path; checked for presence only.")
    parser.add_argument("--out-dir", default=DEFAULT_OUT_DIR, help="Audit output directory.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    root = Path(args.root).resolve()
    paths = Paths(
        root=root,
        draw=(root / args.draw).resolve(),
        harvest=(root / args.harvest).resolve(),
        database=(root / args.database).resolve(),
        expo_xlsx=(root / args.expo_xlsx).resolve(),
        external_expo_pdf=Path(args.external_expo_pdf),
        out_dir=(root / args.out_dir).resolve(),
    )
    for required in [paths.draw, paths.harvest, paths.database]:
        if not required.exists():
            raise FileNotFoundError(required)

    summary, rows, expo_decisions = build_audit(paths)
    base = paths.out_dir / "total_permit_cross_compare_2025"
    write_csv(base.with_suffix(".csv"), rows, REPORT_COLUMNS)
    write_csv(paths.out_dir / "total_permit_cross_compare_2025_expo_match_decisions.csv", expo_decisions, ["raw_heading", "matched_hunt_code", "matched_hunt_name", "permits", "match_status", "score", "matched_tokens"])
    write_json(base.with_suffix(".json"), {"summary": summary, "rows": rows, "expo_match_decisions": expo_decisions})
    write_markdown(base.with_suffix(".md"), summary, rows)
    print(json.dumps(summary, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
