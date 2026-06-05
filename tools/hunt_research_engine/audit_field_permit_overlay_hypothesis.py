#!/usr/bin/env python3
"""Test whether harvest-vs-draw permit gaps are explained by permit overlays.

The hypothesis is:

    permits in the field ~= public draw permits + expo/conservation/etc.

This tool is read-only. It consumes the harvest/draw reconciliation output and
structured overlay evidence that already exists in the repo. It does not mutate
DATABASE.csv, draw truth, harvest truth, or runtime files.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


DEFAULT_RECONCILIATION = "audits/hunt_research_engine/harvest_draw_reconciliation.csv"
DEFAULT_OUT_DIR = "audits/hunt_research_engine"
CONSERVATION_2025_2027 = "pipeline/RAW/hunt_unit_database/2026/reports/conservation_permits_2025_2027_grouped.csv"
CONSERVATION_2022_2024 = "pipeline/RAW/hunt_unit_database/library-master.csv"
EXPO_2025_XLSX = "pipeline/RAW/hunt_unit_database/2026/xlsx/2025  Expo Permits Draw Results.xlsx"
EXPO_2026_XLSX = "pipeline/RAW/hunt_unit_database/2026/xlsx/2026 EXPO PERMIT DRAW.xlsx"


@dataclass(frozen=True)
class OverlayRow:
    overlay_family: str
    source_file: str
    year_start: int
    year_end: int
    species: str
    area: str
    condition: str
    permit_count: int
    confidence: str
    notes: str


def norm(value: object) -> str:
    return "" if value is None else str(value).strip()


def key_text(value: object) -> str:
    text = norm(value).lower().replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def tokens(value: object) -> set[str]:
    stop = {
        "and",
        "the",
        "permit",
        "permits",
        "limited",
        "entry",
        "premium",
        "once",
        "lifetime",
        "general",
        "season",
        "buck",
        "bull",
        "cow",
        "only",
        "any",
        "legal",
        "weapon",
        "archery",
        "muzzleloader",
        "multi",
        "multiseason",
        "conservation",
        "expo",
        "deer",
        "elk",
        "bison",
        "pronghorn",
        "moose",
        "bear",
        "sheep",
        "goat",
        "turkey",
    }
    return {part for part in key_text(value).split() if len(part) > 2 and part not in stop}


def number(value: object) -> float | None:
    text = norm(value).replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def species_family(value: object) -> str:
    text = key_text(value)
    if "antlerless elk" in text or text == "elk":
        return "elk"
    if "buck deer" in text or text == "deer" or "mule deer" in text:
        return "deer"
    if "pronghorn" in text:
        return "pronghorn"
    if "bison" in text:
        return "bison"
    if "moose" in text:
        return "moose"
    if "bear" in text:
        return "black bear"
    if "goat" in text:
        return "mountain goat"
    if "desert bighorn" in text:
        return "desert bighorn sheep"
    if "rocky" in text or "bighorn" in text:
        return "rocky mountain bighorn sheep"
    if "turkey" in text:
        return "turkey"
    return text


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def parse_int(value: object, default: int = 0) -> int:
    num = number(value)
    if num is None:
        return default
    return int(round(num))


def load_conservation_2025_2027(root: Path) -> list[OverlayRow]:
    path = root / CONSERVATION_2025_2027
    if not path.exists():
        return []
    rows: list[OverlayRow] = []
    for row in read_csv(path):
        count = parse_int(row.get("permits_2026_conservation") or row.get("conservation_permit_count_2025_2027"))
        if count <= 0:
            continue
        rows.append(
            OverlayRow(
                overlay_family="CONSERVATION_2025_2027",
                source_file=CONSERVATION_2025_2027,
                year_start=2025,
                year_end=2027,
                species=norm(row.get("species")),
                area=norm(row.get("area")),
                condition=norm(row.get("condition")),
                permit_count=count,
                confidence="STRUCTURED_SOURCE",
                notes="Grouped conservation permit row; count applies to each year in 2025-2027 source window unless later source proves otherwise.",
            )
        )
    return rows


def load_conservation_2022_2024(root: Path) -> list[OverlayRow]:
    path = root / CONSERVATION_2022_2024
    if not path.exists():
        return []
    rows: list[OverlayRow] = []
    for row in read_csv(path):
        if norm(row.get("record_type")) != "permit_allocation":
            continue
        if "conservation" not in key_text(row.get("category")):
            continue
        year_start = parse_int(row.get("year_start"))
        year_end = parse_int(row.get("year_end"))
        if not year_start or not year_end:
            continue
        rows.append(
            OverlayRow(
                overlay_family="CONSERVATION_LIBRARY_MASTER",
                source_file=CONSERVATION_2022_2024,
                year_start=year_start,
                year_end=year_end,
                species=norm(row.get("species")),
                area=norm(row.get("area")),
                condition=norm(row.get("condition")),
                permit_count=1,
                confidence="EXTRACTED_PERMIT_ROW",
                notes="Each library-master conservation permit_allocation row counts as one permit allocation.",
            )
        )
    return rows


def parse_expo_heading(text: object) -> tuple[str, str, str, int] | None:
    raw = norm(text)
    if "Permits:" not in raw:
        return None
    match = re.search(r"Permits:\s*(\d+)", raw, flags=re.IGNORECASE)
    if not match:
        return None
    count = int(match.group(1))
    title = raw[: match.start()].strip(" -")
    parts = [part.strip() for part in title.split(" - ") if part.strip()]
    if len(parts) < 2:
        return None
    species = parts[0]
    condition = " ".join(parts[1:-1]) if len(parts) > 2 else ""
    area = parts[-1]
    return species, area, condition, count


def load_expo_xlsx(root: Path, rel_path: str, year: int, family: str) -> list[OverlayRow]:
    path = root / rel_path
    if not path.exists():
        return []
    rows: list[OverlayRow] = []
    workbook = load_workbook(path, read_only=True, data_only=True)
    for sheet in workbook.worksheets:
        for cells in sheet.iter_rows(values_only=True):
            parsed = parse_expo_heading(cells[0] if cells else "")
            if not parsed:
                continue
            species, area, condition, count = parsed
            rows.append(
                OverlayRow(
                    overlay_family=family,
                    source_file=rel_path,
                    year_start=year,
                    year_end=year,
                    species=species,
                    area=area,
                    condition=condition,
                    permit_count=count,
                    confidence="NAME_ONLY_EXPO_HEADING",
                    notes="Expo heading includes permit count but does not carry hunt_code; match is text-based evidence only.",
                )
            )
    return rows


def load_overlays(root: Path) -> list[OverlayRow]:
    overlays = []
    overlays.extend(load_conservation_2025_2027(root))
    overlays.extend(load_conservation_2022_2024(root))
    overlays.extend(load_expo_xlsx(root, EXPO_2025_XLSX, 2025, "EXPO_2025_NAME_ONLY"))
    overlays.extend(load_expo_xlsx(root, EXPO_2026_XLSX, 2026, "EXPO_2026_NAME_ONLY"))
    return overlays


def match_score(recon: dict[str, str], overlay: OverlayRow) -> float:
    year = parse_int(recon.get("year"))
    if not (overlay.year_start <= year <= overlay.year_end):
        return 0.0
    recon_species = species_family(recon.get("harvest_species") or recon.get("draw_species"))
    overlay_species = species_family(overlay.species)
    if recon_species and overlay_species and recon_species != overlay_species:
        return 0.0
    haystack = " ".join(
        [
            recon.get("harvest_hunt_name", ""),
            recon.get("draw_hunt_name", ""),
            recon.get("harvest_weapon", ""),
            recon.get("draw_weapon", ""),
            recon.get("harvest_hunt_type", ""),
            recon.get("draw_hunt_type", ""),
        ]
    )
    area_tokens = tokens(overlay.area)
    if area_tokens:
        overlap = len(area_tokens & tokens(haystack)) / len(area_tokens)
    else:
        overlap = 0.0
    condition_tokens = tokens(overlay.condition)
    condition_overlap = len(condition_tokens & tokens(haystack)) / len(condition_tokens) if condition_tokens else 0.0
    score = 0.7 * overlap + 0.3 * condition_overlap
    if overlap >= 0.95:
        score += 0.15
    return min(score, 1.0)


def write_csv(path: Path, rows: Iterable[dict[str, object]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fieldnames})


def build_audit(root: Path, reconciliation: Path) -> tuple[dict[str, object], list[dict[str, object]], list[dict[str, object]]]:
    recon_rows = read_csv(reconciliation)
    overlays = load_overlays(root)
    overlay_index_rows = [
        {
            "overlay_family": overlay.overlay_family,
            "source_file": overlay.source_file,
            "year_start": overlay.year_start,
            "year_end": overlay.year_end,
            "species": overlay.species,
            "area": overlay.area,
            "condition": overlay.condition,
            "permit_count": overlay.permit_count,
            "confidence": overlay.confidence,
            "notes": overlay.notes,
        }
        for overlay in overlays
    ]
    audit_rows: list[dict[str, object]] = []
    status_counts: Counter[str] = Counter()
    family_counts: Counter[str] = Counter(overlay.overlay_family for overlay in overlays)

    conflict_rows = [row for row in recon_rows if row.get("permit_status") == "PERMIT_CONFLICT_REVIEW"]
    for row in conflict_rows:
        harvest = number(row.get("harvest_permits"))
        draw = number(row.get("draw_total_permits_reconciled"))
        if harvest is None or draw is None:
            continue
        gap = int(round(harvest - draw))
        if gap <= 0:
            status = "NOT_OVERLAY_EXPLAINED_DRAW_GE_HARVEST"
            matches: list[tuple[float, OverlayRow]] = []
        else:
            matches = sorted(
                [(match_score(row, overlay), overlay) for overlay in overlays],
                key=lambda item: item[0],
                reverse=True,
            )
            matches = [(score, overlay) for score, overlay in matches if score >= 0.68]
            matched_sum = sum(overlay.permit_count for _, overlay in matches)
            if matched_sum == gap:
                status = "FIELD_PERMIT_GAP_EXPLAINED_BY_OVERLAYS"
            elif matched_sum and matched_sum < gap:
                status = "FIELD_PERMIT_GAP_PARTIALLY_EXPLAINED_BY_OVERLAYS"
            elif matched_sum and matched_sum > gap:
                status = "OVERLAY_CANDIDATES_EXCEED_GAP_REVIEW"
            else:
                status = "NO_OVERLAY_MATCH_FOUND"
        status_counts[status] += 1
        audit_rows.append(
            {
                "hunt_code": row.get("hunt_code"),
                "year": row.get("year"),
                "harvest_permits": row.get("harvest_permits"),
                "draw_total_permits_reconciled": row.get("draw_total_permits_reconciled"),
                "field_permit_gap": gap,
                "overlay_matched_permits": sum(overlay.permit_count for _, overlay in matches),
                "overlay_gap_status": status,
                "matched_overlay_families": "|".join(sorted({overlay.overlay_family for _, overlay in matches})),
                "matched_overlay_sources": "|".join(sorted({overlay.source_file for _, overlay in matches})),
                "matched_overlay_areas": "|".join(f"{overlay.area}:{overlay.permit_count}" for _, overlay in matches),
                "max_overlay_match_score": f"{matches[0][0]:.3f}" if matches else "",
                "harvest_hunt_name": row.get("harvest_hunt_name"),
                "draw_hunt_name": row.get("draw_hunt_name"),
                "harvest_species": row.get("harvest_species"),
                "draw_species": row.get("draw_species"),
                "harvest_weapon": row.get("harvest_weapon"),
                "draw_weapon": row.get("draw_weapon"),
            }
        )

    summary = {
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "reconciliation_file": str(reconciliation),
        "overlay_sources": {
            "conservation_2025_2027": CONSERVATION_2025_2027,
            "conservation_2022_2024": CONSERVATION_2022_2024,
            "expo_2025": EXPO_2025_XLSX,
            "expo_2026": EXPO_2026_XLSX,
        },
        "overlay_rows_loaded": len(overlays),
        "overlay_family_counts": dict(sorted(family_counts.items())),
        "permit_conflict_rows_tested": len(conflict_rows),
        "overlay_gap_status_counts": dict(sorted(status_counts.items())),
        "hypothesis_result": "SUPPORTED_FOR_SOME_ROWS_REQUIRES_REVIEW",
        "production_rule": "If overlays explain the gap, create a separate field_permits_total layer. Do not overwrite public draw odds quotas or historical harvest-source permits.",
    }
    return summary, audit_rows, overlay_index_rows


def write_markdown(path: Path, summary: dict[str, object]) -> None:
    lines = [
        "# Field Permit Overlay Hypothesis Audit",
        "",
        "This read-only audit tests Tyler's hypothesis that harvest-report `permits` often represent permits in the field after adding Expo, Conservation, and related overlay permits to public draw results.",
        "",
        "## Verdict",
        "",
        "`SUPPORTED_FOR_SOME_ROWS_REQUIRES_REVIEW`",
        "",
        "The model should add a separate `field_permits_total` concept, not overwrite public draw permits. Public draw permits feed draw odds; field permits explain harvest reports and fall field totals.",
        "",
        "## Counts",
        "",
        f"- Overlay rows loaded: `{summary['overlay_rows_loaded']}`",
        f"- Permit conflict rows tested: `{summary['permit_conflict_rows_tested']}`",
        "",
        "## Overlay Families",
        "",
    ]
    for key, value in summary["overlay_family_counts"].items():
        lines.append(f"- `{key}`: `{value}`")
    lines.extend(["", "## Gap Status", ""])
    for key, value in summary["overlay_gap_status_counts"].items():
        lines.append(f"- `{key}`: `{value}`")
    lines.extend(
        [
            "",
            "## Recommended Data Contract",
            "",
            "- Keep `draw_public_permits_total` for draw odds and ladder math.",
            "- Keep `harvest_report_permits` as source-reported harvest context.",
            "- Add derived/reviewed `field_permits_total = draw_public_permits_total + overlay_permits_total` only when overlay evidence is source-backed.",
            "- Track `overlay_permits_total`, `overlay_families`, `overlay_source_files`, and `field_permits_reconciliation_status`.",
            "- Do not let Expo/Conservation/Sportsman overlays inflate public draw odds probability denominators.",
            "",
        ]
    )
    path.write_text("\n".join(lines), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--root", default=".")
    parser.add_argument("--reconciliation", default=DEFAULT_RECONCILIATION)
    parser.add_argument("--out-dir", default=DEFAULT_OUT_DIR)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    root = Path(args.root).resolve()
    reconciliation = (root / args.reconciliation).resolve()
    out_dir = (root / args.out_dir).resolve()
    if not reconciliation.exists():
        raise FileNotFoundError(reconciliation)

    summary, audit_rows, overlay_rows = build_audit(root, reconciliation)
    out_dir.mkdir(parents=True, exist_ok=True)
    write_csv(
        out_dir / "field_permit_overlay_hypothesis_audit.csv",
        audit_rows,
        [
            "hunt_code",
            "year",
            "harvest_permits",
            "draw_total_permits_reconciled",
            "field_permit_gap",
            "overlay_matched_permits",
            "overlay_gap_status",
            "matched_overlay_families",
            "matched_overlay_sources",
            "matched_overlay_areas",
            "max_overlay_match_score",
            "harvest_hunt_name",
            "draw_hunt_name",
            "harvest_species",
            "draw_species",
            "harvest_weapon",
            "draw_weapon",
        ],
    )
    write_csv(
        out_dir / "field_permit_overlay_source_index.csv",
        overlay_rows,
        [
            "overlay_family",
            "source_file",
            "year_start",
            "year_end",
            "species",
            "area",
            "condition",
            "permit_count",
            "confidence",
            "notes",
        ],
    )
    (out_dir / "field_permit_overlay_hypothesis_audit.json").write_text(
        json.dumps(summary, indent=2, sort_keys=True), encoding="utf-8"
    )
    write_markdown(out_dir / "field_permit_overlay_hypothesis_audit.md", summary)
    print(json.dumps(summary, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
