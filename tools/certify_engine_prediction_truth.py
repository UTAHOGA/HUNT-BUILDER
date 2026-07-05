#!/usr/bin/env python3
"""Certify progressive prediction artifacts against locked yearly hunt-code truth.

This tool does not run engines, rewrite truth files, or promote runtime outputs.
It consumes an existing progressive prediction audit directory and writes a
certification package under audits/engine_certified_prediction_truth/<timestamp>/.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Sequence


REPO = Path(__file__).resolve().parents[1]
OUT_ROOT = Path("audits") / "engine_certified_prediction_truth"
PROGRESSIVE_ROOT = Path("audits") / "progressive_prediction_audit"
LOCKED_ROOT = Path("data_truth") / "hunt_code_universe_truth" / "locked"

PROBABILITY_FIELDS = (
    "p_draw",
    "p_draw_mean",
    "p_preference_draw",
    "p_bonus_pool",
    "p_random_pool",
    "p_sportsman_draw",
    "p_availability",
    "p_reserved_mean",
    "p_random_mean",
    "p_max_pool_mean",
    "p_preference_mean",
    "p_youth_mean",
    "cutoff_bucket_probability",
    "guaranteed_probability",
)
PERCENT_FIELDS = (
    "p_draw_pct",
    "p_bonus_pool_pct",
    "p_random_pool_pct",
    "display_odds_pct",
    "availability_pct",
)
HOLDOUT_FAMILIES = {
    "preference_antlerless_deer",
    "preference_antlerless_elk",
    "preference_doe_pronghorn",
}
YEARLY_CANONICAL_WORKBOOKS = Path("outputs") / "yearly_canonical_workbooks"


def clean(value: Any) -> str:
    return str(value if value is not None else "").strip()


def upper(value: Any) -> str:
    return clean(value).upper()


def norm_code(value: Any) -> str:
    return re.sub(r"\s+", "", upper(value))


def norm_residency(value: Any) -> str:
    text = upper(value)
    if text in {"R", "RES", "RESIDENT"}:
        return "RESIDENT"
    if text in {"NR", "NONRES", "NON-RESIDENT", "NONRESIDENT"}:
        return "NONRESIDENT"
    return text


def norm_points(value: Any) -> str:
    text = clean(value)
    if not text:
        return ""
    try:
        number = float(text)
    except ValueError:
        return text
    if math.isfinite(number) and number.is_integer():
        return str(int(number))
    return text


def norm_pool(value: Any) -> str:
    return upper(value) or "STANDARD"


def read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def read_xlsx_sheet(path: Path, sheet_name: str) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - environment guard
        raise RuntimeError("openpyxl is required to read canonical yearly workbooks") from exc

    if not path.exists():
        return []
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            return []
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = [clean(value) for value in next(rows)]
        except StopIteration:
            return []
        if not any(headers):
            return []
        parsed: list[dict[str, str]] = []
        for values in rows:
            row = {headers[index]: clean(value) for index, value in enumerate(values) if index < len(headers)}
            if any(clean(value) for value in row.values()):
                parsed.append(row)
        return parsed
    finally:
        workbook.close()


def write_csv(path: Path, rows: Iterable[dict[str, Any]], fields: Sequence[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(fields), extrasaction="ignore", lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def latest_progressive_dir(repo: Path) -> Path:
    candidates = [path for path in (repo / PROGRESSIVE_ROOT).glob("*") if path.is_dir()]
    if not candidates:
        raise FileNotFoundError(repo / PROGRESSIVE_ROOT)
    return sorted(candidates, key=lambda item: item.name)[-1]


def rel(repo: Path, path: Path) -> str:
    try:
        return str(path.resolve().relative_to(repo.resolve()))
    except ValueError:
        return str(path)


def row_year(row: dict[str, str]) -> str:
    for field in ("target_year", "prediction_year", "forecast_year", "draw_year", "year"):
        value = clean(row.get(field))
        if value:
            return value
    return ""


def row_family(row: dict[str, str], fallback_path: Path | None = None) -> str:
    for field in ("family", "engine_family", "model_strategy", "draw_system_type"):
        value = clean(row.get(field))
        if value:
            return value
    if fallback_path is not None:
        stem = fallback_path.stem
        match = re.match(r"\d{4}_\d{4}_(.+)$", stem)
        if match:
            return match.group(1)
    return ""


def row_draw_system(row: dict[str, str], fallback_family: str = "") -> str:
    for field in ("draw_system_type", "draw_design", "sportsman_draw_design", "model_strategy"):
        value = upper(row.get(field))
        if value:
            return value
    return upper(fallback_family)


def operational_key(row: dict[str, str], family: str, fallback_path: Path) -> tuple[str, str, str, str, str, str, str]:
    return (
        row_year(row),
        norm_code(row.get("hunt_code")),
        row_draw_system(row, family),
        clean(family),
        norm_residency(row.get("residency")),
        norm_points(row.get("points") or row.get("point") or row.get("point_level")),
        norm_pool(row.get("draw_pool")),
    )


def parse_float(value: Any) -> float | None:
    text = clean(value).replace(",", "").replace("%", "")
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    return number if math.isfinite(number) else None


def probability_check(row: dict[str, str]) -> tuple[str, str, str, str]:
    for field in PROBABILITY_FIELDS:
        if field not in row:
            continue
        value = clean(row.get(field))
        if value == "":
            continue
        number = parse_float(value)
        if number is None:
            return field, value, "FAIL", "NON_NUMERIC_PROBABILITY"
        if 0 <= number <= 1:
            return field, value, "PASS", ""
        return field, value, "FAIL", "PROBABILITY_OUT_OF_BOUNDS"
    for field in PERCENT_FIELDS:
        if field not in row:
            continue
        value = clean(row.get(field))
        if value == "":
            continue
        number = parse_float(value)
        if number is None:
            return field, value, "FAIL", "NON_NUMERIC_PERCENT_PROBABILITY"
        if 0 <= number <= 100:
            return field, value, "PASS", ""
        return field, value, "FAIL", "PERCENT_PROBABILITY_OUT_OF_BOUNDS"
    return "", "", "FAIL", "MISSING_PROBABILITY"


def probability_required(row: dict[str, str]) -> bool:
    status_text = " ".join(
        upper(row.get(field))
        for field in (
            "algorithm_status",
            "classification_status",
            "prediction_status",
            "availability_status",
            "rule_status",
            "bear_bonus_note",
            "turkey_bonus_note",
            "reason_codes",
        )
    )
    if any(
        token in status_text
        for token in (
            "EXCLUDED_NOT_PREDICTIVE_DRAW",
            "REFERENCE_ONLY",
            "HARVEST OBJECTIVE",
            "ZERO PERMITS",
            "NOT A PREDICTIVE PUBLIC DRAW-PROBABILITY TARGET",
            "NO_PUBLIC_DRAW_PROBABILITY",
            "MISSING_FORECAST_QUOTA",
            "SOURCE_DATA_INCOMPLETE_NO_PUBLIC_DRAW_PROBABILITY",
        )
    ):
        return False
    return True


def source_years(row: dict[str, str]) -> list[int]:
    values: list[int] = []
    text = ";".join(clean(row.get(field)) for field in ("source_years_used", "source_year", "latest_source_year"))
    for match in re.finditer(r"\b(20\d{2})\b", text):
        values.append(int(match.group(1)))
    return values


def source_year_for_target(row: dict[str, str], target_year: int) -> int:
    value = parse_float(row.get("source_year"))
    if value is not None:
        return int(value)
    return target_year - 1


def prediction_leakage_check(row: dict[str, str], target_year: int) -> tuple[str, str]:
    allowed_source_year = source_year_for_target(row, target_year)
    future_sources = [year for year in source_years(row) if year > allowed_source_year]
    if future_sources:
        return "FAIL", f"source_years_used contains {future_sources} after allowed source year {allowed_source_year}"
    latest = parse_float(row.get("latest_source_year"))
    if latest is not None and int(latest) > allowed_source_year:
        return "FAIL", f"latest_source_year {int(latest)} after allowed source year {allowed_source_year}"
    return "PASS", ""


def load_locked_year(repo: Path, year: int) -> tuple[dict[str, Any], list[dict[str, str]]]:
    year_dir = repo / LOCKED_ROOT / str(year)
    summary = load_json(year_dir / f"LOCKED_{year}_HUNT_CODE_UNIVERSE_SUMMARY.json")
    active_rows = read_csv(year_dir / f"LOCKED_{year}_ACTIVE_YEAR_TRUTH_WITH_BOUNDARY_ID.csv")
    return summary, active_rows


def code_prefix(row: dict[str, str]) -> str:
    value = upper(row.get("prefix"))
    if value:
        return value
    match = re.match(r"^[A-Z]+", norm_code(row.get("hunt_code")))
    return match.group(0) if match else ""


def family_from_truth_row(row: dict[str, str]) -> str:
    prefix = code_prefix(row)
    species = upper(row.get("species"))
    hunt_type = upper(row.get("hunt_type"))
    sex_type = upper(row.get("sex_type"))
    hunt_class = upper(row.get("hunt_class") or row.get("hunt_draw_class"))
    draw_design = upper(row.get("draw_design") or row.get("draw_system_type")).replace("/", "_").replace(" ", "_")
    text = " ".join([prefix, species, hunt_type, sex_type, hunt_class, draw_design, upper(row.get("hunt_name")), upper(row.get("source_file"))])
    if "TURKEY" in text and "YOUTH" in text:
        return "youth_turkey"
    if "TURKEY" in text:
        return "bonus_turkey"
    if prefix == "CG" or "COUGAR" in text or "MOUNTAIN LION" in text:
        return "cougar"
    if "SPORTSMAN" in text:
        return "sportsman"
    if prefix == "BR" or "BLACK BEAR" in text:
        return "bonus_bear"
    if "YOUTH" in text:
        return "youth_draw"
    if "DEDICATED" in text:
        return "dedicated_hunter"
    if prefix in {"DA"} or ("ANTLERLESS" in text and "DEER" in text):
        return "preference_antlerless_deer"
    if prefix in {"EA"} or ("ANTLERLESS" in text and "ELK" in text):
        return "preference_antlerless_elk"
    if prefix in {"PD"} or ("DOE" in text and "PRONGHORN" in text):
        return "preference_doe_pronghorn"
    if prefix in {"BI", "GO", "MA", "MB", "RS", "DS"}:
        return "bonus_oil_big_game"
    if prefix in {"PB"}:
        return "bonus_ple_big_game"
    if prefix in {"EB", "EL", "LO", "LP"}:
        return "bonus_le_big_game"
    if prefix.startswith("DB") and "DEER" in text:
        if draw_design in {"MAX_WEIGHTED_SPLIT", "BONUS_LE_BIG_GAME", "BONUS_CWMU_BIG_GAME"} or "LIMITED ENTRY" in text or "CWMU" in text:
            return "bonus_le_big_game"
        return "preference_general_deer"
    return "UNKNOWN"


def family_from_locked(row: dict[str, str]) -> str:
    return family_from_truth_row(row)


def locked_scorable_rows(active_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    return [
        row
        for row in active_rows
        if clean(row.get("locked_reconciled_bucket")) == "ACTIVE_YEAR_CANONICAL_TRUTH"
        and clean(row.get("scoring_bucket")) == "CANDIDATE_MODEL_SCORABLE_REQUIRES_ENGINE_GATES"
    ]


def canonical_workbook_path(repo: Path, year: int) -> Path:
    matches = sorted((repo / YEARLY_CANONICAL_WORKBOOKS).glob(f"{year}_PERMITS=*_MODEL__CANONICAL_WORKBOOK.xlsx"))
    return matches[0] if matches else Path()


def canonical_rows_for_year(repo: Path, year: int) -> tuple[str, list[dict[str, str]]]:
    csv_path = repo / "outputs" / str(year) / f"{year}_draw_results_canonical.csv"
    if csv_path.exists():
        return rel(repo, csv_path), read_csv(csv_path)

    if year == 2026:
        scorable_path = repo / "outputs" / "2026" / "2026 scorable draw results.csv"
        if scorable_path.exists():
            return rel(repo, scorable_path), read_csv(scorable_path)

    workbook_path = canonical_workbook_path(repo, year)
    if workbook_path.exists():
        rows = read_xlsx_sheet(workbook_path, "RAW_CANONICAL")
        if rows:
            return f"{rel(repo, workbook_path)}#RAW_CANONICAL", rows
        rows = read_xlsx_sheet(workbook_path, "YEARLY_DATABASE")
        if rows:
            return f"{rel(repo, workbook_path)}#YEARLY_DATABASE", rows

    return "", []


def has_number(row: dict[str, str], *fields: str) -> bool:
    return any(parse_float(row.get(field)) is not None for field in fields)


def source_reference_only(row: dict[str, str], family: str) -> tuple[bool, str]:
    text = " ".join(
        upper(row.get(field))
        for field in (
            "record_type",
            "row_type",
            "page_kind",
            "hunt_type",
            "hunt_class",
            "hunt_draw_class",
            "algorithm_status",
            "qa_status",
            "notes",
            "qa_notes",
        )
    )
    if family == "sportsman":
        return True, "SPORTSMAN_RANDOM_ONLY_NOT_DRAW_PROBABILITY_ROW"
    if "CONTACT OPERATOR" in text or "CWMU_CONTACT" in text:
        return True, "CWMU_CONTACT_OPERATOR_REFERENCE_ONLY"
    if "ALLOCATION" in text or "CONSERVATION" in text or "EXPO" in text:
        return True, "ALLOCATION_OR_CONSERVATION_REFERENCE_ONLY"
    if "BONUS POINT" in text and "PURCHASE" in text:
        return True, "BONUS_POINT_PURCHASE_ONLY"
    if "GUARANTEED" in text or "LIFETIME" in text:
        return True, "GUARANTEED_OR_LIFETIME_REFERENCE_ONLY"
    if "REFERENCE_ONLY" in text or "NOT A PREDICTIVE" in text:
        return True, "REFERENCE_ONLY"
    return False, ""


def source_record_type(row: dict[str, str]) -> str:
    return upper(row.get("record_type") or row.get("row_type"))


def row_identity(row: dict[str, str]) -> str:
    return "|".join(
        [
            clean(row.get("source_file") or row.get("draw_source_file") or row.get("source_scope")),
            clean(row.get("pdf_page") or row.get("official_page")),
            clean(row.get("hunt_code")),
            clean(row.get("residency")),
            norm_points(row.get("points")),
        ]
    )


def add_obligation(
    obligations: list[dict[str, Any]],
    row: dict[str, str],
    source_year: int,
    target_year: int,
    source_path: str,
    family: str,
    residency: str,
    probability_field: str,
    probability_value: str,
) -> None:
    obligations.append(
        {
            "source_year": source_year,
            "target_year": target_year,
            "hunt_code": norm_code(row.get("hunt_code")),
            "hunt_name": clean(row.get("hunt_name")),
            "species": clean(row.get("species")),
            "engine_family": family,
            "residency": norm_residency(residency),
            "points": norm_points(row.get("points")),
            "draw_pool": norm_pool(row.get("draw_pool")),
            "source_probability_field": probability_field,
            "source_probability_value": probability_value,
            "source_file": source_path,
            "source_row_identity": row_identity(row),
        }
    )


def canonical_prediction_obligations(
    repo: Path,
    source_year: int,
    target_year: int,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], str]:
    source_path, rows = canonical_rows_for_year(repo, source_year)
    obligations: list[dict[str, Any]] = []
    skipped_rows: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str, str]] = set()

    for row in rows:
        code = norm_code(row.get("hunt_code"))
        if not code:
            continue
        family = family_from_truth_row(row)
        reference_only, reason = source_reference_only(row, family)
        record_type = source_record_type(row)
        if record_type and "POINT" not in record_type:
            reference_only = True
            reason = reason or f"NON_POINT_RECORD_TYPE:{record_type}"
        if reference_only:
            skipped_rows.append(
                {
                    "source_year": source_year,
                    "target_year": target_year,
                    "hunt_code": code,
                    "hunt_name": clean(row.get("hunt_name")),
                    "species": clean(row.get("species")),
                    "engine_family": family,
                    "prediction_final_classification": "PREDICTION_NOT_REQUIRED_REFERENCE_ONLY",
                    "audit_note": reason,
                }
            )
            continue

        if clean(row.get("residency")):
            if has_number(row, "p_draw", "p_draw_percent", "success_ratio"):
                key = (code, family, norm_residency(row.get("residency")), norm_points(row.get("points")))
                if key not in seen:
                    seen.add(key)
                    add_obligation(
                        obligations,
                        row,
                        source_year,
                        target_year,
                        source_path,
                        family,
                        clean(row.get("residency")),
                        "p_draw",
                        clean(row.get("p_draw") or row.get("p_draw_percent") or row.get("success_ratio")),
                    )
            else:
                skipped_rows.append(
                    {
                        "source_year": source_year,
                        "target_year": target_year,
                        "hunt_code": code,
                        "hunt_name": clean(row.get("hunt_name")),
                        "species": clean(row.get("species")),
                        "engine_family": family,
                        "prediction_final_classification": "PREDICTION_SOURCE_DATA_INCOMPLETE",
                        "audit_note": "Canonical point row has no numeric published probability; not used as accuracy obligation.",
                    }
                )
            continue

        for residency, prob_field, percent_field in (
            ("Resident", "resident_p_draw", "resident_p_draw_percent"),
            ("Nonresident", "nonresident_p_draw", "nonresident_p_draw_percent"),
        ):
            success_field = "resident_success_ratio" if residency == "Resident" else "nonresident_success_ratio"
            if not has_number(row, prob_field, percent_field, success_field):
                continue
            used_field = next(field for field in (prob_field, percent_field, success_field) if has_number(row, field))
            key = (code, family, norm_residency(residency), norm_points(row.get("points")))
            if key in seen:
                continue
            seen.add(key)
            add_obligation(
                obligations,
                row,
                source_year,
                target_year,
                source_path,
                family,
                residency,
                used_field,
                clean(row.get(used_field)),
            )

        if not any(
            has_number(row, field)
            for field in (
                "resident_p_draw",
                "resident_p_draw_percent",
                "resident_success_ratio",
                "nonresident_p_draw",
                "nonresident_p_draw_percent",
                "nonresident_success_ratio",
                "p_draw",
                "p_draw_percent",
                "success_ratio",
            )
        ):
            skipped_rows.append(
                {
                    "source_year": source_year,
                    "target_year": target_year,
                    "hunt_code": code,
                    "hunt_name": clean(row.get("hunt_name")),
                    "species": clean(row.get("species")),
                    "engine_family": family,
                    "prediction_final_classification": "PREDICTION_SOURCE_DATA_INCOMPLETE",
                    "audit_note": "Canonical point row has no numeric published probability; not used as accuracy obligation.",
                }
            )

    return obligations, skipped_rows, source_path


def prediction_files_for_year(audit_dirs_by_target_year: dict[int, Path], target_year: int) -> list[Path]:
    audit_dir = audit_dirs_by_target_year.get(target_year)
    if audit_dir is None:
        return []
    return sorted((audit_dir / "runs" / str(target_year) / "predictions").glob("*.csv"))


def select_audit_dirs_by_target_year(audit_dirs: Sequence[Path], target_years: Sequence[int]) -> dict[int, Path]:
    selected: dict[int, Path] = {}
    for target_year in target_years:
        for audit_dir in audit_dirs:
            prediction_dir = audit_dir / "runs" / str(target_year) / "predictions"
            if prediction_dir.exists() and any(prediction_dir.glob("*.csv")):
                selected[target_year] = audit_dir
                break
    return selected


def selected_rows_by_target_year(audit_dirs_by_target_year: dict[int, Path], file_name: str) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for target_year, audit_dir in sorted(audit_dirs_by_target_year.items()):
        for row in read_csv(audit_dir / file_name):
            if clean(row.get("target_year")) == str(target_year):
                rows.append(row)
    return rows


def audit_predictions(repo: Path, audit_dirs_by_target_year: dict[int, Path], target_years: Sequence[int]) -> dict[str, Any]:
    prediction_rows: list[dict[str, Any]] = []
    probability_rows: list[dict[str, Any]] = []
    leakage_rows: list[dict[str, Any]] = []
    duplicate_rows: list[dict[str, Any]] = []
    family_year_code_sets: dict[tuple[int, str], set[str]] = defaultdict(set)
    year_code_sets: dict[int, set[str]] = defaultdict(set)
    row_key_sets: dict[int, set[tuple[str, str, str, str]]] = defaultdict(set)
    prediction_file_summaries: list[dict[str, Any]] = []

    for target_year in target_years:
        for path in prediction_files_for_year(audit_dirs_by_target_year, target_year):
            rows = read_csv(path)
            key_counts: Counter[tuple[str, str, str, str, str, str, str]] = Counter()
            family_counts: Counter[str] = Counter()
            bad_probabilities = 0
            bad_leakage = 0
            zero_point_rows = 0
            for row_number, row in enumerate(rows, start=2):
                if not any(clean(value) for value in row.values()):
                    continue
                family = row_family(row, path)
                code = norm_code(row.get("hunt_code"))
                key = operational_key(row, family, path)
                key_counts[key] += 1
                if code:
                    family_year_code_sets[(target_year, family)].add(code)
                    year_code_sets[target_year].add(code)
                    row_key_sets[target_year].add(
                        (
                            code,
                            family,
                            norm_residency(row.get("residency") or row.get("metric_scope")),
                            norm_points(row.get("points") or row.get("point") or row.get("point_level")),
                        )
                    )
                family_counts[family] += 1
                if norm_points(row.get("points") or row.get("point") or row.get("point_level")) == "0":
                    zero_point_rows += 1

                field, value, prob_status, prob_issue = probability_check(row)
                prob_required = probability_required(row)
                if prob_status != "PASS" and (prob_issue != "MISSING_PROBABILITY" or prob_required):
                    bad_probabilities += 1
                    probability_rows.append(
                        {
                            "target_year": target_year,
                            "prediction_file": rel(repo, path),
                            "row_number": row_number,
                            "hunt_code": code,
                            "family": family,
                            "operational_key": "|".join(key),
                            "probability_field": field,
                            "probability_value": value,
                            "probability_status": prob_status,
                            "probability_issue": prob_issue,
                            "probability_required": str(prob_required).lower(),
                        }
                    )

                leakage_status, leakage_issue = prediction_leakage_check(row, target_year)
                if leakage_status != "PASS":
                    bad_leakage += 1
                leakage_rows.append(
                    {
                        "target_year": target_year,
                        "prediction_file": rel(repo, path),
                        "row_number": row_number,
                        "hunt_code": code,
                        "family": family,
                        "source_year": clean(row.get("source_year")),
                        "source_years_used": clean(row.get("source_years_used")),
                        "latest_source_year": clean(row.get("latest_source_year")),
                        "leakage_status": leakage_status,
                        "leakage_issue": leakage_issue,
                    }
                )

            for key, count in key_counts.items():
                if count > 1 and key[1]:
                    duplicate_rows.append(
                        {
                            "target_year": target_year,
                            "prediction_file": rel(repo, path),
                            "operational_key": "|".join(key),
                            "duplicate_count": count,
                        }
                    )
            prediction_file_summaries.append(
                {
                    "target_year": target_year,
                    "prediction_file": rel(repo, path),
                    "rows": len(rows),
                    "unique_operational_keys": len(key_counts),
                    "duplicate_key_groups": sum(1 for count in key_counts.values() if count > 1),
                    "unique_hunt_codes": len({norm_code(row.get("hunt_code")) for row in rows if norm_code(row.get("hunt_code"))}),
                    "family_counts": json.dumps(dict(sorted(family_counts.items()))),
                    "probability_fail_rows": bad_probabilities,
                    "leakage_fail_rows": bad_leakage,
                    "zero_point_rows": zero_point_rows,
                }
            )

    return {
        "probability_rows": probability_rows,
        "leakage_rows": leakage_rows,
        "duplicate_rows": duplicate_rows,
        "family_year_code_sets": family_year_code_sets,
        "year_code_sets": year_code_sets,
        "row_key_sets": row_key_sets,
        "prediction_file_summaries": prediction_file_summaries,
    }


def locked_coverage_rows(
    repo: Path,
    prediction_meta: dict[str, Any],
    target_years: Sequence[int],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    coverage_rows: list[dict[str, Any]] = []
    missing_rows: list[dict[str, Any]] = []
    holdout_rows: list[dict[str, Any]] = []
    family_rows: list[dict[str, Any]] = []

    family_year_code_sets: dict[tuple[int, str], set[str]] = prediction_meta["family_year_code_sets"]
    year_code_sets: dict[int, set[str]] = prediction_meta["year_code_sets"]
    row_key_sets: dict[int, set[tuple[str, str, str, str]]] = prediction_meta["row_key_sets"]

    for target_year in target_years:
        source_year = target_year - 1
        summary, _active_rows = load_locked_year(repo, source_year)
        obligations, skipped_source_rows, canonical_source = canonical_prediction_obligations(repo, source_year, target_year)
        scorable_by_family: Counter[str] = Counter()
        required_by_family: Counter[str] = Counter()
        predicted_by_family: Counter[str] = Counter()
        missing_by_family: Counter[str] = Counter()
        held_out_by_family: Counter[str] = Counter()

        for row in skipped_source_rows:
            coverage_rows.append(
                {
                    **row,
                    "residency": "",
                    "points": "",
                    "draw_pool": "",
                    "requires_prediction": "false",
                    "matched_by_prediction_key": "false",
                    "matched_by_hunt_code_target_year": "false",
                    "locked_scoring_bucket": "",
                    "locked_bucket": "",
                    "source_probability_field": "",
                    "source_probability_value": "",
                    "source_file": canonical_source,
                }
            )

        for row in obligations:
            code = norm_code(row.get("hunt_code"))
            family = clean(row.get("engine_family"))
            if not code:
                continue
            scorable_by_family[family] += 1
            is_holdout = source_year == 2026 and target_year == 2027 and family in HOLDOUT_FAMILIES
            if is_holdout:
                held_out_by_family[family] += 1
                holdout_rows.append(
                    {
                        "source_year": source_year,
                        "target_year": target_year,
                        "hunt_code": code,
                        "hunt_name": clean(row.get("hunt_name")),
                        "species": clean(row.get("species")),
                        "engine_family": family,
                        "holdout_reason": "2026 locked truth predicts 2027 antlerless/doe outcomes; official public actuals are unreleased.",
                    }
                )
                continue
            required_by_family[family] += 1
            prediction_key = (
                code,
                family,
                norm_residency(row.get("residency")),
                norm_points(row.get("points")),
            )
            matched_key = prediction_key in row_key_sets.get(target_year, set())
            matched_code = code in family_year_code_sets.get((target_year, family), set()) or code in year_code_sets.get(target_year, set())
            matched = matched_key or matched_code
            classification = "PREDICTION_PRESENT" if matched else "PREDICTION_MISSING_BLOCKER"
            if matched:
                predicted_by_family[family] += 1
            else:
                missing_by_family[family] += 1
                missing_rows.append(
                    {
                        "source_year": source_year,
                        "target_year": target_year,
                        "hunt_code": code,
                        "hunt_name": clean(row.get("hunt_name")),
                        "species": clean(row.get("species")),
                        "engine_family": family,
                        "residency": clean(row.get("residency")),
                        "points": clean(row.get("points")),
                        "draw_pool": clean(row.get("draw_pool")),
                        "locked_scoring_bucket": "CANONICAL_POINT_ROW_WITH_PUBLISHED_PROBABILITY",
                        "prediction_final_classification": classification,
                        "audit_note": "Canonical source-year point/residency probability row was not found in following-year progressive prediction artifacts.",
                    }
                )
            coverage_rows.append(
                {
                    "source_year": source_year,
                    "target_year": target_year,
                    "hunt_code": code,
                    "hunt_name": clean(row.get("hunt_name")),
                    "species": clean(row.get("species")),
                    "engine_family": family,
                    "residency": clean(row.get("residency")),
                    "points": clean(row.get("points")),
                    "draw_pool": clean(row.get("draw_pool")),
                    "requires_prediction": str(not is_holdout).lower(),
                    "prediction_final_classification": "PREDICTION_HELD_OUT_UNRELEASED_ACTUALS" if is_holdout else classification,
                    "matched_by_prediction_key": str(matched_key).lower(),
                    "matched_by_hunt_code_target_year": str(matched_code).lower(),
                    "locked_scoring_bucket": "CANONICAL_POINT_ROW_WITH_PUBLISHED_PROBABILITY",
                    "locked_bucket": "ACTIVE_YEAR_CANONICAL_TRUTH",
                    "source_probability_field": clean(row.get("source_probability_field")),
                    "source_probability_value": clean(row.get("source_probability_value")),
                    "source_file": clean(row.get("source_file")),
                }
            )

        for family in sorted(set(scorable_by_family) | set(family_year_code_sets_family_names(target_year, family_year_code_sets))):
            family_rows.append(
                {
                    "source_year": source_year,
                    "target_year": target_year,
                    "engine_family": family,
                    "locked_scorable_hunt_codes": scorable_by_family.get(family, 0),
                    "prediction_required_hunt_codes": required_by_family.get(family, 0),
                    "predicted_required_hunt_codes": predicted_by_family.get(family, 0),
                    "missing_required_hunt_codes": missing_by_family.get(family, 0),
                    "held_out_hunt_codes": held_out_by_family.get(family, 0),
                    "progressive_prediction_hunt_codes": len(family_year_code_sets.get((target_year, family), set())),
                    "official_source_year_active_hunt_codes": summary.get("official_active_hunt_code_count"),
                }
            )

    return coverage_rows, missing_rows, holdout_rows, family_rows


def family_year_code_sets_family_names(target_year: int, family_year_code_sets: dict[tuple[int, str], set[str]]) -> set[str]:
    return {family for year, family in family_year_code_sets if year == target_year}


def write_summary(
    repo: Path,
    output_dir: Path,
    progressive_dir: Path,
    target_years: Sequence[int],
    counts_rows: list[dict[str, str]],
    progressive_leakage_rows: list[dict[str, str]],
    coverage_rows: list[dict[str, Any]],
    missing_rows: list[dict[str, Any]],
    holdout_rows: list[dict[str, Any]],
    duplicate_rows: list[dict[str, Any]],
    probability_rows: list[dict[str, Any]],
    leakage_rows: list[dict[str, Any]],
) -> dict[str, Any]:
    count_status = Counter(row.get("status") for row in counts_rows)
    classified_rows = [row for row in counts_rows if row.get("status") == "CLASSIFIED"]
    failed_counts = [row for row in counts_rows if row.get("status") == "FAIL"]
    probability_failures = [row for row in probability_rows if row.get("probability_status") != "PASS"]
    leakage_failures = [row for row in leakage_rows if row.get("leakage_status") != "PASS"]
    progressive_leakage_failures = [row for row in progressive_leakage_rows if row.get("leakage_status") != "PASS"]
    required_rows = [row for row in coverage_rows if row.get("requires_prediction") == "true"]
    predicted_rows = [row for row in required_rows if row.get("prediction_final_classification") == "PREDICTION_PRESENT"]

    pass_condition = (
        not failed_counts
        and not missing_rows
        and not duplicate_rows
        and not probability_failures
        and not leakage_failures
        and not progressive_leakage_failures
        and all(row.get("blocker_if_failed") == "SOURCE_NOT_AVAILABLE_NO_PROVEN_YOUTH_TURKEY_HISTORY" for row in classified_rows)
    )
    classification = "ENGINE_CERTIFIED_PREDICTION_TRUTH_PASS" if pass_condition else "ENGINE_CERTIFIED_PREDICTION_TRUTH_BLOCKED"
    summary = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "classification": classification,
        "pass_condition": pass_condition,
        "repo": str(repo),
        "progressive_audit_dir": rel(repo, progressive_dir),
        "target_years": list(target_years),
        "family_year_rows": len(counts_rows),
        "family_year_status_counts": dict(sorted(count_status.items())),
        "classified_family_year_rows": len(classified_rows),
        "failed_family_year_rows": len(failed_counts),
        "prediction_required_locked_code_rows": len(required_rows),
        "prediction_present_locked_code_rows": len(predicted_rows),
        "missing_scorable_prediction_rows": len(missing_rows),
        "holdout_rows": len(holdout_rows),
        "duplicate_prediction_key_groups": len(duplicate_rows),
        "probability_failure_rows": len(probability_failures),
        "row_level_leakage_failures": len(leakage_failures),
        "progressive_leakage_failures": len(progressive_leakage_failures),
        "classified_exemptions": classified_rows,
        "outputs": {
            "summary_md": rel(repo, output_dir / "ENGINE_CERTIFICATION_SUMMARY.md"),
            "coverage_csv": rel(repo, output_dir / "YEAR_BY_YEAR_PREDICTION_COVERAGE.csv"),
            "missing_csv": rel(repo, output_dir / "MISSING_SCORABLE_PREDICTIONS.csv"),
            "duplicates_csv": rel(repo, output_dir / "DUPLICATE_PREDICTION_KEYS.csv"),
            "leakage_csv": rel(repo, output_dir / "NO_LEAKAGE_AUDIT.csv"),
            "probability_csv": rel(repo, output_dir / "PROBABILITY_BOUNDS_AUDIT.csv"),
            "holdouts_csv": rel(repo, output_dir / "HOLDOUT_ROWS.csv"),
            "family_coverage_csv": rel(repo, output_dir / "ENGINE_FAMILY_COVERAGE.csv"),
            "promotion_readiness_json": rel(repo, output_dir / "PROMOTION_READINESS.json"),
        },
    }
    (output_dir / "PROMOTION_READINESS.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
    lines = [
        "# Engine Certified Prediction Truth",
        "",
        f"Generated: `{summary['generated_at']}`",
        f"Classification: `{classification}`",
        "",
        "## Inputs",
        "",
        f"- Progressive audit: `{summary['progressive_audit_dir']}`",
        f"- Locked source truth years: `{min(target_years) - 1}-{max(target_years) - 1}`",
        "",
        "## Gate Results",
        "",
        f"- Family/year status counts: `{dict(sorted(count_status.items()))}`",
        f"- Classified family/year rows: `{len(classified_rows)}`",
        f"- Failed family/year rows: `{len(failed_counts)}`",
        f"- Prediction-required locked scorable code rows: `{len(required_rows)}`",
        f"- Prediction-present locked scorable code rows: `{len(predicted_rows)}`",
        f"- Missing scorable predictions: `{len(missing_rows)}`",
        f"- Held-out unreleased rows: `{len(holdout_rows)}`",
        f"- Duplicate prediction key groups: `{len(duplicate_rows)}`",
        f"- Probability failures: `{len(probability_failures)}`",
        f"- Row-level leakage failures: `{len(leakage_failures)}`",
        f"- Progressive leakage failures: `{len(progressive_leakage_failures)}`",
        "",
        "## Classification Rule",
        "",
        "- `ENGINE_CERTIFIED_PREDICTION_TRUTH_PASS` requires zero missing scorable predictions, zero duplicate prediction keys, zero probability failures, zero leakage failures, and no failed family/year runs.",
        "- The known 2018->2019 youth turkey row is allowed only as `SOURCE_NOT_AVAILABLE_NO_PROVEN_YOUTH_TURKEY_HISTORY` because no proven source history exists for that starting step.",
        "- 2027 antlerless/doe actuals are held out and not penalized.",
        "",
        "## Outputs",
        "",
        "- `YEAR_BY_YEAR_PREDICTION_COVERAGE.csv`",
        "- `MISSING_SCORABLE_PREDICTIONS.csv`",
        "- `DUPLICATE_PREDICTION_KEYS.csv`",
        "- `NO_LEAKAGE_AUDIT.csv`",
        "- `PROBABILITY_BOUNDS_AUDIT.csv`",
        "- `HOLDOUT_ROWS.csv`",
        "- `ENGINE_FAMILY_COVERAGE.csv`",
        "- `PROMOTION_READINESS.json`",
    ]
    if classified_rows:
        lines.extend(["", "## Classified Exemptions", ""])
        for row in classified_rows:
            lines.append(
                f"- `{row.get('source_year')}->{row.get('target_year')} {row.get('family')}`: `{row.get('blocker_if_failed')}`"
            )
    (output_dir / "ENGINE_CERTIFICATION_SUMMARY.md").write_text("\n".join(lines) + "\n", encoding="utf-8")
    return summary


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Certify progressive prediction artifacts against locked hunt-code truth.")
    parser.add_argument("--repo", default=str(REPO))
    parser.add_argument("--progressive-audit-dir", type=Path, action="append")
    parser.add_argument("--target-start", type=int, default=2019)
    parser.add_argument("--target-end", type=int, default=2027)
    parser.add_argument("--timestamp", default=datetime.now().strftime("%Y%m%d_%H%M%S"))
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    repo = Path(args.repo).resolve()
    progressive_dirs = args.progressive_audit_dir
    if not progressive_dirs:
        progressive_dirs = [latest_progressive_dir(repo)]
    resolved_progressive_dirs: list[Path] = []
    for progressive_dir in progressive_dirs:
        if not progressive_dir.is_absolute():
            progressive_dir = repo / progressive_dir
        resolved_progressive_dirs.append(progressive_dir.resolve())
    target_years = list(range(args.target_start, args.target_end + 1))
    output_dir = repo / OUT_ROOT / args.timestamp
    output_dir.mkdir(parents=True, exist_ok=True)

    audit_dirs_by_target_year = select_audit_dirs_by_target_year(resolved_progressive_dirs, target_years)
    missing_target_dirs = [year for year in target_years if year not in audit_dirs_by_target_year]
    if missing_target_dirs:
        raise FileNotFoundError(f"No prediction artifacts found for target years: {missing_target_dirs}")

    counts_rows = selected_rows_by_target_year(audit_dirs_by_target_year, "all_year_family_prediction_counts.csv")
    progressive_leakage_rows = selected_rows_by_target_year(audit_dirs_by_target_year, "leakage_check.csv")
    prediction_meta = audit_predictions(repo, audit_dirs_by_target_year, target_years)
    coverage_rows, missing_rows, holdout_rows, family_rows = locked_coverage_rows(repo, prediction_meta, target_years)

    probability_fail_rows = [row for row in prediction_meta["probability_rows"] if row.get("probability_status") != "PASS"]
    leakage_fail_rows = [row for row in prediction_meta["leakage_rows"] if row.get("leakage_status") != "PASS"]

    write_csv(
        output_dir / "YEAR_BY_YEAR_PREDICTION_COVERAGE.csv",
        coverage_rows,
        [
            "source_year",
            "target_year",
            "hunt_code",
            "hunt_name",
            "species",
            "engine_family",
            "residency",
            "points",
            "draw_pool",
            "requires_prediction",
            "prediction_final_classification",
            "matched_by_prediction_key",
            "matched_by_hunt_code_target_year",
            "locked_scoring_bucket",
            "locked_bucket",
            "source_probability_field",
            "source_probability_value",
            "source_file",
            "audit_note",
        ],
    )
    write_csv(
        output_dir / "MISSING_SCORABLE_PREDICTIONS.csv",
        missing_rows,
        [
            "source_year",
            "target_year",
            "hunt_code",
            "hunt_name",
            "species",
            "engine_family",
            "residency",
            "points",
            "draw_pool",
            "locked_scoring_bucket",
            "prediction_final_classification",
            "audit_note",
        ],
    )
    write_csv(
        output_dir / "DUPLICATE_PREDICTION_KEYS.csv",
        prediction_meta["duplicate_rows"],
        ["target_year", "prediction_file", "operational_key", "duplicate_count"],
    )
    write_csv(
        output_dir / "NO_LEAKAGE_AUDIT.csv",
        leakage_fail_rows,
        [
            "target_year",
            "prediction_file",
            "row_number",
            "hunt_code",
            "family",
            "source_year",
            "source_years_used",
            "latest_source_year",
            "leakage_status",
            "leakage_issue",
        ],
    )
    write_csv(
        output_dir / "PROBABILITY_BOUNDS_AUDIT.csv",
        probability_fail_rows,
        [
            "target_year",
            "prediction_file",
            "row_number",
            "hunt_code",
            "family",
            "operational_key",
            "probability_field",
            "probability_value",
            "probability_status",
            "probability_issue",
            "probability_required",
        ],
    )
    write_csv(
        output_dir / "HOLDOUT_ROWS.csv",
        holdout_rows,
        ["source_year", "target_year", "hunt_code", "hunt_name", "species", "engine_family", "holdout_reason"],
    )
    write_csv(
        output_dir / "ENGINE_FAMILY_COVERAGE.csv",
        family_rows,
        [
            "source_year",
            "target_year",
            "engine_family",
            "locked_scorable_hunt_codes",
            "prediction_required_hunt_codes",
            "predicted_required_hunt_codes",
            "missing_required_hunt_codes",
            "held_out_hunt_codes",
            "progressive_prediction_hunt_codes",
            "official_source_year_active_hunt_codes",
        ],
    )
    write_csv(
        output_dir / "PREDICTION_FILE_SUMMARY.csv",
        prediction_meta["prediction_file_summaries"],
        [
            "target_year",
            "prediction_file",
            "rows",
            "unique_operational_keys",
            "duplicate_key_groups",
            "unique_hunt_codes",
            "family_counts",
            "probability_fail_rows",
            "leakage_fail_rows",
            "zero_point_rows",
        ],
    )

    summary = write_summary(
        repo,
        output_dir,
        resolved_progressive_dirs[0],
        target_years,
        counts_rows,
        progressive_leakage_rows,
        coverage_rows,
        missing_rows,
        holdout_rows,
        prediction_meta["duplicate_rows"],
        prediction_meta["probability_rows"],
        prediction_meta["leakage_rows"],
    )
    print(f"AUDIT_DIR: {output_dir}")
    print(f"CLASSIFICATION: {summary['classification']}")
    print(f"PASS_CONDITION: {summary['pass_condition']}")
    print(f"MISSING_SCORABLE_PREDICTIONS: {summary['missing_scorable_prediction_rows']}")
    print(f"DUPLICATE_PREDICTION_KEY_GROUPS: {summary['duplicate_prediction_key_groups']}")
    print(f"PROBABILITY_FAILURE_ROWS: {summary['probability_failure_rows']}")
    print(f"LEAKAGE_FAILURE_ROWS: {summary['row_level_leakage_failures']}")
    return 0 if summary["pass_condition"] else 2


if __name__ == "__main__":
    raise SystemExit(main())
